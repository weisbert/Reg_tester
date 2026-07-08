#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_testcase.py — 无头序列生成器 + 渲染器（阶段二 M3 核心）。

读 flowgraph/1 + regmap/1 + modes/1 → 生成 testcase/1（唯一事实来源），
再渲染 ate.txt（交付格式）/ debug.html（designer 看）。

算法权威定义见仓库 SCHEMAS.md 第 5、6 节；GUI 的 webapp/generator.js 与本文件
**逐字节一致**（M4 用 node 交叉验证）。只依赖标准库；脚本本身不含任何真实信号名/地址，
全部从 JSON 读入 → 可安全进公开仓库。

用法：
    python gen_testcase.py --project projects/adpll_demo --mode BT_2G_RX
    python gen_testcase.py --flowgraph fg.json --regmap rm.json --mode-file m.json \\
        --out-json tc.json --out-ate ate.txt --out-html debug.html
    python gen_testcase.py --project projects/adpll_demo --mode BT_2G_RX --print
"""
import argparse
import html
import json
import os
import re
import sys

SCHEMA = "testcase/1"

# 排序兜底基座：距 DCO 源头的器件类深度（边缺失时用）。见 SCHEMAS.md Step D。
DEVICE_STAGE = {
    "dco": 0, "logic": 0, "div": 1, "mux": 2, "buf": 2, "inv": 2,
    "blackbox": 2, "route": 3, "group": 0,
}


# ------------------------------------------------------------------ helpers
def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def parse_bit(bit):
    """'11:10' -> (11,10); '4' -> (4,4). 返回 (hi, lo)."""
    bit = str(bit).strip()
    if ":" in bit:
        hi, lo = bit.split(":", 1)
        return int(hi), int(lo)
    n = int(bit)
    return n, n


def field_mask(bit):
    hi, lo = parse_bit(bit)
    width = hi - lo + 1
    return (((1 << width) - 1) << lo), lo, width


def set_field(word, bit, val):
    mask, lo, width = field_mask(bit)
    val = int(val) & ((1 << width) - 1)
    return (word & ~mask) | (val << lo)


def get_field(word, bit):
    mask, lo, _ = field_mask(bit)
    return (word & mask) >> lo


def hex4(v):
    return "0x%04X" % (v & 0xFFFF)


def hex_addr(a):
    """规范化地址成 0x + 大写8位。"""
    s = str(a).lower().replace("0x", "")
    return "0x%08X" % int(s, 16)


def to_int(v, default=0):
    if v is None:
        return default
    if isinstance(v, int):
        return v
    s = str(v).strip()
    try:
        return int(s, 16) if s.lower().startswith("0x") else int(s, 0)
    except ValueError:
        return default


# ------------------------------------------------------------------ core model
class RegView:
    """把 regmap/1 的 signals 按 reg_group 解析成 variant 视图。"""

    def __init__(self, regmap, group):
        self.regmap = regmap
        self.group = group
        self.common = regmap.get("common_group", "COMMON")
        self.by_id = {s["id"]: s for s in regmap.get("signals", [])}

    def variant(self, sig_id):
        s = self.by_id.get(sig_id)
        if not s:
            return None
        variants = s.get("variants", {})
        return variants.get(self.group) or variants.get(self.common)

    def signal(self, sig_id):
        return self.by_id.get(sig_id)


def collect_gates(flowgraph, gate_override=None):
    """建 gate_nodes[signal]=set(node_id)、node_gates[node]=[gate...]。gate 带 node 归属。
    gate_override[node_id]=[signal...]：**用户指定该 block 的关断总闸**——只用这些信号做门，
    其余 off_control 跳过（工具往往在一个 cell 上找到多个使能脚，真·总闸由人指定，存 layout）。"""
    gate_override = gate_override or {}
    gate_nodes = {}
    node_gates = {}
    for n in flowgraph.get("nodes", []):
        nid = n["id"]
        ov = gate_override.get(nid)   # 该节点若指定了总闸，只认这些信号
        gs = []
        for oc in n.get("off_controls", []):
            sig = oc.get("signal_ref")
            if not sig:
                continue
            if ov is not None and sig not in ov:
                continue
            g = {
                "node": nid,
                "signal": sig,
                "pin": oc.get("pin"),
                "off_value": oc.get("off_value", 0),
                "active_high": oc.get("active_high"),
                "polarity_inferred": oc.get("polarity_inferred", False),
                "lane": oc.get("lane"),
            }
            gs.append(g)
            gate_nodes.setdefault(sig, set()).add(nid)
        if gs:
            node_gates[nid] = gs
    return gate_nodes, node_gates


def on_value(gate):
    """enable 门的'开'值：高有效→1，低有效→0；缺失按高有效兜底。"""
    ah = gate.get("active_high")
    if ah is None:
        return 1
    return 1 if ah else 0


def build_edge_maps(flowgraph, visible):
    """建 可见层 前驱表 preds[node]=set(前驱node)。跨 composite 折叠到可见祖先。"""
    node_by_id = {n["id"]: n for n in flowgraph.get("nodes", [])}

    def visible_anc(nid):
        # 折叠到最近的可见节点（inferred 子节点本身可见；隐藏/折叠情况由 visible 集合决定）
        cur = nid
        seen = set()
        while cur is not None and cur not in seen:
            seen.add(cur)
            if cur in visible:
                return cur
            n = node_by_id.get(cur)
            cur = n.get("parent") if n else None
        return nid if nid in visible else None

    preds = {}
    for e in flowgraph.get("edges", []):
        frm = visible_anc(e["from"]["node"])
        for t in e.get("to", []):
            to = visible_anc(t["node"])
            if not frm or not to or frm == to:
                continue
            preds.setdefault(to, set()).add(frm)
    return preds, node_by_id


def shutdown_rank(node_id, node_by_id, preds, memo, stack=None):
    """距 DCO 源头深度（越大=越靠末端）。边缺失时用器件类基座。带环保护。"""
    if node_id in memo:
        return memo[node_id]
    if stack is None:
        stack = set()
    if node_id in stack:  # 环：退回基座
        n = node_by_id.get(node_id, {})
        return DEVICE_STAGE.get(n.get("device"), 1)
    stack.add(node_id)
    n = node_by_id.get(node_id, {})
    base = DEVICE_STAGE.get(n.get("device"), 1)
    best = base
    for p in preds.get(node_id, ()):  # 前驱越深，本节点越深
        best = max(best, 1 + shutdown_rank(p, node_by_id, preds, memo, stack))
    stack.discard(node_id)
    memo[node_id] = best
    return best


# ------------------------------------------------------------------ generator
def generate(flowgraph, regmap, mode, gate_override=None):
    group = mode.get("reg_group", regmap.get("primary_group", "BT"))
    rv = RegView(regmap, group)
    enabled = set(mode.get("enabled_nodes", []))
    baseline_over = mode.get("baseline", {}) or {}

    gate_nodes, node_gates = collect_gates(flowgraph, gate_override)

    # Step B: 每个门信号开/关（共用位：开压倒关）
    def signal_on(sig):
        return any(n in enabled for n in gate_nodes.get(sig, ()))

    warnings = []

    # Step C: 基线寄存器映像 -------------------------------------------------
    # touched addrs = 所有门信号 ∪ baseline 信号 ∪ enabled 节点上任何 control 信号
    touched_sigs = set(gate_nodes.keys())
    touched_sigs |= set(baseline_over.keys())
    for n in flowgraph.get("nodes", []):
        if n["id"] in enabled:
            for c in n.get("controls", []):
                if c.get("signal_ref"):
                    touched_sigs.add(c["signal_ref"])

    images = {}   # addr -> {value, reg, reset}
    fields_set = {}  # addr -> list of field dicts explicitly set by this mode

    def ensure_img(sig):
        v = rv.variant(sig)
        if not v or not v.get("addr"):
            return None
        addr = hex_addr(v["addr"])
        if addr not in images:
            images[addr] = {
                "value": to_int(v.get("reset"), 0),
                "reg": v.get("reg_name"),
                "reset": to_int(v.get("reset"), 0),
            }
            fields_set[addr] = []
        return addr, v

    unresolved = []
    # 0) 预置：所有 touched 信号（含激活节点的 tune/tail 控制）建立 reset 映像，
    #    保证基线写覆盖激活通路相关的全部寄存器（如 DCO 尾电流寄存器）。
    #    不解析（无寄存器的频率/模式信号如 ct/mt/ft）静默跳过——它们不是电流门。
    for sig in sorted(touched_sigs):
        ensure_img(sig)
    # 3) 门信号写 on/off
    for sig in sorted(gate_nodes.keys()):
        # 取该 signal 任一门（同信号 off_value 一致）
        any_gate = None
        for nid in gate_nodes[sig]:
            for g in node_gates.get(nid, []):
                if g["signal"] == sig:
                    any_gate = g
                    break
            if any_gate:
                break
        r = ensure_img(sig)
        if not r:
            unresolved.append(sig)
            continue
        addr, v = r
        val = on_value(any_gate) if signal_on(sig) else any_gate["off_value"]
        images[addr]["value"] = set_field(images[addr]["value"], v["bit"], val)
        fields_set[addr].append({
            "signal": sig, "bit": v["bit"], "value": val,
            "role": "enable", "on": bool(signal_on(sig)),
        })

    # 3.5) mux_sel：把 GUI 的 MUX 选择写进对应 sel 字段（约定 0=上/1=下；真实寄存器编码可能不同，见 SCHEMAS）。
    #      放在显式 baseline 之前，故 baseline 手改值仍可压倒它。
    mux_sel = mode.get("mux_sel", {}) or {}
    node_by_id_all = {n["id"]: n for n in flowgraph.get("nodes", [])}
    for nid in sorted(mux_sel.keys()):
        node = node_by_id_all.get(nid)
        if not node:
            continue
        sel_ctrls = [c for c in node.get("controls", [])
                     if c.get("role") == "sel" or (c.get("pin") and "sel" in c["pin"].lower())]
        if len(sel_ctrls) != 1:
            if sel_ctrls:
                warnings.append("MUX %s 有多个 sel 控制，mux_sel 未落值（请在 baseline 里明确）" % nid)
            continue
        sig = sel_ctrls[0].get("signal_ref")
        r = ensure_img(sig)
        if not r:
            continue
        addr, v = r
        val = to_int(mux_sel[nid], 0)
        images[addr]["value"] = set_field(images[addr]["value"], v["bit"], val)
        fields_set[addr].append({"signal": sig, "bit": v["bit"], "value": val, "role": "mux_sel"})

    # 4) baseline 显式覆盖（tune/ictrl/ct/mt 手改值，压倒门默认与 mux_sel）
    for sig, val in baseline_over.items():
        r = ensure_img(sig)
        if not r:
            unresolved.append(sig)
            continue
        addr, v = r
        val = to_int(val, 0)
        images[addr]["value"] = set_field(images[addr]["value"], v["bit"], val)
        fields_set[addr].append({
            "signal": sig, "bit": v["bit"], "value": val, "role": "override",
        })

    if unresolved:
        warnings.append("基线里有未解析到寄存器的信号（跳过）：%s" % ", ".join(sorted(set(unresolved))))

    # 5) baseline.writes
    baseline_writes = []
    for addr in sorted(images.keys()):
        im = images[addr]
        baseline_writes.append({
            "addr": addr, "reg": im["reg"], "value": hex4(im["value"]),
            "reset": hex4(im["reset"]),
            "fields": fields_set[addr],
        })

    # Step D: 关闭步骤 -------------------------------------------------------
    # active_gate_nodes = enabled 里有至少一个基线为开门的节点
    active_nodes = []
    for nid in enabled:
        gs = [g for g in node_gates.get(nid, []) if signal_on(g["signal"])]
        if gs:
            active_nodes.append(nid)

    # 排序
    visible = set(n["id"] for n in flowgraph.get("nodes", []))  # 全节点可见（生成层不折叠）
    preds, node_by_id = build_edge_maps(flowgraph, visible)
    order_mode = (mode.get("order", {}) or {}).get("mode", "auto")
    # 次级排序键：enabled_nodes 里的位置（设计者按 源→末端 录入）→ 越靠后越靠末端 → 越先关。
    # 这解决"同 rank 同器件类相邻级"用 id 字典序会把上下游关反的问题（composite 内缺 leaf→leaf 边所致）。
    enabled_list = mode.get("enabled_nodes", []) or []
    eidx = {}
    for i, n in enumerate(enabled_list):
        eidx[n] = i
    memo = {}

    def sort_key(n):
        return (-shutdown_rank(n, node_by_id, preds, memo), -eidx.get(n, -1), n)

    if order_mode == "manual":
        manual = list((mode.get("order", {}) or {}).get("manual", []))
        ordered = [n for n in manual if n in active_nodes]
        rest = [n for n in active_nodes if n not in ordered]
        if rest:
            warnings.append("manual 顺序未覆盖的激活节点按 auto 追加：%s" % ", ".join(sorted(rest)))
        rest.sort(key=sort_key)
        ordered += rest
    else:
        ordered = sorted(active_nodes, key=sort_key)

    # 顺序不确定性提示：相邻两级 rank 相同（拓扑没定序，靠 enabled_nodes 录入序兜底）时提醒人工确认。
    ambiguous = []
    for a, b in zip(ordered, ordered[1:]):
        if shutdown_rank(a, node_by_id, preds, memo) == shutdown_rank(b, node_by_id, preds, memo):
            ambiguous.append([a, b])
    if ambiguous:
        warnings.append("以下相邻级的先后由拓扑无法判定，按 enabled_nodes 录入序（源→末端）兜底，请人工确认：%s"
                        % "; ".join("%s→%s" % (a, b) for a, b in ambiguous))

    # 逐节点累积关闭
    reg_image = {addr: images[addr]["value"] for addr in images}
    signals_off = set()
    steps = []
    shared_collateral = []
    for idx, nid in enumerate(ordered, 1):
        node = node_by_id.get(nid, {})
        gs = [g for g in node_gates.get(nid, []) if signal_on(g["signal"])]
        writes_by_addr = {}
        step_gates = []
        step_warn = []
        for g in gs:
            sig = g["signal"]
            v = rv.variant(sig)
            if not v or not v.get("addr"):
                continue
            addr = hex_addr(v["addr"])
            # 该信号还驱动哪些别的激活节点（共用位受害者）
            others = sorted(x for x in gate_nodes.get(sig, ()) if x != nid and x in enabled)
            gate_rec = {
                "signal": sig, "pin": g.get("pin"), "off_value": g["off_value"],
                "shared": len(gate_nodes.get(sig, ())) > 1,
                "polarity_inferred": g.get("polarity_inferred", False),
                "collateral_nodes": others,
            }
            step_gates.append(gate_rec)
            if sig in signals_off:
                shared_collateral.append({"step": idx, "node": nid, "signal": sig})
                continue
            before = get_field(reg_image[addr], v["bit"])
            new = set_field(reg_image[addr], v["bit"], g["off_value"])
            if new != reg_image[addr]:
                rec = writes_by_addr.setdefault(addr, {
                    "addr": addr, "reg": v.get("reg_name"),
                    "prev": hex4(reg_image[addr]), "fields": [],
                })
                rec["fields"].append({
                    "signal": sig, "bit": v["bit"],
                    "before": before, "after": g["off_value"], "role": "enable",
                })
                reg_image[addr] = new
                rec["value"] = hex4(new)
            signals_off.add(sig)
            if others:
                step_warn.append(
                    "共用位 %s 关闭同时波及：%s（这些块的电流一并消失）"
                    % (sig, ", ".join(others)))
        writes = [writes_by_addr[a] for a in sorted(writes_by_addr.keys())]
        note = None
        if not writes:
            note = "本级门已被前面共用位提前关掉（仍是一个测量点）"
        steps.append({
            "index": idx,
            "off_node": nid,
            "off_label": node.get("inst_name") or node.get("name") or nid.split("::")[-1],
            "device": node.get("device"),
            "measure": "关此级后测总电流",
            "gates": step_gates,
            "writes": writes,
            "warnings": step_warn,
            "note": note,
        })

    # Step E: 诊断透传
    uncovered = []
    for u in flowgraph.get("diagnostics", {}).get("uncovered_off_gates", []):
        uncovered.append({"signal": u.get("signal"), "note": "真门但无可挂节点，序列不会自动关，需人工补"})

    return {
        "schema_version": SCHEMA,
        "mode": mode.get("id"),
        "mode_name": mode.get("name"),
        "reg_group": group,
        "base_addr": regmap.get("base_addr"),
        "order_mode": order_mode,
        "baseline": {"note": "建立全开起始态（激活通路开、其余门关、tune/ictrl 取基线值）",
                     "writes": baseline_writes},
        "steps": steps,
        "extra_writes": mode.get("extra_writes", []) or [],
        "warnings": warnings,
        "diagnostics": {"uncovered_off_gates": uncovered, "shared_collateral": shared_collateral},
        "stats": {"baseline_regs": len(baseline_writes), "steps": len(steps),
                  "gates_off": len(signals_off)},
    }


# ------------------------------------------------------------------ renderers
def render_ate(tc):
    L = []
    ap = L.append
    ap("# " + "=" * 66)
    ap("# Test sequence: %s  (reg_group=%s)" % (tc.get("mode"), tc.get("reg_group")))
    if tc.get("mode_name"):
        ap("# %s" % tc["mode_name"])
    ap("# Generated by Reg_tester gen_testcase (%s)" % tc["schema_version"])
    ap("# 语义：累积逐级关闭；每一步先发本段写、再测总电流；相邻步电流差=该级功耗。")
    ap("# 数据行格式：ADDR VALUE MODULE  [; 行内注释]  ——ADDR=0x+大写8位、VALUE=0x+大写4位；")
    ap("#            MODULE 后可有以 ' ; ' 起的行内注释；以 # 起头的整行为纯注释。")
    ap("# " + "=" * 66)
    ap("")
    ap("# --- baseline：建立全开起始态 ---")
    for w in tc["baseline"]["writes"]:
        seg = []
        for f in w.get("fields", []):
            if f.get("role") in ("override", "mux_sel"):
                seg.append("%s=%s" % (f["signal"], f["value"]))
            elif f.get("on"):
                seg.append("%s=on" % f["signal"])
        cmt = ("  ; " + ", ".join(seg)) if seg else ""
        ap("%s %s  baseline:%s%s" % (w["addr"], w["value"], w["reg"] or "?", cmt))
    ap("")
    for st in tc["steps"]:
        ap("# --- step %d: OFF %s (%s) → 测总电流 ---"
           % (st["index"], st["off_label"], st.get("device") or "?"))
        for wn in st.get("warnings", []):
            ap("#   ⚠ %s" % wn)
        if st.get("note"):
            ap("#   · %s" % st["note"])
        for w in st["writes"]:
            ftxt = ", ".join("%s[%s]:%d→%d" % (f["signal"], f["bit"], f["before"], f["after"])
                             for f in w.get("fields", []))
            ap("%s %s  off:%s  ; %s→%s  %s"
               % (w["addr"], w["value"], st["off_node"], w["prev"], w["value"], ftxt))
        ap("")
    if tc.get("extra_writes"):
        ap("# --- extra writes（模式级额外写）---")
        for w in tc["extra_writes"]:
            ap("%s %s  extra  ; %s" % (hex_addr(w["addr"]), hex4(to_int(w["value"])), w.get("note", "")))
        ap("")
    if tc["diagnostics"]["uncovered_off_gates"]:
        ap("# --- ⚠ 未覆盖的门（需人工补写）---")
        for u in tc["diagnostics"]["uncovered_off_gates"]:
            ap("#   %s : %s" % (u["signal"], u["note"]))
    return "\n".join(L) + "\n"


def build_reference(flowgraph, regmap, gate_override=None, group=None, logic_expr=None):
    """Reference 表：每个 block ->
      - **EN 映射(顶层布尔式)**：底层 EN 脚 = 哪些顶层信号怎么组合（如 `dco_en & buf_en`），
        由门级网表方程(logic_expr) + regmap.drives 反查(raw net→顶层信号) 拼出；
      - 布尔式里每个顶层信号 -> 寄存器 addr/bit（**多信号=多行**，block 列跨行合并）。
    ★=gate_override 指定的关断总闸。没门级方程的 block 退化为直接列 off_controls。"""
    gate_override = gate_override or {}
    group = group or regmap.get("primary_group", "BT")
    sig_by_id = {s["id"]: s for s in regmap.get("signals", [])}
    node_by_id = {n["id"]: n for n in flowgraph.get("nodes", [])}
    eqs = (logic_expr or {}).get("equations", {}) or {}
    powered_by = {}   # block_id -> 供电的 power switch 节点（来自电源域边，按 int_vdd 轨连接，模块内唯一）
    for e in flowgraph.get("edges", []):
        if e.get("kind") == "power":
            for t in e.get("to", []):
                powered_by[t["node"]] = e["from"]["node"]

    def picked_oc(node):
        """选定用于 EN 脚/追踪的 off_control：优先用户指定的关断总闸(gate_override)，否则第一个。"""
        ocs = node.get("off_controls") or [{}]
        dz = gate_override.get(node["id"], [])
        for o in ocs:
            if o.get("signal_ref") in dz:
                return o
        return ocs[0]

    def real_en_pin(node):
        """EN脚显示 block **符号上**的真实使能脚；off_control 若经 glue 门(via_glue)，
        A1/A2 是门的脚不是符号脚——回溯到被该门输出驱动的 block 引脚。"""
        oc = picked_oc(node)
        glue = node_by_id.get(oc.get("via_glue")) if oc.get("via_glue") else None
        if glue:
            outs = set(p.get("net") for p in glue.get("pins", []) if p.get("dir") == "output")
            hit = [p["name"] for p in node.get("pins", []) if p.get("net") in outs]
            if hit:
                return hit[0] + "（经 glue %s）" % (oc.get("via_glue", "").split("/")[-1])
        return oc.get("pin") or ""

    def reg_of(sid):
        s = sig_by_id.get(sid)
        if not s:
            return ("(不在regmap)", "", "")
        v = s.get("variants", {})
        vv = v.get(group) or v.get("COMMON")
        if not vv:
            return ("(无%s/COMMON)" % group, "", "")
        return (vv.get("addr"), vv.get("bit"), vv.get("reg_name") or "")

    raw2top = {}   # (chain_mod, raw_net) -> 顶层信号（drives 反查）
    for s in regmap.get("signals", []):
        for dv in s.get("drives", []) or []:
            if "." in dv:
                mod, net = dv.split(".", 1)
                raw2top[(mod, net)] = s["id"]
    master_sig = {}   # 每模块主DCO使能 = 该模块 DCO 核(device=dco) 的 off_control 信号（数据驱动，不硬编码网名）
    for n in flowgraph.get("nodes", []):
        if n.get("device") == "dco":
            for o in n.get("off_controls", []):
                if o.get("signal_ref"):
                    master_sig.setdefault(n.get("module"), set()).add(o["signal_ref"])

    # ---- 完整 EN 布尔追踪：底层 EN 脚 → 穿过 BUF_TOP glue 门(ND2/NR2/INV/…) + DCO_Logic 方程 → 顶层信号 ----
    def cell_op(t):
        t = t or ""
        if re.match(r'^INV', t): return 'inv'
        if re.match(r'^(BUFF?|BUFT|DEL|CK)', t): return 'buf'
        if re.match(r'^ND\d', t): return 'nand'
        if re.match(r'^NR\d', t): return 'nor'
        if re.match(r'^AN\d', t): return 'and'
        if re.match(r'^OR\d', t): return 'or'
        return None

    net2gate = {}   # BUF_TOP 级简单门：输出网 -> (op, [输入网])
    for nn in flowgraph.get("nodes", []):
        op = cell_op(nn.get("inst_type"))
        if not op:
            continue
        outs = [p.get("net") for p in nn.get("pins", []) if p.get("dir") == "output"]
        ins = [p.get("net") for p in nn.get("pins", []) if p.get("dir") == "input" and p.get("role") != "power"]
        if len(outs) == 1 and outs[0]:
            net2gate[outs[0]] = (op, ins)

    def sub_raw(e, mod):
        return re.sub(r'[A-Za-z_][A-Za-z0-9_]*', lambda mo: raw2top.get((mod, mo.group(0)), mo.group(0)), e)

    def expr_net(net, mod, depth=0, seen=None):
        seen = seen or set()
        if depth > 18 or net in seen:
            return net
        if net in eqs.get(mod, {}):
            return sub_raw(eqs[mod][net], mod)          # DCO_Logic 叶子
        g = net2gate.get(net)
        if g:
            op, ins = g
            sub = [expr_net(i, mod, depth + 1, seen | {net}) for i in ins]
            if op == 'nand': return '!(%s)' % ' & '.join(sub)
            if op == 'nor':  return '!(%s)' % ' | '.join(sub)
            if op == 'and':  return '(%s)' % ' & '.join(sub)
            if op == 'or':   return '(%s)' % ' | '.join(sub)
            if op == 'inv':  return '!%s' % sub[0]
            if op == 'buf':  return sub[0]
        return raw2top.get((mod, net), net)             # 叶子：raw→顶层，或原样

    def _bal(s):
        dep = 0
        for ch in s:
            if ch == '(': dep += 1
            elif ch == ')':
                dep -= 1
                if dep < 0: return False
        return dep == 0

    def simplify(e):
        e = e.strip()
        for _ in range(20):
            e2 = re.sub(r'!\(!([A-Za-z0-9_]+)\)', r'\1', e)
            e2 = re.sub(r'\(([A-Za-z0-9_]+)\s*&\s*\1\)', r'\1', e2)
            if e2.startswith('!(!(') and e2.endswith('))') and _bal(e2[4:-2]):
                e2 = e2[4:-2]
            if e2.startswith('(') and e2.endswith(')') and _bal(e2[1:-1]):
                e2 = e2[1:-1]
            if e2 == e:
                break
            e = e2
        return e

    def enable_net(node):
        oc = picked_oc(node)
        glue = node_by_id.get(oc.get("via_glue")) if oc.get("via_glue") else None
        if glue:
            outs = set(p.get("net") for p in glue.get("pins", []) if p.get("dir") == "output")
            for p in node.get("pins", []):
                if p.get("net") in outs:
                    return p.get("net"), p.get("name")
        p = [pp for pp in node.get("pins", []) if pp.get("name") == oc.get("pin")]
        return (p[0].get("net"), p[0].get("name")) if p else (None, oc.get("pin"))

    def en_expr_of(node):
        mod = node.get("module")
        if mod not in eqs:
            return None
        net, pin = enable_net(node)
        if not net:
            return None
        raw = expr_net(net, mod)
        inv = bool(re.search(r'enb$|_b$|nb$', pin or ""))   # enb/复补脚 → ON = !(...)
        return simplify(('!(%s)' % raw) if inv else raw)

    def parse_bit(bs):
        bs = str(bs)
        if ":" in bs:
            hi, lo = bs.split(":", 1)
            return int(lo), int(hi) - int(lo) + 1     # (lo, width)
        return int(bs), 1

    def add_row(blk, seen, sid, role, off, master=False, pwr=False):
        if not sid or sid in seen:
            return
        seen.add(sid)
        a, b, r = reg_of(sid)
        blk["rows"].append({"sig": sid, "addr": a, "bit": b, "reg": r, "role": role,
                            "off": off, "master": master, "pwr": pwr})

    TOP = re.compile(r'd_[A-Za-z0-9_]+')
    GATEABLE = ("buf", "div", "mux", "dco", "power_switch")
    blocks = []
    for n in flowgraph.get("nodes", []):
        if n.get("device") not in GATEABLE or not n.get("off_controls"):
            continue
        mod = n.get("module")
        desig = set(gate_override.get(n["id"], []))
        expr = en_expr_of(n)
        parts = n["id"].split("/")
        blk = {"module": mod or "", "cell": n.get("inst_type") or "?",
               "path": "/".join(parts[1:]) if len(parts) > 1 else n["id"],   # 模块内唯一路径（真正的标识，非 cell 名）
               "inst": n.get("inst_name") or parts[-1],
               "dev": n["device"], "en_pin": real_en_pin(n), "expr": "", "rows": []}
        seen = set()
        # 1) EN映射布尔式里的顶层信号（追踪的那条使能路径）
        if expr:
            blk["expr"] = expr
            for sid in TOP.findall(expr):
                is_master = sid in master_sig.get(mod, set())
                role = ("主DCO使能(整域)" if is_master else "本级使能") + ("★总闸" if sid in desig else "")
                add_row(blk, seen, sid, role, 0, master=is_master)
        # 2) **所有** off_controls 的信号都列出（2.2：多门块不能只显第一个；也兜底 2.1 空表达式）
        for o in n["off_controls"]:
            sid = o.get("signal_ref")
            add_row(blk, seen, sid, "本级使能" + ("★总闸" if sid in desig else ""), o.get("off_value", 0))
        if not blk["expr"]:
            sigs = [o.get("signal_ref") for o in n["off_controls"] if o.get("signal_ref")]
            show = [s for s in sigs if s in desig] or sigs
            blk["expr"] = " & ".join(show) + "  (未过逻辑门 / 直接控制)"
        # 3) 电源域总闸：若本 block 由 power switch 供电（关它=整域掉电）
        psw = node_by_id.get(powered_by.get(n["id"]))
        if psw and psw["id"] != n["id"]:
            for o in psw.get("off_controls", []):
                add_row(blk, seen, o.get("signal_ref"),
                        "电源域总闸(power switch %s)" % (psw.get("inst_name") or ""),
                        o.get("off_value", 0), pwr=True)
        blocks.append(blk)
    # 位定义（per-signal，含 lo/width）+ per-bit 展开（供位编辑器逐 bit 标注）
    bit_defs, bit_lookup = [], []
    for s in regmap.get("signals", []):
        v = s.get("variants", {})
        vv = v.get(group) or v.get("COMMON")
        if not (vv and vv.get("addr") and str(vv.get("bit", "")) != ""):
            continue
        lo, width = parse_bit(vv["bit"])
        addr_key = re.sub(r'(?i)^0x', '', str(vv["addr"])).upper()
        bit_defs.append({"addr": vv["addr"], "addr_key": addr_key, "bit_str": str(vv["bit"]),
                         "lo": lo, "width": width, "sig": s["id"], "reg": vv.get("reg_name") or "",
                         "off": s.get("off_value", 0)})
        for bb in range(lo, lo + width):
            bit_lookup.append({"key": "%s_%d" % (addr_key, bb), "sig": s["id"]})
    return {"group": group, "blocks": blocks, "bit_defs": bit_defs, "bit_lookup": bit_lookup}


def render_xlsx(testcases, path, reference=None):
    """testcases = [(mode_id, tc), ...] -> 一个 .xlsx，**每个模式一张 sheet(tab)**；
    reference（可选）= build_reference 输出，会加一张 Reference sheet 放最前。
    需 openpyxl（延迟导入，只在导出 Excel 时用；核心生成仍 stdlib）。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="D9D9D9")
    base_fill = PatternFill("solid", fgColor="E2EFDA")   # baseline 段浅绿
    step_fill = PatternFill("solid", fgColor="FCE4D6")   # 每步边界浅橙（=测量点）

    if reference:
        pwr_fill = PatternFill("solid", fgColor="DDEBF7")   # 电源域总闸行浅蓝
        ws = wb.create_sheet("Reference")
        ws.cell(1, 1, "Reference — 底层 EN 脚 → 顶层信号布尔式(含 glue 门) → 寄存器  (reg_group=%s；唯一标识=模块+实例路径，非 CELL 类型；★=指定总闸；绿=主DCO使能；蓝=电源域总闸)" % reference["group"]).font = bold
        cols2 = ["模块", "实例路径(唯一标识)", "CELL类型(可重复)", "device", "EN脚", "EN映射(= 顶层布尔式)", "顶层信号", "地址 ADDR", "bit", "关断值", "角色/说明", "寄存器"]
        for c, h in enumerate(cols2, 1):
            cell = ws.cell(3, c, h); cell.font = bold; cell.fill = hdr_fill
        r = 4
        for blk in reference["blocks"]:
            r0 = r
            for row in blk["rows"]:
                ws.cell(r, 7, row["sig"]); ws.cell(r, 8, row["addr"]); ws.cell(r, 9, row["bit"])
                ws.cell(r, 10, row["off"]); ws.cell(r, 11, row["role"]); ws.cell(r, 12, row["reg"])
                fill = pwr_fill if row.get("pwr") else (base_fill if row.get("master") else None)
                if fill:
                    for c in range(7, 13):
                        ws.cell(r, c).fill = fill
                r += 1
            if r == r0:          # 无信号行也占一行，绝不被下一块覆盖(2.1)
                r = r0 + 1
            ws.cell(r0, 1, blk["module"]); ws.cell(r0, 2, blk["path"]); ws.cell(r0, 3, blk["cell"])
            ws.cell(r0, 4, blk["dev"]); ws.cell(r0, 5, blk["en_pin"])
            ws.cell(r0, 6, ("EN = " + blk["expr"]) if blk["expr"] else "")
            if r - 1 > r0:
                for c in range(1, 7):
                    ws.merge_cells(start_row=r0, start_column=c, end_row=r - 1, end_column=c)
        for c, wd in enumerate([8, 26, 22, 8, 13, 48, 30, 13, 5, 7, 22, 14], 1):
            ws.column_dimensions[get_column_letter(c)].width = wd
        ws.freeze_panes = "A4"

        # ---- 位编辑器（Excel 原生公式，无宏）：只改要动的 bit，其余自动保持 ----
        bd = reference.get("bit_defs", [])
        blk_lookup = reference.get("bit_lookup", [])   # per-bit 展开：多位字段的每个 bit 都能查到信号名
        wsd = wb.create_sheet("位定义")     # VLOOKUP 用，隐藏
        wsd.cell(1, 1, "key").font = bold; wsd.cell(1, 2, "信号").font = bold
        for i, x in enumerate(blk_lookup, 2):
            wsd.cell(i, 1, x["key"]); wsd.cell(i, 2, x["sig"])
        wsd.sheet_state = "hidden"

        we = wb.create_sheet("位编辑器")
        in_fill = PatternFill("solid", fgColor="FFF2CC")   # 输入格黄
        out_fill = PatternFill("solid", fgColor="C6E0B4")  # 输出格绿
        we.cell(1, 1, "位编辑器 — 只在「目标」行填你要改的 bit（0/1），其余留空自动保持；新值自动算出").font = bold
        we.cell(3, 1, "当前值(hex) →").font = bold
        c = we.cell(3, 2, "3F4B"); c.fill = in_fill; c.font = bold; c.number_format = "@"   # 文本格式：防 3E48 被当科学计数
        we.cell(4, 1, "地址(可选,标信号) →")
        c = we.cell(4, 2, ""); c.fill = in_fill; c.number_format = "@"
        we.cell(5, 1, "新值(hex) →").font = bold
        cur = 'HEX2DEC(SUBSTITUTE(SUBSTITUTE($B$3,"0x",""),"0X",""))'
        akey = 'UPPER(SUBSTITUTE(SUBSTITUTE($B$4,"0x",""),"0X",""))'
        we.cell(7, 1, "bit"); we.cell(8, 1, "信号名"); we.cell(9, 1, "当前bit")
        we.cell(10, 1, "目标(留空=不变)").font = bold
        we.cell(11, 1, "结果bit")
        bits = list(range(15, -1, -1))
        for i, b in enumerate(bits):
            col = 2 + i
            L = get_column_letter(col)
            we.cell(7, col, b).font = bold
            we.cell(8, col, '=IFERROR(VLOOKUP(%s&"_"&%s$7,\'位定义\'!$A:$B,2,FALSE),"")' % (akey, L))
            we.cell(9, col, '=IFERROR(MOD(INT(%s/2^%s$7),2),"")' % (cur, L))
            cc = we.cell(10, col, None); cc.fill = in_fill; cc.number_format = "@"
            we.cell(11, col, '=IF(%s10="",%s9,%s10)' % (L, L, L))
        f, la = get_column_letter(2), get_column_letter(1 + len(bits))
        c = we.cell(5, 2, '=IF($B$3="","",IF(ISERROR(%s),"⚠当前值需hex",IF(%s>65535,"⚠超16位",'
                          '"0x"&DEC2HEX(SUMPRODUCT(%s11:%s11,2^(%s7:%s7)),4))))' % (cur, cur, f, la, f, la))
        c.fill = out_fill; c.font = bold
        we.column_dimensions["A"].width = 18
        for col in range(2, 2 + len(bits)):
            we.column_dimensions[get_column_letter(col)].width = 5

        # ---- 解码器（反向）：粘贴一串 (地址,值) → 自动列出每个寄存器哪些 bit=1、哪些信号在开 ----
        wx = wb.create_sheet("解码器")
        wx.cell(1, 1, "解码器 — 左边粘贴 dump(地址+值hex)，右表自动算每个信号；单 bit 显 ● ON/off，多位字段显「值 N」；筛「状态」看开着的").font = bold
        wx.cell(2, 1, "地址").font = bold
        wx.cell(2, 2, "值(hex)").font = bold
        wx.cell(2, 3, "key")
        for r in range(3, 63):   # 60 行输入
            ca = wx.cell(r, 1, None); ca.fill = in_fill; ca.number_format = "@"
            cb = wx.cell(r, 2, None); cb.fill = in_fill; cb.number_format = "@"
            # 地址与值都填了才算一条（半行粘贴不会被误判为"全关"，修 1.5）
            wx.cell(r, 3, '=IF(OR(A%d="",B%d=""),"",UPPER(SUBSTITUTE(SUBSTITUTE(A%d,"0x",""),"0X","")))' % (r, r, r))
        wx.column_dimensions["C"].hidden = True
        for j, h in enumerate(["信号", "地址", "bit", "寄存器", "该寄存器值", "字段值", "状态"], 5):  # 列 E..K
            wx.cell(2, j, h).font = bold
        for i, x in enumerate(bd, 3):
            key = 'UPPER(SUBSTITUTE(SUBSTITUTE(F%d,"0x",""),"0X",""))' % i
            lo, wid = x["lo"], x["width"]
            wx.cell(i, 5, x["sig"]); wx.cell(i, 6, x["addr"]); wx.cell(i, 7, x["bit_str"]); wx.cell(i, 8, x["reg"])
            wx.cell(i, 9, '=IFERROR(INDEX($B:$B,MATCH(%s,$C:$C,0)),"")' % key)
            wx.cell(i, 10, '=IFERROR(IF(I%d="","",MOD(INT(HEX2DEC(SUBSTITUTE(SUBSTITUTE(I%d,"0x",""),"0X",""))/2^%d),2^%d)),"?")' % (i, i, lo, wid))
            if wid == 1:
                wx.cell(i, 11, '=IF(J%d="","",IF(J%d=1,"● ON","off"))' % (i, i))
            else:
                wx.cell(i, 11, '=IF(J%d="","","值 "&J%d)' % (i, i))   # 多位字段：显字段值，不误当 ON/off
        wx.auto_filter.ref = "E2:K%d" % (2 + len(bd))
        for cc, w in {"A": 14, "B": 10, "E": 30, "F": 13, "G": 6, "H": 16, "I": 12, "J": 7, "K": 9}.items():
            wx.column_dimensions[cc].width = w
        wx.freeze_panes = "A3"
    cols = ["步骤", "操作", "地址 ADDR", "值 VALUE", "寄存器/模块", "信号变化(bit:前→后)"]
    widths = [9, 22, 13, 10, 22, 44]
    used = set()
    for mode_id, tc in testcases:
        nm = re.sub(r"[\[\]:*?/\\]", "_", str(mode_id or tc.get("mode") or "mode"))[:31] or "mode"
        base = nm
        k = 1
        while nm in used:
            nm = (base[:28] + "_%d" % k); k += 1
        used.add(nm)
        ws = wb.create_sheet(nm)
        ws.cell(1, 1, "%s  %s  |  reg_group=%s" % (tc.get("mode"), tc.get("mode_name") or "", tc.get("reg_group"))).font = bold
        ws.cell(2, 1, "累积逐级关闭：每步先发本段写、再测总电流；相邻步电流差 = 该级功耗。")
        mw = tc.get("warnings") or (tc.get("diagnostics") or {}).get("warnings") or []
        if mw:
            ws.cell(3, 1, "⚠ 模式级：" + "；".join(str(x) for x in mw)).font = bold
        for c, h in enumerate(cols, 1):
            cell = ws.cell(4, c, h); cell.font = bold; cell.fill = hdr_fill
        r = 5
        for i, w in enumerate(tc["baseline"]["writes"]):
            seg = []
            for f in w.get("fields", []):
                if f.get("role") in ("override", "mux_sel"):
                    seg.append("%s=%s" % (f["signal"], f["value"]))
                elif f.get("on"):
                    seg.append("%s=on" % f["signal"])
            row = ["baseline" if i == 0 else "", "建立全开起始态" if i == 0 else "",
                   w["addr"], w["value"], w.get("reg") or "", ", ".join(seg)]
            for c, v in enumerate(row, 1):
                cc = ws.cell(r, c, v)
                if i == 0:
                    cc.fill = base_fill
            r += 1
        for st in tc["steps"]:
            node = (st.get("off_node") or "").split("/")[-1]
            hlabel, hop = "step %d" % st["index"], "OFF %s" % st["off_label"]
            tail = "；".join(list(st.get("warnings") or []) + ([st["note"]] if st.get("note") else []))
            if not st["writes"]:
                # 空步骤：本级门已被前面共用位提前关掉，仍是一个测量点 —— 必须保留(修 1.8)
                cells = [hlabel, hop, "", "", node, "（本级门已被前面共用位提前关，仍是测量点）" + ("；" + tail if tail else "")]
                for c, v in enumerate(cells, 1):
                    ws.cell(r, c, v).fill = step_fill
                r += 1
                continue
            for i, w in enumerate(st["writes"]):
                ftxt = ", ".join("%s[%s]:%d→%d" % (f["signal"], f["bit"], f["before"], f["after"])
                                 for f in w.get("fields", []))
                if i == 0 and tail:
                    ftxt = (ftxt + "   ⚠ " + tail) if ftxt else ("⚠ " + tail)
                row = [hlabel if i == 0 else "", hop if i == 0 else "",
                       w["addr"], w["value"], node, ftxt]
                for c, v in enumerate(row, 1):
                    cc = ws.cell(r, c, v)
                    if i == 0:
                        cc.fill = step_fill
                r += 1
        if tc["diagnostics"].get("uncovered_off_gates"):
            r += 1
            ws.cell(r, 1, "⚠ 未覆盖门（需人工补写）:").font = bold; r += 1
            for u in tc["diagnostics"]["uncovered_off_gates"]:
                ws.cell(r, 1, "%s : %s" % (u["signal"], u.get("note", ""))); r += 1
        for c, wd in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = wd
        ws.freeze_panes = "A5"
    wb.save(path)
    return len(testcases)


def render_debug_html(tc, flowgraph=None):
    def esc(s):
        return html.escape(str(s if s is not None else ""))

    parts = []
    parts.append("""<!doctype html><html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>debug — %s</title><style>
:root{--bg:#0f1216;--fg:#e6e9ef;--mut:#8b93a2;--acc:#6ea8fe;--warn:#f0b400;--off:#ff6b6b;--card:#1a1f27;--line:#2a2f3a}
@media(prefers-color-scheme:light){:root{--bg:#f7f8fa;--fg:#1a1f27;--mut:#5a6270;--acc:#2b6cb0;--card:#fff;--line:#e2e6ec}}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--fg);font:14px/1.5 ui-monospace,SFMono-Regular,Menlo,Consolas,monospace}
.wrap{max-width:1100px;margin:0 auto;padding:24px}
h1{font-size:20px;margin:0 0 4px}.sub{color:var(--mut);margin-bottom:20px}
.card{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:14px 16px;margin:12px 0}
.step h2{font-size:15px;margin:0 0 8px}.badge{display:inline-block;padding:1px 8px;border-radius:20px;font-size:12px;background:var(--acc);color:#000;margin-right:6px}
table{border-collapse:collapse;width:100%%;margin-top:8px;font-size:13px}
th,td{text-align:left;padding:4px 8px;border-bottom:1px solid var(--line)}th{color:var(--mut);font-weight:600}
.addr{color:var(--acc)}.val{font-weight:700}.off{color:var(--off)}.warn{color:var(--warn)}
.bit{color:var(--mut)}.arrow{color:var(--mut)}.mut{color:var(--mut)}
.wbox{background:var(--bg);border:1px solid var(--line);border-radius:6px;padding:2px 6px;font-size:12px}
</style></head><body><div class="wrap">""" % esc(tc.get("mode")))
    parts.append("<h1>%s <span class='mut'>· %s</span></h1>" % (esc(tc.get("mode")), esc(tc.get("mode_name") or "")))
    parts.append("<div class='sub'>reg_group=<b>%s</b> · order=<b>%s</b> · base=%s · %d baseline regs · %d steps</div>"
                 % (esc(tc.get("reg_group")), esc(tc.get("order_mode")), esc(tc.get("base_addr")),
                    tc["stats"]["baseline_regs"], tc["stats"]["steps"]))
    for w in tc.get("warnings", []):
        parts.append("<div class='card warn'>⚠ %s</div>" % esc(w))

    # baseline
    parts.append("<div class='card'><h2>Baseline — 全开起始态</h2><table><tr><th>ADDR</th><th>REG</th><th>VALUE</th><th>reset</th><th>set fields</th></tr>")
    for w in tc["baseline"]["writes"]:
        ff = "; ".join("%s[%s]=%s%s" % (esc(f["signal"]), esc(f["bit"]), esc(f["value"]),
                                        " (on)" if f.get("on") else "")
                       for f in w.get("fields", []))
        parts.append("<tr><td class='addr'>%s</td><td>%s</td><td class='val'>%s</td><td class='mut'>%s</td><td>%s</td></tr>"
                     % (esc(w["addr"]), esc(w["reg"]), esc(w["value"]), esc(w.get("reset", "")), ff))
    parts.append("</table></div>")

    # steps
    for st in tc["steps"]:
        parts.append("<div class='card step'><h2><span class='badge'>step %d</span>OFF <span class='off'>%s</span> <span class='mut'>(%s)</span></h2>"
                     % (st["index"], esc(st["off_label"]), esc(st.get("device") or "")))
        for wn in st.get("warnings", []):
            parts.append("<div class='warn'>⚠ %s</div>" % esc(wn))
        if st.get("note"):
            parts.append("<div class='mut'>· %s</div>" % esc(st["note"]))
        if st["writes"]:
            parts.append("<table><tr><th>ADDR</th><th>REG</th><th>prev→VALUE</th><th>fields</th></tr>")
            for w in st["writes"]:
                ff = "; ".join("%s <span class='bit'>[%s]</span> %d<span class='arrow'>→</span><b class='off'>%d</b>"
                               % (esc(f["signal"]), esc(f["bit"]), f["before"], f["after"])
                               for f in w.get("fields", []))
                parts.append("<tr><td class='addr'>%s</td><td>%s</td><td><span class='mut'>%s</span><span class='arrow'>→</span><b class='val'>%s</b></td><td>%s</td></tr>"
                             % (esc(w["addr"]), esc(w["reg"]), esc(w["prev"]), esc(w["value"]), ff))
            parts.append("</table>")
        parts.append("<div class='mut' style='margin-top:6px'>▸ %s</div></div>" % esc(st["measure"]))

    if tc.get("extra_writes"):
        parts.append("<div class='card'><h2>Extra writes（模式级额外写）</h2><table><tr><th>ADDR</th><th>VALUE</th><th>note</th></tr>")
        for w in tc["extra_writes"]:
            parts.append("<tr><td class='addr'>%s</td><td class='val'>%s</td><td>%s</td></tr>"
                         % (esc(hex_addr(w["addr"])), esc(hex4(to_int(w["value"]))), esc(w.get("note", ""))))
        parts.append("</table></div>")

    if tc["diagnostics"]["uncovered_off_gates"]:
        parts.append("<div class='card warn'><h2>⚠ 未覆盖的门（需人工补写）</h2><ul>")
        for u in tc["diagnostics"]["uncovered_off_gates"]:
            parts.append("<li>%s — %s</li>" % (esc(u["signal"]), esc(u["note"])))
        parts.append("</ul></div>")
    parts.append("</div></body></html>")
    return "".join(parts)


# ------------------------------------------------------------------ project I/O
def load_inputs(args):
    gate_override = {}
    if args.project:
        fg = load_json(os.path.join(args.project, "flowgraph.json"))
        rm = load_json(os.path.join(args.project, "regmap.json"))
        if not args.mode:
            sys.exit("--project 需配 --mode <id>")
        mode = load_json(os.path.join(args.project, "modes", args.mode + ".json"))
        lp = os.path.join(args.project, "layout.json")   # 关断总闸指定（per-block，跨模式复用）
        if os.path.exists(lp):
            gate_override = (load_json(lp) or {}).get("gate_override", {}) or {}
    else:
        if not (args.flowgraph and args.regmap and args.mode_file):
            sys.exit("需 --project 或 (--flowgraph --regmap --mode-file)")
        fg = load_json(args.flowgraph)
        rm = load_json(args.regmap)
        mode = load_json(args.mode_file)
    gate_override = dict(gate_override)
    gate_override.update(mode.get("gate_override", {}) or {})   # 模式内可再覆盖（per-mode 优先）
    return fg, rm, mode, gate_override


def _harden_stdout():
    """非 UTF-8 控制台（如 Windows GBK）打印含 ⚠/中文的 warning 不该崩进程。文件输出一律 utf-8。"""
    for stream in (sys.stdout, sys.stderr):
        try:
            if hasattr(stream, "reconfigure"):
                stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


def main(argv=None):
    _harden_stdout()
    ap = argparse.ArgumentParser(description="生成'电流逐级关闭'测试序列（testcase/1）+ 渲染 ate.txt/debug.html")
    ap.add_argument("--project", help="projects/<name> 目录（含 flowgraph.json/regmap.json/modes/）")
    ap.add_argument("--mode", help="模式 id（配 --project）")
    ap.add_argument("--flowgraph")
    ap.add_argument("--regmap")
    ap.add_argument("--mode-file")
    ap.add_argument("--out-json", help="写 testcase JSON（默认 project/testcases/<mode>.json）")
    ap.add_argument("--out-ate", help="写 ate.txt")
    ap.add_argument("--out-html", help="写 debug.html")
    ap.add_argument("--xlsx", help="导出 Excel：**每个模式一张 sheet**。--project X --xlsx out.xlsx（默认导所有模式；配 --mode a,b 只导指定的）")
    ap.add_argument("--print", dest="do_print", action="store_true", help="打印 ate.txt 到控制台")
    args = ap.parse_args(argv)

    if args.xlsx:                                    # 多模式 → 一个 Excel（每模式一 tab）
        if not args.project:
            sys.exit("--xlsx 需配 --project")
        fg = load_json(os.path.join(args.project, "flowgraph.json"))
        rm = load_json(os.path.join(args.project, "regmap.json"))
        lp = os.path.join(args.project, "layout.json")
        go = ((load_json(lp) or {}).get("gate_override", {}) or {}) if os.path.exists(lp) else {}
        mdir = os.path.join(args.project, "modes")
        ids = ([m.strip() for m in args.mode.split(",") if m.strip()] if args.mode
               else [f[:-5] for f in sorted(os.listdir(mdir)) if f.endswith(".json")] if os.path.isdir(mdir) else [])
        if not ids:
            sys.exit("没找到模式：%s" % mdir)
        tcs = []
        for mid in ids:
            m = load_json(os.path.join(mdir, mid + ".json"))
            g = dict(go); g.update(m.get("gate_override", {}) or {})
            tcs.append((mid, generate(fg, rm, m, g)))
        lep = os.path.join(args.project, "dco_logic_expr.json")   # 门级布尔式（可选，用于 EN 映射列）
        logic_expr = load_json(lep) if os.path.exists(lep) else None
        ref = build_reference(fg, rm, go, logic_expr=logic_expr)   # block→EN布尔式→寄存器 参考表
        render_xlsx(tcs, args.xlsx, reference=ref)
        print("Excel 已写: %s  (Reference + %d 个模式 sheet: %s)" % (args.xlsx, len(tcs), ", ".join(ids)))
        return

    fg, rm, mode, gate_override = load_inputs(args)
    tc = generate(fg, rm, mode, gate_override)

    out_json = args.out_json
    if not out_json and args.project:
        d = os.path.join(args.project, "testcases")
        os.makedirs(d, exist_ok=True)
        out_json = os.path.join(d, (mode.get("id") or "mode") + ".json")
    if out_json:
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(tc, f, ensure_ascii=False, indent=1)

    ate = render_ate(tc)
    if args.out_ate:
        with open(args.out_ate, "w", encoding="utf-8", newline="\n") as f:
            f.write(ate)
    elif args.project and not args.do_print:
        with open(os.path.join(args.project, "testcases", (mode.get("id") or "mode") + ".ate.txt"),
                  "w", encoding="utf-8", newline="\n") as f:
            f.write(ate)

    if args.out_html:
        with open(args.out_html, "w", encoding="utf-8", newline="\n") as f:
            f.write(render_debug_html(tc, fg))
    elif args.project:
        with open(os.path.join(args.project, "testcases", (mode.get("id") or "mode") + ".debug.html"),
                  "w", encoding="utf-8", newline="\n") as f:
            f.write(render_debug_html(tc, fg))

    if args.do_print:
        sys.stdout.write(ate)
    else:
        s = tc["stats"]
        print("[gen_testcase] mode=%s group=%s -> %d baseline regs, %d steps, %d gates off"
              % (tc["mode"], tc["reg_group"], s["baseline_regs"], s["steps"], s["gates_off"]))
        if out_json:
            print("  testcase -> %s" % out_json)
        for w in tc.get("warnings", []):
            print("  ⚠ %s" % w)


if __name__ == "__main__":
    main()
