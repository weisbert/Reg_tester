#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""diff_mode_regs.py — 多个「模式寄存器写序 Excel」的差异审计报告。

⚠ 定位：只给人看的审计工具，**不生成可执行写序**。
   写序是顺序敏感的（同址多写=分级使能/先清后置/触发），任何按"末态差异"
   生成的写入序列都可能破坏时序语义。跨模式切换请全量重放原写序
   （后续 pack 工具负责按工具B格式切块），差异报告只用来回答：
   模式间差多少、哪些地址是"A 写 B 不写"的孤儿、哪些段可能顺序敏感。

表单结构假设（与 grab_regs.py 一致，自动探测表头行，匹配不到按前四列兜底）：
    Register Name | Address | Value | Description

用法：
    python diff_mode_regs.py <xlsx 或 目录>... [--out 审计.xlsx] [--sheet Sheet1]

  位置参数    模式 Excel 文件，或含它们的目录（目录取其中全部 *.xlsx，按文件名序）
  --out       输出审计工作簿路径（默认 模式差异审计.xlsx）
  --sheet     只读各文件的指定 sheet（默认全部 sheet）

输出工作簿 6 个 sheet：
  总览        每模式行数/地址数/多写地址数 + 两两"末态差异数"矩阵
  末态对比    地址并集 × 各模式末值（黄=模式间有差异，灰=该模式没写）
  孤儿清单    没被所有模式覆盖的地址：谁写了、谁没写（切模式残留风险源）
  多写时间线  各模式内同址多写的值历史（顺序敏感段嫌疑清单，禁止压缩）
  公共前缀    所有模式开头完全相同的 (地址,值) 行段（唯一可安全提取的公共段）
  跳过行      有内容但没解析出地址/值的行（防漏检）

依赖 openpyxl；只读输入，不改输入。
"""
import argparse
import io
import re
import sys
from collections import OrderedDict


def norm_addr(x):
    """地址归一：去空白/0x/下划线，大写。非字符串也转。"""
    if x is None:
        return ""
    s = str(x).strip().replace("_", "")
    s = re.sub(r"(?i)^0x", "", s)
    return s.upper()


def norm_value(x):
    """值归一用于比较：同地址归一；若是纯十六进制再去前导零（'0111'=='111'）。显示仍用原始值。"""
    s = norm_addr(x)
    if s and re.fullmatch(r"[0-9A-F]+", s):
        s = s.lstrip("0") or "0"
    return s


def find_header(rows, scan=20):
    """在前 scan 行里找含 'address' 的表头行，返回 (行号idx, {列名->列idx})。找不到 -> (None, 兜底映射)。"""
    for i, row in enumerate(rows[:scan]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(c == "address" or c.replace(" ", "") == "registeraddress" for c in cells):
            m = {}
            for j, c in enumerate(cells):
                key = c.replace(" ", "")
                if key in ("registername", "regname", "name"):
                    m["name"] = j
                elif key in ("address", "registeraddress", "addr"):
                    m["addr"] = j
                elif key in ("value", "registervalue", "writevalue", "val"):
                    m["value"] = j
                elif key in ("description", "desc", "note", "comment"):
                    m["desc"] = j
            if "addr" in m:
                m.setdefault("name", 0)
                m.setdefault("value", m["addr"] + 1)
                m.setdefault("desc", m["value"] + 1)
                return i, m
    return None, {"name": 0, "addr": 1, "value": 2, "desc": 3}


ADDR_RE = re.compile(r"[0-9A-F]{2,}$")


def parse_mode(path, only_sheet=None):
    """读一个模式写序 Excel → {'writes': [逐行写序], 'skipped': [没解析出的行]}。写序严格保持原始顺序。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [only_sheet] if only_sheet else wb.sheetnames
    writes, skipped = [], []
    for sn in sheets:
        if sn not in wb.sheetnames:
            sys.exit("%s 里找不到 sheet: %s（有: %s）" % (path, sn, ", ".join(wb.sheetnames)))
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        hidx, cm = find_header(rows)
        start = (hidx + 1) if hidx is not None else 0
        for i in range(start, len(rows)):
            row = rows[i]
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            def cell(k):
                j = cm[k]
                return ("" if j >= len(row) or row[j] is None else str(row[j]).strip())

            addr = norm_addr(row[cm["addr"]] if cm["addr"] < len(row) else None)
            val = cell("value")
            if not ADDR_RE.fullmatch(addr) or not val:
                skipped.append({"sheet": sn, "row": i + 1,
                                "content": " | ".join(str(c) for c in row if c is not None)[:120]})
                continue
            writes.append({"seq": len(writes) + 1, "sheet": sn, "row": i + 1,
                           "name": cell("name"), "addr": addr,
                           "value": val, "vnorm": norm_value(val), "desc": cell("desc")})
    wb.close()
    return {"writes": writes, "skipped": skipped}


def last_state(writes):
    """写序 → 末态 OrderedDict(addr -> 最后一次写的记录)，保持首次出现顺序。"""
    st = OrderedDict()
    for w in writes:
        st[w["addr"]] = w
    return st


def build_report(labels, modes, out_path):
    """labels: 模式名列表；modes: {label: parse_mode结果}。生成审计工作簿并返回统计摘要。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    BOLD = Font(bold=True)
    FILL_DIFF = PatternFill("solid", fgColor="FFF2A8")    # 黄：模式间有差异
    FILL_MISS = PatternFill("solid", fgColor="D9D9D9")    # 灰：该模式没写
    FILL_HEAD = PatternFill("solid", fgColor="DDEBF7")

    states = {lb: last_state(modes[lb]["writes"]) for lb in labels}
    multi = {lb: OrderedDict((a, hs) for a, hs in _history(modes[lb]["writes"]).items() if len(hs) > 1)
             for lb in labels}

    # 地址并集，按（模式顺序，模式内首现顺序）排
    union = OrderedDict()
    for lb in labels:
        for a, w in states[lb].items():
            union.setdefault(a, w["name"])

    wb = openpyxl.Workbook()

    # ---- 总览 ----
    ws = wb.active
    ws.title = "总览"
    ws.append(["模式", "写序行数", "唯一地址数", "同址多写地址数", "跳过行数"])
    for lb in labels:
        ws.append([lb, len(modes[lb]["writes"]), len(states[lb]), len(multi[lb]),
                   len(modes[lb]["skipped"])])
    ws.append([])
    ws.append(["两两末态差异数（值不同 + 单侧缺失）"])
    ws.append([""] + labels)
    pair_diff = {}
    for la in labels:
        row = [la]
        for lc in labels:
            if la == lc:
                row.append("-")
                continue
            n = sum(1 for a in set(states[la]) | set(states[lc])
                    if (states[la].get(a) or {}).get("vnorm") != (states[lc].get(a) or {}).get("vnorm"))
            pair_diff[(la, lc)] = n
            row.append(n)
        ws.append(row)
    ws.append([])
    ws.append(["⚠ 本报告仅供审计。跨模式切换请全量重放原写序，勿按末态差异生成写入序列",
               ])
    for r in (1, len(labels) + 3):
        for c in ws[r]:
            c.font = BOLD

    # ---- 末态对比 ----
    ws = wb.create_sheet("末态对比")
    head = ["Register Name", "Address"] + labels + ["覆盖", "差异", "多写模式"]
    ws.append(head)
    n_diff_addr = 0
    diff_list = []
    for a, name in union.items():
        vals = [states[lb].get(a) for lb in labels]
        vnorms = set(w["vnorm"] for w in vals if w)
        covered = sum(1 for w in vals if w)
        differ = len(vnorms) > 1
        n_diff_addr += differ
        mw = ",".join(lb for lb in labels if a in multi[lb])
        if differ or covered < len(labels):
            diff_list.append({"addr": a, "name": name, "mw": mw,
                              "vals": [(w["value"] if w else None) for w in vals]})
        r = [name, a] + [(w["value"] if w else "") for w in vals] + \
            ["%d/%d" % (covered, len(labels)), "Y" if differ else "", mw]
        ws.append(r)
        for j, w in enumerate(vals):
            cell = ws.cell(row=ws.max_row, column=3 + j)
            if w is None:
                cell.fill = FILL_MISS
            elif differ:
                cell.fill = FILL_DIFF
    ws.freeze_panes = "C2"

    # ---- 孤儿清单 ----
    ws = wb.create_sheet("孤儿清单")
    ws.append(["Address", "Register Name", "写过的模式", "没写的模式",
               "切换残留风险：从左列模式切到右列模式时该地址保持旧值"])
    n_orphan = 0
    for a, name in union.items():
        have = [lb for lb in labels if a in states[lb]]
        lack = [lb for lb in labels if a not in states[lb]]
        if lack:
            n_orphan += 1
            ws.append([a, name, ",".join(have), ",".join(lack)])
    ws.freeze_panes = "A2"

    # ---- 多写时间线 ----
    ws = wb.create_sheet("多写时间线")
    ws.append(["模式", "Address", "Register Name", "写次数", "值历史（按执行序）", "源行号",
               "⚠ 顺序敏感嫌疑段，禁止 last-write-wins 压缩"])
    for lb in labels:
        for a, hs in multi[lb].items():
            ws.append([lb, a, hs[0]["name"], len(hs),
                       " -> ".join(h["value"] for h in hs),
                       ",".join("%s!r%d" % (h["sheet"], h["row"]) for h in hs)])
    ws.freeze_panes = "A2"

    # ---- 公共前缀 ----
    ws = wb.create_sheet("公共前缀")
    seqs = [modes[lb]["writes"] for lb in labels]
    npfx = 0
    if seqs and all(seqs):
        for tup in zip(*seqs):
            k0 = (tup[0]["addr"], tup[0]["vnorm"])
            if all((w["addr"], w["vnorm"]) == k0 for w in tup[1:]):
                npfx += 1
            else:
                break
    ws.append(["所有 %d 个模式开头完全相同的 (地址,值) 连续段：%d 行" % (len(labels), npfx),
               "", "这是唯一可安全提取为公共段的部分（保持原顺序整段搬运）"])
    ws.append(["#", "Register Name", "Address", "Value", "Description"])
    base = seqs[0] if seqs else []
    for w in base[:npfx]:
        ws.append([w["seq"], w["name"], w["addr"], w["value"], w["desc"]])
    if npfx < min((len(s) for s in seqs), default=0):
        ws.append([])
        ws.append(["首个分歧点（各模式在公共前缀后的下一行）："])
        for lb, s in zip(labels, seqs):
            w = s[npfx] if npfx < len(s) else None
            ws.append([lb] + ([w["name"], w["addr"], w["value"], w["desc"]] if w else ["<已到文件尾>"]))
    ws.freeze_panes = "A3"

    # ---- 跳过行 ----
    ws = wb.create_sheet("跳过行")
    ws.append(["模式", "sheet", "行号", "内容（截断）"])
    for lb in labels:
        for s in modes[lb]["skipped"]:
            ws.append([lb, s["sheet"], s["row"], s["content"]])
    ws.freeze_panes = "A2"

    # 统一表头样式 + 列宽
    for ws in wb.worksheets:
        for c in ws[1]:
            c.font = BOLD
            c.fill = FILL_HEAD
            c.alignment = Alignment(vertical="center")
        widths = {}
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
            for c in row:
                if c.value is not None:
                    widths[c.column_letter] = min(48, max(widths.get(c.column_letter, 8),
                                                          len(str(c.value)) + 2))
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

    multi_lines = []
    for lb in labels:
        for a, hs in multi[lb].items():
            multi_lines.append({"mode": lb, "addr": a, "name": hs[0]["name"],
                                "seq": " -> ".join(h["value"] for h in hs)})

    wb.save(out_path)
    return {"union": len(union), "diff_addr": n_diff_addr, "orphan": n_orphan,
            "prefix": npfx, "pair_diff": pair_diff,
            "diff_list": diff_list, "multi_lines": multi_lines}


def _history(writes):
    hist = OrderedDict()
    for w in writes:
        hist.setdefault(w["addr"], []).append(w)
    return hist


def main(argv=None):
    ap = argparse.ArgumentParser(description="模式寄存器写序 Excel 差异审计（只看不改，不生成执行序）")
    ap.add_argument("inputs", nargs="+", help="模式 xlsx 文件或目录")
    ap.add_argument("--out", default="模式差异审计.xlsx")
    ap.add_argument("--sheet", help="只读各文件的该 sheet（默认全部）")
    args = ap.parse_args(argv)

    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

    import os
    files = []
    for p in args.inputs:
        if os.path.isdir(p):
            files += sorted(os.path.join(p, f) for f in os.listdir(p)
                            if f.lower().endswith(".xlsx") and not f.startswith("~$"))
        else:
            files.append(p)
    if len(files) < 2:
        sys.exit("至少要 2 个模式 Excel 才有差异可比（现在 %d 个）" % len(files))

    labels, modes = [], {}
    for f in files:
        lb = os.path.splitext(os.path.basename(f))[0]
        if lb in modes:
            sys.exit("模式名重复：%s" % lb)
        m = parse_mode(f, args.sheet)
        labels.append(lb)
        modes[lb] = m
        print("读入 %-40s 写序 %4d 行 / 唯一地址 %4d / 同址多写 %2d / 跳过 %d"
              % (lb, len(m["writes"]), len(last_state(m["writes"])),
                 sum(1 for hs in _history(m["writes"]).values() if len(hs) > 1),
                 len(m["skipped"])))

    stat = build_report(labels, modes, args.out)
    print()
    print("地址并集 %d 个；模式间有末态差异的地址 %d 个；未被全部模式覆盖的孤儿地址 %d 个；"
          % (stat["union"], stat["diff_addr"], stat["orphan"]))
    print("所有模式完全相同的公共前缀 %d 行。" % stat["prefix"])

    # 差异清单（切换签名块的原料，直接从控制台复制走）
    if stat["diff_list"]:
        print()
        print("== 差异地址清单（%d 个；末态不同或未全覆盖；∅=该模式没写）==" % len(stat["diff_list"]))
        for i, lb in enumerate(labels, 1):
            print("   M%d = %s" % (i, lb))
        for d in stat["diff_list"][:200]:
            cells = " ".join("M%d=%s" % (i, v if v is not None else "∅")
                             for i, v in enumerate(d["vals"], 1))
            tag = ("   ★撞多写:" + d["mw"]) if d["mw"] else ""
            print("  %-10s %-26s %s%s" % (d["addr"], d["name"][:26], cells, tag))
        if len(stat["diff_list"]) > 200:
            print("  …共 %d 个，超出 200 的看工作簿 末态对比 sheet" % len(stat["diff_list"]))

    if stat["multi_lines"]:
        print()
        print("== 同址多写时间线（顺序敏感段，禁止折叠进签名块）==")
        agg = OrderedDict()
        for m in stat["multi_lines"]:
            agg.setdefault((m["addr"], m["name"], m["seq"]), []).append(m["mode"])
        for (a, nm, seq), lbs in agg.items():
            who = "全部模式" if len(lbs) == len(labels) else ",".join(lbs)
            print("  %-10s %-26s %s   [%s]" % (a, nm[:26], seq, who))

    print()
    print("审计工作簿已写: %s" % args.out)
    print("⚠ 提醒：差异清单需配合每模式 init+lock 行使用；撞多写的地址要整段原序重放。")


if __name__ == "__main__":
    main()
