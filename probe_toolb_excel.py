#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""probe_toolb_excel.py — 探测「工具B流程 Excel」的完整结构（为 pack 工具逆向格式用）。

和 explore_excel.py（看大寄存器簿的数据）不同，这个专抓**格式语义**：
流程表的列定义藏在表头，步骤类型清单往往藏在下拉框（数据有效性）里，
另有合并单元格/隐藏行列/公式/批注/宏 都可能承载工具B的执行语义。

用法：
    python probe_toolb_excel.py <工具B样例.xlsx/.xlsm> [--json out.json] [--rows 30] [--cols 0] [--sheet NAME]

  --json   完整结构落盘 JSON（★把这个文件带回来分析，控制台只是摘要）
  --rows   每个 sheet 顶部 dump 的行数（默认 30，0=不 dump）
  --cols   每行最多探测的列数（默认 0=全部列；行 dump 只存非空格子，宽表不怕大）
  --sheet  只探测指定 sheet（默认全部，含隐藏 sheet）

每个 sheet 报告：
  可见性 | 尺寸 | 冻结窗格 | 合并单元格 | 隐藏行/列 | 数据有效性(下拉框选项★)
  表头候选行 + 重复列组探测（如 Addr/Value ×11 → 一行 11 个寄存器的排布）
  公式清单 | 批注 | 顶部 N 行原值 dump
工作簿级：defined names、是否含 VBA 宏（含宏时宏逻辑本脚本看不了，需另行导出）。

依赖 openpyxl；只读不写。⚠ 输出 JSON 含公司数据，别提交进 git（数据侧已 gitignore）。
"""
import argparse
import io
import json
import re
import sys
import zipfile
from collections import Counter


def s(x, limit=120):
    """单元格值 → 截断字符串。None → ""。"""
    if x is None:
        return ""
    t = str(x)
    return t if len(t) <= limit else t[:limit] + "…"


def detect_repeat_groups(cells):
    """一行表头里找重复列组：去掉尾部编号后同名列 ≥3 次即报。
    如 Addr1..Addr11 / Value1..Value11 → {'addr': 11, 'value': 11}。"""
    base = []
    for c in cells:
        t = str(c).strip() if c is not None else ""
        if not t:
            continue
        b = re.sub(r"[\s_\-#.]*\d+$", "", t)
        if b:
            base.append(b.lower())
    return {b: n for b, n in Counter(base).items() if n >= 3}


def probe_sheet(ws, ws_formula, dump_rows, max_cols=0):
    """探测单个 sheet → dict。ws 来自 data_only=True 加载，ws_formula 来自 data_only=False。"""
    from openpyxl.utils import get_column_letter
    info = {
        "name": ws.title,
        "state": ws.sheet_state,                      # visible / hidden / veryHidden
        "max_row": ws.max_row, "max_col": ws.max_column,
        "freeze_panes": ws.freeze_panes,
        "merged": [str(r) for r in ws.merged_cells.ranges],
        "hidden_cols": sorted(k for k, d in ws.column_dimensions.items() if d.hidden),
        "hidden_rows": sorted(k for k, d in ws.row_dimensions.items() if d.hidden)[:50],
        "validations": [], "header_candidates": [], "formulas": [], "comments": [], "rows": [],
    }

    # 数据有效性（下拉框=步骤类型清单的最大嫌疑地）
    try:
        for dv in ws.data_validations.dataValidation:
            info["validations"].append({
                "sqref": str(dv.sqref), "type": dv.type,
                "formula1": s(dv.formula1, 500), "formula2": s(dv.formula2, 200),
                "allow_blank": dv.allow_blank,
                "prompt": s(dv.prompt, 200), "error": s(dv.error, 200),
            })
    except Exception as e:
        info["validations"].append({"error_reading": repr(e)})

    # 顶部行 dump + 表头候选（非空格子多的行）+ 重复列组
    ncol = ws.max_column or 0
    if max_cols:
        ncol = min(ncol, max_cols)
    top = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, max(dump_rows, 15)),
                                         max_col=ncol, values_only=True), start=1):
        top.append((i, [s(c, 80) for c in row]))
    for i, vals in top[:dump_rows]:
        cells = {}
        for j, v in enumerate(vals, start=1):
            if v:
                cells[get_column_letter(j)] = v
        info["rows"].append({"r": i, "cells": cells})
    for i, vals in top[:15]:
        nz = sum(1 for v in vals if v)
        if nz >= 3:
            cand = {"r": i, "non_empty": nz, "cells": [v for v in vals if v][:40]}
            g = detect_repeat_groups(vals)
            if g:
                cand["repeat_groups"] = g
            info["header_candidates"].append(cand)
    info["header_candidates"].sort(key=lambda c: -c["non_empty"])
    info["header_candidates"] = info["header_candidates"][:3]

    # 公式 & 批注（从 formula 版工作簿读）
    if ws_formula is not None:
        nf = 0
        for row in ws_formula.iter_rows(min_row=1, max_row=min(ws_formula.max_row or 0, 500),
                                        max_col=ncol):
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    if nf < 40:
                        info["formulas"].append({"cell": c.coordinate, "f": s(c.value, 200)})
                    nf += 1
                if c.comment is not None and len(info["comments"]) < 20:
                    info["comments"].append({"cell": c.coordinate, "text": s(c.comment.text, 300)})
        info["formula_count"] = nf
    return info


def main(argv=None):
    ap = argparse.ArgumentParser(description="工具B流程 Excel 结构探测（只读）")
    ap.add_argument("xlsx")
    ap.add_argument("--json", help="完整结构落盘 JSON（带回来分析用）")
    ap.add_argument("--rows", type=int, default=30, help="每 sheet 顶部 dump 行数（默认 30）")
    ap.add_argument("--cols", type=int, default=0, help="每行最多探测列数（默认 0=全部列）")
    ap.add_argument("--sheet", help="只探测该 sheet")
    args = ap.parse_args(argv)

    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

    if args.xlsx.lower().endswith(".xls"):
        sys.exit("这是旧版 .xls，openpyxl 读不了：用 Excel 另存为 .xlsx/.xlsm 再探，或直接把文件发我。")

    # 宏探测（zip 里有没有 vbaProject.bin）
    has_vba = False
    try:
        with zipfile.ZipFile(args.xlsx) as z:
            has_vba = any(n.lower().endswith("vbaproject.bin") for n in z.namelist())
    except Exception:
        pass

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)        # 值
    try:
        wbf = openpyxl.load_workbook(args.xlsx, data_only=False)  # 公式/批注
    except Exception:
        wbf = None

    report = {"file": args.xlsx, "has_vba_macro": has_vba,
              "sheetnames": wb.sheetnames,
              "defined_names": {}, "sheets": []}
    try:
        for name, dn in wb.defined_names.items():
            report["defined_names"][name] = s(dn.value, 300)
    except Exception:
        pass

    targets = [args.sheet] if args.sheet else wb.sheetnames
    for sn in targets:
        if sn not in wb.sheetnames:
            sys.exit("找不到 sheet: %s（有: %s）" % (sn, ", ".join(wb.sheetnames)))
        ws = wb[sn]
        if getattr(ws, "iter_rows", None) is None:   # chartsheet 等
            report["sheets"].append({"name": sn, "state": "non-worksheet", "skipped": True})
            continue
        wsf = wbf[sn] if (wbf is not None and sn in wbf.sheetnames) else None
        report["sheets"].append(probe_sheet(ws, wsf, args.rows, args.cols))

    # ---- 控制台摘要 ----
    print("文件: %s   宏(VBA): %s" % (args.xlsx, "有 ★宏逻辑需另行导出" if has_vba else "无"))
    print("sheet: %s" % ", ".join("%s%s" % (i["name"], "" if i.get("state") == "visible" else "[%s]" % i.get("state"))
                                   for i in report["sheets"]))
    if report["defined_names"]:
        print("defined names: %d 个（详见 JSON）" % len(report["defined_names"]))
    for i in report["sheets"]:
        if i.get("skipped"):
            continue
        print()
        print("== %s ==  %d 行 × %d 列  冻结=%s" % (i["name"], i["max_row"], i["max_col"], i["freeze_panes"]))
        if i["merged"]:
            print("  合并单元格 %d 处: %s%s" % (len(i["merged"]), ", ".join(i["merged"][:8]),
                                            " …" if len(i["merged"]) > 8 else ""))
        if i["hidden_cols"] or i["hidden_rows"]:
            print("  隐藏列: %s   隐藏行: %s" % (",".join(i["hidden_cols"]) or "-",
                                          ",".join(str(r) for r in i["hidden_rows"][:10]) or "-"))
        for dv in i["validations"]:
            print("  ★下拉/有效性 @%s type=%s: %s" % (dv.get("sqref"), dv.get("type"), dv.get("formula1")))
        for c in i["header_candidates"]:
            g = c.get("repeat_groups")
            print("  表头候选 r%d（非空 %d）: %s" % (c["r"], c["non_empty"], " | ".join(c["cells"][:20])))
            if g:
                print("    ★重复列组: %s  ← 一行多寄存器的排布证据" % ", ".join("%s×%d" % (k, v) for k, v in g.items()))
        if i.get("formula_count"):
            print("  公式 %d 个（JSON 存前 40）" % i["formula_count"])
        if i["comments"]:
            print("  批注 %d 条（详见 JSON）" % len(i["comments"]))
        for r in i["rows"][:8]:
            print("   r%-3d | %s" % (r["r"], " | ".join("%s:%s" % (k, v) for k, v in r["cells"].items())[:150]))
        if len(i["rows"]) > 8:
            print("   …顶部共 dump %d 行，全在 JSON 里" % len(i["rows"]))

    if args.json:
        with io.open(args.json, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=1)
        print()
        print("完整结构已写: %s   ★把这个 JSON 带回来" % args.json)
    else:
        print()
        print("提示：加 --json toolb_probe.json 落盘完整结构，把 JSON 带回来分析。")


if __name__ == "__main__":
    main()
