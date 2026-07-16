#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
probe_current_data.py — 电流数据探查脚本 v3（在数据机上运行，单文件，仅依赖 openpyxl）

对数据根目录做一次结构+数据快照，输出紧凑 JSON（同时生成 .gz 压缩版，优先传 .gz）：
  - 目录树 + 每个 xlsx 的 tab 列表（全递归，与层级无关）
  - 仿真长表：不认文件名，任何 tab 表头含 ID+Mode+Current（且有 simulation/Unit/Tier
    之一）即整页 dump
  - 实测 Result*.xlsx：按表头定位 NO./Mode/Current_mA/Temperature 列。
    列匹配带优先级：带单位后缀的 Current_mA 优先于裸 Current（测试项开关列），
    Temperature 开头优先于 Vtemp。只 dump 测量区的行（有电流值或 Init/Lock/OFF/chamber 行），
    测试计划行不要。表头模板跨文件去重。

用法：
  python probe_current_data.py <数据根目录>
  # 生成 <根目录>\probe_dump.json 和 probe_dump.json.gz，拷 .gz 回开发机

不改动任何原文件，只读。
"""
import argparse
import datetime
import fnmatch
import gzip
import json
import os
import re
import sys

import openpyxl

MAX_ROWS = 2000    # 每个 sheet 最多 dump 的数据行数
SCAN_ROWS = 30     # 找表头时扫描的行数
MAX_SAMPLE_CELLS = 120
MAX_CELL_STR = 100
# 本套工具自己的输出，扫描时跳过（防自吞：current_db 导出簿的 Sim_long 页长得像仿真表）
EXCLUDE_GLOBS = ["Current_compare_pivot*.xlsx", "probe_dump*", "current.db"]


def norm(v):
    return str(v).strip().lower() if v is not None else ""


def cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def jval(v):
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return str(v)
    if isinstance(v, str) and len(v) > MAX_CELL_STR:
        return v[:MAX_CELL_STR] + "…"
    return v


def match_sim_header(row):
    """仿真长表表头：ID + Mode + Current*，且有 simulation/Unit/Tier 佐证。"""
    names = [norm(c) for c in row]
    return ("id" in names and "mode" in names
            and any(n.startswith("current") for n in names)
            and (any(n.startswith("simulation") for n in names)
                 or "unit" in names))  # 注意: 不认 tier——本工具导出的 Sim_long 页有 tier 列


def match_result_header(row):
    """实测表表头。带优先级：Current_mA(带单位) > Current(裸)；Temperature* > temp。
    返回列映射 dict（含 unit）或 None。"""
    no_col = mode_col = None
    cur_exact = cur_bare = temp_exact = temp_bare = None
    for j, c in enumerate(row):
        n = norm(c)
        if n in ("no.", "no", "no．") and no_col is None:
            no_col = j
        elif n == "mode" and mode_col is None:
            mode_col = j
        elif re.fullmatch(r"current[_\s]*[munµμ]?a", n) and cur_exact is None:
            cur_exact = j
        elif n == "current" and cur_bare is None:
            cur_bare = j
        elif n.startswith("temperature") and temp_exact is None:
            temp_exact = j
        elif n == "temp" and temp_bare is None:
            temp_bare = j
    cur_col = cur_exact if cur_exact is not None else cur_bare
    temp_col = temp_exact if temp_exact is not None else temp_bare
    if no_col is None or cur_col is None:
        return None
    if mode_col is None:
        mode_col = no_col + 1
    unit = "ma"
    m = re.fullmatch(r"current[_\s]*([munµμ]?a)", norm(row[cur_col]))
    if m and m.group(1):
        unit = m.group(1)
    return {"no": no_col, "mode": mode_col, "cur": cur_col, "temp": temp_col, "unit": unit}


def is_meas_row(no_raw, label, cur):
    """测量区行：有电流值，或是序列结构行（Init/Lock/OFF/chamber）。"""
    if cur is not None:
        return True
    ln, nn = norm(label), norm(no_raw)
    return (ln.startswith(("init", "lock", "off"))
            or "chamber" in ln or "chamber" in nn)


def scan_workbook(path):
    """扫一个工作簿：tab 列表 + 每个 tab 是否命中 仿真/实测 表头。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        info = {"sheets": list(wb.sheetnames), "sim_sheets": [], "result_sheets": {}}
        for sn in wb.sheetnames:
            ws = wb[sn]
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=SCAN_ROWS,
                                                 values_only=True), 1):
                if match_sim_header(row):
                    info["sim_sheets"].append({"sheet": sn, "header_row": i})
                    break
                cols = match_result_header(row)
                if cols is not None:
                    info["result_sheets"][sn] = {
                        "header_row": i,
                        "key_cols_1based": {k: (v + 1 if isinstance(v, int) else v)
                                            for k, v in cols.items()}}
                    break
        return info, None
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"
    finally:
        try:
            wb.close()
        except Exception:
            pass


def dump_sheet_rows(path, sheet):
    """整页 dump（跳过全空行，密集数组）。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows = []
        for r in ws.iter_rows(values_only=True):
            if all(v is None for v in r):
                continue
            rows.append([jval(v) for v in r])
            if len(rows) > MAX_ROWS:
                break
        return rows
    finally:
        wb.close()


def dump_result(path, sheet, hdr_idx, cols_1b):
    """按已定位的表头 dump 实测关键列（只留测量区行）+ 1 行全列样本。"""
    cols = {k: (v - 1 if isinstance(v, int) else v) for k, v in cols_1b.items()}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        hdr_row = None
        rows, samples, n_skipped = [], [], 0
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i == hdr_idx:
                hdr_row = row
                continue
            if i < hdr_idx:
                continue
            no_raw = cell(row, cols["no"])
            label = cell(row, cols["mode"])
            cur = cell(row, cols["cur"])
            temp = cell(row, cols["temp"]) if cols["temp"] is not None else None
            if no_raw is None and label is None and cur is None:
                continue
            if not is_meas_row(no_raw, label, cur):
                n_skipped += 1
                continue
            rows.append([i, jval(no_raw), jval(label), jval(cur), jval(temp)])
            if not samples and cur is not None:
                cells = [[j + 1,
                          str(cell(hdr_row, j)) if cell(hdr_row, j) is not None else "",
                          jval(v)] for j, v in enumerate(row) if v is not None]
                samples.append(cells[:MAX_SAMPLE_CELLS])
            if len(rows) >= MAX_ROWS:
                break
        headers = [[j + 1, str(v)] for j, v in enumerate(hdr_row or []) if v not in (None, "")]
        return {"sheet": sheet, "header_row": hdr_idx,
                "key_cols_1based": cols_1b, "headers": headers,
                "row_fields": ["row_idx", "no_raw", "mode_label", "current", "temp"],
                "rows": rows, "rows_skipped_non_meas": n_skipped,
                "sample_rows_full": samples}
    finally:
        wb.close()


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    ap = argparse.ArgumentParser(description="电流数据探查 v3：紧凑 JSON 快照（全递归）")
    ap.add_argument("root", help="数据根目录")
    ap.add_argument("-o", "--out", help="输出 JSON 路径（默认 <root>\\probe_dump.json）")
    ap.add_argument("--result-glob", default="Result*.xlsx", help="实测文件名通配符")
    args = ap.parse_args()

    root = os.path.abspath(args.root)
    if not os.path.isdir(root):
        raise SystemExit(f"[错误] 目录不存在: {root}")
    out_path = args.out or os.path.join(root, "probe_dump.json")

    dump = {"probe_version": 3,
            "probed_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "root": root, "tree": [], "workbooks": {}, "detected": {},
            "header_profiles": [], "sim": [], "results": []}

    def profile_id(headers):
        for i, p in enumerate(dump["header_profiles"]):
            if p == headers:
                return i
        dump["header_profiles"].append(headers)
        return len(dump["header_profiles"]) - 1

    xlsx_files = []
    for dirpath, dirnames, filenames in os.walk(root):
        rel_dir = os.path.relpath(dirpath, root)
        for f in filenames:
            full = os.path.join(dirpath, f)
            rel = os.path.normpath(os.path.join(rel_dir, f)) if rel_dir != "." else f
            try:
                st = os.stat(full)
                dump["tree"].append({"path": rel, "size": st.st_size,
                                     "mtime": datetime.datetime.fromtimestamp(st.st_mtime)
                                     .strftime("%Y-%m-%d %H:%M:%S")})
            except OSError:
                continue
            if (f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
                    and not any(fnmatch.fnmatch(f, pat) for pat in EXCLUDE_GLOBS)):
                xlsx_files.append((rel, full))

    for rel, full in xlsx_files:
        info, err = scan_workbook(full)
        if info is None:
            dump["workbooks"][rel] = {"error": err}
            print(f"[跳过] {rel}: {err}")
            continue
        dump["workbooks"][rel] = info["sheets"]
        is_result_file = fnmatch.fnmatch(os.path.basename(full), args.result_glob)

        for s in info["sim_sheets"]:
            try:
                rows = dump_sheet_rows(full, s["sheet"])
                dump["sim"].append({"file": rel, "sheet": s["sheet"], "rows": rows,
                                    "truncated": len(rows) > MAX_ROWS})
                print(f"[仿真] {rel} / {s['sheet']} -> {len(rows)} 行")
            except Exception as e:
                dump["sim"].append({"file": rel, "sheet": s["sheet"],
                                    "error": f"{type(e).__name__}: {e}"})

        if is_result_file and info["result_sheets"]:
            for sn, meta in info["result_sheets"].items():
                try:
                    r = dump_result(full, sn, meta["header_row"], meta["key_cols_1based"])
                    r["header_profile"] = profile_id(r.pop("headers"))
                    r["file"] = os.path.basename(rel)
                    r["folder"] = os.path.dirname(rel)
                    dump["results"].append(r)
                    print(f"[实测] {rel} / {sn} -> 测量区 {len(r['rows'])} 行"
                          f"（滤掉计划行 {r['rows_skipped_non_meas']}）"
                          f" 关键列={r['key_cols_1based']}")
                except Exception as e:
                    dump["results"].append({"file": os.path.basename(rel),
                                            "folder": os.path.dirname(rel),
                                            "error": f"{type(e).__name__}: {e}"})
        elif is_result_file:
            dump["results"].append({"file": os.path.basename(rel),
                                    "folder": os.path.dirname(rel),
                                    "error": "找不到含 NO./Current 表头的 sheet",
                                    "sheets": info["sheets"]})
            print(f"[实测] {rel} -> 没找到表头！tab={info['sheets']}")
        elif info["result_sheets"]:
            dump["detected"][rel] = {"result_like_sheets": info["result_sheets"]}

    if not dump["sim"]:
        print("\n[警告] 没有发现仿真长表（表头需含 ID/Mode/Current + simulation|Unit|Tier）！")
        print("       请把仿真数据工作簿放进这个目录（文件名任意）后重跑。")

    blob = json.dumps(dump, ensure_ascii=False, separators=(",", ":"), default=str)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(blob)
    gz_path = out_path + ".gz"
    with gzip.open(gz_path, "wt", encoding="utf-8") as f:
        f.write(blob)
    print(f"\n[完成] {out_path}（{os.path.getsize(out_path)/1024:.0f} KB）")
    print(f"       {gz_path}（{os.path.getsize(gz_path)/1024:.0f} KB）<- 优先传这个")
    print(f"       仿真表 {len(dump['sim'])} 张 / 实测文件 {len(dump['results'])} 个")


if __name__ == "__main__":
    main()
