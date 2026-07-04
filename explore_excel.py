#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
explore_excel.py — 探查一个 Excel 文件的结构

用途：
    在不知道寄存器表 / 控制信号表长什么样的情况下，先把它的结构摸清楚：
    有几个 sheet、每个 sheet 多大、表头在哪几行、有没有合并单元格、
    每一列大概是什么类型的数据（十六进制地址？bit 位？信号名？）。
    这样我们就能在本地复刻一份等价的 Excel 来跑后续任务。

依赖：
    只用 openpyxl（不需要 pandas）。
        pip install openpyxl

用法：
    python explore_excel.py <path-to-xlsx>
    python explore_excel.py <path-to-xlsx> --rows 30        # 每个 sheet 预览多少行 (默认 20)
    python explore_excel.py <path-to-xlsx> --dump summary.json  # 把完整单元格矩阵导出成 JSON，便于本地复刻
    python explore_excel.py <path-to-xlsx> --sheet "Reg Map"   # 只看某一个 sheet

注意：
    .xlsx / .xlsm 用 openpyxl 直接读（.xlsm 的宏不影响读数据）；
    老的 .xls（BIFF）openpyxl 读不了，脚本会提示你先另存为 .xlsx / .xlsm。

    .xlsm 常见坑：若某些值是宏/公式算出来的，默认读的是 Excel 缓存的计算结果；
    如果该文件从没被 Excel 打开保存过，这些格子会读成空。这时加 --formulas
    改看公式原文（如 =A1+1），先弄清结构。
"""

import argparse
import json
import os
import sys

# Windows 控制台默认代码页常常不是 UTF-8，中文会乱码；强制 UTF-8 输出。
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


def col_letter(idx):
    """1 -> A, 27 -> AA"""
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def guess_type(values):
    """根据一列的采样值粗略猜测类型，方便识别 地址/bit/信号名 等。"""
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_empty:
        return "empty"
    kinds = set()
    for v in non_empty:
        if isinstance(v, bool):
            kinds.add("bool")
        elif isinstance(v, int):
            kinds.add("int")
        elif isinstance(v, float):
            kinds.add("float")
        else:
            s = str(v).strip()
            sl = s.lower()
            if sl.startswith("0x") or sl.startswith("0b"):
                kinds.add("hex/bin-str")
            elif s.startswith("'") or "'h" in sl or "'b" in sl or "'d" in sl:
                kinds.add("verilog-num")  # 例如 8'hFF
            else:
                # 纯数字字符串？
                try:
                    int(s)
                    kinds.add("int-str")
                except ValueError:
                    try:
                        float(s)
                        kinds.add("float-str")
                    except ValueError:
                        kinds.add("text")
    if len(kinds) == 1:
        return next(iter(kinds))
    return "/".join(sorted(kinds))


def preview_value(v, width=18):
    if v is None:
        s = ""
    else:
        s = str(v).replace("\n", "\\n").replace("\t", " ")
    if len(s) > width:
        s = s[: width - 1] + "…"
    return s.ljust(width)


def explore_sheet(ws, max_preview_rows, dump_full):
    print("=" * 100)
    print(f"[SHEET] {ws.title!r}")
    print(f"  尺寸(dimensions) : {ws.dimensions}")
    print(f"  最大行 x 最大列   : {ws.max_row} x {ws.max_column}  "
          f"({ws.max_column} 列 = A..{col_letter(ws.max_column)})")
    print(f"  冻结窗格(freeze)  : {ws.freeze_panes}")

    # --- 合并单元格：寄存器表里经常用它把一个寄存器名跨多个 bit 行 ---
    merged = [str(r) for r in ws.merged_cells.ranges]
    if merged:
        print(f"  合并单元格 {len(merged)} 处 : "
              + ", ".join(merged[:15]) + (" ..." if len(merged) > 15 else ""))
    else:
        print("  合并单元格        : 无")

    # 读全部单元格值（用 values_only 快）
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        print("  (空 sheet)")
        return {"title": ws.title, "rows": []}

    ncols = ws.max_column

    # --- 预览前若干行 ---
    print(f"\n  前 {min(max_preview_rows, len(all_rows))} 行预览：")
    header = "   r\\c |" + "|".join(
        col_letter(c + 1).center(18) for c in range(min(ncols, 12))
    )
    print("  " + header + (" ..." if ncols > 12 else ""))
    print("  " + "-" * len(header))
    for ridx, row in enumerate(all_rows[:max_preview_rows], start=1):
        cells = "|".join(
            preview_value(row[c] if c < len(row) else None)
            for c in range(min(ncols, 12))
        )
        print(f"  {ridx:>4} |{cells}" + (" ..." if ncols > 12 else ""))

    # --- 每列类型猜测（用前 200 行采样）---
    print("\n  每列数据类型猜测（采样前 200 行）：")
    sample = all_rows[:200]
    for c in range(ncols):
        col_vals = [r[c] if c < len(r) else None for r in sample]
        t = guess_type(col_vals)
        # 找该列第一个非空值当例子
        example = next((str(v) for v in col_vals if v not in (None, "")), "")
        if len(example) > 30:
            example = example[:29] + "…"
        print(f"    {col_letter(c+1):>3} : {t:<16} 例: {example}")

    sheet_dump = {
        "title": ws.title,
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cells": merged,
    }
    if dump_full:
        # 完整矩阵（值），便于本地复刻。datetime 等转成字符串。
        def norm(v):
            if v is None or isinstance(v, (int, float, bool, str)):
                return v
            return str(v)
        sheet_dump["rows"] = [[norm(v) for v in row] for row in all_rows]
    return sheet_dump


def main():
    ap = argparse.ArgumentParser(description="探查 Excel 结构")
    ap.add_argument("path", help="xlsx / xlsm 文件路径")
    ap.add_argument("--rows", type=int, default=20, help="每个 sheet 预览多少行 (默认 20)")
    ap.add_argument("--sheet", default=None, help="只看指定名字的 sheet")
    ap.add_argument("--dump", default=None, help="把完整内容导出到该 JSON 文件")
    ap.add_argument("--formulas", action="store_true",
                    help="读公式原文而非缓存计算值（宏/公式算出的值读成空时用）")
    args = ap.parse_args()

    if not os.path.isfile(args.path):
        sys.exit(f"找不到文件: {args.path}")

    if args.path.lower().endswith(".xls"):
        sys.exit("openpyxl 读不了老的 .xls 格式，请先在 Excel 里另存为 .xlsx / .xlsm 再试。")

    try:
        import openpyxl
    except ImportError:
        sys.exit("缺少 openpyxl，请先: pip install openpyxl")

    print(f"打开: {args.path}")
    # data_only=True -> 读公式计算后的缓存值；--formulas 时读公式原文。
    # .xlsm 直接这样读即可，宏(VBA)不影响取数据。
    wb = openpyxl.load_workbook(args.path, read_only=False,
                                data_only=not args.formulas)
    print(f"共 {len(wb.sheetnames)} 个 sheet: {wb.sheetnames}\n")

    dump = {"file": os.path.abspath(args.path), "sheets": []}
    targets = [args.sheet] if args.sheet else wb.sheetnames
    for name in targets:
        if name not in wb.sheetnames:
            print(f"!! 没有名为 {name!r} 的 sheet，跳过")
            continue
        ws = wb[name]
        dump["sheets"].append(
            explore_sheet(ws, args.rows, dump_full=bool(args.dump))
        )

    if args.dump:
        with open(args.dump, "w", encoding="utf-8") as f:
            json.dump(dump, f, ensure_ascii=False, indent=2)
        print("\n" + "=" * 100)
        print(f"完整内容已导出到: {args.dump}")


if __name__ == "__main__":
    main()
