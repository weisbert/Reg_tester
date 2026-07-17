#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_all_mode_testitem.py — 合并生成「全模式电流测量 testitem」。

输入两个目录：
  00_Mode_Reg       各模式寄存器写序 Excel（算跨模式差异 → 签名块）
  00_Mode_testitem  各模式已写好的工具B流程表（init/lock/OFF 行 **原样保留**）

输出一份工具B流程表：
  [模式i 签名行(自动生成, 2行) → 模式i 原 testitem 全部数据行(逐格原样拷贝)] × N
  → 末尾统一一行 Chamber_Close（各文件里的 Chamber_Close 行被摘除去重）

签名行 = 差异地址并集 × 该模式末态值（自包含，模式顺序可自由排），
克隆该模式首行的其余参数（频率/SSA/ReadBack 配置），测量开关全 NO（Test=YES）。

用法（在 Analyzer\\excel 目录下）：
    python gen_all_mode_testitem.py 00_Mode_Reg 00_Mode_testitem \
        [--temps 25,105,-40] \
        [--out Current_all_mode_testitem_gen.xlsx] [--order 子串,子串,...] \
        [--template 某testitem子串] [--map "testitem子串=reg子串" ...] [--force] [--overwrite]

  --temps     温度点序列（℃，逗号分隔，按给定顺序执行）。给了就生成全温度流程：
              每个温度段开头插一行 Chamber=YES 的设温行（SET_TEMP_x，只控温不测量），
              段内=完整的全模式链，且所有行 Temperature 列改写为该温度（探查/入库
              按这列区分温度）；若末温≠首温，结尾自动加回首温的设温行再 Chamber_Close
              （避免低温开箱结霜）。不给 --temps 维持单温度旧行为。

  --order     模式顺序（testitem 文件名子串，默认按文件名排序）
  --template  用哪个 testitem 当输出模板（表头/说明sheet/格式，默认顺序第一个）
  --map       手工指定 testitem↔寄存器表 配对（自动 token 匹配失败时用，可多次）
  --force     各 testitem 表头与模板不一致时仍按列位置拷贝
  --overwrite 输出文件已存在时覆盖

注意：输出写的是“值”（FCW 等公式列取其算好的缓存值），不带公式——
要改频点请回单模式 testitem 改完重新合并。依赖 openpyxl；需与 diff_mode_regs.py 同目录。
"""
import argparse
import io
import os
import re
import sys
from collections import OrderedDict

from diff_mode_regs import parse_mode, last_state, _history

FLAG_COLS = ["Test", "IPN", "SpotPN", "ReadBack", "OtherSpur", "Vtune", "Vtemp",
             "Current", "PNTrace", "SpurList", "Chamber"]


# ---------- 模式名 token 匹配（testitem ↔ 寄存器表） ----------

AXES = ("radio", "band", "dir", "dco", "sync", "loop")


def features(name):
    """文件名 → 特征轴 {radio, band, dir, dco, sync, loop}。缺失轴=None。"""
    s = os.path.splitext(os.path.basename(name))[0].lower()
    f = {}
    if "wifi" in s or re.search(r"(^|[^a-z])w[25]g", s):
        f["radio"] = "wifi"
    elif "bt" in s:
        f["radio"] = "bt"
    if "dco2g" in s:
        f["dco"] = "2g"
    elif "dco5g" in s:
        f["dco"] = "5g"
    if "unsync" in s:
        f["sync"] = "unsync"
    elif "sync" in s:
        f["sync"] = "sync"
    if "loopback" in s or re.search(r"(^|[^a-z])loop($|[^a-z])", s):
        f["loop"] = "y"
    t = re.sub(r"dco[25]g", "", s)
    if "2g" in t:
        f["band"] = "2g"
    elif "5g" in t:
        f["band"] = "5g"
    if "rx" in t:
        f["dir"] = "rx"
    elif "tx" in t:
        f["dir"] = "tx"
    return f


def auto_match(titem, regs, overrides):
    """testitem 文件 → 寄存器表文件。冲突轴淘汰，交集计分，歧义即退出。"""
    for pat_t, pat_r in overrides:
        if pat_t.lower() in os.path.basename(titem).lower():
            hit = [r for r in regs if pat_r.lower() in os.path.basename(r).lower()]
            if len(hit) != 1:
                sys.exit("--map '%s=%s' 在寄存器表里匹配到 %d 个" % (pat_t, pat_r, len(hit)))
            return hit[0]
    ft = features(titem)
    scored = []
    for r in regs:
        fr = features(r)
        if any(ft.get(k) and fr.get(k) and ft[k] != fr[k] for k in AXES):
            continue                                    # 任一轴冲突 → 淘汰
        score = sum(1 for k in ft if fr.get(k) == ft[k])
        scored.append((score, r))
    scored.sort(key=lambda x: -x[0])
    if not scored or (len(scored) > 1 and scored[0][0] == scored[1][0]):
        sys.exit("testitem '%s' 自动配对失败/歧义（候选: %s）——用 --map \"%s子串=寄存器表子串\" 指定"
                 % (os.path.basename(titem), [os.path.basename(r) for _, r in scored[:3]],
                    os.path.basename(titem)[:8]))
    return scored[0][1]


# ---------- 工具B流程表读取 ----------

def find_flow_sheet(wb):
    """找含 'Test Item' + 'REG ADDR1' 表头的 sheet → (ws, 表头行号, 表头list)。"""
    for ws in wb.worksheets:
        for r, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            cells = [("" if c is None else str(c).strip()) for c in row]
            low = [c.lower() for c in cells]
            if "test item" in low and "reg addr1" in low:
                return ws, r, cells
    return None, None, None


def col_index(header, name):
    """表头精确匹配（strip 后）→ 0-based 列号；找不到 None。"""
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def read_rows(path):
    """testitem → (表头list, 数据行list[list], chamber行list[list])。值读取(data_only)。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws, hr, header = find_flow_sheet(wb)
    if ws is None:
        sys.exit("%s 里找不到流程 sheet（表头需含 Test Item 和 REG ADDR1）" % path)
    icol = col_index(header, "Test Item")
    data, chamber = [], []
    for row in ws.iter_rows(min_row=hr + 1, values_only=True):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        ti = str(vals[icol]).strip() if icol is not None and icol < len(vals) and vals[icol] is not None else ""
        (chamber if ti.lower() == "chamber_close" else data).append(vals)
    wb.close()
    return header, data, chamber


# ---------- 主流程 ----------

def main(argv=None):
    ap = argparse.ArgumentParser(description="合并生成全模式电流测量 testitem（保留各模式原有 lock/OFF 行）")
    ap.add_argument("reg_dir", help="00_Mode_Reg 目录")
    ap.add_argument("titem_dir", help="00_Mode_testitem 目录")
    ap.add_argument("--out", default="Current_all_mode_testitem_gen.xlsx")
    ap.add_argument("--order", help="模式顺序：testitem 文件名子串，逗号分隔")
    ap.add_argument("--template", help="当模板的 testitem 文件名子串（默认顺序第一个）")
    ap.add_argument("--temps", help="温度点序列（℃，逗号分隔，如 25,105,-40）")
    ap.add_argument("--map", action="append", default=[], metavar="T子串=R子串",
                    help="手工指定 testitem↔寄存器表 配对，可多次")
    ap.add_argument("--force", action="store_true", help="表头不一致仍按列位置拷贝")
    ap.add_argument("--overwrite", action="store_true")
    args = ap.parse_args(argv)

    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

    if os.path.exists(args.out) and not args.overwrite:
        sys.exit("输出文件已存在：%s（换 --out 或加 --overwrite）" % args.out)

    def xlsx_in(d):
        return sorted(os.path.join(d, f) for f in os.listdir(d)
                      if f.lower().endswith(".xlsx") and not f.startswith("~$"))

    regs = xlsx_in(args.reg_dir)
    titems = xlsx_in(args.titem_dir)
    if not regs or not titems:
        sys.exit("目录为空：%s(%d) / %s(%d)" % (args.reg_dir, len(regs), args.titem_dir, len(titems)))

    if args.order:
        ordered = []
        for s in [x.strip() for x in args.order.split(",") if x.strip()]:
            hit = [t for t in titems if s.lower() in os.path.basename(t).lower()]
            if len(hit) != 1:
                sys.exit("--order '%s' 匹配到 %d 个 testitem" % (s, len(hit)))
            if hit[0] in ordered:
                sys.exit("--order '%s' 重复" % s)
            ordered.append(hit[0])
        titems = ordered

    overrides = []
    for m in args.map:
        if "=" not in m:
            sys.exit("--map 格式：testitem子串=寄存器表子串")
        overrides.append(tuple(m.split("=", 1)))

    # 配对 + 打印（人工目检的关卡）
    pairing = OrderedDict((t, auto_match(t, regs, overrides)) for t in titems)
    print("== 模式配对（务必目检）==")
    for t, r in pairing.items():
        print("  %-42s ←→ %s" % (os.path.basename(t), os.path.basename(r)))

    # 差异签名：解析所有配对到的寄存器表
    parsed = OrderedDict()
    for r in pairing.values():
        if r not in parsed:
            m = parse_mode(r)
            parsed[r] = {"state": last_state(m["writes"]),
                         "multi": set(a for a, hs in _history(m["writes"]).items() if len(hs) > 1)}
    union = OrderedDict()
    for r in parsed:
        for a, w in parsed[r]["state"].items():
            union.setdefault(a, w["name"])
    sig_addrs = []
    for a in union:
        ws_ = [parsed[r]["state"].get(a) for r in parsed]
        vn = set(w["vnorm"] for w in ws_ if w)
        if len(vn) > 1 or any(w is None for w in ws_):
            sig_addrs.append(a)
    bad = [a for a in sig_addrs if any(a in parsed[r]["multi"] for r in parsed)]
    if bad:
        sys.exit("✋ 差异地址撞同址多写（过程语义，不能折叠进签名）：%s ——需人工处理" % bad)
    holes = [(a, os.path.basename(r)) for a in sig_addrs for r in parsed if a not in parsed[r]["state"]]
    if holes:
        sys.exit("✋ 以下 (地址,寄存器表) 没有末态值：%s" % holes)
    print("差异地址 %d 个（签名块），零撞多写" % len(sig_addrs))

    # 模板
    tpl_path = titems[0]
    if args.template:
        hit = [t for t in titems if args.template.lower() in os.path.basename(t).lower()]
        if len(hit) != 1:
            sys.exit("--template '%s' 匹配到 %d 个" % (args.template, len(hit)))
        tpl_path = hit[0]

    import openpyxl
    out_wb = openpyxl.load_workbook(tpl_path, data_only=True)
    out_ws, hr, tpl_header = find_flow_sheet(out_wb)
    if out_ws is None:
        sys.exit("模板 %s 里找不到流程 sheet" % tpl_path)
    npairs = sum(1 for h in tpl_header if re.fullmatch(r"REG ADDR\d+", h))
    if npairs == 0:
        sys.exit("模板表头里没找到 REG ADDRn 列")
    c_no, c_mode = col_index(tpl_header, "NO."), col_index(tpl_header, "Mode")
    c_pairs = [(col_index(tpl_header, "REG ADDR%d" % (k + 1)),
                col_index(tpl_header, "REG Value%d" % (k + 1))) for k in range(npairs)]
    if any(a is None or v is None for a, v in c_pairs):
        sys.exit("REG ADDR/Value 列不成对")
    c_flags = {n: col_index(tpl_header, n) for n in FLAG_COLS}

    if out_ws.max_row > hr:
        out_ws.delete_rows(hr + 1, out_ws.max_row - hr)

    # 温度点解析
    c_temp = next((i for i, h in enumerate(tpl_header) if h.startswith("Temperature")), None)
    temps = None
    if args.temps:
        temps = []
        for x in args.temps.split(","):
            x = x.strip()
            if x:
                temps.append(float(x) if "." in x else int(x))
        if not temps:
            sys.exit("--temps 是空的")
        if c_temp is None:
            sys.exit("模板表头没找到 Temperature 列，做不了多温度")

    # 收集各模式块（签名行 + 原行），只读一遍文件
    n_sig_rows = -(-len(sig_addrs) // npairs)
    last_chamber = None
    blocks = []
    tempset_base = None
    chamber_yes_in = []
    print()
    for t in titems:
        header, data, chamber = read_rows(t)
        if chamber:
            last_chamber = chamber[-1]
        mism = [(i + 1, a, b) for i, (a, b) in enumerate(
                zip(tpl_header, header + [""] * (len(tpl_header) - len(header))))
                if a != ("" if b is None else str(b).strip())]
        if mism and not args.force:
            print("✋ %s 表头与模板不一致（前5处）：" % os.path.basename(t))
            for i, a, b in mism[:5]:
                print("   列%d 模板='%s' 此表='%s'" % (i, a, b))
            sys.exit("列位置对不上不能按位置拷贝——确认无碍就加 --force")
        if not data:
            sys.exit("%s 没有数据行" % t)

        label = re.sub(r"_?(current)?_?test_?item", "", os.path.splitext(os.path.basename(t))[0],
                       flags=re.I).strip("_")
        state = parsed[pairing[t]]["state"]
        base = list(data[0]) + [None] * max(0, len(tpl_header) - len(data[0]))
        if tempset_base is None:
            tempset_base = list(base)
        rows = []
        for i in range(n_sig_rows):
            row = list(base)
            if c_no is not None:
                row[c_no] = "SWITCH"
            if c_mode is not None:
                row[c_mode] = "%s_sig%d" % (label, i + 1)
            for name, ci in c_flags.items():
                if ci is not None:
                    row[ci] = "YES" if name == "Test" else "NO"
            chunk = sig_addrs[i * npairs:(i + 1) * npairs]
            for k, (ca, cv) in enumerate(c_pairs):
                if k < len(chunk):
                    w = state[chunk[k]]
                    v = str(w["value"]).strip()
                    row[ca] = chunk[k]
                    row[cv] = v if v.lower().startswith("0x") else "0x" + v
                else:
                    row[ca] = row[cv] = None
            rows.append(row)
        ci = c_flags.get("Chamber")
        for row in data:
            rows.append(list(row))
            if (ci is not None and ci < len(row) and row[ci] is not None
                    and str(row[ci]).strip().upper() == "YES"):
                chamber_yes_in.append(os.path.basename(t))
        blocks.append(rows)
        print("  %-42s 签名 %d 行 + 原行 %d 行（Chamber_Close 摘除 %d 行）"
              % (os.path.basename(t), n_sig_rows, len(data), len(chamber)))
    if chamber_yes_in:
        print("  ⚠ 以下 testitem 的原行里已有 Chamber=YES，会和设温行叠加，请自查：%s"
              % sorted(set(chamber_yes_in)))

    def make_tempset(T):
        """设温行：只控温不写不测。Chamber=YES + Temperature=T。"""
        row = list(tempset_base)
        if c_no is not None:
            row[c_no] = "CHAMBER"
        if c_mode is not None:
            row[c_mode] = "SET_TEMP_%s" % T
        for name, ci2 in c_flags.items():
            if ci2 is not None:
                row[ci2] = "YES" if name in ("Test", "Chamber") else "NO"
        for ca, cv in c_pairs:
            row[ca] = row[cv] = None
        row[c_temp] = T
        return row

    # 拼装输出：多温度 = [设温行 + 全模式链] × N + 回首温 + Chamber_Close
    total = 0
    for T in (temps if temps else [None]):
        if T is not None:
            out_ws.append(make_tempset(T))
            total += 1
        for rows in blocks:
            for r in rows:
                rr = list(r)
                if T is not None:
                    while len(rr) <= c_temp:
                        rr.append(None)
                    rr[c_temp] = T
                out_ws.append(rr)
                total += 1
    if temps and temps[-1] != temps[0]:
        out_ws.append(make_tempset(temps[0]))
        total += 1
        print("  末温 %s ≠ 首温 %s：已加回温行（避免低温开箱结霜）" % (temps[-1], temps[0]))

    if last_chamber is not None:
        lc = list(last_chamber)
        if temps:
            while len(lc) <= c_temp:
                lc.append(None)
            lc[c_temp] = temps[0]
        out_ws.append(lc)
        total += 1
        print("  末尾追加 Chamber_Close ×1")
    else:
        print("  ⚠ 所有 testitem 里都没找到 Chamber_Close 行，末尾未加关箱步骤")

    # 数据有效性范围顺延到新表尾
    last_row = hr + total
    try:
        for dv in out_ws.data_validations.dataValidation:
            parts = []
            for p in str(dv.sqref).split():
                if ":" in p:
                    s, e = p.split(":")
                    parts.append(s + ":" + re.sub(r"\d+$", str(last_row), e))
                else:
                    parts.append(p)
            dv.sqref = " ".join(parts)
    except Exception as e:
        print("  ⚠ 下拉框范围顺延失败（不影响执行，只影响编辑）：%r" % e)

    out_wb.save(args.out)
    print()
    print("✔ 已生成 %s：%d 个模式 × %s，数据行共 %d 行（表头沿用 %s）"
          % (args.out, len(titems),
             ("%d 个温度点 %s" % (len(temps), temps)) if temps else "单温度",
             total, os.path.basename(tpl_path)))
    print("  结构：%s每模式 [签名 %d 行(测量全NO) → 原 init/lock/OFF 行原样] → 末尾 Chamber_Close"
          % ("每温度段 [SET_TEMP 设温行(Chamber=YES) → 全模式链(Temperature 列已改写)] → 回首温 → " if temps else "",
             n_sig_rows))
    print("  ⚠ 输出为纯值（公式列已取算好的值）；上温箱前先常温跑通整链：逐模式判锁 + 电流对表。")


if __name__ == "__main__":
    main()
