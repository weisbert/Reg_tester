#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""gen_signature_rows.py — 从 diff_mode_regs 的交接 JSON 生成各模式「签名块」。

签名块 = 差异地址并集 × 该模式的末态值，自包含（不依赖前一个模式是谁，
模式顺序可自由排列）。输出为工具B流程表的 REG 对布局（地址/值交替），
每 N 对一行，直接整块复制粘贴到流程表的 REG ADDR1 列起始处。

⚠ 签名块只覆盖静态配置差异；每个模式仍需自己的 init+lock 行
  （顺序敏感的分级使能/校准触发在那里处理），锁后用 ReadBack 判锁。

用法：
    python gen_signature_rows.py [交接.json] [--out 签名块.xlsx] \
        [--order 模式1,模式2,...] [--pairs 11] [--no-0x]

  位置参数   diff_mode_regs 生成的交接 JSON（默认 模式差异_交接.json）
  --out     输出 Excel（默认 签名块.xlsx；每模式一段，纯 地址/值 矩形块）
  --order   模式顺序，逗号分隔，支持子串匹配（默认按交接文件顺序）
  --pairs   每行 REG 对数（默认 11，与工具B模板一致）
  --no-0x   值不加 0x 前缀（默认加：地址裸 hex、值 0x 前缀，同平台填写建议）

依赖 openpyxl。
"""
import argparse
import io
import json
import sys


def resolve_order(all_modes, spec):
    """--order 子串匹配解析，歧义/无匹配即退出。"""
    out = []
    for s in [x.strip() for x in spec.split(",") if x.strip()]:
        hit = [m for m in all_modes if m == s] or \
              [m for m in all_modes if s.lower() in m.lower()]
        if len(hit) != 1:
            sys.exit("模式 '%s' 匹配到 %d 个：%s" % (s, len(hit), hit or all_modes))
        if hit[0] in out:
            sys.exit("模式 '%s' 重复" % hit[0])
        out.append(hit[0])
    return out


def main(argv=None):
    ap = argparse.ArgumentParser(description="从差异交接 JSON 生成各模式签名块（工具B REG 对布局）")
    ap.add_argument("handoff", nargs="?", default="模式差异_交接.json")
    ap.add_argument("--out", default="签名块.xlsx")
    ap.add_argument("--order", help="模式顺序，逗号分隔，支持子串匹配")
    ap.add_argument("--pairs", type=int, default=11, help="每行 REG 对数（默认 11）")
    ap.add_argument("--no-0x", dest="no0x", action="store_true", help="值不加 0x 前缀")
    args = ap.parse_args(argv)

    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

    data = json.load(io.open(args.handoff, encoding="utf-8"))
    modes = resolve_order(data["modes"], args.order) if args.order else list(data["modes"])
    diff = data["diff_list"]
    if not diff:
        sys.exit("交接文件里 diff_list 为空，没有差异可生成。")

    # 安全闸 1：撞多写的差异地址不能折叠进签名块
    coll = [d for d in diff if d.get("multi_write_in")]
    if coll:
        print("✋ 以下差异地址在部分模式里是同址多写（过程语义），已拒绝生成——这些地址必须整段原序处理：")
        for d in coll:
            print("   %s %s  多写于: %s" % (d["addr"], d.get("name", ""), ",".join(d["multi_write_in"])))
        sys.exit(1)

    # 安全闸 2：某模式没写过某差异地址（∅）→ 无法给值，点名退出
    holes = [(d["addr"], m) for d in diff for m in modes if d["vals"].get(m) is None]
    if holes:
        print("✋ 以下 (地址, 模式) 没有末态值（该模式的表没写它），需人工决定写什么：")
        for a, m in holes:
            print("   %s @ %s" % (a, m))
        sys.exit(1)

    def fmt_val(v):
        v = str(v).strip()
        if args.no0x:
            return v[2:] if v.lower().startswith("0x") else v
        return v if v.lower().startswith("0x") else "0x" + v

    def fmt_addr(a):
        a = str(a).strip()
        return a[2:] if a.lower().startswith("0x") else a

    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "签名块"
    BOLD = Font(bold=True)

    ws.append(["各模式签名块：每段=纯 地址/值 矩形，整块选中复制 → 粘贴到工具B流程行的 REG ADDR1 格起（每模式 %d 行）"
               % -(-len(diff) // args.pairs)])
    ws.append(["签名块后必须跟该模式的 init + lock 行；ReadBack 判锁；差异地址 %d 个，全部零撞多写" % len(diff)])
    ws.append([])
    n_rows_per_mode = -(-len(diff) // args.pairs)
    for mode in modes:
        c = ws.cell(row=ws.max_row + 1, column=1, value="◆ %s  （%d 对 → %d 行）" % (mode, len(diff), n_rows_per_mode))
        c.font = BOLD
        for i in range(0, len(diff), args.pairs):
            row = []
            for d in diff[i:i + args.pairs]:
                row += [fmt_addr(d["addr"]), fmt_val(d["vals"][mode])]
            ws.append(row)
        ws.append([])

    # 附一张对照表（人审用：地址×名字×各模式值）
    ws2 = wb.create_sheet("差异对照")
    ws2.append(["Address", "Register Name"] + modes)
    for c in ws2[1]:
        c.font = BOLD
    for d in diff:
        ws2.append([d["addr"], d.get("name", "")] + [fmt_val(d["vals"][m]) for m in modes])
    ws2.freeze_panes = "C2"
    for col, w in (("A", 12), ("B", 22)):
        ws2.column_dimensions[col].width = w

    wb.save(args.out)

    print("差异地址 %d 个 → 每模式 %d 行 × 最多 %d 对；模式顺序：%s"
          % (len(diff), n_rows_per_mode, args.pairs, " → ".join(modes)))
    print("签名块已写: %s   （『签名块』sheet 整块复制粘贴；『差异对照』sheet 人审）" % args.out)
    print()
    for mode in modes:
        print("◆ %s" % mode)
        for i in range(0, len(diff), args.pairs):
            print("   " + "  ".join("%s=%s" % (fmt_addr(d["addr"]), fmt_val(d["vals"][mode]))
                                    for d in diff[i:i + args.pairs]))
    print()
    print("⚠ 每个模式块的行序：签名行 → Initail_buf → Lock_step1/2 → OFF 序列；纯写入行的测量开关建议全 NO。")


if __name__ == "__main__":
    main(argv=None)
