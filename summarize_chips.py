#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
summarize_chips.py — 一个目录里的多颗芯片 → 一份给评审看的汇总 Excel

输入是这样一棵树（一层芯片目录，芯片编号＝目录名）：

    <根目录>/
        <芯片1>/  <模块>PLL_Temperature_Sweep_*.xlsx     PLL 温度扫描
                <模块>VCO_*.xlsx                       VCO 开环压控
                <模块>_Current_*.xlsx                  电流（本版只清点，不处理）
        <芯片2>/  ...

输出一份 .xlsx：

    PLL_Summary  上面一张 <模块A> PLL、下面一张 <模块B> PLL 的性能汇总表。
                 列 = 测试项 | Unit | Limit | Spec(Min/Typ/Max) | 仿真(Min/Typ/Max)
                      | 汇总(Min/Typ/Max) | 判定 | 每颗芯片(常温/最低/最高温 + 全温MAX
                      + MAX出现在哪) | 备注
                 Spec 与仿真列给人填，填完判定列（Excel 公式）自动出 PASS/FAIL。
                 填过一次就别再填第二次：`spec_from_xlsx.py <填好的簿子>
                 --merge <配置>` 把它们存成数据，以后每次重出都自带。
    温巡          每颗芯片一个**竖条**，条内每个模块两张图（压控温巡 + 频率漂移），
                 一张图只画一颗芯片一个模块（不叠线）；图下面就是它们的数据源。
    _审计         每个数字出自哪一份文件、哪些文件被跳过。**默认隐藏**

出稿纪律（跟 summarize_pll_sweep / summarize_vco_sweep 一致）
    · 正表里不写使用说明、不写告警、不写排除记录——那些打在控制台。
    · 讲不清的参数不进表：压控电压/片上温度/电流都**不进**汇总表
      （压控看「温巡」页的图，电流另有专门的表格，格式定了再加页）。
    · 判定看 Limit 列的方向 + Spec 的 Min/Max：≤ 只判上限、≥ 只判下限、
      range 两头都判。Typ 与仿真列只作对照，不参与判定。

用法：
    python summarize_chips.py <根目录>
    python summarize_chips.py <根目录> --dry-run          # 只清点+核对识别结果
    python summarize_chips.py <根目录> -o 汇总.xlsx
    python summarize_chips.py <根目录> --chips <芯片1>,<芯片2> --modules <模块>
"""

import argparse
import json
import os
import re
import sys
from collections import OrderedDict

from spec_book import SpecBook
from summarize_vco_sweep import load_vco
from sweep_lib import (
    COLOR_FLAG, COLOR_MUTED, COLOR_PASS, FILL_FAIL, FILL_PASS,
    LEG_STYLE, Columns, SweepError, apply_y, as_text, axis_bounds, blank_policy,
    DEFAULT_ROW_PT, chart_rows, col_px, cols_cm, fit_strip, fmt_num, is_blank,
    leg_series, legend_bottom,
    load_sweep, median, nice_step, num, put, stats_all, styles, style_series, txt,
)
from xlsx_formula_cache import FormulaCache

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

VCACHE = FormulaCache()

EMU_PX = 9525.0            # 1 px = 9525 EMU
DEFAULT_COL_W = 8.43       # Excel 没设过的列就是这个宽度

# ---------------------------------------------------------------- 发现层

KIND_PLL, KIND_VCO, KIND_CUR = "pll", "vco", "current"
KIND_LABEL = {KIND_PLL: "PLL 温扫", KIND_VCO: "VCO 开环", KIND_CUR: "电流"}

# 文件名里的时间戳 _2026-07-28-15-27-19
TS_RE = re.compile(r"_(\d{4})-(\d{2})-(\d{2})-(\d{2})-(\d{2})-(\d{2})")
# 自己的产物 / Excel 临时文件，别把它们当输入读回来
# `_spec.xlsx` 是手填 Spec 的那份（自己产物的副本）——它的文件名里同样带着
# 模块名和 pll/vco 字样，掉进根目录里会被当成一份数据簿读进来
SKIP_RE = re.compile(r"(^~\$)|(_summary\.xlsx$)|(_chips_summary\.xlsx$)"
                     r"|(_spec\.xlsx$)", re.I)


def natkey(s):
    """A015 < A087 < A105：数字段按数值比，别按字符串比。"""
    return [int(t) if t.isdigit() else t.lower() for t in re.split(r"(\d+)", str(s))]


def classify(basename):
    """文件名 -> (模块, 类型)。认不出来的返回 (None, None)。

    命名约定是 `<模块><类型>_...`：拿类型关键字（current / vco / pll）在名字里
    的位置一切两半，**前半就是模块名**——所以工具里不写死任何模块名，
    换个芯片、换个模块照样认（铁律：通用引擎零真实字面量）。
    `<模块>_VCO_...` 这种带下划线的也认，尾部的分隔符剥掉。

    ★ 芯片编号**只认目录名**，文件名里的编号不参与判断。原厂模板复制粘贴
      忘改名是这条线已经出过的事故（同一份模板里 REG ADDR7/8/9 重名同源），
      文件名里的编号跟目录名对不上时按目录名走、控制台报一句。
    """
    n = os.path.splitext(basename)[0].lower()
    for kw, kind in (("current", KIND_CUR), ("vco", KIND_VCO), ("pll", KIND_PLL)):
        i = n.find(kw)
        if i < 0:
            continue
        mod = basename[:i].strip(" _-.")
        return (mod or None), kind
    return None, None


class Book:
    __slots__ = ("chip", "module", "kind", "path", "name", "ts")

    def __init__(self, chip, module, kind, path):
        self.chip, self.module, self.kind, self.path = chip, module, kind, path
        self.name = os.path.basename(path)
        m = TS_RE.search(self.name)
        self.ts = m.group(0) if m else ""

    @property
    def sort_key(self):
        # 时间戳优先（字符串可比：年-月-日-时-分-秒 定宽），没有时间戳的看 mtime
        return (self.ts, os.path.getmtime(self.path))


def _norm_mod(x):
    """模块名比对用的归一化：只抹掉大小写和分隔符，**别的字符一个都不能丢**。

    ★★ 不能用电流那边的 `_norm`（`[^0-9a-z]` 全删）：它会把中文整段抹掉，
      `前缀_模块B` 和 `模块B` 压出来一模一样，"名字被前缀带偏了"这条判据当场失效。
      2026-08-05 第一版就是这么写的，在复现用例上一声不吭。
    """
    return re.sub(r"[\s_\-.#()（）\[\]]+", "", str(x).lower())


def alias_of(mod, alias):
    """文件名前缀改过 -> 认成两个模块。别名表把它并回去。

    别名表是人手写的，`前缀_模块B` / `前缀 模块B` / `前缀-模块b` 该算同一条，
    所以除了原样，再拿归一化后的形式找一遍。
    """
    if not alias:
        return mod
    return alias.get(mod) or alias.get(_norm_mod(mod)) or mod


def near_modules(picked):
    """名字互相包含、而且**从不属于同一颗芯片**的模块对 —— 多半是文件名前缀改过。

    ★★ 真事故（2026-08-05）：`<模块B>PLL_..._<芯片>.xlsx` 更新时改名成
      `前缀_<模块B>PLL_..._<芯片>.xlsx`，模块名就从 `模块B` 变成 `前缀_模块B`。
      那颗芯片在原模块那一栏**整块消失**、另起一个只有它一片的新模块，
      温巡页于是空出两个图位（用户原话"往下空了两个身位才是下一个模块"）。
      表照样出得来，一个字都不报——这正是最该报的那种错。
    ★ 判据要两条都满足才开口：名字包含 **且** 两边的芯片集合不相交。
      同一颗芯片两个名字都有，那就真是两个模块（比如 `A` 和 `A_LP`），
      喊了就是假阳性——而警告有一次假阳性，下次真的那条就被当噪声划过去。
    ★★ 芯片集合要**按类型分开**比，不能拿全部文件混在一起比。改名的往往只有
      一种文件（只有温扫那份加了前缀，VCO 那份还是老名字），混着比的话
      那颗芯片在原模块名下照样出现（来自 VCO），两边就"相交"了，
      这条哨兵当场哑掉——2026-08-05 第一版就是这么写的，在复现用例上没响。
    """
    who = {}
    for (chip, mod, kind) in picked:
        who.setdefault((kind, mod), set()).add(chip)
    out = []
    for (kind, a), ca in sorted(who.items()):
        for (kind2, b), cb in sorted(who.items()):
            if kind2 != kind or a == b:
                continue
            na, nb = _norm_mod(a), _norm_mod(b)
            if not nb or nb not in na or (ca & cb):
                continue
            # 长的那个才是"被前缀带偏的"。压完一模一样时（只差大小写/分隔符）
            # 两个方向都成立，靠字典序只报一次
            if len(na) == len(nb) and a >= b:
                continue
            out.append((a, b, KIND_LABEL[kind],
                        sorted(ca, key=natkey), sorted(cb, key=natkey)))
    return out


def discover(root, only_chips=None, only_modules=None, alias=None):
    """扫目录 -> (选中的, 同类被跳过的, 认不出来的, 散放的, 被 --chips/--modules 滤掉的)。

    ★ 过滤掉的也要报出来。模块名是从文件名前缀认出来的，同一个模块的几份文件
      前缀不一致很常见（`<模块>PLL_...` 但电流那份写成 `<芯片系列>_<模块>_CURRENT_...`），
      这时候一给 --modules / modules 配置，那份就**安静地消失**——
      报表少一页，控制台一个字都没有。
    ★ 别名在**模块过滤之前**换：不然配了 `modules` 之后，被前缀带偏的那几份
      先被滤掉，别名再对也救不回来（顺序反了就是白写）。
    """
    dirs = sorted((d for d in os.listdir(root)
                   if os.path.isdir(os.path.join(root, d))), key=natkey)
    loose = []
    if not dirs:
        # 根目录本身就装着一颗芯片的文件：目录名当芯片号
        dirs = [""]
    else:
        # ★ 有芯片目录时只扫子目录，根目录下的散放文件一律不读——但要报出来。
        #   重组目录结构之前的旧文件常常还躺在根目录里，静默跳过等于少算一颗芯片。
        loose = [f for f in sorted(os.listdir(root))
                 if os.path.isfile(os.path.join(root, f))
                 and f.lower().endswith((".xlsx", ".xlsm")) and not SKIP_RE.search(f)]
    picked, dropped, unknown, filtered = {}, [], [], []
    for d in dirs:
        chip = d or os.path.basename(os.path.abspath(root))
        if only_chips and chip not in only_chips:
            filtered.append((chip, "*", f"芯片 {chip} 不在指定的芯片清单里"))
            continue
        sub = os.path.join(root, d) if d else root
        for f in sorted(os.listdir(sub)):
            if not f.lower().endswith((".xlsx", ".xlsm")) or SKIP_RE.search(f):
                continue
            mod, kind = classify(f)
            if mod is None or kind is None:
                unknown.append((chip, f, "模块" if kind is not None else "类型"))
                continue
            mod = alias_of(mod, alias)
            if only_modules and mod not in only_modules:
                filtered.append((chip, f, f"模块认成 {mod!r}，不在指定的模块清单里"))
                continue
            b = Book(chip, mod, kind, os.path.join(sub, f))
            key = (chip, mod, kind)
            old = picked.get(key)
            if old is None:
                picked[key] = b
            elif b.sort_key > old.sort_key:
                picked[key] = b
                dropped.append((old, b))          # 旧的那份让位
            else:
                dropped.append((b, old))
    return picked, dropped, unknown, loose, filtered


def print_excluded(excluded, indent="     ", cap=14):
    """排除的行逐行列出原因，不静默丢弃。

    ★ 只打一个"排除 N 行"是不够的：N 对不上预期时没法判断到底丢了什么。
      这条纪律单簿脚本里一直有，跨芯片脚本一开始漏了——第一次跑真数据
      就撞上了（预期 3 行、实际 6 行，只能靠猜）。
    """
    for xl, why in excluded[:cap]:
        print(f"{indent}行{xl}: {why}")
    if len(excluded) > cap:
        print(f"{indent}…还有 {len(excluded) - cap} 行（全部内容在隐藏的 _审计 页）")


# ---------------------------------------------------------------- 指标挑选

# 只有这些进汇总表。★ 压控/片上温度/电流一律不进：
#   压控的结论是「距轨还剩多少」，一个 Min/Max 说不清，去「温巡」页看图；
#   电流另有专门的测试表格，格式定了再单独加页。
WANT_EXACT = ["Freq_MHz", "Power_dBm"]
# `IPN*` 按前缀收：换算成 DSB 之后行名会从 IPN_SSB 变成 IPN_DSB，
# 写死名字的话那两行会当场从表里消失（而且不报错）。
WANT_PREFIX = ["IPN", "SpotPN@", "Spur@"]
# 表里的分组带（顺序＝出现在页面上的顺序）。
# `xxx@` = 按 offset 排序收；`xxx*` = 按前缀收、保持识别顺序
BANDS = [("Frequency / Output", ["Freq_MHz", "Power_dBm"]),
         ("Phase Noise", ["IPN*", "SpotPN@"]),
         ("Spur", ["Spur@"])]
# 单位 -> 小数位
ND = {"MHz": 3}


def _off_key(label):
    """SpotPN@100kHz / Spur@26MHz -> 按 offset 数值排序。"""
    m = re.search(r"@([\d.]+)(k|M)Hz", label)
    if not m:
        return 1e9
    return float(m.group(1)) * (0.001 if m.group(2) == "k" else 1.0)


def canon_items(sweeps):
    """把各颗芯片各自识别出的指标对齐成一份表的行序。

    ★ 跨簿子只能按**标签**对齐，不能按列位置：同一个模板不同批次导出，
      列位有可能挪，名字才是身份。某颗芯片缺某个指标就在那一列留空。
    """
    seen, rows = {}, []
    for sw in sweeps:
        for it in sw.items:
            lb = it.label
            if lb not in WANT_EXACT and not any(lb.startswith(p) for p in WANT_PREFIX):
                continue
            if lb not in seen:
                seen[lb] = (lb, it.unit)
    out = []
    for band, keys in BANDS:
        picked = []
        for k in keys:
            if k.endswith("@"):
                picked += sorted((v for lb, v in seen.items() if lb.startswith(k)),
                                 key=lambda v: _off_key(v[0]))
            elif k.endswith("*"):
                picked += [v for lb, v in seen.items() if lb.startswith(k[:-1])]
            elif k in seen:
                picked.append(seen[k])
        if picked:
            out.append((band, picked))
    rows = out
    return rows


# ---------------------------------------------------------------- 频率折算

def unit_scaling(unit):
    """这个单位的量，载波乘 N 之后自己怎么变。

    · 频率维（MHz / MHz/V / MHz/code）——直接 ×N。
    · 相噪与杂散是**相对载波**的量：载波乘 N，相位起伏跟着乘 N，
      落到 dB 域就是 +20·log10(N)。IPN 是积分相噪，同理；
      Spur 的 dBc 也是同一条（杂散来自同一个相位调制，跟着载波一起被倍频）。
    · Power_dBm / Vtune_V / Current_mA / 温度 / CT 码都不跟着变——
      倍频器改的是频率与相位，不是这些量。
    """
    u = txt(unit).lower()
    if u.startswith("mhz"):
        return "lin"
    if u.startswith("dbc"):
        return "db"
    return None


def parse_scale(spec):
    """`模块:类型=倍数`，逗号分隔；模块与类型都可以写 `*`，只写模块＝该模块全部类型。

    倍数**只能从命令行进来**：它是"这颗芯片的测试点在真实频点的几分之一"，
    是芯片事实，写进公开仓就等于把架构写进去了（铁律：通用引擎零真实字面量）。
    """
    out = []
    for part in txt(spec).split(","):
        part = part.strip()
        if not part:
            continue
        if "=" not in part:
            sys.exit(f"--scale 这一段少了 '='：{part!r}（写法 模块:类型=倍数）")
        key, _, val = part.rpartition("=")
        try:
            n = float(val)
        except ValueError:
            sys.exit(f"--scale 的倍数不是数：{part!r}")
        if n <= 0:
            sys.exit(f"--scale 的倍数必须为正：{part!r}")
        mod, sep, kind = key.partition(":")
        if not sep:
            mod, kind = key, "*"
        out.append((mod.strip() or "*", kind.strip().lower() or "*", n))
    return out


def scale_of(rules, mod, kind):
    """后写的规则盖前面的——可以先写一条通配、再写某个模块的例外。"""
    n = 1.0
    for m, k, v in rules:
        if m in ("*", mod) and k in ("*", kind):
            n = v
    return n


def scale_book(sw, n):
    """把一份读完的簿子里所有跟载波成比例的量折算到 N 倍频点上。

    ★ 为什么在**读完之后**动 r.vals，而不是写表的时候再乘：表和图读的是同一份
      r.vals。只改写表那一路，同一页上表里写折算后的频点、图上还画着实测频点。
    ★ Kvco 不用单独处理：它是逐区间 ΔF/ΔVtune 算出来的，F 乘了 N 它自然就 ×N。
      Fmin/Fmax/Tuning Range/温漂同理，全是拿 F 算的。
    ★ 只动值不动横轴：Vtune 与 CT 码是自变量，倍频器不改它们。
    """
    import math
    if not n or abs(n - 1.0) < 1e-12:
        return None
    db = 20.0 * math.log10(n)
    fi = getattr(sw, "freq_item", None)
    if fi is None and hasattr(sw, "item"):
        fi = sw.item("Freq_MHz")
    before = median([r.vals.get(fi.col) for r in sw.rows
                     if r.vals.get(fi.col) is not None]) if fi is not None else None
    n_lin = n_db = 0
    for it in sw.items:
        how = unit_scaling(it.unit)
        if how is None:
            continue
        hit = False
        for r in sw.rows:
            v = r.vals.get(it.col)
            if v is None:
                continue
            r.vals[it.col] = (v * n) if how == "lin" else (v + db)
            hit = True
        if hit:
            n_lin, n_db = (n_lin + 1, n_db) if how == "lin" else (n_lin, n_db + 1)
    info = {"n": n, "db": db, "before": before,
            "after": (before * n) if before is not None else None,
            "n_lin": n_lin, "n_db": n_db}
    # 挂回簿子上：条件行要报"折算之前实测的是哪个频点"——
    # 他要的是「测的是哪个、报的是哪个」两个都看得见，不是只看见折算后的那个。
    sw.scale_info = info
    return info


def note_scale(sinfo, kind, mod, chip, info):
    """记下折算结果，并当场把折算前后的频点打出来。

    ★ 折算是最容易"看着对、其实错一个倍数"的改动：差 2 倍，表和图都还长得
      像那么回事。所以每份簿子都报一句实测中位频点 → 折算后频点，
      对不对一眼就知道。
    """
    if not info:
        return
    sinfo.setdefault(kind, {})[mod] = info
    n = fmt_num(info["n"], 4)
    b, a = info["before"], info["after"]
    print(f"     · 折算 ×{n}：频率类 {info['n_lin']} 项 ×{n}"
          + (f"（实测中位 {fmt_num(b)} → {fmt_num(a)} MHz）" if b is not None else "")
          + f"，相噪/杂散 {info['n_db']} 项 {info['db']:+.2f} dB")


# 单边带 → 双边带：10·log10(2)
DSB_DB = 3.010299956639812


def to_dsb(sw):
    """把**积分**相噪从单边带换算到双边带（+3.01 dB），并把行名改掉。

    ★ 为什么只动 IPN：
      · 逐点相噪 L(f) 的定义本身就是单边带——载波一侧、偏移 f 处、1 Hz 带宽内
        相对载波的功率。它是"某一点的密度"，没有"两边加起来"这回事
        （真要说双边带那是相位谱密度 S_φ = 2L，单位 rad²/Hz，不是 dBc/Hz）。
      · 积分相噪回答的是"总共抖了多少"。相位调制的 +f 与 −f 两个边带是同一个
        相位起伏的两半，算总相位误差必须两边都算：σ_φ² = ∫S_φ df = 2·∫L df，
        落到 dB 上就是 +10·log10(2)。
      · 杂散跟逐点相噪同理：dBc 的惯例就是报单边带那一根。
      佐证：原表的列名本来就叫 `IPN_SSB` / `IPN_Omit_SSB`（仪器自己标了单边带），
      而 `SpotPNResult` / `OtherSpurResult` 没有这个后缀。
    ★ **改了值就必须改名**：留着 `IPN_SSB` 这个名字写 DSB 的数，
      跟"两个都叫频点却差 4 倍"是同一类错。
    """
    n = 0
    for it in sw.items:
        if not it.label.startswith("IPN"):
            continue
        for r in sw.rows:
            v = r.vals.get(it.col)
            if v is not None:
                r.vals[it.col] = v + DSB_DB
        it.label = (it.label.replace("_SSB", "_DSB") if "_SSB" in it.label
                    else it.label + "_DSB")
        n += 1
    return n


def ipn_order_check(sw, chip, tol=0.5):
    """IPN_Omit 只可能比 IPN **好**（剔掉杂散再积分）。反过来就是个信号。

    ★ 反向差多少，就是这两次积分之间的散布下限——它们本该是同一段谱算出来的
      两个数，一个比另一个"更差"只能来自重复性。真数据上：一颗片子 −40℃ 开环
      的反向差到 5 dB，另一颗只有 0.7 dB。**散布 5 dB 的时候，3 dB 的片间差
      就说明不了任何事**，而表上那两个数看着一样体面。
    ★ 走控制台 + 审计页，不进正表（正表不写告警）。
    """
    ipn = next((i for i in sw.items
                if i.label.startswith("IPN") and "Omit" not in i.label), None)
    omt = next((i for i in sw.items if i.label.startswith("IPN")
                and "Omit" in i.label), None)
    if ipn is None or omt is None:
        return []
    by_t = {}
    for r in sw.rows:
        a_, b_ = r.vals.get(ipn.col), r.vals.get(omt.col)
        t = getattr(r, "temp", None)
        if a_ is None or b_ is None:
            continue
        d = b_ - a_                       # >0 ＝ Omit 更差 ＝ 不该发生
        st = by_t.setdefault(t, [[], 0])
        st[1] += 1
        if d > tol:
            st[0].append(d)
    seg = []
    for t, (ds, tot) in sorted(by_t.items(), key=lambda x: (x[0] is None, x[0])):
        if not ds:
            continue
        # ★ 报中位数，不只报最大值：275 行里的最大值是极值统计，
        #   拿它当"重复性"会把这批数据说得比实际更糟。
        seg.append(f"{fmt_num(t)}℃ {len(ds)}/{tot}（中位 {median(ds):.1f}、"
                   f"最多 {max(ds):.1f} dB）")
    if not seg:
        return []
    return [f"{chip} {omt.label} 比 {ipn.label} **还差**：" + "；".join(seg)
            + "。剔掉杂散只可能更好，反过来说明这两个积分不是同一次测出来的，"
              "反向差就是这一趟的重复性下限——比它小的片间差不作数"]


def note_dsb(n, chip):
    if n:
        print(f"     · 积分相噪 SSB→DSB：{n} 项 {DSB_DB:+.2f} dB，行名改成 IPN_DSB")


def scale_title(info, dsb=False):
    """折算这件事必须**印在表上**——它是取数条件，不是给填表人看的说明。

    放大标题那一行：一张表一句，不在每一行的备注里重复 40 遍。
    """
    parts = []
    if info:
        n = fmt_num(info["n"], 4)
        parts.append(f"折算 ×{n}：频率 ×{n}，相噪/杂散 {info['db']:+.1f} dB")
    if dsb:
        parts.append(f"积分相噪 SSB→DSB {DSB_DB:+.2f} dB")
    return f"（{'；'.join(parts)}）" if parts else ""


def spur_note(data_or_sw, label):
    """杂散那两行的备注：这个数到底是在哪个偏移上取的。

    ★ 值不再来自模板那个固定频点了（那一格量到的是噪底），所以"取自哪儿"
      成了这一行**怎么算出来的**的一部分——评审第一句就是"你这 26M 是在哪测的"。
      各片实测偏移不一样就都列出来，别只报第一片的。
    """
    sws = data_or_sw.values() if isinstance(data_or_sw, dict) else [data_or_sw]
    offs, tgt, tol = [], None, None
    for sw in sws:
        if sw is None:
            continue
        it = next((x for x in sw.items if x.label == label), None)
        if it is None or not it.pick_stats:
            continue
        tgt, tol = it.pick_stats.get("target"), it.pick_stats.get("tol")
        o = it.pick_stats.get("off")
        if o is not None and fmt_num(o) not in offs:
            offs.append(fmt_num(o))
    if not offs or tgt is None:
        return ""
    return (f"实测偏移 {' / '.join(str(x) for x in offs)} MHz"
            f"（标称 {fmt_num(tgt)}，取 ±{fmt_num(tol)} MHz 内最大的一条）")


def temp_view(sw, item):
    """{温度: [该温度经过的每一次的值]}。整趟温巡四段各走一遍，同温会有 2~4 个值。"""
    d = {}
    if item is None:
        return d
    for lg in sw.legs:
        for t, v in leg_series(lg, item).items():
            d.setdefault(t, []).append(v)
    return d


def val_at(view, t):
    """某个温度点的代表值＝该温度**全部经过点**的中位数。

    取中位数比随便挑一次稳，跟「常温 Typ 取中位」是同一条规矩。
    """
    vs = view.get(t)
    return median(vs) if vs else None
def chip_stat(sw, label):
    """一颗芯片一个指标的 (min, typ, max, min_t, max_t)；没这个指标返回 None。"""
    it = sw.item(label)
    if it is None:
        return None
    s = stats_all(sw.legs, it, sw.room_t)
    if not s:
        return None
    return s


# ---------------------------------------------------------------- 汇总页

C_ITEM, C_UNIT, C_LIMIT = 1, 2, 3
C_SPEC, C_GAP1 = 4, 7
C_SIM, C_GAP2 = 8, 11
C_SUM, C_JUDGE, C_GAP3 = 12, 15, 16
C_CHIP0 = 17           # 第一颗芯片的第一列
# ★★ 每颗芯片只有三列：常温 / 最低温 / 最高温。
#   曾经是 7 列（再带全温 MIN/MAX 各配一个 @℃），一轮轮加出来的，每列单看都有
#   理由，加在一起就成了"要读说明才看得懂的表"——而这份簿子的第一条要求就是
#   「不能出现给人 debug 用的信息」。被砍掉的是：@℃（极值在哪个温度，对下判断
#   没有决策价值，还会写出"等 7 处"这种噪声）、逐片 MIN/MAX（跟三个温度列是
#   两种统计量，摆在一行里读起来自相矛盾，得靠 caption 解释）、同温重复性。
#   现在一行里只有一种数：**该温度的实测值**。汇总列就是对这三列取，
#   谁都能用眼睛验证，不需要任何脚注。
CHIP_AX = 3
CHIP_W = CHIP_AX + 1   # +1 = 组间间隔列
CHIP_NUMCOL = (0, 1, 2)

AXES = ["Min", "Typ", "Max"]


def _cl(i):
    from openpyxl.utils import get_column_letter
    return get_column_letter(i)


def _align(h, wrap=True):
    # 不用 cell.alignment.copy()：openpyxl 3.1 起那个 copy 是 deprecated，
    # 每调一次在控制台喷一行 DeprecationWarning
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def chip_col(n):
    return C_CHIP0 + n * CHIP_W


def note_col(n_chips):
    return chip_col(n_chips)


def rail_cols(n_chips):
    """组与组之间的"竖栏"列号。填成中灰＝把每一片界定成一个视觉区域。

    ★ 40 列宽的表里最先丢的信息是"我现在看的是哪一颗芯片"。
      一条实心竖栏比给整块上底色更省视觉预算（Gestalt 的 common region），
      而且灰度打印下照样成立。
    """
    return [C_GAP1, C_GAP2, C_GAP3] + [chip_col(k) + CHIP_AX for k in range(n_chips)]


def _edges(ws, r0, r1, c_first, c_last):
    """给一个列组的左右两侧加中等粗细竖边框。

    ★ 汇总组是全表唯一要"跳出来"的（判定就是对着它做的）。只靠底色不够：
      灰度打印和色弱下几种浅色会撞，边框不会。
    """
    from openpyxl.styles import Border, Side
    th, md = Side(style="thin", color="FF000000"), Side(style="medium", color="FF000000")
    for r in range(r0, r1 + 1):
        for c, which in ((c_first, "l"), (c_last, "r")):
            b = ws.cell(r, c).border
            ws.cell(r, c).border = Border(left=md if which == "l" else (b.left or th),
                                          right=md if which == "r" else (b.right or th),
                                          top=b.top, bottom=b.bottom)


def _hguide(ws, r, c0, c1):
    """一条横向导引线（中等粗细的下边框）。

    ★ 人一眼能数清的上限是 4 个（subitizing）。SpotPN 有连续 8 行，
      在 40 列宽的表上横着扫很容易串行。每 4 行给一条横线，不上底色——
      竖向已经分区了，再加行斑马就成了网格噪声。
    """
    from openpyxl.styles import Border, Side
    md = Side(style="medium", color="FF000000")
    for c in range(c0, c1 + 1):
        b = ws.cell(r, c).border
        ws.cell(r, c).border = Border(left=b.left, right=b.right, top=b.top, bottom=md)


def _chip_axes(tlabels):
    """芯片组的轴名就是三个温度，别的什么都不放。"""
    return [(t or "—") for t in tlabels]


def fold_temp_cols(ws, col_of, n_chips, nax, keep=0):
    """--slim：每片只留常温那一列，其余温度列收进 Excel 大纲（默认折起）。

    ★ 为什么是"折"不是"删"：汇总列的 Min/Max 就是这些格子里的值。真删掉，
      汇总那几个数在表上就没了出处（"导出来的数必须能用表上的格子推出来"），
      脚本自查第①条会当场 exit 1。折起来一个数都没少，点 ＋ 就能核。
      6 片时实测区 24 列 → 12 列，一屏看得完。
    ★ keep 不写死 0：PLL 页的第一列就是常温，VCO 页的三列是升序温度
      （−40 在最左），留错列的话折完剩下的是低温那列。
    """
    from openpyxl.worksheet.properties import Outline
    # summaryRight=False ＝ 汇总列在明细列**左边**，＋/− 按钮就画在留下的那列头上；
    # 默认的 True 会把按钮画到右边的灰竖栏上（点得到，但看着像点在缝里）
    ws.sheet_properties.outlinePr = Outline(summaryBelow=True, summaryRight=False)
    n = 0
    for k in range(n_chips):
        c0 = col_of(k)
        for j in range(nax):
            if j == keep:
                continue
            d = ws.column_dimensions[_cl(c0 + j)]
            d.outlineLevel = 1
            d.hidden = True
            n += 1
        ws.column_dimensions[_cl(c0 + keep)].collapsed = True
    return n


def _header(ws, r0, chips, st, title, n_chips, tlabels):
    """三行表头：大标题 / 组名 / 轴名。"""
    last = note_col(n_chips)
    c = put(ws, r0, C_ITEM, title, st, st["f_sep"], bold=True, align="left", size=12)
    ws.merge_cells(start_row=r0, start_column=C_ITEM, end_row=r0, end_column=last)
    c.alignment = _align("left", wrap=False)

    hr, ar = r0 + 1, r0 + 2
    for col, name in ((C_ITEM, "测试项"), (C_UNIT, "Unit"),
                      (C_LIMIT, "Limit"), (C_JUDGE, "判定")):
        put(ws, hr, col, name, st, st["f_head"], bold=True)
        put(ws, ar, col, None, st, st["f_head"])
        ws.merge_cells(start_row=hr, start_column=col, end_row=ar, end_column=col)
    chip_ax = _chip_axes(tlabels)
    # ★ 表头只写这一列**是什么**，不写"该怎么用"。"（留空，填完自动判定）"
    #   "(各片最小)" 这类是写给填表的人的，评审看见只会觉得奇怪——
    #   designer 本来就知道 Spec 列填什么，Min 是什么也不用括号解释。
    groups = [(C_SPEC, "Spec", AXES),
              (C_SIM, "仿真", AXES),
              (C_SUM, f"汇总 · {n_chips} 片", AXES)]
    for n, chip in enumerate(chips):
        groups.append((chip_col(n), chip, chip_ax))
    for col, name, axes in groups:
        put(ws, hr, col, name, st, st["f_head"], bold=True)
        for j in range(1, len(axes)):
            put(ws, hr, col + j, None, st, st["f_head"])
        ws.merge_cells(start_row=hr, start_column=col, end_row=hr,
                       end_column=col + len(axes) - 1)
        fill = st["f_in"] if col in (C_SPEC, C_SIM) else st["f_head"]
        for j, lb in enumerate(axes):
            put(ws, ar, col + j, lb, st, fill, bold=True, size=9)
    # 竖栏只画在两行表头上——r0 是横跨整表的大标题，已经 merge 过，
    # 往合并区里写会抛 MergedCell read-only
    for col in rail_cols(n_chips):
        for r in (hr, ar):
            put(ws, r, col, None, st, st["f_rail"])
    put(ws, hr, last, "备注", st, st["f_head"], bold=True)
    put(ws, ar, last, None, st, st["f_head"])
    ws.merge_cells(start_row=hr, start_column=last, end_row=ar, end_column=last)
    ws.row_dimensions[ar].height = 30
    return ar + 1


# ★ 这里原来有一行 `_caption()`：「每片三列＝该温度的实测值。汇总列 Min / Max 就是
#   这些格子里的最小 / 最大…填 Spec 的 Min / Max，判定列自动出 PASS / FAIL。」
#   2026-08-04 整行删掉——**从头到尾都是在教人怎么读表和怎么填表**，评审要的是数。
#   列名已经写着 25℃ / Min / Spec，再解释一遍只会招来"这行字是干嘛的"。
#   口径要留档就进隐藏的 _审计 页和控制台，正表里不写。


def _band(ws, r, name, st, n_chips):
    put(ws, r, C_ITEM, name, st, st["f_sep"], bold=True, align="left")
    for c in range(C_ITEM + 1, note_col(n_chips) + 1):
        put(ws, r, c, None, st, st["f_sep"])
    return r + 1


def _cond_rows(ws, r, chips, data, st, n_chips):
    """条件行：这组数在什么条件下取的。一行一个条件，每颗芯片给自己的值。"""
    def per_chip(label, fn):
        nonlocal r
        put(ws, r, C_ITEM, label, st, st["f_group"], align="left")
        for c in (C_UNIT, C_LIMIT, C_JUDGE):
            put(ws, r, c, None, st, st["f_group"])
        for c in list(range(C_SPEC, C_SUM + 3)) + [C_JUDGE]:
            put(ws, r, c, None, st, st["f_group"])
        for c in rail_cols(n_chips):
            put(ws, r, c, None, st, st["f_rail"])
        for n, chip in enumerate(chips):
            sw = data.get(chip)
            v = fn(sw) if sw is not None else "未测"
            c0 = chip_col(n)
            put(ws, r, c0, as_text(v), st, st["f_group"], size=9)
            for j in range(1, CHIP_AX):
                put(ws, r, c0 + j, None, st, st["f_group"])
            ws.merge_cells(start_row=r, start_column=c0, end_row=r,
                           end_column=c0 + CHIP_AX - 1)
        put(ws, r, note_col(n_chips), None, st, st["f_group"])
        r += 1

    def temp_range(sw):
        ts = sw.temps
        return f"{fmt_num(ts[0])} ~ {fmt_num(ts[-1])}" if ts else ""

    def n_points(sw):
        n = sum(1 for lg in sw.legs for x in lg.rows if x.kind != "lock")
        return f"{len(sw.temps)} 档 / {n} 测点"

    def cond_val(sw, col, nd=3):
        ci = sw.cols.idx(col)
        if ci is None:
            return ""
        vals = {num(x.raw[ci]) for lg in sw.legs for x in lg.rows
                if ci < len(x.raw) and num(x.raw[ci]) is not None}
        vals.discard(None)
        return "/".join(str(fmt_num(v, nd)) for v in sorted(vals)) if vals else ""

    per_chip("温度范围 (℃)", temp_range)
    per_chip("温度点数", n_points)
    # ★ 这两行必须一眼分得出谁是测出来的、谁是设计上的数。原来叫
    #   「测试频点 fLO」和「实测频点」——两个都长得像测量结果，
    #   而它们差 4 倍，看的人只会以为哪儿算错了（用户原话："让人很误解"）。
    #   fLO 是原表声明的**标称** LO，不是量出来的。
    per_chip("fLO 标称 (MHz)", lambda sw: cond_val(sw, "fLO_MHz"))
    # 折算之后，"仪器实际测的是哪个频点"就只剩这一行还看得见了
    # （结果行里的 Freq 已经是折算后的载波）。没折算就不出这一行。
    if any(getattr(s, "scale_info", None) for s in data.values() if s is not None):
        per_chip("仪器实测频点 (MHz)", lambda sw: (
            fmt_num(getattr(sw, "scale_info", {}).get("before"))
            if getattr(sw, "scale_info", None) else ""))
    per_chip("参考 fXO (MHz)", lambda sw: cond_val(sw, "fXO_MHz"))
    per_chip("锁定方式", lambda sw: f"{len(sw.legs)} 段，每段开头重锁一次")
    return r


def _cells(r):
    """判定要用到的四个格子：Spec 的 Min/Max、汇总的 Min/Max。"""
    return (f"${_cl(C_SPEC)}{r}", f"${_cl(C_SPEC + 2)}{r}",
            f"{_cl(C_SUM)}{r}", f"{_cl(C_SUM + 2)}{r}")


def _bound(r, kind):
    """单边指标的那一道界在哪一格。

    `≤` 优先看 Max 格、`≥` 优先看 Min 格；那一格空着就用另一格。
    ★ 不强求填在"对"的格子里：`Fmin ≤ 2884` 这种要求，人会很自然地把 2884
      写进 **Min** 格（它说的是频率能下探到哪）。硬要求填对格子，代价是
      那一行**悄悄不判**——比判错还难发现。外面有 COUNT 两格的总护栏，
      两格都空时这个表达式取到的 0 不会被用上。
    """
    dmin, dmax, _l, _h = _cells(r)
    a, b = (dmax, dmin) if kind == "≤" else (dmin, dmax)
    return f"IF(COUNT({a})>0,{a},{b})"


def _judge_formula(r):
    """判定公式：**Limit 列说的方向** + Spec 的 Min/Max。两格都空就不判（留空）。

    ★★ 2026-08-05 修的错：原来只按"必须落在 [Min, Max] 窗口里"判，**完全不看
      Limit 列**（当时把它当成"只是给填表人提示方向"）。可 `Fmin ≤ 2884` 的意思
      是"至少要能压到 2884"——实测 2800 是覆盖更宽、该 PASS，却被判成"低于 Min"；
      `Fmax ≥ 3423` 实测 3500 同理。**能往好的方向超出去的量，只判一头。**
    ★ 方向读的是格子 `$C{r}` 而不是生成时定死：Limit 是带下拉的可填列，
      在 Excel 里改完方向，判定得当场跟着变，不然那个下拉就是个摆设。
    判的对象是**汇总列**（全部芯片全温的最差值），不是逐片判：
    上限看汇总 Max（最大的那个最坏），下限看汇总 Min。Typ 与仿真列不参与判定。
    """
    dmin, dmax, smin, smax = _cells(r)
    lim = f"${_cl(C_LIMIT)}{r}"
    le = (f"IF(AND(COUNT({smax})>0,{smax}>{_bound(r, '≤')}),"
          f"\"FAIL\",\"PASS\")")
    ge = (f"IF(AND(COUNT({smin})>0,{smin}<{_bound(r, '≥')}),"
          f"\"FAIL\",\"PASS\")")
    over = f"AND(COUNT({dmax})>0,COUNT({smax})>0,{smax}>{dmax})"
    under = f"AND(COUNT({dmin})>0,COUNT({smin})>0,{smin}<{dmin})"
    both = f"IF(OR({over},{under}),\"FAIL\",\"PASS\")"
    return (f"=IF(COUNT({dmin},{dmax})=0,\"\","
            f"IF({lim}=\"≤\",{le},IF({lim}=\"≥\",{ge},{both})))")


def _fill_spec(ws, r, st, nd, sp, fill=None):
    """把 Spec / 仿真 那六格填上（没有 spec 就还是留空给人手填）。

    ★ 写的是**字面数字**，不是公式：下一轮 spec_from_xlsx.py 还要把这份簿子
      再读回去（人在这份上接着改），公式读回来就成了一串 "=…"。
    """
    fill = fill or st["f_in"]
    d = sp or {}
    for base, key in ((C_SPEC, "spec"), (C_SIM, "sim")):
        vals = d.get(key) or {}
        for j, ax in enumerate(AXES):
            v = vals.get(ax.lower())
            cell = put(ws, r, base + j, v, st, fill)
            if isinstance(v, (int, float)):
                cell.number_format = "0." + "0" * nd


def _result_row(ws, r, label, unit, chips, data, st, n_chips, tpick,
                band="", spec=None):
    nd = ND.get(unit, 2)
    put(ws, r, C_ITEM, label, st, st["f_res"], align="left")
    put(ws, r, C_UNIT, unit, st, st["f_res"], size=9)
    # Limit ＝ 判定方向（不是提示）：相噪/杂散越小越好 -> ≤，只判上限；
    # 频率/功率两头都可能有要求 -> range。见 _judge_formula
    sp = spec.row(band, label) if spec is not None else None
    lim = (sp or {}).get("limit") or (
        "≤" if any(label.startswith(p) for p in ("IPN", "SpotPN@", "Spur@"))
        else "range")
    put(ws, r, C_LIMIT, lim, st, st["f_in"], size=9)
    _fill_spec(ws, r, st, nd, sp)
    for c in rail_cols(len(chips)):
        put(ws, r, c, None, st, st["f_rail"])

    mins, typs, maxs, marks = [], [], [], []
    for n, chip in enumerate(chips):
        sw = data.get(chip)
        s = chip_stat(sw, label) if sw is not None else None
        c0 = chip_col(n)
        vals = [None] * CHIP_AX
        if s:
            # 三个代表温度点：该温度**全部经过点**的中位数。整趟温巡会多次经过
            # 同一个温度（四段各走一遍），取中位比随便挑一次稳。
            view = temp_view(sw, sw.item(label))
            vals = [fmt_num(val_at(view, t), nd) for t in tpick]
            # ★ 汇总列就是对**表上这几个格子**取最小/最大/常温。用满精度、或者
            #   改成对逐点取极值，都会出现"表里的格子跟汇总对不上"——那种表
            #   没人敢信，而且解释起来要写一段 caption。宁可自洽。
            got = [v for v in vals if v is not None]
            if got:
                mins.append(min(got))
                maxs.append(max(got))
            if vals[0] is not None:
                typs.append(vals[0])          # tpick[0] 是常温
            marks.append((chip, s))
        for j, v in enumerate(vals):
            cell = put(ws, r, c0 + j, v, st, st["f_res"])
            if v is not None:
                cell.number_format = "0." + "0" * nd

    # 三个数都是从上面那些格子里直接取的，看的人用眼睛就能验证
    agg = [min(mins) if mins else None, median(typs) if typs else None,
           max(maxs) if maxs else None]
    for j, v in enumerate(agg):
        cell = put(ws, r, C_SUM + j, v, st, st["f_sum"], bold=True)
        if v is not None:
            cell.number_format = "0." + "0" * nd
    put(ws, r, C_JUDGE, _judge_formula(r), st, st["f_res"], bold=True)

    # ★ 备注只在真的缺东西时才写。以前这里写「Min 出自 XX / 同温重复性 0.8@65℃」
    #   那类工程内部信息——那是 debug 用的，不该出现在给人 review 的表上。
    gone = [c for c in chips if c not in {x[0] for x in marks}]
    note = spur_note(data, label) if label.startswith("Spur@") else ""
    if gone:
        note = (note + "；" if note else "") + f"未测: {', '.join(gone)}"
    put(ws, r, note_col(n_chips), as_text(note), st, st["f_res"],
        align="left", size=9, color=COLOR_MUTED)
    return r + 1


def _over_spec_cf(ws, r0, r1, st):
    """超规的那一格自己标红（report-forge 的视觉语言：红粗体＝超规）。

    只有判定列变红的话，一行 20 多个格子里到底是哪头超了还得自己比，
    评审时那一秒的迟疑就会变成一个问题。

    ★ 这里的判据必须跟 `_judge_formula` **逐条对应**：单边指标只有"坏"的那头
      会红（`≤` 只红汇总 Max，`≥` 只红汇总 Min）。两边一漂移就会出现
      "判定 PASS、格子却是红的"，那种表比没有标色更糟。
    """
    from openpyxl.formatting.rule import FormulaRule
    if r1 < r0:
        return
    red = st["Font"](bold=True, color=COLOR_FLAG)
    dmin, dmax, smin, smax = _cells(r0)
    lim = f"${_cl(C_LIMIT)}{r0}"
    two = f"AND({lim}<>\"≤\",{lim}<>\"≥\")"      # range / 留空 ＝ 两头都判
    have = f"COUNT({dmin},{dmax})>0"
    for col, conds in (
            (C_SUM, [f"AND({two},COUNT({dmin})>0,COUNT({smin})>0,{smin}<{dmin})",
                     f"AND({lim}=\"≥\",{have},COUNT({smin})>0,"
                     f"{smin}<{_bound(r0, '≥')})"]),
            (C_SUM + 2, [f"AND({two},COUNT({dmax})>0,COUNT({smax})>0,{smax}>{dmax})",
                         f"AND({lim}=\"≤\",{have},COUNT({smax})>0,"
                         f"{smax}>{_bound(r0, '≤')})"])):
        for cond in conds:
            ws.conditional_formatting.add(
                f"{_cl(col)}{r0}:{_cl(col)}{r1}",
                FormulaRule(formula=[cond], font=red, stopIfTrue=False))


def _pass_fail_cf(ws, col_letter, r0, r1, st):
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    if r1 < r0:
        return
    rng = f"{col_letter}{r0}:{col_letter}{r1}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{col_letter}{r0}="FAIL"'],
        fill=PatternFill("solid", fgColor=FILL_FAIL, bgColor=FILL_FAIL),
        font=st["Font"](bold=True, color=COLOR_FLAG), stopIfTrue=False))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'{col_letter}{r0}="PASS"'],
        fill=PatternFill("solid", fgColor=FILL_PASS, bgColor=FILL_PASS),
        font=st["Font"](bold=True, color=COLOR_PASS), stopIfTrue=False))


def pick_temps(sweeps, room_target=25.0):
    """挑三个代表温度：常温 / 最低 / 最高。从数据里取，不写死。

    取全部芯片温度点的**交集**——只有大家都测过的温度，同一列横着比才有意义。
    交集空了（各片温度栅格对不上）退回并集，缺的格子自然留空。
    """
    grids = [set(sw.temps) for sw in sweeps if sw.temps]
    if not grids:
        return [None] * 3, [None] * 3
    common = set.intersection(*grids) or set.union(*grids)
    ts = sorted(common)
    room = min(ts, key=lambda t: abs(t - room_target))
    pick = [room, ts[0], ts[-1]]
    # 去重但保住列数（只测过 1~2 档温度时后面几列空着）
    out, seen = [], set()
    for t in pick:
        out.append(None if t in seen else t)
        seen.add(t)
    return out, [f"{fmt_num(t)}℃" if t is not None else None for t in out]


def _limit_dropdown(ws, r0, r1):
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"≤,≥,range"', allow_blank=True)
    dv.error = "只能填 ≤ / ≥ / range"
    # ★ 这段提示只在**点进格子要填的时候**弹出来，不打印、不占版面——
    #   正表不写给填表人看的话，说的是纸面上的字。而"填 ≤ 会怎么判"
    #   恰恰是填的那一秒最该知道的事（判错方向就是 2026-08-05 那个 bug）。
    dv.promptTitle = "判定方向"
    dv.prompt = ("≤ 只判上限（往小超出算好，PASS）\n"
                 "≥ 只判下限（往大超出算好，PASS）\n"
                 "range 两头都判（必须落在 Min~Max 里）\n"
                 "单边时界填 Min 格或 Max 格都认")
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{_cl(C_LIMIT)}{r0}:{_cl(C_LIMIT)}{r1}")


def write_summary(wb, tables, chips, st, slim=False, sinfo=None, dsb=False,
                  spec=None):
    """tables = [(模块名, {芯片: Sweep}), ...]，一个模块一张表，从上到下排。"""
    ws = wb.create_sheet("PLL_Summary")
    n = len(chips)
    ws.column_dimensions[_cl(C_ITEM)].width = 24
    ws.column_dimensions[_cl(C_UNIT)].width = 8
    ws.column_dimensions[_cl(C_LIMIT)].width = 8
    for c in list(range(C_SPEC, C_SPEC + 3)) + list(range(C_SIM, C_SIM + 3)) + \
            list(range(C_SUM, C_SUM + 3)):
        ws.column_dimensions[_cl(c)].width = 10
    ws.column_dimensions[_cl(C_JUDGE)].width = 8
    for c in (C_GAP1, C_GAP2, C_GAP3):
        ws.column_dimensions[_cl(c)].width = 2
    for k in range(n):
        for j in range(CHIP_AX):
            ws.column_dimensions[_cl(chip_col(k) + j)].width = \
                10 if j in CHIP_NUMCOL else 11        # @℃ 列要放得下"等 N 处"
        ws.column_dimensions[_cl(chip_col(k) + CHIP_AX)].width = 2
    ws.column_dimensions[_cl(note_col(n))].width = 26

    r = 1
    judged = []
    for mod, data in tables:
        sweeps = [s for s in data.values() if s is not None]
        items = canon_items(sweeps)
        tpick, tlabels = pick_temps(sweeps)
        t0 = r                                   # 这张表的第一行（大标题）
        # 表键＝没加折算注解那截标题：折算倍数换了 spec 也还认得出是同一张表
        base = f"{mod} PLL 性能汇总"
        full = base + scale_title((sinfo or {}).get(mod), dsb)
        sp = spec.table(ws.title, base, full) if spec is not None else None
        r = _header(ws, r, chips, st, full, n, tlabels)
        r = _cond_rows(ws, r, chips, data, st, n)
        j0 = r
        for band, rows in items:
            r = _band(ws, r, band, st, n)
            b0 = r
            for label, unit in rows:
                r = _result_row(ws, r, label, unit, chips, data, st, n, tpick,
                                band=band, spec=sp)
            # 组内超过 4 行就每 4 行给一条横向导引线（SpotPN 有 8 行）
            if r - b0 > 4:
                for rr in range(b0 + 3, r - 1, 4):
                    _hguide(ws, rr, C_ITEM, note_col(n))
        judged.append((j0, r - 1))
        _limit_dropdown(ws, j0, r - 1)
        # 汇总组＝判定的依据，给它左右两道中等粗细竖框（灰度打印下也分得开）
        _edges(ws, t0, r - 1, C_SUM, C_SUM + 2)
        r += 2
    for a, b in judged:
        _pass_fail_cf(ws, _cl(C_JUDGE), a, b, st)
        _over_spec_cf(ws, a, b, st)
    # ★ 只冻结列不冻结行：一页上下两张表，冻住行的话滚到下面那张时，
    #   顶上钉着的还是上面那张的标题（"<模块A> PLL 性能汇总"），指着下面那张
    #   模块的数写着上面那个模块的名字——比丢掉表头更容易看错。表本身只 20 来行，
    #   在一张表里滚动表头不会跑掉。
    ws.freeze_panes = f"{_cl(C_CHIP0)}1"
    if slim:
        # tpick = [常温, 最低温, 最高温]，留第 0 列
        fold_temp_cols(ws, chip_col, n, CHIP_AX, keep=0)
    return ws


# ---------------------------------------------------------------- 温巡页

# ★★ 竖条宽度与图宽必须由**同一个数**推出来。原来图宽写死 14.5cm、列宽写死
#   5/9×7（＝14.15cm），两边各改各的——每张图往右边邻居里压 13px，
#   三颗芯片的图连成一片（2026-08-04 用户报的"两个温巡的图片重叠了"）。
#   现在：给一个目标图宽 → fit_strip 等比撑开数据列 → 图宽再由实际列宽反算，
#   谁都跑不掉。图占几行也按图高算，不写死。
JCOLS = ["序", "温度℃", "事件", "Vtune_V", "重锁点", "Freq_MHz", "Δf (kHz)", "重锁点"]
JW_BASE = [5] + [9] * 7        # 撑开之前的列宽（比例）
GAP_W = 2                      # 竖条之间的间隔列
STRIP_W = len(JCOLS) + 1
CHART_W_CM = 15.5              # 默认图宽（--chart-w 可改）
CHART_AR = 8.2 / 14.5          # 高宽比，沿用原来那张图的比例


def _journey_rows(sw):
    """按实际测试顺序摊平：[(序, 温度, 事件, vtune, freq, Δf, 是否重锁), ...]

    Δf 相对**本颗芯片自己**的第一个测点——各片绝对频率不同，
    跨片比的是漂移量，不是绝对值。
    """
    vt = sw.item("Vtune_V") or next((i for i in sw.items
                                     if "vtune" in i.label.lower()), None)
    fq = sw.item("Freq_MHz") or next((i for i in sw.items
                                      if i.cat == "Frequency"), None)
    seq = [(lg, x) for lg in sw.legs for x in lg.rows]
    f0 = None
    if fq:
        for _lg, x in seq:
            if x.vals.get(fq.col) is not None:
                f0 = x.vals[fq.col]
                break
    out = []
    for i, (lg, x) in enumerate(seq):
        is_lock = x.kind == "lock"
        v = x.vals.get(vt.col) if vt else None
        f = x.vals.get(fq.col) if fq else None
        df = None if (f is None or f0 is None) else round((f - f0) * 1000.0, 3)
        out.append((i + 1, x.temp,
                    f"锁@{fmt_num(lg.lock_temp)}℃" if is_lock else None,
                    v, f, df, is_lock))
    return out, (vt.unit if vt else "V"), f0


def _jchart(ws, kind, chip, mod, col0, r_data, n_rows, bounds, st, title_extra="",
            rows=None, size=(14.5, 8.2)):
    """一张温巡图：横轴=测试顺序（刻度标当时的温度），重锁点单独一条红三角。

    为什么不按"vs 温度"画：温巡是有先后的，同一个温度会经过好几次；
    按温度画就把先后这一维压没了，看不出"锁完一路漂到哪、下次重锁拉回多少"。
    ★ 同一模块所有芯片用**同一个纵轴范围**，否则并排的两张图没法比。
    """
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import ChartLines

    c_val = col0 + (3 if kind == "vt" else 6)      # 值列（数据块内第 4 / 第 7 列）
    ch = LineChart()
    ch.title = (f"{chip} · {mod} " +
                ("压控电压 温巡过程" if kind == "vt" else "输出频率漂移 温巡过程")
                + title_extra)
    ch.style = 13
    ch.width, ch.height = size
    blank_policy(ch)
    ch.y_axis.title = "Vtune (V)" if kind == "vt" else "Δf (kHz)"
    ch.x_axis.title = "温度 (℃)　—　从左到右＝实际测试先后"
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.y_axis.majorGridlines = ChartLines()
    for c in (c_val, c_val + 1):
        ch.add_data(Reference(ws, min_col=c, min_row=r_data - 1,
                              max_row=r_data + n_rows - 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=col0 + 1, min_row=r_data,
                                max_row=r_data + n_rows - 1))
    style_series(ch.series[0], LEG_STYLE[0][0], "circle", size=4)
    style_series(ch.series[1], COLOR_FLAG, "triangle", line=False, size=10)
    apply_y(ch, bounds)
    legend_bottom(ch)
    skip = max(1, n_rows // 14)
    ch.x_axis.tickLblSkip = skip
    ch.x_axis.tickMarkSkip = skip
    # ★ 压控温巡图标出全程最高/最低两点的数值：这张图的全部意义就是"压控走到哪了、
    #   离轨还剩多少"，那两个数就是结论本身，不该让人在网格里用眼睛估。
    #   （频率漂移那张不标：它故意用了远宽于数据的窗口——记录精度只有 1 kHz，
    #   标出来是 ±1~2 kHz，等于给一条平线打标签。）
    if kind == "vt" and rows:
        from summarize_vco_sweep import _label_points
        pts = [(x[3], i) for i, x in enumerate(rows) if x[3] is not None]
        if pts:
            _label_points(ch.series[0], [min(pts)[1], max(pts)[1]], numfmt="0.0###")
    return ch


def _chart_hole(ws, c0, r0, w, h, text, st):
    """这颗芯片没有这个模块的数据 —— 图位空着，但**必须写明空的是什么**。

    ★★ 图位不能挪：横着一条 band 就是"同一个模块同一张图的各片对照"，
      往上顶一格，右边那颗芯片的图就跟左边错开一行，整页对照关系当场作废。
      所以只能留洞。
    ★★ 但**留白不写字**就是 2026-08-05 用户报的这条："某颗芯片前两张图正常，
      再往下空了两个身位才是下一个模块"——一块无字空白在版面上读起来是
      "图没画出来 / 布局有 bug"，而不是"这颗片子没有这个模块的文件"。
      下面的数据块本来就写着「模块：芯片 未测」，图区却是哑的，两边说法不一致。
      这不算"正表写说明"：它陈述的是**这个位置为什么没有数**，是事实不是用法。
    """
    for r in range(r0, r0 + h):
        for c in range(c0, c0 + w):
            put(ws, r, c, None, st, st["f_group"])
    cell = put(ws, r0, c0, text, st, st["f_group"], size=11, color=COLOR_MUTED)
    ws.merge_cells(start_row=r0, start_column=c0,
                   end_row=r0 + h - 1, end_column=c0 + w - 1)
    cell.alignment = _align("center")


def write_journey(wb, tables, chips, st, no_charts=False, chart_w=CHART_W_CM):
    """一页里：每颗芯片一个竖条；条内每模块两张图 + 两块数据。

    横着一条 band = 同一个模块同一个指标的各芯片对照；竖着一条 = 同一颗芯片。
    """
    ws = wb.create_sheet("温巡")
    n = len(chips)
    jw, cw = fit_strip(JW_BASE, chart_w)      # 列宽撑到放得下图，图宽再由列宽反算
    chart_h = cw * CHART_AR
    CHART_H = chart_rows(chart_h)
    for k in range(n):
        c0 = 1 + k * STRIP_W
        for j in range(len(JCOLS)):
            ws.column_dimensions[_cl(c0 + j)].width = jw[j]
        ws.column_dimensions[_cl(c0 + STRIP_W - 1)].width = GAP_W

    # ★ 只写这一页是什么，不写"该怎么看"。原来这里有一整句版式说明
    #   （一片一竖条／横轴按测试先后／共用纵轴范围）——芯片名就写在每条竖条上面，
    #   横轴刻度和图例自己会说话，那句话是写给我自己的。
    c = put(ws, 1, 1, "温度巡回过程", st, st["f_sep"], bold=True, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n * STRIP_W)
    c.alignment = _align("left", wrap=True)
    ws.row_dimensions[1].height = 30
    for k, chip in enumerate(chips):
        put(ws, 2, 1 + k * STRIP_W, chip, st, st["f_head"], bold=True, size=12)
        for j in range(1, STRIP_W):
            put(ws, 2, 1 + k * STRIP_W + j, None, st, st["f_head"])
        ws.merge_cells(start_row=2, start_column=1 + k * STRIP_W,
                       end_row=2, end_column=k * STRIP_W + STRIP_W)

    # 先把数据块写在下半页（图放上半页，图的引用照样指得到）
    prepared = {}
    r_data = 3 + (0 if no_charts else 2 * len(tables) * CHART_H)
    for mod, data in tables:
        head_row = r_data + 1
        maxn = 0
        for k, chip in enumerate(chips):
            sw = data.get(chip)
            c0 = 1 + k * STRIP_W
            put(ws, r_data, c0, f"{mod} 温巡数据（按测试顺序）" if sw is not None
                else f"{mod}：{chip} 未测", st, st["f_sep"], bold=True, align="left")
            for j in range(1, STRIP_W):
                put(ws, r_data, c0 + j, None, st, st["f_sep"])
            ws.merge_cells(start_row=r_data, start_column=c0,
                           end_row=r_data, end_column=c0 + STRIP_W - 1)
            for j, h in enumerate(JCOLS):
                put(ws, head_row, c0 + j, h, st, st["f_head"], bold=True, size=9)
            if sw is None:
                continue
            rows, vunit, f0 = _journey_rows(sw)
            maxn = max(maxn, len(rows))
            for i, (seq, temp, ev, v, f, df, is_lock) in enumerate(rows):
                r = head_row + 1 + i
                fill = st["f_group"] if is_lock else st["f_res"]
                put(ws, r, c0 + 0, seq, st, fill, size=9)
                put(ws, r, c0 + 1, fmt_num(temp), st, fill, size=9)
                put(ws, r, c0 + 2, ev, st, fill, size=9, bold=is_lock,
                    color=COLOR_FLAG if is_lock else None)
                for j, val in ((3, v), (4, v if is_lock else None),
                               (5, f), (6, df), (7, df if is_lock else None)):
                    cell = put(ws, r, c0 + j, fmt_num(val, 4 if j < 5 else 3), st, fill,
                               size=9)
                    if val is not None:
                        cell.number_format = "0.0000" if j < 5 else \
                            ("0.000" if j == 5 else "0.0")
            prepared.setdefault(mod, {})[chip] = (head_row + 1, len(rows), rows, f0)
        r_data = head_row + 1 + maxn + 1

    if no_charts:
        return ws

    # 纵轴范围：同一模块所有芯片一起算，横向才可比
    row = 3
    for m, (mod, data) in enumerate(tables):
        got = prepared.get(mod, {})
        vt_all = [x[3] for ch_ in got.values() for x in ch_[2]]
        df_all = [x[5] for ch_ in got.values() for x in ch_[2]]
        b_vt = axis_bounds(vt_all)
        dfs = [abs(x) for x in df_all if x is not None]
        # ★ 频率漂移图故意用宽窗：记录精度只有 1 kHz，贴着数据画会把量化噪声
        #   放大成满屏方波，评审第一句必然是"频率为什么在跳"。
        lim = max(3.0 * max(dfs), 5.0) if dfs else 5.0
        step = nice_step(2 * lim)
        lim = -(-lim // step) * step
        b_df = (-lim, lim, step)
        for kind, bounds, band in (("vt", b_vt, 0), ("df", b_df, 1)):
            for k, chip in enumerate(chips):
                if chip not in got:
                    _chart_hole(ws, 1 + k * STRIP_W, row + band * CHART_H,
                                STRIP_W - 1, CHART_H, f"{mod}：{chip} 未测", st)
                    continue
                first, cnt, _rows, f0 = got[chip]      # _rows 用来定位极值点
                extra = (f"（相对首点 {fmt_num(f0, 6)} MHz）"
                         if kind == "df" and f0 is not None else "")
                ch = _jchart(ws, kind, chip, mod, 1 + k * STRIP_W, first, cnt,
                             bounds, st, extra, rows=_rows, size=(cw, chart_h))
                ws.add_chart(ch, f"{_cl(1 + k * STRIP_W)}{row + band * CHART_H}")
        row += 2 * CHART_H
    return ws


# ---------------------------------------------------------------- VCO 页

# 结论行的定义全部复用 summarize_vco_sweep.build_conclusion（Fmin/Fmax 怎么由两段
# 扫描拼出来、Margin 怎么算、Kvco 怎么逐区间取——都是逐轮改出来的，不重新发明）。
# 这里只做两件事：删掉不进评审表的行、把它按芯片横排。
VCO_DROP = {
    "CT Band Step (avg)",       # 平均步长下不了判断，spec 只约束最坏情况
    "Drift in CT Codes",        # 与 Freq Drift + CT Band Step 强冗余（≈ 前者÷后者）
    "Current_mA",               # 电流另有专门的测试表格
    "Margin to fVCO (low)",     # 用户 2026-07-30 删：Fmin/Fmax 直接对 Spec 判就够了
    "Margin to fVCO (high)",
    "Kvco max/min",             # 用户 2026-07-30 删
    # ★★ 下面两行一起删（用户两次说 Sub-band Tuning Range 看不懂）。
    #   它们唯一回答的问题是"有没有锁不上的盲区"（子带宽度必须大于相邻码步长），
    #   而**这份测试已经用实测回答了那个问题**：闭环锁定行在三个温度都锁上了、
    #   PLL 温扫在 16 个温度里一直保持锁定。真有盲区落在目标频率上，它就锁不上。
    #   所以这两行是设计余量的话题，不是这颗芯片的判定项。
    #   只删分子（子带宽度）会留下一个没有分子的分母，比两个都留更糟——
    #   所以 CT Band Step (max) 一起走。它还有个毛病：只有常温算得出来，
    #   一行三格里两格是空的，本身就长得像"缺数据"。
    #   要查随时有：单簿脚本 summarize_vco_sweep.py 的结论页两行都还在。
    "Sub-band Tuning Range",
    "CT Band Step (max)",
    "Target fVCO",              # 用户 2026-07-30 删（Margin 那两行已经没了，它成了孤立参照）
    # ★ Kvco min/max 删（用户 2026-07-30）：它们是"逐区间斜率的极值"，回答的是
    #   「带内增益平不平」——那是个**形状**问题，一条曲线一秒看明白、两个数字反而
    #   看不明白（用户看着它们问"这到底是啥"就是证据）。而且表里那两个数**就是**
    #   VCO压控 页 Kvco-vs-Vtune 那条曲线的最低点和最高点，同一件事说了两遍。
    #   形状交给图（图上给最高/最低点打了数值标注），表里换成两个能下判断的数：
    #   Kvco average（可核）+ Kvco @工作点（环路带宽真正用的那个）。
    "Kvco min",
    "Kvco max",
}
# ★★ 2026-08-04 评审意见：Fmin/Fmax 改回**通行定义**＝Vtune 钉在 CT 扫那个电压上、
#   把电容阵列扫一遍的两端（口径改在 build_conclusion 里，单簿脚本跟着一起变）。
#   随之：`Frequency Range` 整组没了（它的三行都挪进 CT Band）、
#   `CT Band Coverage` 删了（新口径下跟 Tuning Range 是同一个数）、
#   Kvco 组前面多两行实测端点 `F(Vtune=最低/最高)`（Kvco average 的原料，可核）。
# 判定方向的补丁：build_conclusion 里 Kvco 那几行 dir 是空的（单簿页只作对照），
# 跨芯片表要判定就得写明方向。Kvco 的 spec 通常给一个范围：
# 太小则环路带宽不够/锁不住，太大则杂散与噪声恶化。
VCO_LIMIT_FIX = {"Kvco average": "range"}
# 备注补一句"这是什么"。★ 只补定义，不补解读——评审要的是"这个数怎么来的"。
VCO_NOTE_ADD = {}      # 解释性的尾巴全去掉了，备注只留算式
# ★ 备注整句改写：让算式里的每一项都**正好是表上某一行的名字**，
#   这样导出来的数用眼睛就能核（用户原话："缺少了计算他们的中间值，我有些没安全感"）。
# ★ 备注里**只留算式**。原来每条后面还跟着一句解释（"＝整个电容阵列能覆盖多宽"
#   "两项都在上面的实测端点行里，可以直接核"）——评审知道 Tuning Range 是什么，
#   也不用我告诉他去哪儿核；那些话跟表头里的"（各片最小）"是同一类。
VCO_NOTE_SET = {
    # 纯出处（"原表 Temperature 列"）不进正表：那是给我自己对数用的
    "Temperature": "",
}
# ★ 空的：`Kvco @工作点` 请回来了。我当初提议删它的理由是"它落在 Kvco min/max
#   之间、判定上冗余"——那个理由依赖 min/max 存在。min/max 一走，它就是环路带宽
#   真正用的那个数，而且是全表唯一说得清"工作点增益"的行。
# ★★ 开环 VCO 不报 IPN（2026-08-05，用户师父指出，物理上站得住）：
#   自由振荡的 VCO 相噪在低 offset 上按 1/f³、1/f² **发散**，积分完全被最低那个
#   offset 支配——那个数说的是"从哪儿起积"，不是 VCO 好坏；而且自由振荡会**漂**，
#   测量期间载波自己在走，近端"相噪"其实是频率漂移。这一段恰恰是闭环里被环路
#   压掉的，所以开环 IPN 既不反映 VCO、也不预测系统。
#   ★这条一次解释掉三个"数据怪怪的"：IPN 出现**正值**（>1 弧度 RMS）、
#   IPN 与 IPN_Omit 散布 5~10 dB、跟工作频率走 11.6 dB。都是发散积分 + 漂移。
#   开环该看的是**指定 offset 的单点相噪**（SpotPN@* 八档都在），删掉不丢东西。
#   PLL 页照旧报 IPN——闭环下它是真指标。单簿脚本的结论页也还留着。
VCO_DROP_PREFIX = ("IPN",)


def vco_rows(sw, ref_temp, op_vtune):
    """一颗芯片一个模块的结论行 -> [(cat, item, unit, dir, kind, {温度: 值})]"""
    from summarize_vco_sweep import build_conclusion
    rows, temps = build_conclusion(sw.by_kind, sw.items, sw.freq_item, ref_temp,
                                   sw.fvco, sw.fvco_ref, {}, op_vtune)
    out = []
    for d in rows:
        if d["item"] in VCO_DROP or d["item"].startswith(VCO_DROP_PREFIX):
            continue
        vals = dict(d["vals"])
        note = VCO_NOTE_SET.get(d["item"], d.get("note") or "")
        if d["item"].startswith("Spur@"):
            note = spur_note(sw, d["item"]) or note
        add = VCO_NOTE_ADD.get(d["item"]) if d["item"] not in VCO_NOTE_SET else None
        if add:
            note = f"{note}；{add}" if note else add
        out.append({"cat": d["cat"], "item": d["item"], "unit": d["unit"],
                    "dir": VCO_LIMIT_FIX.get(d["item"], d["dir"]),
                    "kind": d["kind"], "vals": vals, "note": note})
    out = _drift_at_op(out, ref_temp)
    return out, temps


def _drift_at_op(out, ref_temp):
    """温漂改成**工作点**的漂移：Freq@T − Freq@常温，两项都是表上 `Freq` 那一行。

    ★ 原来是"14 个 Vtune 点各算一个差、取绝对值最大的那个"——14 个中间量全在表外，
      看的人只能选择信。改成工作点之后零新增行、完全可核，而且符合这条线早就定过的
      规矩：性能按工作点报，不跨扫描取最差。
      漂移量在不同 Vtune 上真差很多的话，那是另一个更该被单独看见的现象，
      不该被一个"取最坏"悄悄吸收掉。
    """
    fr = next((x for x in out if x["cat"].startswith("@ Vtune")
               and x["item"] == "Freq"), None)
    di = next((k for k, x in enumerate(out)
               if x["item"].startswith("Freq Drift")), None)
    if fr is None or di is None:
        return out
    ts = [t for t, v in fr["vals"].items() if isinstance(v, (int, float))]
    if not ts:
        return out
    ref = min(ts, key=lambda t: abs(t - ref_temp))
    base = fr["vals"][ref]
    out[di] = dict(out[di],
                   item=f"Freq Drift vs {fmt_num(ref)}℃",
                   vals={t: fr["vals"][t] - base for t in ts if t != ref},
                   note=f"@工作点 那一行的 Freq：F(T) − F({fmt_num(ref)}℃)")
    return out


def flat_vtune_temps(sw, ratio=0.1):
    """哪些温度的 Vtune 扫「频率几乎不动」。返回 [(温度, 该温度频率跨度, 各温中位)]。

    ★ 为什么要专门查这个：Kvco 是逐区间 ΔF/ΔVtune 算的。如果某个温度实际上没真正
      切到开环（环路还锁着 / DAC 没驱动到管子 / testmux 没切），那个温度的频率
      就不跟 Vtune 走，Kvco 会算出接近 0 的值——而 0.00 MHz/V 这种数看着像
      "低温增益低"，其实是"这个温度的扫描没生效"。工具不能安静地报它。
      判据用相对量（跟其他温度的中位数比），不写死绝对阈值：不同 VCO 差几十倍。
    """
    from summarize_vco_sweep import group_series
    spans = {}
    for kind, groups in sw.by_kind:
        if kind != "vtune":
            continue
        for g in groups:
            if g.temp is None:
                continue
            ser = group_series(g, sw.freq_item)
            if len(ser) >= 2:
                xs = sorted(ser)
                spans[g.temp] = abs(ser[xs[-1]] - ser[xs[0]])
    if len(spans) < 2:
        return []
    med = median(list(spans.values()))
    if not med:
        return []
    return [(t, sp, med) for t, sp in sorted(spans.items()) if sp < ratio * med]


def coarse_temps(sw):
    """CT 扫的码不是逐码密扫的那些温度。相邻码步长 > 1 就算粗扫。"""
    from summarize_vco_sweep import group_series
    out = set()
    for kind, groups in sw.by_kind:
        if kind != "ct":
            continue
        for g in groups:
            xs = sorted(group_series(g, sw.freq_item))
            if len(xs) < 2:
                continue
            if min(b - a for a, b in zip(xs, xs[1:])) > 1:
                out.add(g.temp)
    return out


def op_vtune_of(sw):
    """CT 扫把 Vtune 钉在哪个值上——工作点就取它。"""
    for kind, groups in sw.by_kind:
        if kind == "ct":
            for g in groups:
                xs = [r.vt for r in g.rows if r.vt is not None]
                if xs:
                    return xs[0]
    return None


def write_vco_summary(wb, vtables, chips, st, vtemps, slim=False,
                      sinfo=None, dsb=False, spec=None):
    """VCO 汇总表：每颗芯片的组内轴＝三个温度（不是 Min/Typ/Max）。

    ★ 为什么这里用温度当轴、PLL 那页用极值当轴：VCO 这些量本来就是"一个温度
      算出一个数"，而温度只有 3 档。把 3 个数压成 Min/Typ/Max 还是占 3 列，
      却把"哪个温度"丢了——纯信息损失。PLL 那边是 49 个测点压成 3 个数，
      压缩有真收益。
    """
    return write_grouped_page(wb, "VCO_Summary", "{mod} VCO 开环特性汇总",
                              vtables, chips, st, vtemps, slim=slim,
                              sinfo=sinfo, dsb=dsb, spec=spec)


def write_grouped_page(wb, sheet, title_fmt, vtables, chips, st, vtemps,
                       slim=False, item_w=26, note_w=46, sinfo=None, dsb=False,
                       spec=None):
    """「一行一个量 × 每片若干温度」这种页的通用写法。

    ★ VCO 汇总页和电流页是同一个形状：行是结论量、组内轴是温度、汇总列对
      全部片全温取。所以只留这一份实现——不是为了少打字，是因为出稿自查
      （selfcheck）按这个骨架逐格反推汇总列，两份实现一漂移，自查就只护得住一页。
      两页的差别只有：页名、大标题、行从哪儿来。
    """
    ws = wb.create_sheet(sheet)
    n = len(chips)
    nax = len(vtemps)
    cw = nax + 1

    def vchip(k):
        return C_CHIP0 + k * cw

    def vnote():
        return vchip(n)

    def vrails():
        return [C_GAP1, C_GAP2, C_GAP3] + [vchip(k) + nax for k in range(n)]

    ws.column_dimensions[_cl(C_ITEM)].width = item_w
    ws.column_dimensions[_cl(C_UNIT)].width = 9
    ws.column_dimensions[_cl(C_LIMIT)].width = 8
    for c in list(range(C_SPEC, C_SPEC + 3)) + list(range(C_SIM, C_SIM + 3)) + \
            list(range(C_SUM, C_SUM + 3)):
        ws.column_dimensions[_cl(c)].width = 10
    ws.column_dimensions[_cl(C_JUDGE)].width = 8
    for c in (C_GAP1, C_GAP2, C_GAP3):
        ws.column_dimensions[_cl(c)].width = 2
    for k in range(n):
        for j in range(nax):
            ws.column_dimensions[_cl(vchip(k) + j)].width = 11
        ws.column_dimensions[_cl(vchip(k) + nax)].width = 2
    ws.column_dimensions[_cl(vnote())].width = note_w

    r = 1
    judged = []
    for mod, data in vtables:
        t0 = r
        # ---- 三行表头 ----
        base = title_fmt.format(mod=mod)
        full = base + scale_title((sinfo or {}).get(mod), dsb)
        sp = spec.table(sheet, base, full) if spec is not None else None
        c = put(ws, r, C_ITEM, full,
                st, st["f_sep"], bold=True, align="left", size=12)
        ws.merge_cells(start_row=r, start_column=C_ITEM, end_row=r, end_column=vnote())
        c.alignment = _align("left", wrap=False)
        hr, ar = r + 1, r + 2
        for col, name in ((C_ITEM, "测试项"), (C_UNIT, "Unit"),
                          (C_LIMIT, "Limit"), (C_JUDGE, "判定")):
            put(ws, hr, col, name, st, st["f_head"], bold=True)
            put(ws, ar, col, None, st, st["f_head"])
            ws.merge_cells(start_row=hr, start_column=col, end_row=ar, end_column=col)
        groups = [(C_SPEC, "Spec", AXES, 3),
                  (C_SIM, "仿真", AXES, 3),
                  (C_SUM, f"汇总 · {n} 片", AXES, 3)]
        for k, chip in enumerate(chips):
            groups.append((vchip(k), chip,
                           [f"{fmt_num(t)}℃" for t in vtemps], nax))
        for col, name, axes, w in groups:
            put(ws, hr, col, name, st, st["f_head"], bold=True)
            for j in range(1, w):
                put(ws, hr, col + j, None, st, st["f_head"])
            ws.merge_cells(start_row=hr, start_column=col, end_row=hr,
                           end_column=col + w - 1)
            fill = st["f_in"] if col in (C_SPEC, C_SIM) else st["f_head"]
            for j, lb in enumerate(axes):
                put(ws, ar, col + j, lb, st, fill, bold=True, size=9)
        for col in vrails():
            for rr in (hr, ar):
                put(ws, rr, col, None, st, st["f_rail"])
        put(ws, hr, vnote(), "备注", st, st["f_head"], bold=True)
        put(ws, ar, vnote(), None, st, st["f_head"])
        ws.merge_cells(start_row=hr, start_column=vnote(), end_row=ar,
                       end_column=vnote())
        ws.row_dimensions[ar].height = 30
        r = ar + 1

        # ★ 这里原来有一行口径说明（每片三列是什么、汇总怎么取、判定看哪头、
        #   CT 全码扫只在常温）。2026-08-04 整行删掉：全是读表说明。
        #   "CT 全码扫只在常温" 这条**是测试条件、不是说明**，它已经在下面的
        #   条件行 `CT Code Range (points)` 里，不必在正表上再说一遍。

        # ---- 行序：按 (cat, item) 对齐各片；cat 变了插一条分组带 ----
        order, seen = [], set()
        notes_of = {}
        for chip in chips:
            for d in (data.get(chip) or []):
                key = (d["cat"], d["item"])
                notes_of.setdefault(key, d["note"])
                if key not in seen:
                    seen.add(key)
                    order.append((d["cat"], d["item"], d["unit"], d["dir"],
                              d["kind"], d.get("nd"), d.get("strong", False)))
        j0 = None
        cur_cat = None
        band0 = None
        for cat, item, unit, dr, kind, nd_, strong in order:
            if cat != cur_cat:
                if band0 is not None and r - band0 > 4:
                    for rr in range(band0 + 3, r - 1, 4):
                        _hguide(ws, rr, C_ITEM, vnote())
                put(ws, r, C_ITEM, cat, st, st["f_sep"], bold=True, align="left")
                for cc in range(C_ITEM + 1, vnote() + 1):
                    put(ws, r, cc, None, st, st["f_sep"])
                r += 1
                cur_cat, band0 = cat, r
            is_res = kind == "result"
            if is_res and j0 is None:
                j0 = r
            nd = nd_ or (3 if unit == "MHz" else 2)
            body = st["f_res"] if is_res else st["f_group"]
            spr = sp.row(cat, item) if (sp is not None and is_res) else None
            put(ws, r, C_ITEM, item, st, body, align="left", bold=strong)
            put(ws, r, C_UNIT, unit, st, body, size=9)
            put(ws, r, C_LIMIT, ((spr or {}).get("limit") or dr) if is_res
                else None, st, st["f_in"] if is_res else body, size=9)
            _fill_spec(ws, r, st, nd, spr, st["f_in"] if is_res else body)
            for cc in vrails():
                put(ws, r, cc, None, st, st["f_rail"])

            allv, roomv = [], []
            for k, chip in enumerate(chips):
                got = {(d["cat"], d["item"]): d["vals"] for d in (data.get(chip) or [])}
                vals = got.get((cat, item), {})
                for j, t in enumerate(vtemps):
                    v = vals.get(t)
                    disp = fmt_num(v, nd) if isinstance(v, (int, float)) else v
                    cell = put(ws, r, vchip(k) + j, disp, st, body, bold=strong,
                               size=10 if isinstance(disp, (int, float)) else 9)
                    if isinstance(disp, (int, float)):
                        cell.number_format = "0." + "0" * nd
                        allv.append(disp)
                        if t == _room_of(vtemps):
                            roomv.append(disp)
                put(ws, r, vchip(k) + nax, None, st, st["f_rail"])

            agg = [min(allv) if allv else None,
                   median(roomv) if roomv else None,
                   max(allv) if allv else None] if is_res else [None] * 3
            for j, v in enumerate(agg):
                cell = put(ws, r, C_SUM + j, v, st,
                           st["f_sum"] if is_res else body, bold=True)
                if v is not None:
                    cell.number_format = "0." + "0" * nd
            put(ws, r, C_JUDGE, _judge_formula(r) if is_res else None, st,
                st["f_res"] if is_res else body, bold=True)
            # 备注只写"哪一片整行没数"。不要去说某一个温度格为什么空：
            # 有些空格是设计上的（温漂对参考温自己恒等于 0，那格不是测量结果；
            # 粗码温度算不出步长），写成"缺数据"是误报。
            # ★ 备注 = 这个数怎么算出来的（算式，来自单簿脚本的 note）。
            #   评审第一句就是"这怎么算出来的"；不写算式就得每次口头解释一遍。
            note = notes_of.get((cat, item), "")
            if is_res:
                miss = []
                for c_ in chips:
                    got = {(d["cat"], d["item"]): d["vals"]
                           for d in (data.get(c_) or [])}
                    vals_ = got.get((cat, item)) or {}
                    if not any(isinstance(v, (int, float)) for v in vals_.values()):
                        miss.append(c_)
                if miss:
                    note = (note + "；" if note else "") +                            f"没有这一项: {', '.join(miss)}"
            put(ws, r, vnote(), as_text(note), st, body, align="left", size=9,
                color=COLOR_MUTED)
            r += 1
        if band0 is not None and r - band0 > 4:
            for rr in range(band0 + 3, r - 1, 4):
                _hguide(ws, rr, C_ITEM, vnote())
        if j0 is not None:
            judged.append((j0, r - 1))
            _limit_dropdown(ws, j0, r - 1)
        _edges(ws, t0, r - 1, C_SUM, C_SUM + 2)
        r += 2
    for a, b in judged:
        _pass_fail_cf(ws, _cl(C_JUDGE), a, b, st)
        _over_spec_cf(ws, a, b, st)
    ws.freeze_panes = f"{_cl(C_CHIP0)}1"
    if slim and vtemps:
        # ★ 这里的列是**升序温度**，第 0 列是最低温。留常温＝留 _room_of 那一列，
        #   照搬 PLL 页的 keep=0 会折得只剩 −40℃。
        fold_temp_cols(ws, vchip, n, nax, keep=vtemps.index(_room_of(vtemps)))
    return ws


def _room_of(vtemps, target=25.0):
    return min(vtemps, key=lambda t: abs(t - target)) if vtemps else None


# ---- VCO 图 --------------------------------------------------------------

VSTRIP_W = 9            # 一颗芯片一竖条：8 列数据 + 1 列间隔
VW_BASE = [10] * (VSTRIP_W - 1)
# 四张图：Vtune 轴给"值 + 斜率"，CT 轴也给"值 + 斜率"，对称
VCO_TAGS = ("v", "k", "c", "d")
VCO_TITLE = {"v": "频率 vs Vtune", "k": "Kvco vs Vtune", "c": "频率 vs CT 码",
             "d": "δf vs CT 码（仅常温）"}
VCO_XLABEL = {"v": "Vtune (V)", "k": "Vtune (V)", "c": "CT code", "d": "CT code"}
VCO_YHEAD = {"v": "F", "k": "Kvco", "c": "F", "d": "δf"}
VCO_YAXIS = {"v": "F (MHz)", "k": "Kvco (MHz/V)", "c": "F (MHz)",
             "d": "|δf| (MHz/code)"}
# 四张图都给**全图最低/最高两点**打数值标注。值图（F-vs-Vtune / F-vs-CT）的曲线
# 是单调的，全图最低/最高点正好落在两个端点上，要的就是它们；
# 斜率图上则是最平和最陡的那两个区间。
# 只标两点是这条线定过的规矩——每条线各标首末点会在同一个横轴端点上叠成一坨。
VCO_LABELED = ("v", "k", "c", "d")


def _vco_series(sw, temps):
    """四组曲线：F-vs-Vtune / Kvco-vs-Vtune / F-vs-CT码 / ΔF-vs-CT码。

    ★ ΔF-vs-CT码（相邻两个码的频率差）**只给密扫的那个温度**。高低温只测几个
      粗码，"相邻两点"跨了几十个码，算出来是那几十个码的平均，跟常温的逐码步长
      不是一个量——混在一张图上就是骗人。这也正是表里那一行被删掉的原因。
    """
    from summarize_vco_sweep import group_series, slopes
    fv, kv, fc, fd = {}, {}, {}, {}
    coarse = coarse_temps(sw)
    for kind, groups in sw.by_kind:
        for g in groups:
            if g.temp is None:
                continue
            ser = group_series(g, sw.freq_item)
            pts = sorted(ser.items())
            if kind == "vtune":
                fv[g.temp] = pts
                kv[g.temp] = [(sl[0], abs(sl[1])) for sl in slopes(g, sw.freq_item)]
            elif kind == "ct":
                fc[g.temp] = pts
                if g.temp not in coarse:
                    fd[g.temp] = [(sl[0], abs(sl[1]))
                                  for sl in slopes(g, sw.freq_item)]
    return fv, kv, fc, fd


def write_vco_charts(wb, vtables, chips, st, vtemps, no_charts=False,
                     chart_w=CHART_W_CM):
    """一页里：每颗芯片一竖条，条内三张图（F-Vtune / Kvco-Vtune / F-CT码）+ 数据源。

    横着一条 band = 同一张图的各片对照；同一模块各片共用纵轴范围，才能直接比。
    """
    ws = wb.create_sheet("VCO压控")
    n = len(chips)
    vw, cw = fit_strip(VW_BASE, chart_w)      # 跟温巡页同一套：列宽撑到放得下图
    chart_h = cw * CHART_AR
    VCHART_H = chart_rows(chart_h)
    for k in range(n):
        c0 = 1 + k * VSTRIP_W
        for j in range(VSTRIP_W - 1):
            ws.column_dimensions[_cl(c0 + j)].width = vw[j]
        ws.column_dimensions[_cl(c0 + VSTRIP_W - 1)].width = GAP_W

    c = put(ws, 1, 1, "VCO 开环压控", st, st["f_sep"], bold=True, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n * VSTRIP_W)
    c.alignment = _align("left", wrap=True)
    ws.row_dimensions[1].height = 32
    for k, chip in enumerate(chips):
        put(ws, 2, 1 + k * VSTRIP_W, chip, st, st["f_head"], bold=True, size=12)
        for j in range(1, VSTRIP_W):
            put(ws, 2, 1 + k * VSTRIP_W + j, None, st, st["f_head"])
        ws.merge_cells(start_row=2, start_column=1 + k * VSTRIP_W,
                       end_row=2, end_column=(k + 1) * VSTRIP_W)

    nb = len(VCO_TAGS) * len(vtables)      # 每个模块几个 band
    r_data = 3 + (0 if no_charts else nb * VCHART_H)
    prepared = {}
    for mod, data in vtables:
        for tag in VCO_TAGS:
            head = [VCO_XLABEL[tag]] + [f"{VCO_YHEAD[tag]}@{fmt_num(t)}℃"
                                        for t in vtemps]
            title = VCO_TITLE[tag]
            head_row = r_data + 1
            maxn = 0
            for k, chip in enumerate(chips):
                c0 = 1 + k * VSTRIP_W
                ser = (data.get(chip) or {}).get(tag) or {}
                put(ws, r_data, c0, f"{mod} · {title}" if ser
                    else f"{mod} · {title}：{chip} 未测", st, st["f_sep"],
                    bold=True, align="left")
                for j in range(1, VSTRIP_W):
                    put(ws, r_data, c0 + j, None, st, st["f_sep"])
                ws.merge_cells(start_row=r_data, start_column=c0,
                               end_row=r_data, end_column=c0 + VSTRIP_W - 1)
                for j, h in enumerate(head):
                    put(ws, head_row, c0 + j, h, st, st["f_head"], bold=True, size=9)
                if not ser:
                    continue
                xs = sorted({x for t in vtemps for x, _v in ser.get(t, [])})
                maxn = max(maxn, len(xs))
                idx = {x: i for i, x in enumerate(xs)}
                for i, x in enumerate(xs):
                    put(ws, head_row + 1 + i, c0, x, st, st["f_res"], size=9)
                for j, t in enumerate(vtemps):
                    for x, v in ser.get(t, []):
                        cell = put(ws, head_row + 1 + idx[x], c0 + 1 + j,
                                   fmt_num(v, 3), st, st["f_res"], size=9)
                        cell.number_format = "0.000"
                prepared.setdefault((mod, tag), {})[chip] = (head_row + 1, len(xs),
                                                            ser, xs)
            r_data = head_row + 1 + maxn + 1

    if no_charts:
        return ws

    row = 3
    for mod, _data in vtables:
        for tag in VCO_TAGS:
            got = prepared.get((mod, tag), {})
            allv = [v for ch in got.values() for t in vtemps
                    for _x, v in ch[2].get(t, [])]      # ch[2] = ser
            bounds = axis_bounds(allv)
            for k, chip in enumerate(chips):
                if chip not in got:
                    _chart_hole(ws, 1 + k * VSTRIP_W, row, VSTRIP_W - 1,
                                VCHART_H, f"{mod} · {VCO_TITLE[tag]}：{chip} 未测", st)
                    continue
                first, cnt, ser, xs = got[chip]
                ws.add_chart(_vco_chart(ws, tag, chip, mod, 1 + k * VSTRIP_W,
                                        first, cnt, vtemps, bounds, ser, xs,
                                        size=(cw, chart_h)),
                             f"{_cl(1 + k * VSTRIP_W)}{row}")
            row += VCHART_H
    return ws


def _vco_chart(ws, tag, chip, mod, col0, r_data, n_rows, vtemps, bounds, ser, xs,
               size=(14.5, 8.2)):
    from openpyxl.chart import Reference, ScatterChart, Series
    # ★ 借单簿脚本那份数值标注（只标全图最低/最高两点——三条温度曲线在同一个
    #   横轴端点上值挨得很近，各标各的会叠成一坨）。放在那边是因为它先写出来的，
    #   这里不再抄一份，抄了必然漂移。
    from summarize_vco_sweep import _label_points, axis_numfmt, label_pos

    ch = ScatterChart()
    ch.title = f"{chip} · {mod} " + VCO_TITLE[tag]
    ch.style = 13
    ch.width, ch.height = size
    blank_policy(ch)
    ch.x_axis.title = VCO_XLABEL[tag]
    ch.y_axis.title = VCO_YAXIS[tag]
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    xref = Reference(ws, min_col=col0, min_row=r_data, max_row=r_data + n_rows - 1)
    built = {}
    for j, t in enumerate(vtemps):
        pts = ser.get(t, [])
        if not pts:
            continue
        yref = Reference(ws, min_col=col0 + 1 + j, min_row=r_data - 1,
                         max_row=r_data + n_rows - 1)
        sr = Series(yref, xref, title_from_data=True)
        color, sym = LEG_STYLE[j % len(LEG_STYLE)]
        # ★ 点数少的那几条只打记号不连线：高低温 CT 只测几个粗码，
        #   跨几十个码直连出来是一条不存在的曲线，看图的人会以为中间测过。
        sparse = tag == "c" and len(pts) < max(8, n_rows // 4)
        style_series(sr, color, sym, line=not sparse,
                     size=8 if sparse else (3 if len(pts) > 60 else 5))
        ch.series.append(sr)
        built[t] = (sr, color)
    if tag in VCO_LABELED and built:
        pos = {x: i for i, x in enumerate(xs)}
        if built:
            # ★ 同一个系列上的多个标注要**一次交给它**：_label_points 每次调用都
            #   重建 s.dLbls，分两次调等于第二次把第一次的覆盖掉（曲线是直线时
            #   最低点和最高点落在同一条系列上，就会只剩一个标注）。
            # ★★ 只标**一条线**（常温那条）的最低/最高。原来标的是全图两个极值：
            #   最高点落在 −40℃ 那条、最低点落在 105℃ 那条，两个标签一蓝一绿
            #   分属两条线——用户："为什么最高点和最低点的颜色不一样？感觉怪怪的，
            #   我觉得同一条曲线的最高最低才合理"。同一条线的两端才是一对能读出
            #   意思的数（这条线从哪走到哪）；跨温度的包络表里有专门的行。
            want = {}
            for t_lab in sorted(built, key=lambda t: abs(t - 25.0)):
                pts_ = [(y, x) for x, y in ser.get(t_lab, []) if x in pos]
                if not pts_:
                    continue                    # 常温那条没点就顺延
                lo_, hi_ = min(pts_), max(pts_)
                want[t_lab] = [(pos[lo_[1]], False), (pos[hi_[1]], True)]
                break
            for t, items in want.items():
                sr_, color = built[t]
                # 带系列名 + 染成本条线的颜色：三条线只有两个标签时，
                # 不这么做就看不出这两个数是哪条线上的
                _label_points(sr_, [i for i, _m in items], numfmt="0.##",
                              pos={i: label_pos(m) for i, m in items},
                              color=color, sername=True)
    apply_y(ch, bounds)
    # ★ 横轴自己留边距：标注放在极值点的正上/正下，而极值点几乎总在横轴两端，
    #   Excel 自动定的范围常常刚好卡在数据末端，标注就被右边框切掉半截。
    if xs and len(xs) > 1 and max(xs) > min(xs):
        xb = axis_bounds(list(xs), pad=0.12)
        lo_x = 0.0 if (min(xs) >= 0 and xb[0] < 0) else xb[0]
        # 数据非负就别把轴画到负数去：Vtune / CT 码没有负的，轴上冒出 "-0.2 V"
        # 一眼就假。左边那个标注是往**下**放的，不吃左边距，夹掉不影响它。
        ch.x_axis.scaling.min, ch.x_axis.scaling.max = lo_x, xb[1]
        ch.x_axis.majorUnit = xb[2]
    nf = axis_numfmt(bounds)
    if nf:
        ch.y_axis.numFmt = nf
    legend_bottom(ch)
    return ch


# ---------------------------------------------------------------- 审计页

def write_audit(wb, picked, dropped, unknown, failed, notes, st,
                excl_all=(), quality_all=(), spec_rows=()):
    """每个数出自哪份文件 + 哪些文件没用上。**隐藏页**——正表不写这些。"""
    ws = wb.create_sheet("_审计")
    ws.sheet_state = "hidden"
    for i, w in enumerate((44, 10, 12, 70, 14), 1):
        ws.column_dimensions[_cl(i)].width = w
    r = 1
    put(ws, r, 1, "数据来源（每颗芯片每个模块用的是哪一份文件）", st, st["f_sep"],
        bold=True, align="left")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    for i, h in enumerate(["芯片", "模块", "类型", "文件", "规模"], 1):
        put(ws, r, i, h, st, st["f_head"], bold=True)
    r += 1
    for (chip, mod, kind), b in sorted(picked.items(), key=lambda x: natkey(x[0])):
        put(ws, r, 1, chip, st, st["f_res"])
        put(ws, r, 2, mod, st, st["f_res"])
        put(ws, r, 3, KIND_LABEL[kind], st, st["f_res"])
        put(ws, r, 4, b.name, st, st["f_res"], align="left", size=9)
        put(ws, r, 5, notes.get(id(b), "未读（本版不处理这类）"), st, st["f_res"], size=9)
        r += 1
    r += 1
    for title, rowsrc in (
            ("同类多份，只用了时间戳最新的那份（下面是被跳过的）",
             [(b.chip, b.module, KIND_LABEL[b.kind], b.name, f"让位给 {w.ts or w.name}")
              for b, w in dropped]),
            ("文件名认不出模块/类型，没读", [(c, "", "", f, f"认不出{why}")
                                            for c, f, why in unknown]),
            ("读失败", [(b.chip, b.module, KIND_LABEL[b.kind], b.name, why)
                        for b, why in failed]),
            ("排除的行（逐行原因；这些行不进统计）",
             [(chip, mod, kind, f"行{xl}", why)
              for chip, mod, kind, ex in excl_all for xl, why in ex]),
            ("数据可信度（这些只进这里和控制台，不进正表）",
             [(chip, mod, kind, what, why)
              for chip, mod, kind, ex in quality_all for what, why in ex]),
            # 存着 spec、这次却没落到格子上的行。不报出来的话，表上那一行的
            # Spec 列是空的、判定列也是空的——看着像"这项没定 spec"，
            # 其实是名字对不上，静悄悄地漏判了
            ("Spec 没对上的行（配置里填了，这次表上没有这一行）",
             [(sheet, tkey, cat, item, why)
              for sheet, tkey, cat, item, why in spec_rows])):
        if not rowsrc:
            continue
        put(ws, r, 1, title, st, st["f_sep"], bold=True, align="left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1
        for tup in rowsrc:
            for i, v in enumerate(tup, 1):
                put(ws, r, i, as_text(v), st, st["f_res"], align="left", size=9)
            r += 1
        r += 1
    return ws


# ---------------------------------------------------------------- 电流（逐级关断）

class CurBook:
    """一份逐级关断电流簿读出来的样子。"""
    __slots__ = ("runs", "temps", "warnings", "excluded", "n_rows", "n_cols",
                 "key_name", "val_name", "unit")

    def __init__(self):
        self.runs, self.temps, self.warnings, self.excluded = {}, [], [], []
        self.n_rows = self.n_cols = 0
        self.key_name = self.val_name = self.unit = ""


def load_current(path, temp_col=None, key_col=None, val_col=None):
    """读一份逐级关断电流簿 → 每个温度一趟「测量点阶梯」。

    ★★ 值列装的是**每关掉一步之后的总电流**，不是那个模块自己的电流。
      模块电流 ＝ 关它之前那个测量点 − 关它之后那个测量点。
    ★★ 阶梯到哪儿为止**不靠"restore"这类字样去认**（那是模板里的人话，随时会改）：
      逐级关断的总电流是**单调下降**的，第一个明显回升的测量点就是"恢复后复测"。
      用形状判不用文字判——顺带这个复测点还是这趟数据的可信度证据
      （它应该回到全开基线附近）。
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if len(rows) < 2:
        raise SweepError("表里没有数据行")
    out = CurBook()
    out.n_rows, out.n_cols = len(rows), max(len(r) for r in rows)
    cols = Columns(rows[0])
    out.warnings = list(cols.dup_report())

    tname, ti = (temp_col, cols.idx(temp_col)) if temp_col else \
        cols.find(r"temperature", r"^temp")
    if ti is None:
        raise SweepError("找不到温度列，用 --temp-col 指定")
    vname, vi = (val_col, cols.idx(val_col)) if val_col else \
        cols.find(r"current.*m\s*a", r"^i(dd|cc)", r"电流")
    if vi is None:
        raise SweepError("找不到电流值列，用 --cur-col 指定")
    kname, ki = (key_col, cols.idx(key_col)) if key_col else (None, None)
    if ki is None:
        # 键列＝第一列那种"一步一个短名字"的列（模板里表头常叫 NO.）。
        # 找不到就退回第一列，并报一句。
        kname, ki = cols.find(r"^no\.?$", r"^item$", r"^block$", r"^step$")
        if ki is None:
            kname, ki = (txt(rows[0][0]) or "第一列"), 0
            out.warnings.append(f"没认出步骤名那一列，按第一列（{kname}）用；"
                                f"不对的话加 --key-col")
    mname, mi = cols.find(r"^mode$", r"^desc", r"说明")
    out.key_name, out.val_name = kname, vname
    out.unit = "mA"

    # 按温度分趟；同一温度里保持文件顺序
    by_t = OrderedDict()
    for n, raw in enumerate(rows[1:], start=2):
        if all(is_blank(v) for v in raw):
            continue
        key = txt(raw[ki]) if ki < len(raw) else ""
        t = num(raw[ti]) if ti < len(raw) else None
        if t is None:
            out.excluded.append((n, "没有温度值（多半是页脚行）", key))
            continue
        i_tot = num(raw[vi]) if vi < len(raw) else None
        mode = txt(raw[mi]) if (mi is not None and mi < len(raw)) else ""
        by_t.setdefault(t, []).append((n, key, mode, i_tot))

    for t, seq in by_t.items():
        pts = [(n, k, m, i) for n, k, m, i in seq if i is not None]
        if not pts:
            out.excluded.append((0, f"{fmt_num(t)}℃ 这一趟一个测量点都没有", ""))
            continue
        base = pts[0][3]
        steps, prev, cut = [], base, None
        for idx, (n, k, m, i) in enumerate(pts[1:], start=1):
            # 回升超过基线 5% ＝ 恢复段开始了，阶梯到此为止
            if i - prev > abs(base) * 0.05:
                cut = idx
                break
            steps.append((k, m, i, prev - i))
            prev = i
        # ★★ 恢复后复测取**这一趟的最后一个测量点**，不是第一个回升的那个。
        #   恢复常常是分几步写回的（真数据里阶梯后面还挂着十来行），
        #   取第一个回升点＝拿"刚恢复了一小半"的状态去跟全开基线比，
        #   算出来是"差了 58%"这种吓人的数，其实只是取错了行。
        #   （真数据上就是这么报出 −9.58 mA 的。）
        tail = pts[cut:] if cut is not None else []
        out.runs[t] = {"baseline": base, "baseline_key": pts[0][1],
                       "steps": steps, "recheck": (tail[-1] if tail else None),
                       "tail_n": len(tail)}
    out.temps = sorted(out.runs)
    return out


def _norm(x):
    return re.sub(r"[^0-9a-z]", "", txt(x).lower())


def guess_key(missing, avail):
    """清单里这个键没对上，文件里最像的是哪个。avail = {键: Mode}。

    ★ 最常见的写错就是**把 Mode 列的话当成键**（`OFF L5 LO PreBUF` 对应的键
      其实叫 `L5_LOPRE`）。所以候选要拿 键 和 Mode 两个都比一遍。
    """
    import difflib
    t = _norm(missing)
    if not t:
        return None
    best, score = None, 0.0
    for k, m in avail.items():
        for cand in (k, m):
            c = _norm(cand)
            if not c:
                continue
            r = 1.0 if (t in c or c in t) else                 difflib.SequenceMatcher(None, t, c).ratio()
            if r > score:
                best, score = k, r
    return best if score >= 0.6 else None


def _mode_note(key, mode):
    """Mode 列只在它**不是步骤名的换皮**时才写进备注。

    `gtcxo_core` 配 `OFF gtcxo_core_en` 说的是同一件事，写出来只是把行名
    换个写法再念一遍。判据：把两边都压成小写字母数字，步骤名整个落在
    Mode 里就当没信息。
    """
    m = txt(mode)
    if not m:
        return ""
    k = re.sub(r"[^0-9a-z]", "", txt(key).lower())
    return "" if (k and k in re.sub(r"[^0-9a-z]", "", m.lower())) else m


def run_quality(cur, chip):
    """这一趟数据的可信度：恢复后复测有没有回到全开基线。

    ★ 逐级关断测出来的是**差值**，差值的误差下限就是"同一个状态测两次差多少"。
      唯一能观测到它的地方就是恢复后复测 vs 全开基线——两者本该相等。
    ★★ 这句话是**整趟一句**，不是每个模块各说一遍。真数据上复测比基线差
      9.58 mA（基线的 17%），比 27 个模块每一个的 ΔI 都大——那不是
      "27 个模块都不可信"，那是"这一趟没恢复回去"。挂在每一行上只会让人
      问"这个 ⚠ 是什么"，而且正表里根本不该出现告警。
    返回控制台/审计页用的字符串列表；一条都没有就是这趟干净。
    """
    out = []
    for t in cur.temps:
        run = cur.runs[t]
        base, rc = run["baseline"], run.get("recheck")
        if not rc:
            out.append(f"{chip} {fmt_num(t)}℃: 这一趟**没有恢复后复测点**，"
                       f"没法判断差值的误差下限")
            continue
        d = rc[3] - base
        if not base:
            continue
        small = [k for k, _m, _i, dd in run["steps"] if abs(dd) < abs(d)]
        pct = abs(d) / abs(base) * 100.0
        if abs(d) < abs(base) * 0.01 and not small:
            continue                       # 回到基线的 1% 以内，且没有更小的台阶
        msg = (f"{chip} {fmt_num(t)}℃: 恢复后复测 {fmt_num(rc[3], 4)} mA，"
               f"全开基线 {fmt_num(base, 4)} mA，差 {d:+.4f}（{pct:.1f}%）"
               f"——差值的误差下限就是这个数")
        if run.get("tail_n", 0) > 1:
            msg += f"（阶梯后面共 {run['tail_n']} 个点，取的是最后一个）"
        if small:
            msg += (f"；{len(small)}/{len(run['steps'])} 个台阶比它还小"
                    + (f"（{', '.join(small[:6])}"
                       + (" …" if len(small) > 6 else "") + "）" if small else ""))
        out.append(msg)
    return out


def current_rows(cur, parts, total_name="", chip_note=""):
    """一份电流簿 → 一张表的行（跟 VCO 页同形：cat/item/unit/dir/kind/vals/note）。

    parts = [(部件名, [步骤键…], 部件备注), …]。一个部件一条分组带，带里逐步一行、
    末尾一行 `<部件名> 合计`；部件多于一个时最后再来一条 `总计` 带，
    放一行 total_name ＝ 各部件合计相加。

    ★ 分组是**加在原来那张全量表之外**的，不是替换（用户 2026-08-04：
      "我不是让你把原来的删了……原来的我也要，只是在原来的基础上添加"）。
      全量表由调用方另起一张，逻辑就是"一个部件、键给 None"。
    """
    # 每个键在各温度下的 ΔI
    dmap, order, modes = {}, [], {}
    for t in cur.temps:
        for k, m, _i, d in cur.runs[t]["steps"]:
            dmap.setdefault(k, {})[t] = d
            modes.setdefault(k, m)
            if k not in order:
                order.append(k)

    # ★★ 备注只写**名字说不出来**的事。行名已经讲明白的，写在备注里就是噪声
    #   （用户 2026-08-04：「这些都是常识性质的东西没必要反复强调」）。
    #   于是「全开基线电流＝所有模块都开着时的总电流」「复测 − 基线＝恢复后复测
    #   − 全开基线」这类同义反复整片删掉，只留"关断步数"那句范围提示——
    #   它是唯一一句名字里看不出来的（步数是整趟的，不是本表列出来的行数）。
    out = []
    cond = [("全开基线电流", {t: cur.runs[t]["baseline"] for t in cur.temps}),
            ("恢复后复测", {t: (cur.runs[t]["recheck"][3] if cur.runs[t].get("recheck")
                            else None) for t in cur.temps}),
            ("复测 − 基线", {t: ((cur.runs[t]["recheck"][3] - cur.runs[t]["baseline"])
                              if cur.runs[t].get("recheck") else None)
                          for t in cur.temps})]
    for item, vals in cond:
        out.append({"cat": "Condition", "item": item, "unit": "mA", "dir": "",
                    "kind": "cond", "vals": vals, "note": "", "nd": 4})
    out.append({"cat": "Condition", "item": "关断步数", "unit": "",
                "dir": "", "kind": "cond",
                "vals": {t: len(cur.runs[t]["steps"]) for t in cur.temps},
                "note": "整趟的步数，不只本表列出来的这些", "nd": 0})

    sums = []                       # [(部件名, {温度: 合计})]
    for pname, keys, pnote in parts:
        ks = order if keys is None else keys
        n_in = 0
        for k in ks:
            vals = dmap.get(k, {})
            # "关它之前 − 关它之后的总电流" 每行写一遍＝把一张表的口径说 27 遍，
            # 而且逐级关断本来就是这么算的；Mode 只在它不是步骤名换皮时才有信息量。
            note = _mode_note(k, modes.get(k))
            # ★★ 这里**不写告警**。原来每行前面挂一句"⚠ 比本趟重复性还小，只能当
            #   上限看"——真数据上复测比基线差了 9.58 mA，于是 27 行全挂上同一句话，
            #   成了满屏噪声，而且正表里出现 ⚠ 一定会被追问（用户 2026-08-04）。
            #   更要紧的是**它本来就该是整趟一句话**："这一趟恢复后没回到基线"，
            #   不是"每个模块各自不可信"。判据挪到 run_quality()，走控制台 + 审计页。
            out.append({"cat": pname, "item": k, "unit": "mA", "dir": "≤",
                        "kind": "result", "vals": vals, "note": note, "nd": 4})
            if k in dmap:
                n_in += 1
        if not n_in:
            continue
        psum = {}
        for t in cur.temps:
            got = [dmap[k][t] for k in ks
                   if k in dmap and isinstance(dmap[k].get(t), (int, float))]
            if got:
                psum[t] = sum(got)
        if not psum:
            continue
        # ★ 一个部件只有一行就不补合计：合计跟那一行一模一样，同一个数写两遍，
        #   正是「一行只有一种统计量」那条规矩要挡的。总计那行的备注就直接
        #   引它自己那一行的名字。
        if n_in < 2:
            sums.append((pname, psum, next(k for k in ks if k in dmap)))
            continue
        sums.append((pname, psum, f"{pname} 合计"))
        # 「本组各行相加」也是同义反复：行叫「X 合计」、就在 X 这一带的末尾。
        # 只有清单给了这一组特有的话（比如共用器件）才写。
        out.append({"cat": pname, "item": f"{pname} 合计", "unit": "mA", "dir": "≤",
                    "kind": "result", "vals": psum, "note": pnote, "nd": 4,
                    "strong": True})

    # ---- 总计 ----
    # 备注里每一项都**正好是表上某一行的名字**，评审能用眼睛把它加出来。
    if total_name and len(sums) >= 2:
        tot = {}
        for t in cur.temps:
            got = [d[t] for _p, d, _l in sums if t in d]
            if got:
                tot[t] = sum(got)
        if tot:
            terms = " + ".join(l for _p, _d, l in sums)
            out.append({"cat": "总计", "item": total_name, "unit": "mA", "dir": "≤",
                        "kind": "result", "vals": tot, "note": terms, "nd": 4,
                        "strong": True})
    return out


# ---------------------------------------------------------------- 配置文件

CONFIG_NAME = "chips.json"


def load_config(root, path=None):
    """跟着**数据**走的配置：默认找 `<根目录>/chips.json`。

    ★ 为什么要它：折算倍数、电流分组、模块顺序这些是**这个项目的固定事实**，
      不是每次跑都要重想的参数。写在命令行上就意味着"谁记得住谁才跑得对"——
      换个人、隔一阵子、开一段新对话，命令就少一截，而少一截**不会报错**，
      只会安安静静出一份数全错的报表。
    ★ 放数据目录旁边而不是仓库里：里面有真实模块名（zero-IP），
      而且它本来就属于"这批数据"，跟着数据一起搬。
    """
    p = path or os.path.join(root, CONFIG_NAME)
    if not os.path.isfile(p):
        return {}, (p if path else None)
    with open(p, encoding="utf-8") as f:
        cfg = json.load(f)
    if not isinstance(cfg, dict):
        sys.exit(f"配置文件最外层要是一个对象: {p}")
    return cfg, p


def scale_from_cfg(m):
    """配置里的 {"模块:类型": 倍数} -> parse_scale 那套三元组。"""
    out = []
    for k, v in (m or {}).items():
        mod, sep, kind = str(k).partition(":")
        if not sep:
            mod, kind = k, "*"
        out.append((mod.strip() or "*", kind.strip().lower() or "*", float(v)))
    return out


def _part_list(m):
    out = []
    for pn, ps in (m or {}).items():
        if isinstance(ps, dict):
            out.append((pn, list(ps.get("keys") or []), txt(ps.get("note") or "")))
        else:
            out.append((pn, list(ps), ""))
    return out


def groups_from_cfg(m):
    """清单 -> [(报告名, [(部件名, 键表, 部件备注)…], 总计行的名字)]。

    三种写法都认：
      "组名":   [键…]                                 一张表一个组
      "组名":   {"keys": [键…], "note": "…"}           同上，带一句组备注
      "报告名": {"parts": {"部件名": [键…] | {…}}}      一张表分几块，末尾一行总计
    第三种是「<某模块> 总功耗 ＝ A + B + C」那种报告（用户 2026-08-04 要的）。
    """
    out = []
    for g, spec in (m or {}).items():
        if isinstance(spec, dict) and spec.get("parts"):
            out.append((g, _part_list(spec["parts"]), g))
        elif isinstance(spec, dict):
            out.append((g, [(g, list(spec.get("keys") or []),
                             txt(spec.get("note") or ""))], ""))
        else:
            out.append((g, [(g, list(spec), "")], ""))
    return out


def implied_scale(sw, kind):
    """簿子**自己声明**的目标频点 ÷ 实测频点 ＝ 这份数据该乘几。

    PLL 簿看 `fLO_MHz`，VCO 簿看 `fVCO_MHz`（load_vco 已经读成 sw.fvco）。

    ★ 只拿它**核对**，不拿它当默认值。折算猜错一个倍数，整本报表每个数都错，
      而且每个数看着都正常——这种错必须由人拍板，不能由工具猜。
      但反过来，工具有义务在"配的倍数跟簿子自己说的对不上"、或者"压根没配
      而簿子看着需要折算"的时候**当场喊出来**。
    """
    fi = getattr(sw, "freq_item", None)
    if fi is None and hasattr(sw, "item"):
        fi = sw.item("Freq_MHz")
    if fi is None:
        return None
    meas = median([r.vals.get(fi.col) for r in sw.rows
                   if r.vals.get(fi.col) is not None])
    if not meas:
        return None
    tgt = None
    if kind == KIND_VCO:
        tgt = getattr(sw, "fvco", None)
    else:
        ci = sw.cols.idx("fLO_MHz")
        if ci is not None:
            tgt = median([num(r.raw[ci]) for r in sw.rows
                          if ci < len(r.raw) and num(r.raw[ci]) is not None])
    if not tgt:
        return None
    return {"target": tgt, "meas": meas, "ratio": tgt / meas}


def check_scale(imp, n, chip, kind_label, tol=0.10):
    """配的倍数 vs 簿子自己声明的比值。对不上就喊，没配也喊。

    ★ 只在比值**贴近一个整数**时才开口。VCO 簿的实测频率本来就扫了一整个调谐
      范围，中位数当分母只能算个毛估——比值落在 1.6 这种地方，说明这份簿子
      压根没有"整数倍频"这层关系，那就没有判据，闭嘴比瞎喊强。
      警告一旦有假阳性，下次真的那条也会被当噪声划过去。
    """
    if not imp:
        return
    r = imp["ratio"]
    n0 = round(r)
    if n0 < 1 or abs(r - n0) > tol * n0:
        return
    if n and n != 1.0:
        if abs(n0 - n) > 1e-9:
            print(f"     ⚠⚠ {chip}: 折算配的是 ×{fmt_num(n, 4)}，但这份簿子自己说的是 "
                  f"{fmt_num(imp['target'])} ÷ 实测 {fmt_num(imp['meas'])} ≈ "
                  f"×{n0} —— 两个只有一个是对的，别就这么发出去")
    elif n0 > 1:
        print(f"     ⚠⚠ {chip}: **没配折算**，但这份 {kind_label} 簿子自己写着目标频点 "
              f"{fmt_num(imp['target'])} MHz、实测才 {fmt_num(imp['meas'])} MHz "
              f"（≈ ×{n0}）。相噪/杂散没折算到真实载波上就跟 spec 对不上——"
              f"要么在 {CONFIG_NAME} 里配 scale，要么确认这份就该按实测报")


# ---------------------------------------------------------------- 追溯

def _xfrm(sinfo, kind, mod, dsb, label, unit):
    """这一行在折算里被加了多少 dB / 乘了多少。返回 (加法项列表, 乘数)。"""
    add, mul = [], 1.0
    info = (sinfo.get(kind) or {}).get(mod)
    how = unit_scaling(unit)
    if info and how == "lin":
        mul = info["n"]
    elif info and how == "db":
        add.append((f"×{fmt_num(info['n'], 4)} 折算", info["db"]))
    if dsb and label.startswith("IPN"):
        add.append(("SSB→DSB", DSB_DB))
    return add, mul


def _book_cell(wb, label, chip, temp, sheets, mod):
    """在产出的簿子里找到 (行名, 芯片, 温度) 那一格。**按内容找，不按坐标算**——
    算坐标就等于把版式规则抄第二遍，版式一改追溯就悄悄指错格子。"""
    # ★ 必须指定在哪一页找：同一个行名（IPN_DSB）PLL 页和 VCO 页都有，
    #   按页序碰运气就会拿 PLL 的格子去对 VCO 的数，然后报一片 ✗。
    #   （这个 bug 就是被这条检查自己抓出来的。）
    #   还得认是**哪个模块那一块**：一页上下叠着两张表（模块 A / 模块 B），
    #   行名一模一样、芯片列名也一模一样，只按行名找必然撞到上面那张。
    #   大标题那一行的特征＝它下一行是「测试项」。
    tag = f"{fmt_num(temp)}℃"
    for name in sheets:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        cur = ""
        for r in range(1, ws.max_row + 1):
            v1 = txt(ws.cell(row=r, column=C_ITEM).value)
            if v1 and txt(ws.cell(row=r + 1, column=C_ITEM).value) == "测试项":
                cur = v1
                continue
            if v1 != label or not cur.startswith(str(mod)):
                continue
            for hr in range(r - 1, 0, -1):          # 往上找最近那行表头
                c0 = next((c for c in range(1, ws.max_column + 1)
                           if txt(ws.cell(row=hr, column=c).value) == chip), None)
                if c0 is None:
                    continue
                for c in range(c0, ws.max_column + 1):
                    if txt(ws.cell(row=hr + 1, column=c).value) == tag:
                        return f"{name}!{_cl(c)}{r}", ws.cell(row=r, column=c).value
                break
    return None, None


def _say(chip, t, srcname, rows, agg, add, mul, unit, how, wb=None, label=None,
         sheets=(), mod=""):
    """把一格的来龙去脉打出来：原表哪几行 → 原始值 → 加了什么 → 等于多少。"""
    from openpyxl.utils import get_column_letter as gl
    base = agg
    line = f"      {how} = {round(base, 6)}"
    for name, d in add:
        base += d
        line += f"  {name} {d:+.2f}"
    if mul != 1.0:
        base *= mul
        line += f"  ×{fmt_num(mul, 4)}"
    line += f"  →  {round(base, 6)} {unit}"
    where = got = None
    if wb is not None and label:
        where, got = _book_cell(wb, label, chip, t, sheets, mod)
        if where is not None:
            ok = got is not None and abs(float(got) - base) < 0.006
            line += f"   簿子 {where} = {got}  {'✓' if ok else '✗ 对不上！'}"
    # ★ 表上没有这个温度就整段不打。PLL 温扫有 16 档温度、表上只列 3 档，
    #   其余 13 档每档刷四行「表上不列」——追溯是拿来核对的，不是拿来倒数据的。
    if wb is not None and where is None:
        return False
    print(f"  {chip}  {fmt_num(t)}℃")
    for xl, col, raw in rows[:8]:
        print(f"      原表 行{xl} 列{gl(col + 1)}  {raw}")
    if len(rows) > 8:
        print(f"      …共 {len(rows)} 行（{how}）")
    print(line)
    return True


def trace_item(label, tables, vsweeps, sinfo, dsb, op_cfg, out_path=None):
    """把某一行的数一路追回原表。**用来分清是程序算错了还是数据本身就这样。**

    ★ 这个问题会反复出现：某一格看着离谱（IPN 是正的、杂散差 60 dB、
      复测差 58%），第一反应总是"是不是脚本弄错了"。争论没用，把原表的
      行号列号、原始值、加了几 dB、加完等不等于表上那个数，一路摊开就完了。
    """
    import openpyxl
    from summarize_vco_sweep import group_series
    wb = openpyxl.load_workbook(out_path, data_only=True) if out_path else None
    print()
    print(f"=== 追溯「{label}」——  原表 → 折算 → 簿子上那一格，三头对一遍 ===")
    hit = False
    for mod, data in tables:                      # PLL 页：该温度全部经过点的中位数
        for chip, sw in data.items():
            it = sw.item(label)
            if it is None:
                continue
            hit = True
            print(f"[{mod} PLL 温扫] {os.path.basename(sw.path)}")
            add, mul = _xfrm(sinfo, KIND_PLL, mod, dsb, label, it.unit)
            skipped = 0
            for t in sw.temps:
                pts = [(r.xl, r.col_of(it), r.raw[r.col_of(it)], r.vals[it.col])
                       for lg in sw.legs for r in lg.rows
                       if r.kind != "lock" and r.temp == t
                       and r.vals.get(it.col) is not None]
                if not pts:
                    continue
                # 表上那格 = 折算后各点的中位数；反推回折算前好跟原表对
                med_after = median([p[3] for p in pts])
                med_before = (med_after / mul) - sum(d for _n, d in add)
                if not _say(chip, t, "", [(a, b, c) for a, b, c, _d in pts],
                            med_before, add, mul, it.unit,
                            f"{len(pts)} 个点的中位数", wb, label,
                            ("PLL_Summary",), mod):
                    skipped += 1
            if skipped:
                print(f"      （另有 {skipped} 档温度表上不列，略）")
    for mod, chips_ in vsweeps.items():           # VCO 页：工作点那一个点
        for chip, sw in chips_.items():
            it = next((x for x in sw.items if x.label == label), None)
            if it is None:
                continue
            hit = True
            print(f"[{mod} VCO 开环] {os.path.basename(sw.path)}")
            add, mul = _xfrm(sinfo, KIND_VCO, mod, dsb, label, it.unit)
            op = op_cfg if op_cfg else op_vtune_of(sw)
            for kind, groups in sw.by_kind:
                if kind != "vtune":
                    continue
                for g in groups:
                    ser = group_series(g, it)
                    if not ser or g.temp is None or op is None:
                        continue
                    x = min(ser, key=lambda z: abs(z - op))
                    r = next((rr for rr in g.rows
                              if rr.vt is not None and abs(rr.vt - x) < 1e-9
                              and rr.vals.get(it.col) is not None), None)
                    if r is None:
                        continue
                    after = r.vals[it.col]
                    before = (after / mul) - sum(d for _n, d in add)
                    _say(chip, g.temp, "", [(r.xl, r.col_of(it), r.raw[r.col_of(it)])],
                         before, add, mul, it.unit,
                         f"Vtune={fmt_num(x, 4)} 那一点", wb, label,
                         ("VCO_Summary",), mod)
    if not hit:
        print(f"  没有哪份簿子有叫「{label}」的行。表上的行名照抄即可"
              f"（IPN_DSB / SpotPN@1kHz / Spur@26MHz / Freq_MHz …）")


# ---------------------------------------------------------------- 出稿自查

def selfcheck(path):
    """把刚写出来的簿子读回来自查两条。**跑完就查，不给人多敲一条命令的机会。**

    ① 汇总列的每个数字都必须能在同一行的格子里找到。
       （2026-07-30 翻过车：温度列放中位数、极值却对逐点取，出现「-40℃ 那格
       写 -54.19，MAX 写 -53.45 @-40℃」，用户当场判成算错了。）
    ② 判定列的公式格不能留下**空的缓存值**。openpyxl 有的版本会写
       `<f>公式</f><v></v>`＝在文件里写着"这公式的结果就是空"，自己机器上打开
       正常（触发了重算），发给别人就是一片空白。xlsx_formula_cache 负责清掉，
       这里负责确认真的清掉了——不同 openpyxl 版本行为不一样，不能只信一台机器。
    """
    import openpyxl
    wb = openpyxl.load_workbook(path)
    wv = openpyxl.load_workbook(path, data_only=True)
    fails, n_agg, n_f = [], 0, 0

    for name in ("PLL_Summary", "VCO_Summary", "Current_Summary"):
        if name not in wb.sheetnames:
            continue
        ws, wsv = wb[name], wv[name]
        ax = 0
        while ws.cell(3, C_CHIP0 + ax).value not in (None, ""):
            ax += 1
        w = ax + 1
        n = 0
        while ws.cell(2, C_CHIP0 + n * w).value not in (None, "", "备注"):
            n += 1
        # 常温在组里的第几列（轴头写的就是温度，挑最接近 25 的）
        ri, best = 0, None
        for j in range(ax):
            try:
                d = abs(float(str(ws.cell(3, C_CHIP0 + j).value).replace("℃", "")) - 25)
            except ValueError:
                continue
            if best is None or d < best:
                ri, best = j, d
        for r in range(1, ws.max_row + 1):
            item = ws.cell(r, 1).value
            got = [ws.cell(r, C_SUM + j).value for j in range(3)]
            if not isinstance(item, str) or not all(
                    v is None or isinstance(v, (int, float)) for v in got):
                continue
            jf = ws.cell(r, C_JUDGE).value
            if isinstance(jf, str) and jf.startswith("="):
                n_f += 1
                if wsv.cell(r, C_JUDGE).value is not None:
                    fails.append(f"{name} 第{r}行 判定列留了缓存值 "
                                 f"{wsv.cell(r, C_JUDGE).value!r}——别人打开会看到它")
            if all(v is None for v in got):
                continue
            cells = [v for v in (ws.cell(r, C_CHIP0 + k * w + j).value
                                 for k in range(n) for j in range(ax))
                     if isinstance(v, (int, float))]
            room = [v for v in (ws.cell(r, C_CHIP0 + k * w + ri).value
                                for k in range(n)) if isinstance(v, (int, float))]
            exp = [min(cells) if cells else None,
                   median(room) if room else None,
                   max(cells) if cells else None]
            for j, (x, y) in enumerate(zip(exp, got)):
                n_agg += 1
                if (x is None) != (y is None) or (
                        x is not None and abs(float(x) - float(y)) > 1e-9):
                    fails.append(f"{name}「{item}」第{j+1}个汇总格: "
                                 f"表上的格子给出 {x}，写的是 {y}")
    fails += chart_overlaps(wb)
    return fails, n_agg, n_f


def chart_overlaps(wb):
    """③ 图和图不许压在一起——**算出来**，不靠眼睛看。

    ★ 2026-08-04 事故：图宽写死 14.5cm，竖条宽度是列宽算出来的 14.15cm，
      每张图往右边邻居里压 13px。这种事读 chart XML 一点看不出来
      （每张图自己都完全正常），肉眼在 Excel 里也只是"边框好像贴着"。
      三套单位（列宽格 / 行高磅 / 图厘米）换算到像素，重叠就是个几何题。
    """
    from openpyxl.utils import get_column_letter as gl
    out = []
    for ws in wb.worksheets:
        boxes = []
        for ch in getattr(ws, "_charts", []):
            a_ = getattr(ch, "anchor", None)
            frm = getattr(a_, "_from", None)
            ext = getattr(a_, "ext", None)
            if frm is None or ext is None:
                continue                      # 双格锚点：位置由格子定，不会压
            x = sum(col_px(ws.column_dimensions[gl(c)].width
                           or DEFAULT_COL_W) if gl(c) in ws.column_dimensions
                    else col_px(DEFAULT_COL_W) for c in range(1, frm.col + 1))
            y = sum((ws.row_dimensions[r].height
                     if r in ws.row_dimensions and ws.row_dimensions[r].height
                     else DEFAULT_ROW_PT) * 4.0 / 3.0 for r in range(1, frm.row + 1))
            boxes.append((_chart_name(ch), x + frm.colOff / EMU_PX,
                          y + frm.rowOff / EMU_PX,
                          ext.cx / EMU_PX, ext.cy / EMU_PX))
        for i in range(len(boxes)):
            for j in range(i + 1, len(boxes)):
                t1, x1, y1, w1, h1 = boxes[i]
                t2, x2, y2, w2, h2 = boxes[j]
                ox = min(x1 + w1, x2 + w2) - max(x1, x2)
                oy = min(y1 + h1, y2 + h2) - max(y1, y2)
                if ox > 1 and oy > 1:
                    out.append(f"{ws.title}: 两张图压在一起（横 {ox:.0f}px / "
                               f"纵 {oy:.0f}px）—— {t1} ×× {t2}")
    return out


def _chart_name(ch):
    try:
        return "".join(r.t or "" for p in ch.title.tx.rich.p for r in (p.r or ()))[:40]
    except Exception:                          # noqa: B902
        return "(无标题)"


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(
        description="多芯片测试目录 → 一份给评审看的汇总 Excel")
    ap.add_argument("root", help="根目录（下面一层是芯片目录，目录名＝芯片编号）")
    ap.add_argument("-o", "--out", default=None,
                    help="输出路径（默认 <根目录名>_chips_summary.xlsx，"
                         "放在根目录**旁边**，免得下次扫描把它当输入读回来）")
    ap.add_argument("--modules", default="",
                    help="只处理这几个模块，并按这个顺序上下排（逗号分隔）。"
                         "不给＝从文件名前缀自动认出全部模块，按名字排序")
    ap.add_argument("--chips", default="", help="只处理这几颗（逗号分隔）")
    ap.add_argument("--module-alias", default="",
                    help="把认成两个的同一个模块并回去：`认出来的名字=真名`，逗号分隔。"
                         "文件名前缀改过（`补充_XXXPLL_...`）就会多认出一个模块，"
                         "那颗芯片在原模块里整块消失。配置里写 module_alias 也行")
    ap.add_argument("--leg-col", default="Mode", help="判断重锁用的列（默认 Mode）")
    ap.add_argument("--lock-pattern", default=r"_lock$",
                    help="该列匹配这个正则的行 = 一次重锁（默认 _lock$）")
    ap.add_argument("--temp-col", default=None, help="温度列（默认自动找 Temperature）")
    ap.add_argument("--no-charts", action="store_true", help="图都不画，只出数据块")
    ap.add_argument("--slim", action="store_true",
                    help="两张汇总表每片只显示常温列，其余温度列折进 Excel 大纲"
                         "（点 ＋ 展开，数一个没少）；芯片多了用")
    ap.add_argument("--groups", default=None,
                    help="电流页的分组清单 JSON（{\"groups\": {\"组名\": [步骤键…]}}）。"
                         "每组末尾自动出一行合计。要给某组加一句备注（"
                         "比如两个模块共用同一个器件、相加时别算两遍），"
                         "把那一组写成 {\"keys\": [步骤键…], \"note\": \"…\"}。"
                         "★含真实模块名，放黄区本地/private，别提交。"
                         "不给＝所有关断步骤按文件顺序排成一组")
    ap.add_argument("--trace", default="",
                    help="把某一行的数一路追到原表：给行名（如 IPN_DSB / Freq_MHz / "
                         "SpotPN@1kHz），逐芯片逐温度打出「原表第几行第几列、原始值多少、"
                         "折算加了多少、加完等不等于表上那个数」。"
                         "用来分清「是程序算错了」还是「数据本身就这样」")
    ap.add_argument("--chart-w", type=float, default=None,
                    help=f"每张图的宽度 cm（默认 {CHART_W_CM}）。芯片竖条的列宽会跟着"
                         "等比撑开，保证放得下——图宽和列宽是同一个数推出来的，"
                         "不会再出现相邻两颗芯片的图压在一起")
    ap.add_argument("--dsb", dest="dsb", action="store_true", default=None,
                    help="把**积分**相噪从单边带换算成双边带（+3.01 dB），"
                         "行名 IPN_SSB→IPN_DSB。逐点相噪与杂散不动"
                         "（L(f) 与 dBc 的定义本来就是单边带）")
    ap.add_argument("--no-dsb", dest="dsb", action="store_false",
                    help="强制按单边带报（盖掉配置文件里的 dsb）")
    ap.add_argument("--config", default=None,
                    help=f"配置文件（默认自动找 <根目录>/{CONFIG_NAME}）。"
                         "里面可以放 scale / groups / modules / chips / "
                         "op_vtune / ref_temp / spur_tol，命令行给了就以命令行为准。"
                         "★含真实模块名，跟数据放一起，别提交")
    ap.add_argument("--spec", default=None,
                    help="Spec / 仿真 / Limit 那七列的来源 JSON（spec_from_xlsx.py "
                         "从填好的簿子里读出来的）。不给＝用配置文件里的 spec 块")
    ap.add_argument("--no-spec", action="store_true",
                    help="七列一律留空给人手填（盖掉配置里的 spec）")
    ap.add_argument("--spur-tol", type=float, default=None,
                    help="杂散取值窗口 ±MHz（默认 2）：真实杂散不落在标称频点上，"
                         "从表尾的杂散清单里在标称频点这个窗口内取幅度最大的一条。"
                         "认不出清单就退回读模板那一格，并在控制台说原因")
    ap.add_argument("--scale", default="",
                    help="把结果折算到倍频点上：`模块:类型=倍数`，逗号分隔，"
                         "类型＝pll/vco，模块与类型都可以写 *。"
                         "例：\"A:pll=4,A:vco=8,B:vco=2\"。"
                         "频率与 Kvco ×倍数，相噪/杂散/IPN +20log10(倍数) dB；"
                         "功率/压控电压/电流/温度不动。倍数印在表的大标题上")
    ap.add_argument("--cur-col", default=None, help="电流值列（默认自动找）")
    ap.add_argument("--key-col", default=None, help="步骤名那一列（默认自动找）")
    ap.add_argument("--no-current", action="store_true", help="不做电流页")
    ap.add_argument("--no-vco", action="store_true",
                    help="不做 VCO 两页（只出 PLL 温扫那两页）")
    ap.add_argument("--ref-temp", type=float, default=None,
                    help="VCO 温漂的参考温度（默认 25）")
    ap.add_argument("--op-vtune", type=float, default=None,
                    help="VCO 工作点调谐电压 V（默认取 CT 扫钉住的那个值）")
    ap.add_argument("--no-audit", action="store_true", help="连隐藏的 _审计 页都不要")
    ap.add_argument("--dry-run", action="store_true", help="只清点和核对识别结果")
    args = ap.parse_args()

    root = os.path.abspath(args.root)
    if not os.path.isdir(root):
        sys.exit(f"不是目录: {root}")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        sys.exit("缺少 openpyxl，请先: pip install openpyxl")

    # ---- 配置：跟着数据走，命令行给了就以命令行为准 ----
    cfg, cfg_path = load_config(root, args.config)
    if args.config and not cfg:
        sys.exit(f"找不到配置文件: {args.config}")
    print(f"配置    : {cfg_path}" if cfg else
          f"配置    : 没有（找过 {os.path.join(root, CONFIG_NAME)}）")
    def pick_opt(cli, key, default=None):
        if cli is not None:
            return cli
        return cfg[key] if key in cfg else default

    ref_temp = float(pick_opt(args.ref_temp, "ref_temp", 25.0))
    spur_tol = float(pick_opt(args.spur_tol, "spur_tol", 2.0))
    op_vtune_cfg = pick_opt(args.op_vtune, "op_vtune")
    dsb = bool(pick_opt(args.dsb, "dsb", False))
    chart_w = float(pick_opt(args.chart_w, "chart_w", CHART_W_CM))

    # Spec / 仿真 / Limit：手填一次，存成数据，以后每次重出都自带
    spec = None
    if not args.no_spec:
        sd, stt, ssrc = cfg.get("spec"), \
            (cfg.get("spec_meta") or {}).get("titles"), \
            os.path.basename(cfg_path or "") or CONFIG_NAME
        if args.spec:
            if not os.path.isfile(args.spec):
                sys.exit(f"找不到 spec 文件: {args.spec}")
            with open(args.spec, encoding="utf-8") as f:
                sj = json.load(f)
            sd = sj.get("spec", sj)
            stt = (sj.get("spec_meta") or {}).get("titles")
            ssrc = os.path.basename(args.spec)
        if sd:
            spec = SpecBook(sd, stt, ssrc)

    want_mod = ([m.strip() for m in args.modules.split(",") if m.strip()]
                or [str(m).strip() for m in (cfg.get("modules") or [])])
    only = ({c.strip() for c in args.chips.split(",") if c.strip()}
            or {str(c).strip() for c in (cfg.get("chips") or [])} or None)
    alias = {}
    for k, v in (cfg.get("module_alias") or {}).items():
        alias[str(k)] = str(v)
        alias[_norm_mod(k)] = str(v)
    for pair in args.module_alias.split(","):
        a, sep, b = pair.partition("=")
        if sep and a.strip() and b.strip():
            alias[a.strip()] = b.strip()
            alias[_norm_mod(a)] = b.strip()
    picked, dropped, unknown, loose, filtered = discover(root, only,
                                                        set(want_mod) or None,
                                                        alias)
    if not picked:
        sys.exit(f"{root} 下没找到能认出来的 .xlsx —— 文件名要长成 "
                 f"`<模块><类型>_...`（类型＝ PLL / VCO / Current）")

    chips = sorted({k[0] for k in picked}, key=natkey)
    # 模块名是从文件名前缀认出来的，不写死在代码里；--modules 可以钉死顺序
    modules = want_mod or sorted({k[1] for k in picked})
    print(f"根目录: {root}")
    print(f"芯片 {len(chips)} 颗: {', '.join(chips)}")
    print(f"模块 {len(modules)} 个: {', '.join(modules)}"
          + ("" if want_mod else "（从文件名认出来的，表的上下顺序按名字排；"
                                 "要换顺序用 --modules）"))
    grid = {}
    for (chip, mod, kind), b in picked.items():
        grid.setdefault(kind, {}).setdefault(mod, {})[chip] = b
    for kind in (KIND_PLL, KIND_VCO, KIND_CUR):
        for mod in modules:
            got = grid.get(kind, {}).get(mod, {})
            if got:
                miss = [c for c in chips if c not in got]
                print(f"  {mod:<6} {KIND_LABEL[kind]:<8} {len(got)} 份"
                      + (f"   缺: {', '.join(miss)}" if miss else ""))
    for a, b, klabel, ca, cb in near_modules(picked):
        print(f"  ⚠⚠ {klabel}：模块「{a}」只有 {', '.join(ca)}，"
              f"「{b}」只有 {', '.join(cb)}，"
              f"两边没有一颗芯片重合——「{a}」多半就是「{b}」，"
              f"只是那几份文件名前缀不一样（模块名＝文件名里 PLL/VCO/Current "
              f"前面那一截）。**这样出表，{', '.join(ca)} 在「{b}」那几栏是空的**"
              f"（温巡页会空出图位）。要么把文件名改回去，要么在 "
              f"{os.path.basename(cfg_path or CONFIG_NAME)} 里写："
              f'  "module_alias": {{"{a}": "{b}"}}')
    for b, w in dropped:
        print(f"  ↷ 跳过 {b.chip}/{b.name}（同类里有更新的 {w.ts or w.name}）")
    for chip, f, why in unknown:
        print(f"  ? 认不出{why}，没读: {chip}/{f}")
    for chip, f, why in filtered:
        print(f"  ⚠ 被过滤掉没读: {chip}/{f} —— {why}")
    if loose:
        print(f"  ⚠ 根目录下还散放着 {len(loose)} 个 .xlsx **没有被扫**"
              f"（有芯片目录时只扫子目录）——如果它们也是要算的芯片，"
              f"各自挪进对应的芯片目录：")
        for f in loose:
            print(f"      {f}")
    n_cur = sum(len(v) for v in grid.get(KIND_CUR, {}).values())

    # ---- 读 PLL 温扫 ----
    tables, failed, notes, warn_seen, vcharts = [], [], {}, {}, []
    excl_all = []
    quality_all = []          # 数据可信度（重复性/自相矛盾）：控制台 + 审计页，不进正表
    scale_rules = parse_scale(args.scale) or scale_from_cfg(cfg.get("scale"))
    sinfo = {}
    if scale_rules:
        print()
        print("折算规则: " + "，".join(
"%s:%s ×%s" % (m, k, fmt_num(v, 4)) for m, k, v in scale_rules)
              + ("" if args.scale else f"   ←{os.path.basename(cfg_path or CONFIG_NAME)}"))
    else:
        print()
        print("折算规则: 没有（相噪/杂散/频率按实测频点报）")
    for mod in modules:
        books = grid.get(KIND_PLL, {}).get(mod, {})
        if not books:
            continue
        data = {}
        print(f"\n=== {mod} PLL 温扫 ===")
        for chip in chips:
            b = books.get(chip)
            if b is None:
                print(f"  {chip}: 没有这个模块的温扫文件")
                continue
            try:
                sw = load_sweep(b.path, leg_col=args.leg_col,
                                lock_pattern=args.lock_pattern,
                                temp_col=args.temp_col, keep_original=False,
                                spur_tol=spur_tol)
            except Exception as e:                    # noqa: B902
                failed.append((b, f"{type(e).__name__}: {e}"))
                print(f"  {chip}: 读失败 —— {type(e).__name__}: {e}")
                continue
            data[chip] = sw
            for _q in ipn_order_check(sw, chip):
                print(f"     ⚠ {_q}")
                quality_all.append((chip, mod, KIND_LABEL[KIND_PLL],
                                    [("IPN vs IPN_Omit", _q)]))
            imp = implied_scale(sw, KIND_PLL)          # 必须在折算之前算
            n_scale = scale_of(scale_rules, mod, KIND_PLL)
            note_scale(sinfo, KIND_PLL, mod, chip, scale_book(sw, n_scale))
            check_scale(imp, n_scale, chip, KIND_LABEL[KIND_PLL])
            if dsb:
                note_dsb(to_dsb(sw), chip)
            n_meas = sum(1 for lg in sw.legs for x in lg.rows if x.kind != "lock")
            notes[id(b)] = f"{sw.n_rows}行×{sw.n_cols}列"
            print(f"  {chip}: {len(sw.legs)} 段 / {len(sw.temps)} 档温度 / "
                  f"{n_meas} 测点 / 指标 {len(sw.items)} 个 / 排除 {len(sw.excluded)} 行"
                  f"   [{b.name}]")
            for sn in getattr(sw, "spur_notes", ()):
                print(f"     · {sn}")
            print_excluded(sw.excluded)
            excl_all.append((chip, mod, KIND_LABEL[KIND_PLL], sw.excluded))
            flos = {txt(x.raw[sw.cols.idx('fLO_MHz')])
                    for lg in sw.legs for x in lg.rows
                    if sw.cols.idx('fLO_MHz') is not None} - {""}
            if len(flos) > 1:
                print(f"     ⚠ 这份簿里有多个 fLO：{sorted(flos)}"
                      f"——汇总把它们并成一行了，要分频点报的话得分块（本版未做）")
            for w in sw.warnings:
                warn_seen.setdefault(w, []).append(chip)
        if data:
            tables.append((mod, data))
        # 告警按条汇总打印：重名列这种模板老问题每份簿都会报，
        # 逐份逐条打 = 30 行同样的话，真正要看的东西反而被冲掉
        for w, who in warn_seen.items():
            print(f"  ⚠ {w}"
                  + (f"   （{len(who)}/{len(data)} 份都有）" if len(who) > 1
                     else f"   （{who[0]}）"))
        warn_seen.clear()

    # ---- 读 VCO 开环 ----
    vtables, vtemps, vsweeps = [], [], {}
    for mod in modules:
        books = grid.get(KIND_VCO, {}).get(mod, {})
        if not books or args.no_vco:
            continue
        vdata, vrows = {}, {}
        print()
        print("=== %s VCO 开环 ===" % mod)
        op = None
        for chip in chips:
            b = books.get(chip)
            if b is None:
                print(f"  {chip}: 没有这个模块的开环文件")
                continue
            try:
                sw = load_vco(b.path, temp_col=args.temp_col, keep_original=False,
                              spur_tol=spur_tol)
            except Exception as e:                    # noqa: B902
                failed.append((b, f"{type(e).__name__}: {e}"))
                print(f"  {chip}: 读失败 —— {type(e).__name__}: {e}")
                continue
            # ★ 折算必须在 vco_rows 之前：Fmin/Fmax/Kvco/温漂全是拿 F 现算的，
            #   先折算，它们自己就跟着 ×N 了。
            # ★ VCO 页已经不报 IPN 了（开环下这个指标没意义，见 VCO_DROP_PREFIX），
            #   再为它喊重复性就是噪声。这条哨兵只留给 PLL 页。
            imp = implied_scale(sw, KIND_VCO)          # 必须在折算之前算
            n_scale = scale_of(scale_rules, mod, KIND_VCO)
            note_scale(sinfo, KIND_VCO, mod, chip, scale_book(sw, n_scale))
            check_scale(imp, n_scale, chip, KIND_LABEL[KIND_VCO])
            if dsb:
                note_dsb(to_dsb(sw), chip)          # 要在 vco_rows 之前
            # 工作点用第一颗芯片的（CT 扫钉住的那个 Vtune）统一喂给全部芯片：
            # 否则各片的组名会变成「@ Vtune 0.4V」「@ Vtune 0.45V」，行对不齐
            if op is None:
                op = op_vtune_cfg if op_vtune_cfg else op_vtune_of(sw)
            rows, temps = vco_rows(sw, ref_temp, op)
            fv, kv, fc, fd = _vco_series(sw, temps)
            vrows[chip] = rows
            vdata[chip] = {"v": fv, "k": kv, "c": fc, "d": fd}
            # ★ 只有真要 --trace 才留着整份簿子。每份 VCO 簿子在内存里是几百 MB
            #   （4.7MB 的 xlsx 展开几十倍），六份攒着直接 MemoryError——
            #   而 MemoryError 的 str() 是**空字符串**，报出来就是"读失败 ——"，
            #   什么线索都没有。2026-08-05 就是这么崩的。
            if args.trace:
                vsweeps.setdefault(mod, {})[chip] = sw
            notes[id(b)] = f"{sw.ws_val.max_row}行×{sw.ws_val.max_column}列"
            coarse = coarse_temps(sw)
            print(f"  {chip}: 温度 {[fmt_num(t) for t in temps]} / "
                  f"结论行 {sum(1 for x in rows if x['kind'] == 'result')} 条 / "
                  f"指标 {len(sw.items)} 个   [{b.name}]")
            if coarse:
                print(f"     · CT 粗码温度 {[fmt_num(t) for t in sorted(coarse)]}"
                      f"（只测几个码）")
            for t, sp, med in flat_vtune_temps(sw):
                print(f"     ⚠⚠ {fmt_num(t)}℃ 的 Vtune 扫**频率几乎没动**："
                      f"全程只变了 {fmt_num(sp, 4)} MHz，其他温度是 {fmt_num(med, 1)} MHz。"
                      f"这个温度多半没真正切到开环（环路还锁着／DAC 没驱动／testmux "
                      f"没切），它的 Kvco 会算出接近 0 的值——那不是低温增益低。"
                      f"对照 Kvco 组里 F(Vtune=…) 那两行就能确认")
            print(f"     排除 {len(sw.excluded)} 行:")
            for sn in getattr(sw, "spur_notes", ()):
                print(f"     · {sn}")
            print_excluded(sw.excluded)
            excl_all.append((chip, mod, KIND_LABEL[KIND_VCO], sw.excluded))
            for w in sw.warnings:
                warn_seen.setdefault(w, []).append(chip)
            for t in temps:
                if t not in vtemps:
                    vtemps.append(t)
        for w, who in warn_seen.items():
            print(f"  ⚠ {w}   （{'/'.join(who)}）")
        warn_seen.clear()
        if vrows:
            vtables.append((mod, vrows))
            vcharts.append((mod, vdata))
    vtemps.sort()

    if not tables and not vtables:
        sys.exit("没有一份 PLL 温扫 / VCO 开环文件读成功")

    if args.dry_run:
        print("\n--dry-run：没有写文件。")
        return

    # ---- 读电流（逐级关断）----
    ctables, ctemps = [], []
    if not args.no_current and grid.get(KIND_CUR):
        wanted = []
        gj = None
        if args.groups:
            with open(args.groups, encoding="utf-8") as f:
                gj = json.load(f)
        elif cfg.get("groups"):
            gj = {"groups": cfg["groups"]}
        if gj is not None:
            # 写法见 groups_from_cfg。带一句组备注、或者分成几块出一行总计，
            # 都从清单进来——工具里一个真实模块名都不许写死。
            wanted = groups_from_cfg(gj.get("groups") or gj)
            print()
            print(f"分组清单: {args.groups or cfg_path} —— "
                  + "；".join(
                      f"{g}[" + "+".join(f"{pn} {len(ks)}行"
                                         for pn, ks, _x in parts) + "]"
                      for g, parts, _t in wanted))
        cdata = {}
        print()
        print("=== 逐级关断电流 ===")
        for mod, books in sorted(grid.get(KIND_CUR, {}).items()):
            for chip in chips:
                b = books.get(chip)
                if b is None:
                    continue
                try:
                    cur = load_current(b.path, temp_col=args.temp_col,
                                       key_col=args.key_col, val_col=args.cur_col)
                except Exception as e:                # noqa: B902
                    failed.append((b, f"{type(e).__name__}: {e}"))
                    print(f"  {chip}: 读失败 —— {type(e).__name__}: {e}")
                    continue
                notes[id(b)] = f"{cur.n_rows}行×{cur.n_cols}列"
                # ★ 一个测量点都没有就别建页：模板复制品（表头齐、值全空）也能
                #   一路读到底，出来是一张整页空表——比没有这页更糟
                if not cur.temps:
                    print(f"  {chip}: 认出了电流列但**一个测量点都没有**，"
                          f"这份不进表   [{b.name}]")
                    for n, why, _k in cur.excluded[:3]:
                        print(f"     行{n}: {why}")
                    continue
                # ★ 再确认这份**像不像逐级关断**：一趟至少得有几步、而且总电流
                #   要往下走。表头对得上不代表内容对得上——模板复制品、或者别的
                #   测试用了同一套模板，都会一路读到底，出一张似是而非的表。
                mx = max(len(cur.runs[t]["steps"]) for t in cur.temps)
                down = sum(1 for t in cur.temps for _k, _m, _i, d in
                           cur.runs[t]["steps"] if d > 0)
                tot = sum(len(cur.runs[t]["steps"]) for t in cur.temps)
                if mx < 3:
                    print(f"  {chip}: 每趟只有 {mx} 个测量点，不像逐级关断"
                          f"（多半是别的测试用了同一套模板），这份不进表"
                          f"   [{b.name}]")
                    continue
                if tot and down < tot * 0.5:
                    print(f"  {chip}: ⚠ {tot} 步里有 {tot - down} 步的电流不降反升——"
                          f"这不像逐级关断，出来的数先别信")
                cdata[chip] = cur
                for t in cur.temps:
                    if t not in ctemps:
                        ctemps.append(t)
                    run = cur.runs[t]
                    rc = run.get("recheck")
                    print(f"  {chip} {fmt_num(t)}℃: 基线 {fmt_num(run['baseline'], 4)} mA"
                          f" / 关断 {len(run['steps'])} 步"
                          + (f" / 恢复后复测 {fmt_num(rc[3], 4)} mA"
                             f"（与基线差 {fmt_num(rc[3] - run['baseline'], 4)}）"
                             if rc else " / **没有恢复后复测点**"))
                # ★ 可信度那句话走这里，不进正表（正表出现 ⚠ 一定被追问）
                q = run_quality(cur, chip)
                for line in q:
                    print(f"     ⚠ {line}")
                if q:
                    quality_all.append((chip, mod, KIND_LABEL[KIND_CUR],
                                        [("这一趟的可信度", line) for line in q]))
                if cur.excluded:
                    print(f"     排除 {len(cur.excluded)} 行:")
                    print_excluded([(n, why) for n, why, _k in cur.excluded])
                    excl_all.append((chip, mod, KIND_LABEL[KIND_CUR],
                                     [(n, why) for n, why, _k in cur.excluded]))
                for w in cur.warnings:
                    print(f"     ⚠ {w}")
        ctemps.sort()
        if cdata:
            # 电流簿一份就同时装着两个 PLL 的步骤，所以这里的"组"来自清单、
            # 不是来自文件名的模块
            miss = {}
            # ★ 报告表在前、**原来那张全量表照旧留在最后**。分组是加出来的，
            #   不是替换（用户 2026-08-04："我不是让你把原来的删了……
            #   原来的我也要，只是在原来的基础上添加"）。
            #   全量表 ＝ 一个部件、键给 None ＝ 文件里全部关断步骤按原顺序。
            plan = list(wanted) + [("全部关断步骤",
                                    [("逐级关断", None, "")], "")]
            for gname, parts, total in plan:
                rowsg = {}
                for chip, cur in cdata.items():
                    rows_ = current_rows(cur, parts, total_name=total)
                    rowsg[chip] = rows_
                    have = {r["item"] for r in rows_ if r["kind"] == "result"
                            and any(isinstance(v, (int, float))
                                    for v in r["vals"].values())}
                    for _pn, ks, _x in parts:
                        for k in (ks or ()):
                            if k not in have and k not in miss.get(chip, ()):
                                miss.setdefault(chip, []).append(k)
                ctables.append((gname, rowsg))
            # ★ 没找到的键，顺手猜一个最像的报出来。真事故：清单里写了
            #   `L5 LO PreBUF`（那是 Mode 列的话），文件里的**键**叫 `L5_LOPRE`——
            #   只说"没找到"就得再跑一趟探查、再来回一轮。
            for chip, ks in miss.items():
                avail = {}
                for t in cdata[chip].temps:
                    for k, m, _i, _d in cdata[chip].runs[t]["steps"]:
                        avail.setdefault(k, m)
                tips = []
                for k in ks:
                    g = guess_key(k, avail)
                    tips.append(f"{k}" + (f"（是不是 {g}？）" if g else ""))
                print(f"  ⚠ {chip}: 清单里这些步骤在文件里没找到（或没有电流值）: "
                      + "，".join(tips[:10])
                      + (f" …共 {len(ks)} 个" if len(ks) > 10 else ""))
            # ★ 反方向也要说：**文件里有、清单里没有**的步骤。它们仍然在
            #   「全部关断步骤」那张全量表上，但不进任何一份总功耗——
            #   该进而漏掉了的话，两份报告就少算了电流，而且一声不响。
            #   这是"排除了什么逐条说出来"那条纪律的另一半。
            if wanted:
                inlist = {k for _g, parts, _t in wanted
                          for _pn, kk, _x in parts for k in (kk or ())}
                for chip, cur in cdata.items():
                    outside = []
                    for t in cur.temps:
                        for k, _m, _i, _d in cur.runs[t]["steps"]:
                            if k not in inlist and k not in outside:
                                outside.append(k)
                    if outside:
                        print(f"  · {chip}: 文件里另有 {len(outside)} 步不在任何一份"
                              f"总功耗里（只出现在「全部关断步骤」表上）: "
                              + "，".join(outside[:12])
                              + (" …" if len(outside) > 12 else ""))

    # ★★ 每一页只列**这一页真有数据**的芯片，不拿全局芯片表去铺列。
    #   性能和电流常常不是同一个人测的、测的也不是同一颗 die（一个目录只有温扫、
    #   另一个只有电流）。用全局表铺列的话，只测了电流的那颗会在 PLL / VCO 四张页上
    #   各占一整列"未测"——一列全是"未测"不是信息，是噪声，还会让人以为那颗片子
    #   测坏了。反过来电流页也一样。
    #   ★ 这不是把差异藏起来：哪颗片子有哪类数据，隐藏的 _审计 页逐份列着，
    #     控制台每次也打。正表上只摆有数的列。
    def _with_data(tbls):
        return [c for c in chips
                if any(d.get(c) is not None for _m, d in tbls)]

    chips_pll = _with_data(tables)
    chips_vco = _with_data(vtables)
    chips_cur = _with_data(ctables)

    # ---- 写出 ----
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    st = styles()
    if tables:
        write_summary(wb, tables, chips_pll, st, slim=args.slim,
                      sinfo=sinfo.get(KIND_PLL), dsb=dsb, spec=spec)
        write_journey(wb, tables, chips_pll, st, no_charts=args.no_charts,
                      chart_w=chart_w)
    if vtables:
        write_vco_summary(wb, vtables, chips_vco, st, vtemps, slim=args.slim,
                          sinfo=sinfo.get(KIND_VCO), dsb=dsb, spec=spec)
        write_vco_charts(wb, vcharts, chips_vco, st, vtemps,
                         no_charts=args.no_charts, chart_w=chart_w)
    # 电流页放最后：前四页是成对的（汇总+过程），别插进它们中间
    if ctables:
        write_grouped_page(wb, "Current_Summary", "{mod}（逐级关断电流）",
                           ctables, chips_cur, st, ctemps, slim=args.slim,
                           item_w=24, note_w=52, spec=spec)
    if not args.no_audit:
        write_audit(wb, picked, dropped, unknown, failed, notes, st, excl_all,
                    quality_all, spec.audit_rows() if spec else ())
    wb.calculation.fullCalcOnLoad = True

    out = args.out or os.path.join(os.path.dirname(root),
                                   os.path.basename(root) + "_chips_summary.xlsx")
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    n_fill, n_strip = VCACHE.inject(out)
    print(f"\n已写出: {os.path.abspath(out)}")
    print("  可见页: " + " / ".join(s.title for s in wb.worksheets
                                    if s.sheet_state == "visible"))
    for nm, cs in (("PLL_Summary / 温巡", chips_pll), ("VCO 两页", chips_vco),
                   ("Current_Summary", chips_cur)):
        if cs and set(cs) != set(chips):
            gone = [c for c in chips if c not in cs]
            print(f"  {nm} 只列了 {', '.join(cs)}；{', '.join(gone)} 没有这类数据，"
                  f"不给它留空列（谁有哪类数据见隐藏的 _审计 页）。")
    if spec:
        print()
        for line in spec.report():
            print("  " + line)
        print("  改 spec: 在这份簿子上直接改 → "
              "python spec_from_xlsx.py <这份簿子> --merge <配置>")
    else:
        print("  各汇总页的 Spec / 仿真 / Limit 列留空，填进 Spec Min/Max "
              "判定列自动出 PASS/FAIL 并上色。填完用 spec_from_xlsx.py 存进配置，"
              "以后每次重出都自带。")
    if args.slim:
        print("  --slim: 两张汇总表每片只显示常温列，其余温度列已折起——"
              "点表头上方的 ＋（或左上角的「2」）展开，数一个都没少。"
              "温巡页 / VCO压控页的竖条不受影响。")
    elif len(chips) >= 5:
        print(f"  提示: {len(chips)} 颗芯片＝实测区 {len(chips) * CHIP_W} 列，"
              f"横着容易数不清第几片。加 --slim 每片只显示常温列"
              f"（其余折起来，点 ＋ 就能展开核对）。")
    if args.trace:
        trace_item(args.trace, tables, vsweeps, sinfo, dsb, op_vtune_cfg, out)
    fails, n_agg, n_f = selfcheck(out)
    print(f"  自查: {n_agg} 个汇总格子都能在同一行的格子里找到；"
          f"{n_f} 个判定公式没有留下缓存值"
          + (f"（清掉了 {n_strip} 个空缓存）" if n_strip else "") + "。")
    if fails:
        print()
        print("  ✗ 自查没过，这份簿子先别发出去：")
        for f in fails[:20]:
            print(f"      {f}")
        sys.exit(1)


if __name__ == "__main__":
    main()
