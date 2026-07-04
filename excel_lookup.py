#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
excel_lookup.py — 拿一份控制信号 list，在寄存器 Excel 里反查它们所在的行

用途：
    我们从网表连接确定了需要的控制信号(reg_net 名, 见 control_signals.json)。
    这个脚本在寄存器簿的每个 sheet、每个单元格里搜这些名字，命中就把整行(裁剪后)
    抓出来——这样就能看到每个控制信号对应的 地址/bit/值/寄存器 是怎么写的。
    输出只含我们关心的信号, 体积小, 可直接贴回。

依赖: openpyxl。 支持 .xlsx / .xlsm。

用法:
    python excel_lookup.py <寄存器.xlsm> --signals control_signals.json --json hits.json
    python excel_lookup.py <寄存器.xlsm> --signals sig1,sig2,sig3
    python excel_lookup.py <寄存器.xlsm> --signals terms.txt --sheets RegMapDesign,regmap,Topout
    python excel_lookup.py <寄存器.xlsm> --signals control_signals.json --formulas   # 值是公式时读公式

匹配: 大小写不敏感的子串匹配; 自动去掉信号名里的位宽后缀 [3:0]。
"""
import argparse
import json
import os
import re
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


def col_letter(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def norm_term(t):
    """去掉位宽后缀与首尾空白, 供匹配。"""
    return re.sub(r'\[[^\]]*\]', '', str(t)).strip()


def load_terms(src):
    """从 JSON(取所有 reg_net) / .txt(每行一个) / 逗号串 读搜索词。"""
    if os.path.isfile(src):
        with open(src, encoding="utf-8") as f:
            data = f.read()
        if src.lower().endswith(".json"):
            obj = json.loads(data)
            terms = []

            def walk(o):
                if isinstance(o, dict):
                    for k, v in o.items():
                        if k == "reg_net" and isinstance(v, str):
                            terms.append(v)
                        else:
                            walk(v)
                elif isinstance(o, list):
                    for x in o:
                        walk(x)
                elif isinstance(o, str):
                    pass
            walk(obj)
            if not terms and isinstance(obj, list):        # 纯字符串列表
                terms = [str(x) for x in obj]
            return terms
        return [ln.strip() for ln in data.splitlines() if ln.strip()]
    return [t.strip() for t in src.split(",") if t.strip()]


def cell_norm(v, maxlen=28):
    if v is None:
        return None
    s = str(v).replace("\n", "\\n")
    return s if len(s) <= maxlen else s[: maxlen - 1] + "…"


def main():
    ap = argparse.ArgumentParser(description="按控制信号 list 反查 Excel")
    ap.add_argument("path", help="寄存器 .xlsx/.xlsm")
    ap.add_argument("--signals", required=True, help="control_signals.json / terms.txt / 逗号分隔")
    ap.add_argument("--json", default=None, help="导出命中到该 JSON")
    ap.add_argument("--sheets", default=None, help="只搜这些 sheet(逗号分隔)")
    ap.add_argument("--formulas", action="store_true", help="读公式原文而非缓存值")
    ap.add_argument("--max-hits", type=int, default=25, help="每个信号最多保留多少命中行")
    args = ap.parse_args()

    if not os.path.isfile(args.path):
        sys.exit(f"找不到文件: {args.path}")
    try:
        import openpyxl
    except ImportError:
        sys.exit("缺少 openpyxl: pip install openpyxl")

    raw_terms = load_terms(args.signals)
    # (原名, 归一化小写) ; 去重、丢掉过短的
    terms = []
    seen = set()
    for t in raw_terms:
        n = norm_term(t).lower()
        if len(n) < 4 or n in seen:
            continue
        seen.add(n)
        terms.append((t, n))
    print(f"搜索词 {len(terms)} 个 (来自 {args.signals})")

    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=not args.formulas)
    target_sheets = ([s.strip() for s in args.sheets.split(",")] if args.sheets
                     else wb.sheetnames)

    results = {orig: [] for orig, _ in terms}
    for sname in target_sheets:
        if sname not in wb.sheetnames:
            print(f"!! 无 sheet {sname!r}, 跳过")
            continue
        ws = wb[sname]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # 先把整行转小写文本, 快速判断这行有没有可能命中
            row_l = [("" if v is None else str(v).lower()) for v in row]
            joined = "".join(row_l)
            for orig, n in terms:
                if len(results[orig]) >= args.max_hits:
                    continue
                if n in joined:
                    cols = [col_letter(i + 1) for i, cell in enumerate(row_l) if n in cell]
                    # 裁掉尾部空列
                    vals = list(row)
                    while vals and (vals[-1] is None or str(vals[-1]).strip() == ""):
                        vals.pop()
                    results[orig].append({
                        "sheet": sname, "row": r_idx, "cols": cols,
                        "values": [cell_norm(v) for v in vals],
                    })

    found = {k: v for k, v in results.items() if v}
    not_found = [k for k, v in results.items() if not v]
    print(f"命中 {len(found)}/{len(terms)} 个信号; 未命中 {len(not_found)} 个")
    if not_found:
        print("未命中: " + ", ".join(not_found[:40]) + (" ..." if len(not_found) > 40 else ""))

    payload = {"file": os.path.basename(args.path),
               "n_terms": len(terms), "n_found": len(found),
               "hits": found, "not_found": not_found}
    if args.json:
        text = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        with open(args.json, "w", encoding="utf-8") as f:
            f.write(text)
        nb = len(text.encode("utf-8"))
        print(f"已导出: {args.json}  ({nb/1024:.1f} KB)")
    else:
        print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
