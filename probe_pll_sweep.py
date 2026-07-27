#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
probe_pll_sweep.py — 探查「性能扫描簿」的结构（默认脱敏，只出结构不出性能数值）

场景：
    一份仪器/脚本导出的性能扫描簿（例如 PLL 温度扫描：每个温度点测
    功率 / 积分相噪 / 若干 offset 的点相噪 / 杂散 / Vtune / 电流）。
    要给它写"汇总脚本"之前，得先摸清结构：几个 sheet、表是横着堆还是竖着堆、
    表头在第几行、哪几列是"条件"、哪几列是"性能"、一个测量点占几行、
    哪些格子系统性地是空的。

    但性能数据本身是敏感的。所以本工具把列分成两类：
      · 条件列 (cond)   —— 温度 / 角 / 模式 / 通道 / 序号 / offset 频点 …
                           这些是"测试怎么摆的"，给【全部去重值】。
      · 性能列 (perf)   —— 功率 / 相噪 / 杂散 / Vtune / 电流 / 输出频率 …
                           默认【只给非空计数 + 类型 + 空格分布】，不出任何数值。

    需要时再逐列放开（--reveal），而不是一上来全导出来。

依赖：
    只用 openpyxl。   pip install openpyxl

用法（先看整体，再按需放开）：
    python probe_pll_sweep.py <xlsx>                          # 保守探查，打印 JSON 到控制台
    python probe_pll_sweep.py <xlsx> --out                    # 同上但写文件（文件名自动取输入名）
    python probe_pll_sweep.py <xlsx> --out probe.json         # 写到指定文件
    python probe_pll_sweep.py <xlsx> --sheet "Sheet1"         # 只看某个 sheet
    python probe_pll_sweep.py <xlsx> --classify-only          # 只看"哪列被判成条件/性能"，最小输出
    python probe_pll_sweep.py <xlsx> --conds --out            # 只出条件面(取值+次数+逐行条件表)，零性能数值
    python probe_pll_sweep.py <xlsx> --reveal Temperature,Mode  # 额外把这几列当条件列出值
    python probe_pll_sweep.py <xlsx> --reveal "REG ADDR*,*_ADDR*"  # 支持 * 通配（宽表几十列时用）
    python probe_pll_sweep.py <xlsx> --hide Freq_MHz            # 把某列强制按性能藏起来
    python probe_pll_sweep.py <xlsx> --level stats            # 性能列补 min/max/mean（聚合，仍不出原始行）
    python probe_pll_sweep.py <xlsx> --level full --sheet X   # 迫不得已时才用：出样本行原值

输出体积：结束会打印字节数；>60KB 时用 --out 写文件，或用 --sheet / --classify-only 收窄。

⚠ --reveal 放开寄存器地址/值这类列之后，输出里就有真实 IP 了：只落 private/ 或黄区本地，
   不要进公开仓库（zero-IP 铁律：push 前扫文件内容 + commit message）。
"""

import argparse
import json
import os
import re
import sys
from fnmatch import fnmatch

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


# ---------------------------------------------------------------- 列分类

# 「条件」= 描述测试怎么摆的，不是被测性能。给全值。
COND_PATTERNS = [
    r"temp", r"tempe?rature", r"^t$", r"corner", r"process",
    r"vdd", r"vcc", r"avdd", r"dvdd", r"volt", r"supply", r"vbat",
    r"mode", r"band", r"chan", r"^ch\d*$", r"path", r"state",
    r"index", r"^idx", r"^no\.?$", r"^num", r"^id$", r"seq", r"step",
    r"item", r"^name$", r"label", r"comment", r"remark", r"note",
    r"date", r"time", r"site", r"chip", r"die", r"wafer", r"sample", r"unit",
    r"config", r"setting", r"sweep", r"point", r"cond", r"target", r"spec",
    r"divid", r"ndiv", r"fref", r"fcw", r"refclk", r"xtal", r"tcxo",
    # 仪器/测试台设置：怎么摆的，不是量出来的
    r"^system$", r"^test$", r"chamber", r"^f(lo|vco|xo)\b|^f(lo|vco|xo)_",
    r"_start|^start", r"_stop|^stop", r"correlat", r"^average$", r"resolution",
]
# 「设定频点」= offset 频率 / 杂散考察频点，属于测试怎么摆的，也给全值。
SETPOINT_PATTERNS = [
    r"spot\w*freq", r"offset", r"spurfreq", r"\w*spur\w*freq\d*", r"freq\d+$",
]
# 寄存器地址/值：既不是条件也不是性能，是 IP。默认一律藏，只有 --reveal 点名才出。
# 单靠上面的条件词表挡不住——Vtemp_ADDR1 命中 "temp"、REG_Value_Step1 命中 "step"，
# 都会被判成条件列，然后把真实地址原样打印出来。
SENSITIVE_PATTERNS = [
    r"addr", r"^reg\b", r"^reg[_ ]", r"value\d+$", r"^readback", r"_dump$",
]
# 明确是被测性能，即使名字里带上面的词也按性能藏起来（优先级最高）。
PERF_FORCE_PATTERNS = [
    r"result", r"ipn", r"phase\s*noise", r"^pn", r"_pn", r"jitter",
    r"power", r"dbm", r"dbc", r"vtune", r"current", r"_ma$", r"_ua$",
    r"gain", r"^p\d", r"harm", r"lock\s*time", r"settl",
]

UNIT_HINTS = [
    (r"dbm", "dBm"), (r"dbc/?hz", "dBc/Hz"), (r"dbc", "dBc"),
    (r"_mhz|freq_mhz", "MHz"), (r"_ghz", "GHz"), (r"_khz", "kHz"),
    (r"_ma\b|current_ma", "mA"), (r"_ua\b", "uA"),
    (r"_v\b|vtune", "V"), (r"_c\b|temp", "degC"),
]


def _match_any(name, patterns):
    n = name.strip().lower()
    return any(re.search(p, n) for p in patterns)


def _in_set(name, names):
    """--reveal/--hide 的匹配：大小写不敏感，支持 * ? 通配（宽表几十列时按前缀放开）。"""
    n = name.strip().lower()
    return any(fnmatch(n, p.strip().lower()) for p in names)


def classify_header(header, reveal, hide):
    """返回 (class, why)。class ∈ {'cond','setpoint','sensitive','perf'}

    只有 cond/setpoint 会出值；sensitive(寄存器地址/值=IP) 与 perf 一律不出，
    要看得显式 --reveal 点名。"""
    if header is None or str(header).strip() == "":
        return "perf", "无表头，保守按性能"
    h = str(header).strip()
    if _in_set(h, hide):
        return "perf", "--hide 指定"
    if _in_set(h, reveal):
        return "cond", "--reveal 指定"
    if _match_any(h, SENSITIVE_PATTERNS):
        return "sensitive", "寄存器地址/值=IP，默认藏"
    if _match_any(h, PERF_FORCE_PATTERNS):
        return "perf", "命中性能词"
    if _match_any(h, SETPOINT_PATTERNS):
        return "setpoint", "命中设定频点词"
    if _match_any(h, COND_PATTERNS):
        return "cond", "命中条件词"
    return "perf", "未识别，保守按性能"


def guess_unit(header):
    if not header:
        return None
    h = str(header).lower()
    for pat, unit in UNIT_HINTS:
        if re.search(pat, h):
            return unit
    return None


# ---------------------------------------------------------------- 基础工具

def col_letter(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)


def is_blank(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def is_numberish(v):
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if s in ("", "-", "--"):
            return False
        try:
            float(s)
            return True
        except ValueError:
            return False
    return False


def as_float(v):
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip().replace(",", ""))
        except ValueError:
            return None
    return None


def norm(v, maxlen=40):
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    s = str(v).replace("\n", "\\n").strip()
    return s if len(s) <= maxlen else s[: maxlen - 1] + "…"


def guess_type(vals):
    kinds = set()
    for v in vals:
        if is_blank(v):
            continue
        if isinstance(v, bool):
            kinds.add("bool")
        elif isinstance(v, int):
            kinds.add("int")
        elif isinstance(v, float):
            kinds.add("float")
        elif is_numberish(v):
            kinds.add("num-str")
        else:
            kinds.add("text")
    if not kinds:
        return "empty"
    return "/".join(sorted(kinds))


def dup_headers(header):
    """同名列 -> 它出现的所有列字母。

    宽表模板扩列时复制粘贴不改序号很常见（如 REG ADDR7/8/9 出现两遍）。
    汇总脚本要是"按表头名找列"，就会静默地一直取到第一个，后面那份永远读不到。
    先报出来，让写脚本的人改用列位置或"第几次出现"定位。
    """
    pos = {}
    for i, h in enumerate(header):
        if is_blank(h):
            continue
        pos.setdefault(str(norm(h)), []).append(col_letter(i + 1))
    return {k: v for k, v in pos.items() if len(v) > 1}


# ---------------------------------------------------------------- 块切分

def row_shape(row):
    """一行的形状：非空数 / 文本数 / 数字数。"""
    n_text = n_num = 0
    for v in row:
        if is_blank(v):
            continue
        if is_numberish(v):
            n_num += 1
        else:
            n_text += 1
    return n_text + n_num, n_text, n_num


def looks_like_header(row):
    """表头行：≥2 个非空，且非空里大多数是非数字文本。"""
    n, n_text, n_num = row_shape(row)
    if n < 2:
        return False
    return n_text >= max(2, int(0.6 * n))


def looks_like_label(row):
    """段落标签行：只有 1 个非空格子且是文本（如 "Temp = -40" / "Vtune_V"）。"""
    n, n_text, n_num = row_shape(row)
    return n == 1 and n_text == 1


def _header_here(cur, row):
    """这一行虽然"长得像表头"，但真的是新表头吗？

    坑：纯文本的数据行（如 寄存器名/地址 表）每行都"长得像表头"，
    会被切成一堆空块。规则：
      · 还没有块 → 是表头
      · 当前块还一行数据都没有 → 这是数据行（连着的第二行文本不当表头）
      · 当前块的数据行以数字为主，而这一行全是文本 → 是新表头（中途重复表头）
      · 否则 → 数据行
    """
    if cur is None:
        return True
    if not cur["data_rows"]:
        return False
    return cur.get("_num_heavy", False)


def split_blocks(all_rows):
    """把一个 sheet 切成若干「块」= 一个表头行 + 其下连续数据行。

    这类导出簿常见几种排布，都要认出来：
      ① 一个大表（表头 1 行 + N 行数据）
      ② 竖着堆几张小表（表头/数据/空行/表头/数据…）
      ③ 中途重复表头（没有空行分隔）
      ④ 单格标签行分段（如 "Temp=-40" 一行、"Vtune_V" 一行后面跟数据）
    """
    blocks = []
    cur = None
    pending_label = None

    def close():
        if cur and cur["data_rows"]:
            cur["data_row_end"] = cur["data_rows"][-1]
            blocks.append(cur)

    for i, row in enumerate(all_rows, start=1):  # 1 基行号
        n, n_text, n_num = row_shape(row)
        if n == 0:
            close()
            cur = None
            continue
        if looks_like_label(row):
            close()
            cur = None
            pending_label = {"row": i,
                             "text": norm(next(v for v in row if not is_blank(v)))}
            continue
        if looks_like_header(row) and _header_here(cur, row):
            close()
            cur = {"header_row": i, "header": list(row), "data_rows": [],
                   "label": pending_label}
            pending_label = None
            continue
        if cur is None:
            # 数据行在前、没表头（无名块，常见于标签行后直接跟一个数）
            cur = {"header_row": None, "header": [], "data_rows": [],
                   "label": pending_label}
            pending_label = None
        cur["data_rows"].append(i)
        cur["_num"] = cur.get("_num", 0) + n_num
        cur["_text"] = cur.get("_text", 0) + n_text
        cur["_num_heavy"] = cur["_num"] > cur["_text"]
    close()
    return blocks


# ---------------------------------------------------------------- 块分析

def block_signature(blk):
    """同一张表在多个段落里重复出现（每个温度一份）时，用表头文字做归并键。"""
    hdr = [norm(h) for h in blk["header"]]
    while hdr and is_blank(hdr[-1]):
        hdr.pop()
    if hdr:
        return ("hdr",) + tuple(str(h) for h in hdr)
    # 无表头块（标签行后跟一个数）：用标签文字区分，别把 Vtune 和 Current 混成一类
    return ("nohdr", str((blk.get("label") or {}).get("text")))


def analyse_block(group, all_rows, level, reveal, hide,
                  max_cond_values=40, max_empty_names=25, max_patterns=8):
    """group = 共享同一表头的若干块（每个温度段一个）。合并统计，出现位置逐个列。"""
    blk = group[0]
    header = blk["header"]
    row_ids = [r for b in group for r in b["data_rows"]]
    rows = [all_rows[r - 1] for r in row_ids]
    width = 0
    for r in [header] + rows:
        for c in range(len(r) - 1, -1, -1):
            if not is_blank(r[c]):
                width = max(width, c + 1)
                break

    cols_out = []
    keep_cols = []          # 参与"填充模式"的列下标
    for c in range(width):
        head = header[c] if c < len(header) else None
        vals = [r[c] if c < len(r) else None for r in rows]
        nonempty = sum(1 for v in vals if not is_blank(v))
        if is_blank(head) and nonempty == 0:
            continue        # 全空列，跳过
        keep_cols.append(c)
        cls, why = classify_header(head, reveal, hide)
        item = {
            "col": col_letter(c + 1),
            "header": norm(head),
            "class": cls,
            "why": why,
            "type": guess_type(vals),
            "nonempty": nonempty,
            "empty": len(vals) - nonempty,
        }
        unit = guess_unit(head)
        if unit:
            item["unit_guess"] = unit

        distinct = []
        seen = set()
        for v in vals:
            if is_blank(v):
                continue
            k = str(v)
            if k not in seen:
                seen.add(k)
                distinct.append(v)
        item["distinct"] = len(distinct)

        if cls in ("cond", "setpoint"):
            # 条件列：给全值（去重、有上限）+ 每个值出现几次。
            # 计数很关键：同一个温度出现多次 = 该温度点测了多轮（例如重锁一次再测），
            # 汇总时就不能"一个温度一行"。
            shown = distinct[:max_cond_values]
            item["values"] = [norm(v) for v in shown]
            if len(distinct) > len(shown):
                item["values_truncated"] = len(distinct) - len(shown)
            if len(distinct) <= max_cond_values:
                counts = {}
                for v in vals:
                    if is_blank(v):
                        continue
                    counts[str(norm(v))] = counts.get(str(norm(v)), 0) + 1
                if any(n > 1 for n in counts.values()):
                    item["value_counts"] = counts
        elif cls == "sensitive":
            pass                # 寄存器地址/值：连聚合都不给，只留列名和计数
        else:
            # 性能列：默认一个数值都不出
            if level in ("stats", "full"):
                nums = [x for x in (as_float(v) for v in vals) if x is not None]
                if nums:
                    item["stats"] = {
                        "n": len(nums),
                        "min": round(min(nums), 6),
                        "max": round(max(nums), 6),
                        "mean": round(sum(nums) / len(nums), 6),
                    }
        cols_out.append(item)

    # --- 填充模式：一个测量点占几行、哪些格子系统性为空 ---
    sigs = {}
    for r_idx, r in zip(row_ids, rows):
        sig = tuple(not is_blank(r[c]) if c < len(r) else False for c in keep_cols)
        g = sigs.setdefault(sig, {"n_rows": 0, "example_rows": []})
        g["n_rows"] += 1
        if len(g["example_rows"]) < 3:
            g["example_rows"].append(r_idx)
    patterns = []
    for sig, g in sorted(sigs.items(), key=lambda kv: -kv[1]["n_rows"])[:max_patterns]:
        empty_names = []
        for filled, c in zip(sig, keep_cols):
            if filled:
                continue
            h = header[c] if c < len(header) else None
            empty_names.append(norm(h) if not is_blank(h) else col_letter(c + 1))
        patterns.append({
            "n_rows": g["n_rows"],
            "example_rows": g["example_rows"],
            "n_empty_cols": len(empty_names),
            "empty_cols": empty_names[:max_empty_names]
            + (["…(截断)"] if len(empty_names) > max_empty_names else []),
        })

    occ = [{"label": (b.get("label") or {}).get("text"),
            "header_row": b["header_row"],
            "rows": [b["data_rows"][0], b.get("data_row_end")]} for b in group]

    out = {
        "occurrences": len(group),
        "where": occ[:60] + ([{"…": f"另 {len(occ)-60} 处"}] if len(occ) > 60 else []),
        "n_data_rows_total": len(rows),
        "n_cols": len(cols_out),
        "columns": cols_out,
        "row_fill_patterns": patterns,
        "distinct_fill_patterns": len(sigs),
    }
    dups = dup_headers(header)
    if dups:
        out["duplicate_headers"] = dups
        out["duplicate_headers_note"] = "同名列，按表头名定位会取错，改用列字母/第几次出现"

    # 多行表头会被误当成第一行数据，这里显式提醒
    if rows and looks_like_header(rows[0]):
        out["warn"] = "首行数据也像表头 → 可能是多行表头，请人工核对"

    if level == "full":
        out["sample_rows"] = [[norm(v) for v in r[:width]] for r in rows[:6]]
    return out


def group_blocks(blocks):
    """按表头归并：返回 [(signature, [blk,...]), ...]，保持首次出现顺序。"""
    order, groups = [], {}
    for b in blocks:
        sig = block_signature(b)
        if sig not in groups:
            groups[sig] = []
            order.append(sig)
        groups[sig].append(b)
    return [(s, groups[s]) for s in order]


def analyse_sheet(ws, level, reveal, hide, max_groups=25):
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    blocks = split_blocks(all_rows)
    groups = group_blocks(blocks)
    out = {
        "title": ws.title,
        "n_rows": ws.max_row,
        "n_cols": ws.max_column,
        "merged_cells_count": len(ws.merged_cells.ranges),
        "merged_cells": [str(r) for r in list(ws.merged_cells.ranges)[:20]],
        "n_blocks": len(blocks),
        "n_block_kinds": len(groups),
        "block_kinds": [analyse_block(g, all_rows, level, reveal, hide)
                        for _, g in groups[:max_groups]],
    }
    if len(groups) > max_groups:
        out["block_kinds_truncated"] = len(groups) - max_groups
    return out


def classify_only(ws, reveal, hide, max_patterns=4, max_empty_names=20):
    """最小输出：每种块只报一次 表头 -> 分类 -> 非空计数，先核对脱敏边界对不对。

    表头后面的数字＝该列有几行非空。计数不是数值，不破脱敏，但能一眼看出
    哪些结果列压根没测（占位列）、以及哪几种行填得不一样（例如重锁行不测相噪）。
    """
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    res = []
    for _, g in group_blocks(split_blocks(all_rows)):
        header = g[0]["header"]
        row_ids = [r for b in g for r in b["data_rows"]]
        rows = [all_rows[r - 1] for r in row_ids]
        cols, keep = [], []
        for c, head in enumerate(header):
            if is_blank(head):
                continue
            cls, _why = classify_header(head, reveal, hide)
            n = sum(1 for r in rows if c < len(r) and not is_blank(r[c]))
            cols.append(f"{norm(head)}[{cls}]={n}")
            keep.append(c)
        labels = [(b.get("label") or {}).get("text") for b in g]
        item = {
            "occurrences": len(g),
            "labels": [l for l in labels if l][:60],
            "first_header_row": g[0]["header_row"],
            "n_data_rows_each": sorted({len(b["data_rows"]) for b in g}),
            "n_data_rows_total": len(rows),
            "cols_note": "表头[分类]=非空行数",
            "cols": cols,
        }
        # 行填充模式：几种行、各几行、每种行哪些列是空的（看重锁行/占位行）
        sigs = {}
        for rid, r in zip(row_ids, rows):
            sig = tuple(c < len(r) and not is_blank(r[c]) for c in keep)
            grp = sigs.setdefault(sig, [])
            grp.append(rid)
        pats = []
        for sig, ids in sorted(sigs.items(), key=lambda kv: -len(kv[1]))[:max_patterns]:
            empt = [str(norm(header[c])) for f, c in zip(sig, keep) if not f]
            pats.append({"n_rows": len(ids), "example_rows": ids[:3],
                         "n_empty_cols": len(empt),
                         "empty_cols": empt[:max_empty_names]
                         + (["…(截断)"] if len(empt) > max_empty_names else [])})
        item["distinct_fill_patterns"] = len(sigs)
        item["row_fill_patterns"] = pats
        dups = dup_headers(header)
        if dups:
            item["duplicate_headers"] = dups
        res.append(item)
    return {"title": ws.title, "n_block_kinds": len(res), "block_kinds": res}


def conds_only(ws, reveal, hide, max_rows=200, max_row_cols=14, max_card=24):
    """只出条件面：条件列的取值+出现次数，外加**逐行的条件表**。

    为什么要逐行表：光有"温度有哪几个值"不够，还得知道这些条件是怎么组合的——
    56 行到底是 几温度 × 几 mode × 几测试项，同一温度是不是被测了多轮
    （例如中途重锁一次再测一遍）。有了它才能定汇总页一行放什么。
    整张表零性能数值，按用户的口径「条件信息可以完整获取」。
    """
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    res = []
    for _, g in group_blocks(split_blocks(all_rows)):
        header = g[0]["header"]
        row_ids = [r for b in g for r in b["data_rows"]]
        rows = [all_rows[r - 1] for r in row_ids]

        cond_cols, cols_out = [], []
        for c, head in enumerate(header):
            if is_blank(head):
                continue
            cls, _why = classify_header(head, reveal, hide)
            if cls not in ("cond", "setpoint"):
                continue
            vals = [r[c] if c < len(r) else None for r in rows]
            counts = {}
            for v in vals:
                if is_blank(v):
                    continue
                counts[str(norm(v))] = counts.get(str(norm(v)), 0) + 1
            cond_cols.append((c, str(norm(head)), len(counts)))
            entry = {
                "col": col_letter(c + 1),
                "header": norm(head),
                "class": cls,
                "distinct": len(counts),
                "nonempty": sum(1 for v in vals if not is_blank(v)),
            }
            if len(counts) == len(rows) and len(rows) > 1:
                entry["note"] = "每行都不同（行号/流水号一类），值省略"
            elif len(counts) <= max_card:
                entry["value_counts"] = counts
            else:
                entry["values_head"] = list(counts)[:max_card]
                entry["values_truncated"] = len(counts) - max_card
            cols_out.append(entry)

        # 逐行表选列：优先"会变但不是每行都不同"的列（真正的扫描维度），
        # 再把 NO./Index 这种唯一编号列留作行标签。
        idx_cols = [(c, h) for c, h, d in cond_cols
                    if d == len(rows) and re.match(r"^(no\.?|index|idx|序号)$", h.strip().lower())]
        vary = [(c, h) for c, h, d in cond_cols if 2 <= d <= max_card]
        pick = (idx_cols[:1] + [x for x in vary if x not in idx_cols])[:max_row_cols]
        pick.sort(key=lambda x: x[0])

        item = {
            "occurrences": len(g),
            "n_data_rows": len(rows),
            "cond_columns": cols_out,
        }
        if pick:
            item["cond_rows"] = {
                "note": "只含条件列，无任何性能数值；row=表内行号",
                "cols": ["row"] + [h for _, h in pick],
                "rows": [[rid] + [norm(r[c] if c < len(r) else None) for c, _ in pick]
                         for rid, r in list(zip(row_ids, rows))[:max_rows]],
            }
            if len(rows) > max_rows:
                item["cond_rows"]["truncated"] = len(rows) - max_rows
        dups = dup_headers(header)
        if dups:
            item["duplicate_headers"] = dups
        res.append(item)
    return {"title": ws.title, "n_block_kinds": len(res), "block_kinds": res}


# ---------------------------------------------------------------- 输出

_INNER_ARR = re.compile(r"\[\s*\n((?:[^\[\]{}]|\n)*?)\n\s*\]")


def compact_inner_arrays(text):
    """把最内层的"纯标量数组"折成一行：indent=2 下一个 56×9 的表会占 500 多行，
    折完一行就是一行数据，既好读体积又小。折的只是空白，JSON 语义不变；
    万一某个字符串里带方括号导致折错，解析不回来就整个放弃折叠。"""
    def rep(m):
        return "[" + re.sub(r"\s+", " ", m.group(1).strip()) + "]"
    out, prev = text, None
    while prev != out:
        prev, out = out, _INNER_ARR.sub(rep, out)
    try:
        json.loads(out)
    except ValueError:
        return text
    return out


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(
        description="探查性能扫描簿结构（默认脱敏：条件列出值，性能列只出计数）")
    ap.add_argument("path", help="xlsx / xlsm 文件路径")
    ap.add_argument("--sheet", default=None, help="只看指定 sheet")
    ap.add_argument("--out", nargs="?", const="__auto__", default=None,
                    help="写 JSON 文件（不给则打印到控制台）；"
                         "只写 --out 不给文件名则自动取 <输入名>.<档位>.probe.json")
    ap.add_argument("--level", choices=["strict", "stats", "full"], default="strict",
                    help="strict=性能列零数值(默认); stats=补 min/max/mean 聚合; "
                         "full=出样本行原值(迫不得已才用)")
    ap.add_argument("--reveal", default="",
                    help="逗号分隔的表头名，强制当条件列出全值")
    ap.add_argument("--hide", default="",
                    help="逗号分隔的表头名，强制当性能列藏起来")
    ap.add_argument("--classify-only", action="store_true",
                    help="只输出「表头 -> 条件/性能」分类，最小体积，先核对脱敏边界")
    ap.add_argument("--conds", action="store_true",
                    help="只出条件面：条件列取值+出现次数 + 逐行条件表（零性能数值）；"
                         "用来看扫描维度怎么组合、同一条件是否重复测了多轮")
    args = ap.parse_args()

    if not os.path.isfile(args.path):
        sys.exit(f"找不到文件: {args.path}")
    if args.path.lower().endswith(".xls"):
        sys.exit("openpyxl 读不了老的 .xls，请先在 Excel 里另存为 .xlsx 再试。")

    try:
        import openpyxl
    except ImportError:
        sys.exit("缺少 openpyxl，请先: pip install openpyxl")

    reveal = {s.strip() for s in args.reveal.split(",") if s.strip()}
    hide = {s.strip() for s in args.hide.split(",") if s.strip()}

    wb = openpyxl.load_workbook(args.path, read_only=False, data_only=True)
    names = [args.sheet] if args.sheet else wb.sheetnames
    for n in names:
        if n not in wb.sheetnames:
            sys.exit(f"没有名为 {n!r} 的 sheet；实际有: {wb.sheetnames}")

    obj = {
        "file": os.path.basename(args.path),
        "n_sheets": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
        "redact_level": ("classify-only" if args.classify_only
                         else "conds-only" if args.conds else args.level),
        "note": "cond/setpoint 列出全值；perf 列默认不出任何数值",
    }
    if args.classify_only:
        obj["sheets"] = [classify_only(wb[n], reveal, hide) for n in names]
    elif args.conds:
        obj["sheets"] = [conds_only(wb[n], reveal, hide) for n in names]
    else:
        obj["sheets"] = [analyse_sheet(wb[n], args.level, reveal, hide) for n in names]

    text = compact_inner_arrays(json.dumps(obj, ensure_ascii=False, indent=2))
    nbytes = len(text.encode("utf-8"))
    out_path = args.out
    if out_path == "__auto__":
        stem = os.path.splitext(os.path.basename(args.path))[0]
        tag = ("classify" if args.classify_only
               else "conds" if args.conds else args.level)
        out_path = f"{stem}.{tag}.probe.json"
    if out_path:
        with open(out_path, "w", encoding="utf-8", newline="\n") as f:
            f.write(text)
        print(f"已写出: {os.path.abspath(out_path)}  "
              f"({nbytes} bytes = {nbytes/1024:.1f} KB)")
    else:
        print(text)
        print(f"\n---- {nbytes} bytes = {nbytes/1024:.1f} KB ----", file=sys.stderr)
        if nbytes > 60 * 1024:
            print("⚠ 输出偏大：加 --out 写文件，或用 --sheet / --classify-only 收窄。",
                  file=sys.stderr)


if __name__ == "__main__":
    main()
