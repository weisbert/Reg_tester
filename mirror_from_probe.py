#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
mirror_from_probe.py — 用 probe_current_data.py(v2) 的 JSON 快照在本地重建数据镜像

重建内容（够 current_db.py 跑通即可，不还原 Result 里的寄存器列）：
  - 仿真工作簿（按原相对路径、原 tab 名，整页写回）
  - Result 文件（按原相对路径，NO./Mode/Current/Temperature 按原列号原行号写回）

用法：
  python mirror_from_probe.py private\\probe_dump.json [-o private\\current_mirror]
  python current_db.py build --root private\\current_mirror --chip C1
"""
import argparse
import gzip
import json
import os
import sys

import openpyxl


def ensure_dir(path):
    d = os.path.dirname(path)
    if d:
        os.makedirs(d, exist_ok=True)


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    ap = argparse.ArgumentParser(description="从 probe_dump.json 重建本地数据镜像")
    ap.add_argument("dump", help="probe_current_data.py 生成的 JSON")
    ap.add_argument("-o", "--out", default=os.path.join("private", "current_mirror"),
                    help="镜像输出根目录（默认 private/current_mirror，已 gitignore）")
    args = ap.parse_args()

    opener = gzip.open if args.dump.lower().endswith(".gz") else open
    with opener(args.dump, "rt", encoding="utf-8") as f:
        dump = json.load(f)
    if dump.get("probe_version", 1) < 3:
        print("[警告] 这是旧版探查快照，建议 git pull 后用 v3 脚本重新探查")
    profiles = dump.get("header_profiles", [])
    out_root = os.path.abspath(args.out)
    os.makedirs(out_root, exist_ok=True)

    # 仿真工作簿：同一文件的多个 sim tab 合进一个工作簿
    sim_books = {}
    for sim in dump.get("sim", []):
        if "error" in sim or not sim.get("rows"):
            print(f"[仿真] {sim.get('file')} 跳过（{sim.get('error', '无数据')}）")
            continue
        sim_books.setdefault(sim["file"], []).append(sim)
    for rel, sheets in sim_books.items():
        wb = openpyxl.Workbook()
        for k, sim in enumerate(sheets):
            ws = wb.active if k == 0 else wb.create_sheet()
            ws.title = sim["sheet"]
            for row in sim["rows"]:
                ws.append(row)
        path = os.path.join(out_root, rel)
        ensure_dir(path)
        wb.save(path)
        print(f"[仿真] {rel} -> {sum(len(s['rows']) for s in sheets)} 行 / {len(sheets)} tab")

    # Result 文件
    for res in dump.get("results", []):
        if "error" in res:
            print(f"[实测] {res.get('folder')}/{res.get('file')} 跳过（{res['error']}）")
            continue
        rel = os.path.join(res.get("folder", ""), res["file"])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = res.get("sheet", "Sheet1")
        hdr_row = res["header_row"]
        headers = res.get("headers")
        if headers is None:
            headers = profiles[res["header_profile"]]
        for col_1b, name in headers:
            ws.cell(row=hdr_row, column=col_1b, value=name)
        cols = res["key_cols_1based"]
        for row_idx, no_raw, label, cur, temp in res["rows"]:
            if cols.get("no"):
                ws.cell(row=row_idx, column=cols["no"], value=no_raw)
            if cols.get("mode"):
                ws.cell(row=row_idx, column=cols["mode"], value=label)
            if cols.get("cur") and cur is not None:
                ws.cell(row=row_idx, column=cols["cur"], value=cur)
            if cols.get("temp") and temp is not None:
                ws.cell(row=row_idx, column=cols["temp"], value=temp)
        path = os.path.join(out_root, rel)
        ensure_dir(path)
        wb.save(path)
        print(f"[实测] {rel} -> {len(res['rows'])} 行")

    print(f"\n[完成] 镜像根目录: {out_root}")
    print(f"下一步: python current_db.py build --root \"{out_root}\" --chip C1")


if __name__ == "__main__":
    main()
