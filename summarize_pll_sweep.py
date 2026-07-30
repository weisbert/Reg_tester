#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
summarize_pll_sweep.py — 性能扫描簿 → 带汇总页的 Excel

输入：一份仪器/脚本导出的性能扫描宽表（一行 = 一个测试项，列里有条件也有结果）。
输出：一份新的 .xlsx，**是拿去 review 的成品**——
    汇总       **这一页就是结论**。compliance 表：上半「条件行」说明这组数在什么
               条件下取的，下半「结果行」一行一个指标 × MIN/TYP/MAX，Spec 与
               limit 留空给人填，填完 PASS/FAIL 与超规标红由公式+条件格式自动出。
               不另起「结论页」——那只会把同样的判定换个说法复述一遍，
               看的人反而要问"这两页什么关系"。
    原始数据   原封不动，就是输入那一页
    温巡过程   按实际测试顺序摊平的全表 + 过程图（重锁点标出来）
    温度明细   指标 × 温度矩阵（按段分列），汇总的数就是对它取 MIN/MAX/MEDIAN
    图表       关键指标的 值-vs-温度 图（每段一条线）
    相噪-offset 相噪曲线，图和数据同页
    _审计      排除了哪些行、有哪些告警。**默认隐藏**

这份簿要回答的问题，以及各页的分工
    一颗锁定的 PLL 走完一整趟温度（中途不重锁，只在端点重锁）：
      ① 各项性能的全温最坏值满不满足 spec        -> 汇总页
      ② 不重锁跑完全温，还锁着吗、压控走到轨没    -> 汇总页的 Freq / Vtune 两行
         （压控余量＝给 Vtune 填 limit=range + Spec 上下轨，判定即余量够不够）
                                                    ＋ 温巡过程页的图
      ③ 重锁有没有效、复位一致不一致              -> 温巡过程图的重锁点 + 页顶那两句结论
    ★ 「锁在哪、锁了几次」是取数的**条件**，不是被考核的性能：
      它在汇总页里只占一行条件行，重锁点本身不计入性能统计。
      按温度段切分只用来看回滞（温度明细/图表），不进 compliance 表的骨架。

两条出稿纪律
    · **正表里不写使用说明、不写告警。** 这份簿要给人 review，翻到"怎么用""⚠ 告警"
      只会招来"这是什么"的追问。这些进隐藏的 _审计 页，控制台每次也完整打印。
    · **格子里不写死数。** 明细页每格是 ='原始页'!XX 引用，汇总页是对明细取
      MIN/MAX/MEDIAN 的公式。原始数据改了汇总跟着变，也不会有人怀疑
      "这个数是手敲的"。代价是 openpyxl 读不到算好的值，只有 Excel 打开才算。

依赖：只用 openpyxl。   pip install openpyxl

用法：
    python summarize_pll_sweep.py <扫描簿.xlsx>                  # 出 <原名>_summary.xlsx
    python summarize_pll_sweep.py <扫描簿.xlsx> -o 汇总.xlsx
    python summarize_pll_sweep.py <扫描簿.xlsx> --dry-run        # 只打印识别结果，不写文件
    python summarize_pll_sweep.py <扫描簿.xlsx> --leg-col Mode --lock-pattern "_lock$"
    python summarize_pll_sweep.py <扫描簿.xlsx> --keep-test-item PLL_Test

先跑 --dry-run 看三件事对不对：列映射 / 分了几段 / 排除了哪些行。
排除的行不会被悄悄丢掉，汇总页底部会逐行列出原因。
"""

import argparse
import os
import re
import sys
from collections import OrderedDict

from xlsx_formula_cache import FormulaCache

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# 读数 / 分段 / 统计 / 画格子 / 图表原语全在公共层 sweep_lib——
# 跨芯片汇总 summarize_chips.py 用的是同一份，口径只能有一处。
from sweep_lib import (                                          # noqa: E402
    FILL_FAIL, FILL_PASS, COLOR_FLAG, COLOR_PASS,
    is_blank, num, txt, fmt_num, leg_series, stats, stats_all,
    load_sweep, SweepError, styles as _styles, put,
    nice_step, axis_bounds, apply_y as _apply_y, legend_bottom as _legend_bottom,
    style_series as _style, blank_policy, LEG_STYLE,
)

# 公式格的缓存值：openpyxl 只写空的 <v></v>，不补真值的话别人打开是一片空白。
# 详见 xlsx_formula_cache 的模块说明（那个坑值得读一遍）。
VCACHE = FormulaCache()


def cache(ws, row, col, value):
    VCACHE.remember(ws, row, col, value)


def sref(title, col_idx0, xl_row, coerce=False):
    """指向原始页某个格子的引用串，如 ='Sheet1'!CD5。

    汇总里的每个数都做成引用而不是写死：原始页改了值，汇总跟着变；
    也免得 review 的人怀疑"这个数是不是手敲进去的"。
    ⚠ 只在源格子确实有值时才引用——Excel 里引用空格子会显示成 0。

    coerce=True 时发 `=--'页'!CD5`（双负号）。原始表里某些列会存成
    **文本型数字**：Python 的 float() 解得动，Excel 的 COUNT/MIN/MATCH 却一律
    不认，于是整列被当成没有数字——MIN 返回 0、MATCH 找不到 0，
    极值温度那格就是 #N/A。双负号把文本转成数值，是数值则原样。
    """
    from openpyxl.utils import get_column_letter
    ref = (f"'{str(title).replace(chr(39), chr(39) * 2)}'"
           f"!{get_column_letter(col_idx0 + 1)}{xl_row}")
    return f"=--{ref}" if coerce else f"={ref}"


# ---------------------------------------------------------------- 汇总页

def build_conditions(legs, cols, rows, items, room_t):
    """从数据里生成「条件行」——这组数是在什么条件下取的。

    ★ 重锁写在这里，只占一行。它是取数的条件，不是被考核的性能：
    上一版把四个温度段做成了性能表的列分组，等于拿锁的过程去当性能结论的骨架。
    """
    def vals_of(name):
        i = cols.idx(name)
        if i is None:
            return []
        seen = []
        for lg in legs:
            for r in lg.rows:
                if r.kind == "lock":
                    continue      # 重锁行的 Mode 带 _lock 后缀，会污染"这组数是什么模式"
                v = txt(r.raw[i]) if i < len(r.raw) else ""
                if v and not is_blank(v) and v not in seen:
                    seen.append(v)
        return seen

    def joined(*names, sep=" / ", unit=""):
        parts = []
        for n in names:
            vs = vals_of(n)
            if vs:
                parts.append("/".join(vs[:4]))
        return sep.join(parts) + (f" {unit}" if parts and unit else "")

    temps = sorted({t for lg in legs for t in lg.temps})
    n_pts = sum(len(leg_series(lg, items[0])) for lg in legs) if items else 0
    conds = []

    if temps:
        conds.append(("Condition", "Temperature", "℃",
                      (temps[0], room_t, temps[-1]),
                      f"{len(temps)} 个温度点，步进最小 "
                      f"{fmt_num(min(b - a for a, b in zip(temps, temps[1:])))}℃"
                      if len(temps) > 1 else ""))
    for label, name in (("Mode", "Mode"), ("Test Item", "Test Item"),
                        ("System", "System")):
        v = vals_of(name)
        if v:
            conds.append(("Condition", label, "", "/".join(v[:4]), ""))
    f = joined("fLO_MHz", "fVCO_MHz", "fXO_MHz")
    if f:
        conds.append(("Condition", "fLO / fVCO / fXO", "MHz", f, ""))

    # 锁定方式：这一行就是"锁的过程"该待的地方
    lt = "/".join(str(fmt_num(lg.lock_temp)) for lg in legs if lg.lock_temp is not None)
    if lt:
        conds.append(("Condition", "锁定方式", "",
                      f"段首重锁 {len([l for l in legs if l.lock_temp is not None])} 次 @{lt}℃，段内全程不重锁",
                      "重锁点不计入下方性能统计，见「重锁对比」页"))
    conds.append(("Condition", "计入统计的测量点", "", f"{n_pts} 点",
                  f"{len(legs)} 段温巡去重后"))

    # 仪器/测量设置
    offs = sorted({num(r.raw[cols.idx(f"SpotPNFreq{i}")])
                   for i in range(1, 12) if cols.idx(f"SpotPNFreq{i}") is not None
                   for lg in legs for r in lg.rows
                   if num(r.raw[cols.idx(f"SpotPNFreq{i}")]) is not None})
    if offs:
        conds.append(("Setup", "相噪 offset 点", "MHz",
                      " / ".join(str(fmt_num(o)) for o in offs), ""))
    spurs = sorted({num(r.raw[cols.idx(f"OtherSpurFreq{i}")])
                    for i in range(1, 12) if cols.idx(f"OtherSpurFreq{i}") is not None
                    for lg in legs for r in lg.rows
                    if num(r.raw[cols.idx(f"OtherSpurFreq{i}")]) is not None})
    if spurs:
        conds.append(("Setup", "杂散考察点", "MHz",
                      " / ".join(str(fmt_num(s)) for s in spurs), ""))
    span = joined("Freq_Start_MHz", "Freq_Stop_MHz", sep=" ~ ")
    if span:
        conds.append(("Setup", "相噪分析带宽", "MHz", span, ""))
    ca = joined("Correlation", "Average")
    if ca:
        conds.append(("Setup", "Correlation / Average", "", ca, ""))
    return conds


def build_drift_rows(jinfo):
    """「相对锁定点偏离」——每段一行，摆成和别的结果行一样的 MIN/TYP/MAX 形状。

    这是这套温巡唯一一个别处算不出来的量：锁定在某个温度之后，一路跑到温度
    另一端，工作点最多往负/正两个方向各偏出去多少、分别偏在哪个温度。
    压控余量够不够就看它。
    MIN = 本段最低点 − 锁定点，MAX = 本段最高点 − 锁定点（所以锁定点自己是 0）。
    给这几行填 limit=range + Spec 上下限，就是"允许偏离多少"的判定。
    """
    from openpyxl.utils import get_column_letter as L
    if not jinfo or not jinfo.get("c_vt") or not jinfo.get("legs"):
        return []
    vt = jinfo["vt_item"]
    J, cv = f"'{jinfo['sheet']}'", L(jinfo["c_vt"])
    out = []
    for lg in jinfo["legs"]:
        f0, f1, k = lg["first"], lg["last"], lg["lock_row"]
        seg = f"{J}!${cv}${f0}:${cv}${f1}"
        lock = f"{J}!${cv}${k}"
        vv = lg.get("vt_vals") or []
        lo_hi = ((min(vv) - vv[0], max(vv) - vv[0]) if vv else (None, None))
        out.append({
            "_min": lo_hi[0], "_max": lo_hi[1],
            "cat": f"{vt.label} 温漂",
            "item": f"锁@{fmt_num(lg['lock_temp'])}℃ 后 {lg['stage']}",
            "unit": vt.unit,
            "m_min": f'=IF(COUNT({seg})=0,"",MIN({seg})-{lock})',
            "m_max": f'=IF(COUNT({seg})=0,"",MAX({seg})-{lock})',
            # 三个代表温度点那几列留空：偏离是"每段一个数", 不是"每个温度一个数"
        })
    return out


def write_summary(wb, legs, items, meta, st, room_t, dref, jinfo):
    """compliance 版式的汇总页：上半条件行、下半结果行，轴 = MIN / TYP / MAX。

    照 report-forge 的 compliance table 来：
      · 条件行(米色 setting)说明这组数在什么条件下取的，不参与判定；
      · 结果行(白底 result)一行一个指标，MIN/TYP/MAX 各自对 spec 判定；
      · limit 给方向：le 比 spec MAX、ge 比 spec MIN、range 两头都比。
    Spec 三格和 limit 都留空给人填，判定与标红是 Excel 公式+条件格式，
    填完立刻出结果，不用重跑脚本。
    """
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter as L
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet("汇总")
    # 实测组：先给几个代表温度点的值（最低 / 常温 / 最高），再给全温 MIN/MAX/Δ。
    # 原先那两列「@℃」（极值出现在哪个温度）看表的人反映碍眼——想知道趋势
    # 直接看三个温度点的值就够了，极值落在哪一格要细究再去「温度明细」。
    all_t = sorted({t for lg in legs for t in lg.temps})
    show_t = sorted({all_t[0], room_t, all_t[-1]} - {None}) if all_t else []
    plan = [("cat", "Category", 15), ("item", "Item", 24), ("unit", "Unit", 9),
            ("sep", "", 2),
            ("s_min", "MIN", 10), ("s_typ", "TYP", 10), ("s_max", "MAX", 10),
            ("limit", "limit", 8), ("sep", "", 2),
            ("sim", "仿真值", 12), ("sep", "", 2)]
    tkeys = []
    for i, t in enumerate(show_t):
        k = f"t{i}"
        tkeys.append((k, t))
        plan.append((k, f"{fmt_num(t)}℃", 11))
    plan += [("m_min", "MIN", 11), ("m_max", "MAX", 11), ("m_d", "Δ", 10),
             ("sep", "", 2),
             ("judge", "判定 / 备注", 26)]
    C = {k: i + 1 for i, (k, _l, _w) in enumerate(plan) if k != "sep"}
    for i, (_k, _l, w) in enumerate(plan):
        ws.column_dimensions[L(i + 1)].width = w

    n_pts = sum(len(leg_series(lg, items[0])) for lg in legs) if items else 0
    temps = sorted({t for lg in legs for t in lg.temps})
    groups = [("Spec", "MIN / TYP / MAX　判据 limit",
               [C["s_min"], C["s_typ"], C["s_max"], C["limit"]]),
              ("实测",
               f"代表温度点 ＋ 全温 {fmt_num(temps[0])} ~ {fmt_num(temps[-1])}℃ "
               f"{n_pts} 点的极值" if temps else "",
               [C[k] for k, _t in tkeys] + [C["m_min"], C["m_max"], C["m_d"]])]

    for r in range(1, 4):
        for c in range(1, len(plan) + 1):
            put(ws, r, c, None, st, st["f_head"])
    for key, label in (("cat", "Category"), ("item", "Item"), ("unit", "Unit"),
                       ("sim", "仿真值"), ("judge", "判定 / 备注")):
        ws.merge_cells(start_row=1, start_column=C[key], end_row=3, end_column=C[key])
        put(ws, 1, C[key], label, st, st["f_head"], bold=True)
    for title, stage, cc in groups:
        ws.merge_cells(start_row=1, start_column=cc[0], end_row=1, end_column=cc[-1])
        put(ws, 1, cc[0], title, st, st["f_head"], bold=True)
        ws.merge_cells(start_row=2, start_column=cc[0], end_row=2, end_column=cc[-1])
        put(ws, 2, cc[0], stage, st, st["f_head"], bold=True, size=9)
        for ci in cc:
            put(ws, 3, ci, plan[ci - 1][1], st, st["f_head"], bold=True, size=9)
    for i, (k, _l, _w) in enumerate(plan):
        if k == "sep":
            for r in range(1, 4):
                put(ws, r, i + 1, None, st, st["f_sep"])

    r = 4
    # ---- 条件行 ----
    cond_first = r
    for cat, item, unit, val, note in meta["conditions"]:
        for i, (k, _l, _w) in enumerate(plan):
            put(ws, r, i + 1, None, st,
                st["f_sep"] if k == "sep" else st["f_group"])
        put(ws, r, C["cat"], cat, st, st["f_group"], bold=True)
        put(ws, r, C["item"], item, st, st["f_group"], align="left")
        put(ws, r, C["unit"], unit, st, st["f_group"])
        if isinstance(val, tuple):
            for ci, v in zip([C[k] for k, _ in tkeys], val):
                put(ws, r, ci, fmt_num(v), st, st["f_group"], bold=True)
        else:
            c0 = C[tkeys[0][0]] if tkeys else C["m_min"]   # 实测组的第一列
            ws.merge_cells(start_row=r, start_column=c0,
                           end_row=r, end_column=C["m_d"])
            put(ws, r, c0, val, st, st["f_group"], bold=True, align="left")
        if note:
            put(ws, r, C["judge"], note, st, st["f_group"], size=8, align="left")
        r += 1

    # ---- 结果行 ----
    # 结果行 = 常规指标 + 「相对锁定点偏离」派生行。后者插在压控电压那一行后面，
    # 因为它就是压控余量的量化。两种行在表里长得一模一样，读的人不用分辨。
    specs = []
    for it in items:
        b = dref.get(it.label)
        s = {"cat": it.cat, "item": it.label, "unit": it.unit}
        # 同时把 Python 侧算好的数留下来，存盘后补进公式格的缓存值
        sa = stats_all(legs, it)
        if sa:
            s["_min"], s["_max"] = sa["min"], sa["max"]
        for k, tt in tkeys:
            vv = sorted(v for lg in legs
                        for t, v in leg_series(lg, it).items() if t == tt)
            if vv:
                s["_" + k] = (vv[len(vv) // 2] if len(vv) % 2
                              else (vv[len(vv) // 2 - 1] + vv[len(vv) // 2]) / 2.0)
        if b:
            D, f0, f1 = b["sheet"], b["first"], b["last"]
            lo = f"{D}!${L(b['c_lo'])}${f0}:${L(b['c_lo'])}${f1}"
            hi = f"{D}!${L(b['c_hi'])}${f0}:${L(b['c_hi'])}${f1}"
            # 全部带护栏：一个数字都没有时给空白, 而不是 MIN 返回 0、
            # MATCH 再找不到 0 而甩出 #N/A。报告里不该出现错误值。
            s["m_min"] = f'=IF(COUNT({lo})=0,"",MIN({lo}))'
            s["m_max"] = f'=IF(COUNT({hi})=0,"",MAX({hi}))'
            # 每个代表温度点：该温度在各段里被测了好几次, 取中位数
            for k, tt in tkeys:
                dr = (b.get("rows_by_temp") or {}).get(tt)
                if dr:
                    rr = f"{D}!${L(2)}${dr}:${L(1 + b['n_legs'])}${dr}"
                    s[k] = f'=IF(COUNT({rr})=0,"",MEDIAN({rr}))'
        specs.append(s)
    drift = build_drift_rows(jinfo)
    if drift:
        vt_label = jinfo["vt_item"].label
        at = next((i for i, s in enumerate(specs) if s["item"] == vt_label), len(specs) - 1)
        specs[at + 1:at + 1] = drift

    res_first = r
    fill_in = {"s_min", "s_typ", "s_max", "limit", "sim"}   # 留给人填的格子
    for sp in specs:
        for i, (k, _l, _w) in enumerate(plan):
            put(ws, r, i + 1, None, st,
                st["f_sep"] if k == "sep" else
                st["f_in"] if k in fill_in else st["f_res"])
        put(ws, r, C["item"], sp["item"], st, st["f_res"], align="left")
        put(ws, r, C["unit"], sp["unit"], st, st["f_res"])
        for key in [k for k, _ in tkeys] + ["m_min", "m_max"]:
            if sp.get(key):
                put(ws, r, C[key], sp[key], st, st["f_res"])
                cache(ws, r, C[key], sp.get("_" + key.replace("m_", "")
                                            if key.startswith("m_") else "_" + key))
        put(ws, r, C["m_d"],
            f'=IF(OR({L(C["m_min"])}{r}="",{L(C["m_max"])}{r}=""),"",'
            f'{L(C["m_max"])}{r}-{L(C["m_min"])}{r})', st, st["f_res"])
        if sp.get("_min") is not None and sp.get("_max") is not None:
            cache(ws, r, C["m_d"], sp["_max"] - sp["_min"])
        mn, mx = f"{L(C['m_min'])}{r}", f"{L(C['m_max'])}{r}"
        sn, sx, lim = f"${L(C['s_min'])}{r}", f"${L(C['s_max'])}{r}", f"${L(C['limit'])}{r}"
        # 实测为空时不判（Excel 里 ""<=数字 会算成 FALSE，不挡住会误判成 FAIL）
        put(ws, r, C["judge"],
            f'=IF(OR({lim}="",{mn}="",{mx}=""),"",'
            f'IF({lim}="le",IF({sx}="","",IF({mx}<={sx},"PASS","FAIL")),'
            f'IF({lim}="ge",IF({sn}="","",IF({mn}>={sn},"PASS","FAIL")),'
            f'IF(AND({sn}<>"",{mn}<{sn}),"FAIL",'
            f'IF(AND({sx}<>"",{mx}>{sx}),"FAIL","PASS")))))',
            st, st["f_res"], bold=True)
        r += 1
    res_last = r - 1

    # Category 纵向合并（条件块和结果块各自按类合并）
    def merge_cat(first, last, key):
        i = first
        while i <= last:
            j = i
            while j + 1 <= last and key(j + 1) == key(i):
                j += 1
            if j > i:
                ws.merge_cells(start_row=i, start_column=1, end_row=j, end_column=1)
            i = j + 1
    merge_cat(cond_first, res_first - 1, lambda rr: meta["conditions"][rr - cond_first][0])
    for n, sp in enumerate(specs):
        put(ws, res_first + n, C["cat"], sp["cat"], st, st["f_res"], bold=True)
    merge_cat(res_first, res_last, lambda rr: specs[rr - res_first]["cat"])

    # limit 下拉，省得手打错
    dv = DataValidation(type="list", formula1='"le,ge,range"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{L(C['limit'])}{res_first}:{L(C['limit'])}{res_last}")

    # 超规标红：MIN / TYP / MAX 三格各自判，语义照 report-forge 的 le/ge/range。
    # ⚠ 逐列加，不能拿一个连续区间盖过去——中间夹着「@℃」列装的是温度，
    #   被拿去跟 spec 比会莫名其妙标红。
    a = res_first
    sn, sx, lim = f"${L(C['s_min'])}{a}", f"${L(C['s_max'])}{a}", f"${L(C['limit'])}{a}"
    for key in [k for k, _ in tkeys] + ["m_min", "m_max"]:
        col = L(C[key])
        cur = f"{col}{a}"
        ws.conditional_formatting.add(
            f"{col}{res_first}:{col}{res_last}", FormulaRule(
                formula=[f'IF({lim}="le",AND({sx}<>"",{cur}>{sx}),'
                         f'IF({lim}="ge",AND({sn}<>"",{cur}<{sn}),'
                         f'IF({lim}="range",OR(AND({sn}<>"",{cur}<{sn}),'
                         f'AND({sx}<>"",{cur}>{sx})),FALSE)))'],
                font=st["Font"](bold=True, color=COLOR_FLAG),
                fill=PatternFill("solid", fgColor=FILL_FAIL, bgColor=FILL_FAIL),
                stopIfTrue=False))
    jr = f"{L(C['judge'])}{res_first}:{L(C['judge'])}{res_last}"
    ws.conditional_formatting.add(jr, FormulaRule(
        formula=[f'{L(C["judge"])}{res_first}="FAIL"'],
        font=st["Font"](bold=True, color=COLOR_FLAG),
        fill=PatternFill("solid", fgColor=FILL_FAIL, bgColor=FILL_FAIL), stopIfTrue=False))
    ws.conditional_formatting.add(jr, FormulaRule(
        formula=[f'{L(C["judge"])}{res_first}="PASS"'],
        font=st["Font"](bold=True, color=COLOR_PASS),
        fill=PatternFill("solid", fgColor=FILL_PASS, bgColor=FILL_PASS), stopIfTrue=False))
    ws.freeze_panes = f"D{4}"

    # 汇总页到此为止：给 review 看的东西不掺任何使用说明/告警，
    # 那些进独立的隐藏审计页（见 write_audit）。
    # 这一页就是这份簿的结论表——条件行说明在什么条件下取的数，
    # 结果行一行一个指标、填了 Spec 与 limit 就自动判定，不需要另起一页复述。
    return ws


# ---------------------------------------------------------------- 温度明细

def write_detail(wb, legs, items, st, src_title, room_t):
    """每个指标一块：行=温度（所有段的温度并集，升序），列=段。图表和汇总都吃这个。

    格子里放的是**指向原始页的引用**，不是抄过来的数。
    右边两列「各段最小/最大」是同一温度下各段之间的散布（回滞有多大），
    汇总页的全温极值就是对这两列取 MIN/MAX——这样汇总也不写死。
    """
    from openpyxl.utils import get_column_letter as L
    ws = wb.create_sheet("温度明细")
    ws.column_dimensions["A"].width = 12
    for i in range(len(legs) + 2):
        ws.column_dimensions[L(2 + i)].width = 13

    temps = sorted({t for lg in legs for t in lg.temps})
    c_lo, c_hi = 2 + len(legs), 3 + len(legs)          # 两列辅助
    blocks = []
    r = 1
    for it in items:
        put(ws, r, 1, f"{it.label}  [{it.unit}]", st, st["f_group"], bold=True, align="left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=c_hi)
        r += 1
        put(ws, r, 1, "温度℃", st, st["f_head"], bold=True)
        for i, lg in enumerate(legs):
            put(ws, r, 2 + i, f"{lg.title} {lg.direction}", st, st["f_head"], bold=True, size=9)
        put(ws, r, c_lo, "各段最小", st, st["f_head"], bold=True, size=9)
        put(ws, r, c_hi, "各段最大", st, st["f_head"], bold=True, size=9)
        head = r
        r += 1
        first = r
        # 温度 -> 该段里取到这个值的那一行在原始页的行号
        srcs = []
        for lg in legs:
            m = OrderedDict()
            for row in lg.rows:
                if row.kind == "lock":
                    continue
                if row.temp is not None and row.vals.get(it.col) is not None:
                    m[row.temp] = row.xl
            srcs.append(m)
        rows_by_temp = {}
        series = [leg_series(lg, it) for lg in legs]
        for t in temps:
            put(ws, r, 1, fmt_num(t), st, st["f_res"])
            rows_by_temp[t] = r
            row_vals = []
            for i, _lg in enumerate(legs):
                xl = srcs[i].get(t)
                put(ws, r, 2 + i,
                    sref(src_title, it.col, xl, it.text_src) if xl else None,
                    st, st["f_res"])
                if xl:
                    cache(ws, r, 2 + i, series[i].get(t))
                    row_vals.append(series[i].get(t))
            rng = f"{L(2)}{r}:{L(1 + len(legs))}{r}"
            put(ws, r, c_lo, f"=IF(COUNT({rng})=0,\"\",MIN({rng}))", st, st["f_res"], size=9)
            put(ws, r, c_hi, f"=IF(COUNT({rng})=0,\"\",MAX({rng}))", st, st["f_res"], size=9)
            if row_vals:
                cache(ws, r, c_lo, min(row_vals))
                cache(ws, r, c_hi, max(row_vals))
            r += 1
        blocks.append({"item": it, "head": head, "first": first, "last": r - 1,
                       "c_lo": c_lo, "c_hi": c_hi, "rows_by_temp": rows_by_temp})
        r += 1
    ws.freeze_panes = "B1"
    return ws, blocks, temps


def write_audit(wb, meta, st):
    """审计页：排除了哪些行、有哪些告警。默认隐藏。

    这份簿是要拿去 review 的，正表里不能掺"怎么用""⚠ 告警"这类给写脚本的人
    看的东西——老板翻到会问这是什么。但也不能真丢掉：哪些行没算进来必须留痕，
    所以单开一页并 hidden，需要时右键取消隐藏。控制台每次都会完整打印一份。
    """
    ws = wb.create_sheet("_审计")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 100
    put(ws, 1, 1, "本页仅供数据核对，不属于报告正文", st, st["f_group"],
        bold=True, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    r = 3
    put(ws, r, 1, "原表行号", st, st["f_head"], bold=True)
    put(ws, r, 2, "未计入汇总的原因", st, st["f_head"], bold=True)
    for xl, why in meta["excluded"]:
        r += 1
        put(ws, r, 1, xl, st)
        put(ws, r, 2, why, st, align="left")
    if meta["warnings"]:
        r += 2
        put(ws, r, 1, "告警", st, st["f_head"], bold=True)
        put(ws, r, 2, "说明", st, st["f_head"], bold=True)
        for w in meta["warnings"]:
            r += 1
            put(ws, r, 1, "⚠", st)
            put(ws, r, 2, w, st, align="left")
    ws.sheet_state = "hidden"
    return ws


# ---------------------------------------------------------------- 温巡过程

def write_journey(wb, legs, items, st, yranges, src_title):
    """按实际测试顺序把整趟温巡摊平成一张表，作为过程图的数据源。

    汇总页和温度明细都是「按温度」看的，那一维把先后顺序压没了。
    可这类测试要回答的恰恰是有先后的问题：锁完一次之后一路漂到哪、
    下一次重锁又拉回多少、频率在整趟里有没有变。所以单开这一页。
    """
    def pick(*names):
        for n in names:
            for it in items:
                if it.label == n:
                    return it
        return None

    vt = pick("Vtune_V") or next((i for i in items if "vtune" in i.label.lower()), None)
    fq = pick("Freq_MHz") or next((i for i in items if i.cat == "Frequency"), None)

    ws = wb.create_sheet("温巡过程")
    # 重锁点那条线的表头就是图例文字，顺手把锁在哪几个温度写进去——
    # 等于把标注做进图例里，看图的人不用再去翻数据页
    lt = "/".join(str(fmt_num(lg.lock_temp)) for lg in legs if lg.lock_temp is not None)
    lock_label = f"重锁点（锁@{lt}℃）" if lt else "重锁点"
    head = ["序号", "原表行", "段", "温度℃", "事件"]
    cols = {}
    if vt:
        cols["vt"] = len(head) + 1
        head += [vt.label, lock_label]
    if fq:
        cols["fq"] = len(head) + 1
        head += ["Δf (kHz)", lock_label, fq.label]
    for i, h in enumerate(head, 1):
        put(ws, 2, i, h, st, st["f_head"], bold=True)
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = \
            14 if i > 2 else 8

    seq = [(lg, r) for lg in legs for r in lg.rows]
    f0, f0_row = None, None
    if fq:
        for _lg, r in seq:
            v = r.vals.get(fq.col)
            if v is not None:
                f0, f0_row = v, r.xl
                break

    r0 = 3
    dfs = []
    for n, (lg, row) in enumerate(seq):
        r = r0 + n
        is_lock = row is lg.rows[0] and lg.lock_temp is not None
        fill = st["f_group"] if is_lock else st["f_res"]
        put(ws, r, 1, n + 1, st, fill)
        put(ws, r, 2, row.xl, st, fill, size=9)
        put(ws, r, 3, lg.n, st, fill)
        put(ws, r, 4, fmt_num(row.temp), st, fill)
        put(ws, r, 5, f"锁@{fmt_num(lg.lock_temp)}℃" if is_lock else None, st, fill,
            bold=is_lock, color=COLOR_FLAG if is_lock else None)
        if vt:
            v = row.vals.get(vt.col)
            ref = sref(src_title, vt.col, row.xl, vt.text_src) if v is not None else None
            put(ws, r, cols["vt"], ref, st, fill)
            put(ws, r, cols["vt"] + 1, ref if (is_lock and ref) else None, st, fill)
            cache(ws, r, cols["vt"], v)
            if is_lock:
                cache(ws, r, cols["vt"] + 1, v)
        if fq:
            v = row.vals.get(fq.col)
            df = None if (v is None or f0 is None) else round((v - f0) * 1000.0, 3)
            if df is not None:
                dfs.append(abs(df))
            # Δf 也做成公式：(本行频率 - 首点频率) × 1000
            if v is not None and f0_row:
                # 每项各自加括号：文本列会带双负号，不括起来就拼成 A---B 这种
                # 一串连续减号，能不能按预期解析全看 Excel 心情
                a = sref(src_title, fq.col, row.xl, fq.text_src)[1:]   # 去掉开头的 '='
                b = sref(src_title, fq.col, f0_row, fq.text_src)[1:]
                dformula = f"=(({a})-({b}))*1000"
            else:
                dformula = None
            put(ws, r, cols["fq"], dformula, st, fill)
            put(ws, r, cols["fq"] + 1,
                dformula if (is_lock and dformula) else None, st, fill)
            put(ws, r, cols["fq"] + 2,
                sref(src_title, fq.col, row.xl, fq.text_src) if v is not None else None,
                st, fill, size=9)
            cache(ws, r, cols["fq"], df)
            cache(ws, r, cols["fq"] + 2, v)
            if is_lock:
                cache(ws, r, cols["fq"] + 1, df)

    note = ("整趟温巡按测试顺序排；米色行 = 重锁点。"
            + (f"  Δf = 相对第 1 个测点（{fmt_num(f0, 6)}）的偏差，"
               f"全程 |Δf| ≤ {max(dfs) if dfs else 0} kHz" if fq and dfs else ""))
    if vt:
        # 自动算两句结论：重锁复位得一致不一致、两次重锁之间最多漂多少。
        # 这两个数就是这类温巡测试要的答案，不该让人自己去表里减。
        resets = [lg.rows[0].vals.get(vt.col) for lg in legs
                  if lg.rows and lg.lock_temp is not None]
        resets = [v for v in resets if v is not None]
        drifts = [(s["delta"], lg.n) for lg in legs for s in [stats(lg, vt)] if s]
        if resets:
            note += (f"\n重锁后 {vt.label} = "
                     + ", ".join(str(fmt_num(v, 5)) for v in resets)
                     + f"（极差 {fmt_num(max(resets) - min(resets), 5)} {vt.unit}"
                     + "，越小说明重锁复位越一致）")
        if drifts:
            d, ln = max(drifts)
            note += f"；两次重锁之间最大漂移 {fmt_num(d, 5)} {vt.unit}（段{ln}）"
    c = put(ws, 1, 1, note, st, st["f_group"], bold=True, align="left")
    c.alignment = c.alignment.copy(wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(head))
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A3"

    charts = []
    lblskip = max(1, len(seq) // 18)          # 横轴最多摆 ~18 个刻度标签，多了糊成一片
    if vt:
        b = yranges.get(vt.label) or axis_bounds([r.vals.get(vt.col) for _l, r in seq])
        charts.append({"title": f"{vt.label} 温巡过程（含重锁点）",
                       "ytitle": f"{vt.label} ({vt.unit})", "col": cols["vt"],
                       "bounds": b, "lblskip": lblskip})
    if fq and dfs:
        # ★这一张的纵轴故意不贴着数据算。记录精度只有 0.001 MHz，Δf 只能取
        # 0/±1 kHz 这几档；贴着数据画就把这点量化噪声放大成满屏方波，
        # 评审第一句话必然是"频率为什么在跳"。给一个比噪声宽得多的窗，
        # 它才如实显示成一条贴零的平线＝频率没变。窗宽仍由数据算。
        lim = max(3.0 * max(dfs), 5.0)
        _s = nice_step(2 * lim)
        lim = -(-lim // _s) * _s                  # 向上取到整刻度，别出现 ±7.269 这种轴
        b = yranges.get("Δf") or (-lim, lim, _s)
        charts.append({"title": f"输出频率漂移 温巡过程（相对首点；"
                                f"全程 |Δf| ≤ {max(dfs)} kHz，记录精度 1 kHz）",
                       "ytitle": "Δf (kHz)", "col": cols["fq"],
                       "bounds": b, "lblskip": lblskip})
    # 汇总页的「相对锁定点偏离」那几行要按段引用这里的格子，把坐标交出去
    lg_rows, i = [], 0
    for lg in legs:
        n = len(lg.rows)
        if n and lg.lock_temp is not None:
            # vt_vals[0] 就是锁定点（本段第一行＝重锁行），后面是段内各点
            vv = [r.vals.get(vt.col) for r in lg.rows] if vt else []
            lg_rows.append({"title": lg.title, "stage": lg.stage,
                            "lock_temp": lg.lock_temp, "first": r0 + i,
                            "last": r0 + i + n - 1, "lock_row": r0 + i,
                            "vt_vals": [v for v in vv if v is not None]})
        i += n
    jinfo = {"first": r0, "last": r0 + len(seq) - 1, "c_temp": 4,
             "sheet": ws.title, "legs": lg_rows, "c_vt": cols.get("vt"),
             "vt_item": vt, "charts": charts} if charts else None
    return ws, jinfo


# ---------------------------------------------------------------- 图表

# 默认只给这几个指标画 值-vs-温度 图。全画会出十几张，评审翻不动也问不出重点；
# 其余指标的极值/Δ 汇总表里都有，要补画用 --chart-items。
DEFAULT_CHART_ITEMS = ["Vtune_V", "IPN_SSB", "Current_mA", "Power_dBm"]


def write_charts(wb, ws_detail, blocks, legs, items, pn_items,
                 ws_j, jinfo, chart_items, yranges, st):
    from openpyxl.chart import Reference, ScatterChart, Series

    ws = wb.create_sheet("图表")
    for i, line in enumerate([
        "上面两张 = 温巡过程：横轴按实际测试先后排，刻度标的是当时的温度。"
        "看整趟温巡里量怎么走、每次重锁把它拉回到哪。红三角 = 重锁点。",
        "下面几张 = 指标 vs 温度（横轴 = 温度，每段一条线）：同一个温度上几条线分叉 = 回滞。"
        "相噪 vs offset 那张在「相噪-offset」页，图和数据放一起。",
    ]):
        put(ws, 1 + i, 1, line, st, st["f_group"] if i == 0 else None, bold=(i == 0),
            align="left")
        ws.merge_cells(start_row=1 + i, start_column=1, end_row=1 + i, end_column=14)

    row = 4
    if jinfo:
        row = _journey_charts(ws, ws_j, jinfo, st, row)

    # 值 vs 温度：只画点名的几个
    wanted = [b for b in blocks if b["item"].label in chart_items]
    for n, blk in enumerate(wanted):
        it, head, first, last = blk["item"], blk["head"], blk["first"], blk["last"]
        ch = ScatterChart()
        ch.title = f"{it.label} vs 温度"
        ch.style = 13
        ch.x_axis.title = "温度 (℃)"
        ch.y_axis.title = it.unit                 # 量名已在标题里，轴上只留单位，省地方
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        ch.height, ch.width = 9, 15
        blank_policy(ch)
        xref = Reference(ws_detail, min_col=1, min_row=first, max_row=last)
        vals = []
        for i, _lg in enumerate(legs):
            yref = Reference(ws_detail, min_col=2 + i, min_row=head, max_row=last)
            s = Series(yref, xref, title_from_data=True)
            color, sym = LEG_STYLE[i % len(LEG_STYLE)]
            _style(s, color, sym)
            ch.series.append(s)
        # 明细页现在放的是公式，openpyxl 读不到算完的值，纵轴范围直接用内存里的数
        vals = [v for lg in legs for v in leg_series(lg, it).values()]
        _apply_y(ch, yranges.get(it.label) or axis_bounds(vals))
        _legend_bottom(ch)
        ws.add_chart(ch, f"{'A' if n % 2 == 0 else 'K'}{row + (n // 2) * 19}")
    row += ((len(wanted) + 1) // 2) * 19

    if pn_items:
        _pn_offset_chart(wb, legs, pn_items, st, yranges)
    return ws


def _journey_charts(ws, ws_j, jinfo, st, row):
    """温巡过程图：横轴=温度（按测试顺序从左到右排），重锁点单独一条只有记号的线。

    为什么不按"vs 温度"画：温巡是有先后的，同一个温度会经过好几次；
    按温度画就把"先后"这一维压没了，看不出"锁完一次之后一路漂到哪、
    下一次重锁又拉回多少"。
    横轴直接用温度当刻度标签＝既保留先后顺序，又能一眼看出走到哪个温度了；
    温度本身不再单独画一条曲线（那条线不带信息，还要占一根次坐标轴、
    把图例和轴标题挤到重叠）。
    """
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.axis import ChartLines

    first, last = jinfo["first"], jinfo["last"]
    for spec in jinfo["charts"]:
        c = LineChart()
        c.title = spec["title"]
        c.style = 13
        c.height, c.width = 9.5, 30
        blank_policy(c)
        c.y_axis.title = spec["ytitle"]
        c.x_axis.title = "温度 (℃)　—　从左到右＝实际测试先后"
        c.x_axis.delete = False
        c.y_axis.delete = False
        c.y_axis.majorGridlines = ChartLines()
        c.add_data(Reference(ws_j, min_col=spec["col"], min_row=first - 1, max_row=last),
                   titles_from_data=True)
        c.add_data(Reference(ws_j, min_col=spec["col"] + 1, min_row=first - 1, max_row=last),
                   titles_from_data=True)
        c.set_categories(Reference(ws_j, min_col=jinfo["c_temp"],
                                   min_row=first, max_row=last))
        _style(c.series[0], "1F77B4", "circle", size=5)
        _style(c.series[1], "D62728", "triangle", line=False, size=11)
        _apply_y(c, spec.get("bounds"))
        _legend_bottom(c)
        c.x_axis.tickLblSkip = spec.get("lblskip", 3)
        c.x_axis.tickMarkSkip = spec.get("lblskip", 3)
        ws.add_chart(c, f"A{row}")
        row += 21
    return row


def _pn_offset_chart(wb, legs, pn_items, st, yranges):
    """相噪-vs-offset：数据和图放在同一页，别让人看见一张孤零零的表不知道干嘛的。

    取哪几条温度：最低 / 最接近常温 25℃ / 最高。之前取"排序后的中位数"，
    结果常温那条挑到了 35℃——中位数是统计意义上的中间，不是工程上的常温。
    """
    from openpyxl.chart import Reference, ScatterChart, Series

    lg = max(legs, key=lambda x: len(x.rows))            # 取点最多的那一段
    temps = sorted({r.temp for r in lg.rows if r.temp is not None})
    if not temps:
        return
    room = min(temps, key=lambda t: abs(t - 25))
    pick = sorted({temps[0], room, temps[-1]})

    ws = wb.create_sheet("相噪-offset")
    put(ws, 1, 1, f"相噪 vs offset —— 取「{lg.title} {lg.stage}」这一段，"
                  f"最低 / 最接近常温 / 最高 三个温度各一条线。右边就是图。",
        st, st["f_group"], bold=True, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.column_dimensions["A"].width = 13
    put(ws, 2, 1, "offset_MHz", st, st["f_head"], bold=True)
    for j, t in enumerate(pick):
        put(ws, 2, 2 + j, f"{fmt_num(t)}℃", st, st["f_head"], bold=True)
        ws.column_dimensions[ws.cell(row=2, column=2 + j).column_letter].width = 12
    vals = []
    for i, it in enumerate(pn_items):
        m = re.search(r"@([\d.]+)(k|M)Hz", it.label)
        off = float(m.group(1)) / (1000.0 if m and m.group(2) == "k" else 1.0) if m else None
        put(ws, 3 + i, 1, off, st, st["f_res"])
        for j, t in enumerate(pick):
            v = None
            for row in lg.rows:
                if row.temp == t and row.vals.get(it.col) is not None:
                    v = row.vals[it.col]
            put(ws, 3 + i, 2 + j, fmt_num(v), st, st["f_res"])
            if v is not None:
                vals.append(v)

    ch = ScatterChart()
    ch.title = "相噪 vs offset"
    ch.style = 13
    ch.x_axis.title = "offset (MHz，对数)"
    ch.y_axis.title = "dBc/Hz"
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.x_axis.scaling.logBase = 10
    ch.height, ch.width = 10, 17
    blank_policy(ch)
    xref = Reference(ws, min_col=1, min_row=3, max_row=2 + len(pn_items))
    for j in range(len(pick)):
        yref = Reference(ws, min_col=2 + j, min_row=2, max_row=2 + len(pn_items))
        s = Series(yref, xref, title_from_data=True)
        color, sym = LEG_STYLE[j % len(LEG_STYLE)]
        _style(s, color, sym)
        ch.series.append(s)
    _apply_y(ch, yranges.get("PN") or axis_bounds(vals))
    _legend_bottom(ch)
    ws.add_chart(ch, "F2")


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(description="性能扫描簿 → 带汇总页的 Excel")
    ap.add_argument("path", help="扫描簿 .xlsx")
    ap.add_argument("-o", "--out", default=None, help="输出路径（默认 <原名>_summary.xlsx）")
    ap.add_argument("--sheet", default=None, help="数据所在 sheet（默认第一个）")
    ap.add_argument("--header-row", type=int, default=1, help="表头行号（1 基，默认 1）")
    ap.add_argument("--leg-col", default="Mode", help="判断重锁用的列（默认 Mode）")
    ap.add_argument("--lock-pattern", default=r"_lock$",
                    help="该列匹配这个正则的行 = 一次重锁（默认 _lock$）")
    ap.add_argument("--temp-col", default=None, help="温度列（默认自动找含 Temperature 的列）")
    ap.add_argument("--keep-test-item", default=None,
                    help="只保留 Test Item 等于该值的行（默认取出现最多的那个值）")
    ap.add_argument("--keep-mode", default=None,
                    help="只保留该模式的行（默认取出现最多的那个值）；重锁行始终保留。"
                         "挡的是夹在扫描序列里的旁路/自检行——它们也带部分结果值，"
                         "不挡就会被算进相邻那一段，把段的走向和极值都带偏")
    ap.add_argument("--chart-items", default=",".join(DEFAULT_CHART_ITEMS),
                    help="哪些指标画 值-vs-温度 图（逗号分隔；all=全画，空=一张都不画）。"
                         "温巡过程图和相噪-offset 图不受这个控制，始终画。"
                         f"默认 {','.join(DEFAULT_CHART_ITEMS)}")
    ap.add_argument("--y-range", default="",
                    help="手工钉死某张图的纵轴，如 \"Vtune_V=0.1:0.8,IPN_SSB=-60:-40\"；"
                         "键还可用 Δf（频率漂移图）/ PN（相噪-offset 图）。"
                         "不给就按数据自动算范围")
    ap.add_argument("--no-audit", action="store_true",
                    help="连隐藏的「_审计」页都不要（排除行/告警只留在控制台输出）")
    ap.add_argument("--dry-run", action="store_true", help="只打印识别结果，不写文件")
    args = ap.parse_args()

    if not os.path.isfile(args.path):
        sys.exit(f"找不到文件: {args.path}")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        sys.exit("缺少 openpyxl，请先: pip install openpyxl")

    # 读取/过滤/分段/识别指标全在 sweep_lib.load_sweep 里（跨芯片汇总用同一份）
    try:
        sw = load_sweep(args.path, sheet=args.sheet, header_row=args.header_row,
                        leg_col=args.leg_col, lock_pattern=args.lock_pattern,
                        temp_col=args.temp_col, keep_test_item=args.keep_test_item,
                        keep_mode=args.keep_mode, keep_original=True)
    except SweepError as e:
        sys.exit(str(e))

    wb, ws, cols = sw.wb, sw.ws, sw.cols
    rows, items, legs, dropped = sw.rows, sw.items, sw.legs, sw.dropped
    excluded, warnings, room_t = sw.excluded, sw.warnings, sw.room_t
    tname, keep_ti, keep_mode, ti_col = sw.temp_name, sw.keep_ti, sw.keep_mode, sw.ti_col
    meta = {
        "excluded": excluded,
        "warnings": warnings,
        "conditions": build_conditions(legs, cols, rows, items, room_t),
    }

    # ---- 打印识别结果 ----
    print(f"源文件 : {os.path.basename(args.path)}   sheet={ws.title}  "
          f"{ws.max_row} 行 × {ws.max_column} 列")
    print(f"温度列 : {tname}    分段列: {args.leg_col}  规则: /{args.lock_pattern}/")
    if ti_col is not None:
        print(f"主测试项: {keep_ti!r}（其余行排除）")
    if keep_mode is not None:
        print(f"主模式  : {keep_mode!r}（+ 重锁行；其余行排除）")
    print(f"识别指标 {len(items)} 个:")
    for it in items:
        print(f"    [{it.cat:<12}] {it.label:<18} {it.unit:<7} <- 列 {it.src}"
              + ("   ⚠ 原始列是文本型数字，引用已加双负号转数值" if it.text_src else ""))
    txts = [it.src for it in items if it.text_src]
    if txts:
        warnings.append(f"原始表里这些结果列存成了文本型数字（Excel 的 COUNT/MIN 不认，"
                        f"不处理会让极值温度出 #N/A）：{', '.join(txts)}；"
                        f"引用已加双负号转成数值")
    if dropped:
        print(f"  跳过的空列 {len(dropped)} 个: " + ", ".join(k for k, _ in dropped))
    print(f"分段 {len(legs)} 段:")
    for lg in legs:
        print(f"    {lg.title:<14} {lg.stage:<16} {len(lg.rows):>3} 点  "
              f"行 {lg.rows[0].xl}~{lg.rows[-1].xl}")
    if excluded:
        print(f"排除 {len(excluded)} 行:")
        for xl, why in excluded:
            print(f"    行{xl}: {why}")
    for w in warnings:
        print(f"  ⚠ {w}")
    if args.dry_run:
        print("\n--dry-run：没有写文件。")
        return

    # ---- 写出 ----
    # --y-range "Vtune_V=0.1:0.8,IPN_SSB=-60:-40"；键还可以是 Δf / PN
    yranges = {}
    for part in args.y_range.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            k, rng = part.split("=", 1)
            lo, hi = (float(x) for x in rng.split(":"))
            yranges[k.strip()] = (lo, hi, nice_step(hi - lo))
        except ValueError:
            sys.exit(f"--y-range 格式应为 指标名=下限:上限，逗号分隔；解析不了: {part!r}")

    st = _styles()
    # 明细页要先建：汇总页的实测列是对它取 MIN/MAX/MEDIAN 的公式，不是写死的数
    ws_detail, blocks, _temps = write_detail(wb, legs, items, st, ws.title, room_t)
    dref = {b["item"].label: {"sheet": f"'{ws_detail.title}'", "first": b["first"],
                              "last": b["last"], "c_lo": b["c_lo"], "c_hi": b["c_hi"],
                              "rows_by_temp": b["rows_by_temp"], "n_legs": len(legs)}
            for b in blocks}
    # 温巡过程要先建：汇总页的「相对锁定点偏离」几行按段引用它的格子
    ws_j, jinfo = write_journey(wb, legs, items, st, yranges, ws.title)
    write_summary(wb, legs, items, meta, st, room_t, dref, jinfo)
    pn_items = [it for it in items if it.label.startswith("SpotPN@")]
    if args.chart_items.strip().lower() == "all":
        chart_items = [it.label for it in items]
    elif args.chart_items.strip():
        chart_items = [s.strip() for s in args.chart_items.split(",") if s.strip()]
    else:
        chart_items = []
    write_charts(wb, ws_detail, blocks, legs, items, pn_items,
                 ws_j, jinfo, chart_items, yranges, st)
    if not args.no_audit:
        write_audit(wb, meta, st)

    # 按阅读顺序排页：汇总（＝这份簿的结论表）在最前，支撑数据靠后
    order = ["汇总", ws.title, "温巡过程", "图表", "相噪-offset",
             "温度明细", "_审计"]
    wb._sheets.sort(key=lambda s: order.index(s.title)
                    if s.title in order else len(order))

    out = args.out or os.path.splitext(args.path)[0] + "_summary.xlsx"
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    n_fill, n_strip = VCACHE.inject(out)
    print(f"\n已写出: {os.path.abspath(out)}")
    print(f"  第 1 页「{ws.title}」= 原始数据原样保留；新增 "
          + " / ".join(n for n in wb.sheetnames if n != ws.title))
    print("  汇总页的 Spec MIN/TYP/MAX 与 limit 列留空，填进去判定自动出、超规自动标红。")
    print(f"  公式格补了 {n_fill} 个缓存值、清掉 {n_strip} 个空缓存 —— "
          f"发给别人打开就能看见数，不用敲回车重算。")


if __name__ == "__main__":
    main()
