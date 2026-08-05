#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
current_db.py — 电流数据库（v4：全模式单文件多温度 + 仿真 tier/stage 过滤 + 可读汇总簿）

把两类数据统一进一个 SQLite 库，并导出对比/汇总 Excel：
  1) 仿真长表：某工作簿的 Current_data 页（ID/Module/Trim/Mode/simulation/Tier/Current/Unit；
     simulation=pre/post，Tier 为电流档位——config.sim_tier 选定与实测同档的数据参与对比）
     ★模块编号取自 **Module 列的数字前缀**（'12_Xxx' -> 12），不是 ID 列——ID 列是全表行
     流水号（同一模块跨 Mode/Tier/stage 逐行递增），拿它当模块编号会与实测 NO. 整体错位。
  2) 实测结果：Result*.xlsx。两种形态自动识别：
     a. 全模式单文件（可含多温度段）：按行分段——`*_sigN` 签名行 / Init 行 NO.=模式名 开段，
        SET_TEMP·Chamber 行闭段；同段温度取自 Temperature 列 → 每 (模式,温度) 一个 run
     b. 旧单模式文件：整表一段（Init 行 NO. 给模式名，文件夹名兜底）
     模式名与仿真表 Mode 自动对齐：大小写/下划线无关、UNSYNC≡NOSYNC、尾部裸 SYNC 可省
     （BT2GRX_unSync≡BT_2G_RX_noSYNC；BT2GRX_sync≡BT_2G_RX）；config.mode_map 可强制指定。

实测解析规则（在每个模式段内独立执行）：
  - 一个序列从 Init 行开始；基线 = 第一个 OFF 行之前最后一行的电流（通常是最后一个 Lock_step）
  - 模块电流 = 上一行电流 - 本行电流（逐级关断做差），统一换算成 uA
  - 第二个及以后的 Init 段 = 锁定复验，忽略（原始行仍入库审计；全模式文件按段分割后天然不会触发）
  - SET_TEMP / chamber 行只控温箱，不参与做差
  - NO. 列多个编号（如 "45,46"）= 该步同时关断的一组模块，按组对比（仿真侧求和）
  - LDO 归并（config.ldo_reparent，如 28->26）：子模块不在被测 LDO 下，
    其实测 delta 并入父模块组；对比时仿真侧同样求和（meas(26)+meas(28) vs sim(26)+sim(28)）
  - NO. 列非数字标签（如 "DCO5G"）：config.label_groups 映射到仿真模块 ID；
    值可以是 ID 列表（同模式），也可以是 {"mode": "CK_ADPLL_DCO2G", "ids": "*"}
    （跨模式取该仿真 Mode 的全部/指定模块合计）

用法（在数据所在机器上）：
  python current_db.py build   --root D:\\Excel --chip C1
    首次运行会在 root 下生成 current_config.json（sim_tier/模式映射/LDO 归并等都在里面改），
    并输出 current.db + Current_compare_pivot.xlsx
  python current_db.py summary --db current.db --out 各模式功耗表.xlsx
    人直接读的汇总簿：总览矩阵(模块×模式×温度+仿真对比) / 温度趋势图 / 对比明细

  也可分步：
  python current_db.py ingest-sim --db current.db --xlsx Current_all_mode_v2.xlsx
  python current_db.py ingest-run --db current.db --xlsx <Result文件> --mode BT_5G_TX --chip C1
  python current_db.py export     --db current.db --out out.xlsx [--all-runs]

依赖：openpyxl（与本仓库其余工具一致，无其他第三方依赖）
"""
import argparse
import datetime
import fnmatch
import json
import os
import re
import sqlite3
import sys
import types

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.text import RichText, Text
from openpyxl.chart.title import Title
from openpyxl.drawing.text import (CharacterProperties, Font as DrawFont, Paragraph,
                                   ParagraphProperties, RegularTextRun)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- 常量/工具

UNIT_TO_UA = {"ua": 1.0, "µa": 1.0, "μa": 1.0, "ma": 1000.0, "na": 0.001, "a": 1e6, "": 1.0}

DEFAULT_CONFIG = {
    "_说明": {
        "sim_workbook": "仿真长表所在工作簿（相对 root 或绝对路径）；null=自动找 Current_all_mode*.xlsx",
        "sim_sheet": "仿真长表的 tab 名",
        "result_glob": "匹配实测文件的通配符；**可以给一个数组**（命名换过词序就再加一条，"
                       "不用改代码）。没匹配上的 xlsx 会在 build/inspect 里报出来，不静默跳过",
        "result_sheet": "实测数据所在 tab；null=自动扫描含 NO./Current 表头的第一个 tab",
        "skip_dirs": "扫描 root 子目录时跳过的文件夹",
        "mode_map": "文件夹名 -> 仿真表 Mode 名 的映射（同名可省略）",
        "ldo_reparent": "子模块ID -> 父模块ID：子模块不在被测 LDO 下，其实测 delta 并入父模块组",
        "ldo_reparent_sim_add_child": "归并时仿真侧是否把子模块电流也加进对比和（子模块电流不在被测轨上时应为 false）",
        "label_groups": "NO. 列非数字标签 -> 仿真模块ID列表如 {\"DCO5G\": [21]}，"
                        "或跨模式 {\"DCO2G\": {\"mode\": \"CK_ADPLL_DCO2G\", \"ids\": \"*\"}}",
        "exclude_globs": "扫描时按文件名跳过的通配符（本工具自己的输出必须在内，防自吞）",
        "sim_label_ids": "仿真表里 Module 无数字前缀的合计/标签行 名->指派编号；留空则从 901 起按名排序自动指派（实测侧用 label_groups 把非数字标签指到该编号）",
        "sim_stage": "主对比列取哪个仿真阶段：post(后仿，默认) 或 pre(前仿)；若某阶段整片为 0，inspect 的零值审计会指出来",
        "sim_stage_fallback": "主阶段(sim_stage)某模块为 0/缺失时，是否改用另一阶段的值补上(逐模块判断，两边都是 0 则保持 0)；补过的格子在报告里标出",
        "sim_zero_ua": "判「这一格等于没有」的阈值 µA（默认 1）：后仿漏接的模块未必是干净的 0，"
                       "见过 0.006µA 这种数值残渣——真实模块电流都在几十~几百 µA，"
                       "≤阈值一律当缺项处理（触发跨阶段补值 + inspect 零值审计）",
        "sim_tier": "参与对比的仿真电流档位（如 Tier2，与实测一致）；空=不过滤（多档共存会重复求和！）",
        "sim_temp_note": "仿真数据的温度/corner 标注，只用于表头展示（如 55C/TT/0.9V）",
        "delta_flag_pct": "汇总簿标红：|偏差%| 超过该值才可能标红（默认 20）",
        "delta_flag_abs_ua": "汇总簿标红的绝对偏差下限 µA（默认 40）：|偏差%| 与 |ΔµA| 双双"
                             "超阈才标红，避免小电流模块被百分比放大成假红",
        "sim_temp_c": "仿真数据的温度（℃，默认从 sim_temp_note 解析数字）；偏差列把实测线性"
                      "插值到该温度再对仿真，消除单点仿真 vs 多温实测的系统性温差",
        "mode_freq": "汇总簿测试频率条件行：模式名 -> 显示文本；不在表里的按名字推断"
                     "（含 2G -> 2.5GHz，含 5G -> 5.8GHz）",
    },
    "sim_workbook": None,
    "sim_sheet": "Current_data",
    "result_glob": "Result*.xlsx",
    "result_sheet": None,
    "skip_dirs": ["Simulation", "自动化"],
    "mode_map": {},
    "ldo_reparent": {"8": "6", "28": "26"},
    "ldo_reparent_sim_add_child": False,
    "label_groups": {},
    "exclude_globs": ["Current_compare_pivot*.xlsx", "probe_dump*", "*功耗表*.xlsx"],
    "sim_label_ids": {},
    "sim_stage": "post",
    "sim_stage_fallback": True,
    "sim_zero_ua": 1.0,
    "sim_tier": "Tier2",
    "sim_temp_note": "55C",
    "delta_flag_pct": 20,
    "delta_flag_abs_ua": 40,
    "delta_ref_temp": None,
    "sim_temp_c": None,
    "mode_freq": {},
}


def norm(v):
    return str(v).strip().lower() if v is not None else ""


def cell(row, idx):
    """read_only 模式下行元组可能比表头短，安全取值。"""
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def as_float(v):
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_ids(no_raw):
    """NO. 单元格 -> 模块ID列表；解析失败返回 None（说明是标签）。"""
    if no_raw is None:
        return None
    if isinstance(no_raw, (int, float)) and not isinstance(no_raw, bool):
        return [int(no_raw)] if float(no_raw).is_integer() else None
    parts = re.split(r"[,，、;；\s]+", str(no_raw).strip())
    ids = []
    for p in parts:
        if not p:
            continue
        if not re.fullmatch(r"\d+", p):
            return None
        ids.append(int(p))
    return ids or None


def canon_mode(name):
    """模式名规范化键：大小写/下划线等分隔无关；UNSYNC≡NOSYNC；尾部裸 SYNC（默认态）可省。
    例：BT2GRX_unSync 与 BT_2G_RX_noSYNC 同键；BT2GRX_sync 与 BT_2G_RX 同键。"""
    s = re.sub(r"[^0-9A-Za-z]", "", str(name or "")).upper()
    s = s.replace("UNSYNC", "NOSYNC")
    return re.sub(r"(?<!NO)SYNC$", "", s)


def resolve_mode(label, sim_modes, mode_map, folder=None):
    """实测段标签 -> 仿真表 Mode 名。优先 config.mode_map（键可以是段标签或文件夹名），
    其次 canon 规范化唯一匹配。返回 (resolved, how)，how ∈ config/auto/ambig/none。"""
    mode_map = mode_map or {}
    for key in (label, folder):
        if key and key in mode_map:
            return mode_map[key], "config"
    c = canon_mode(label)
    hits = sorted({m for m in sim_modes or () if canon_mode(m) == c})
    if len(hits) == 1:
        return hits[0], "auto"
    return label, ("ambig" if hits else "none")


def now_iso():
    return datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


# ---------------------------------------------------------------- 配置

def load_config(root, path=None, create=True):
    """create=False：配置不存在时只返回默认值，不落盘（inspect 是只读体检，不该留痕）。"""
    cfg_path = path or os.path.join(root, "current_config.json")
    if os.path.exists(cfg_path):
        # utf-8-sig：这个文件是给人在 Windows 上手改的，PowerShell 5.1 的
        # `Set-Content -Encoding utf8`、记事本"UTF-8"另存都会写 BOM，严格 utf-8 会直接抛
        # "Unexpected UTF-8 BOM"。utf-8-sig 有 BOM 吃掉、没 BOM 也照常读。
        with open(cfg_path, "r", encoding="utf-8-sig") as f:
            cfg = json.load(f)
        merged = dict(DEFAULT_CONFIG)
        merged.update(cfg)
        return merged, cfg_path, False
    if not create:
        return dict(DEFAULT_CONFIG), cfg_path, False
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(DEFAULT_CONFIG, f, ensure_ascii=False, indent=2)
    return dict(DEFAULT_CONFIG), cfg_path, True


# ---------------------------------------------------------------- 数据库

SCHEMA = """
CREATE TABLE IF NOT EXISTS runs(
    run_id INTEGER PRIMARY KEY,
    mode TEXT, chip TEXT, temp_c REAL,
    src_file TEXT, run_ts TEXT, ingested_ts TEXT, mode_raw TEXT);
CREATE TABLE IF NOT EXISTS meas_raw(
    id INTEGER PRIMARY KEY,
    run_id INTEGER, row_idx INTEGER, seq_idx INTEGER, kind TEXT,
    no_raw TEXT, mode_label TEXT, current_ma REAL, delta_ma REAL,
    temp_c REAL, note TEXT);
CREATE TABLE IF NOT EXISTS meas_module(
    id INTEGER PRIMARY KEY,
    run_id INTEGER, step_order INTEGER,
    group_disp TEXT, step_name TEXT,
    module_ids TEXT, sim_ids TEXT,
    current_ua REAL, note TEXT, sim_mode TEXT);
CREATE TABLE IF NOT EXISTS sim_current(
    id INTEGER PRIMARY KEY,
    module_id INTEGER, module_name TEXT, trim TEXT, mode TEXT,
    stage TEXT, tier TEXT, current_ua REAL, unit_raw TEXT, src_file TEXT);
"""


def open_db(path):
    conn = sqlite3.connect(path)
    conn.executescript(SCHEMA)
    for alter in ("ALTER TABLE runs ADD COLUMN mode_raw TEXT",
                  "ALTER TABLE meas_module ADD COLUMN sim_mode TEXT"):
        try:
            conn.execute(alter)  # v4 前建的库补列
        except sqlite3.OperationalError:
            pass
    return conn


# ---------------------------------------------------------------- 仿真表导入

def match_sim_header(row):
    """仿真长表表头：ID + Module + Mode + Current*，且有 simulation/Unit 佐证。
    ★必须有 Module 列：同一工作簿里的相位噪声长表表头是
      ID/PN_Point/Mode/simulation/Tier/Current/Current_Unit/...，
      ID/Mode/Current/simulation 全都满足，只差 Module。不卡这一条就会静默把
      PN 表当电流表用——它的 Current 列是真电流、单位也对，出来的报告看不出破绽。"""
    names = [norm(c) for c in row]
    return ("id" in names and "mode" in names and "module" in names
            and any(n.startswith("current") for n in names)
            and (any(n.startswith("simulation") for n in names)
                 or "unit" in names))  # 注意: 不认 tier——本工具导出的 Sim_long 页有 tier 列


MODULE_PREFIX_RE = re.compile(r"^\s*(\d+)\s*[_\-.．、 ]")


def module_id_from_name(name):
    """模块编号取自 Module 列的数字前缀（'12_Xxx' -> 12）。

    ★为什么不用 ID 列：真实仿真长表的 ID 列是**全表行流水号**——同一个模块在不同 Mode /
    Tier / stage 下各占一行，ID 逐行递增（观测：ID=1..5 全是同一个模块的 5 个 Mode）。
    拿它当模块编号去对实测 NO. 编号，会整体错位，而且错得看起来正常（有值、非空、量级对）。
    模块身份只在 Module 列的数字前缀里。"""
    m = MODULE_PREFIX_RE.match(str(name or ""))
    return int(m.group(1)) if m else None


def find_sim_header(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        names = [norm(c) for c in row]
        if "id" in names and "mode" in names and any(n.startswith("current") for n in names):
            cols = {}
            for j, n in enumerate(names):
                if n == "id":
                    cols["id"] = j
                elif n == "module":
                    cols["module"] = j
                elif n == "trim":
                    cols["trim"] = j
                elif n == "mode":
                    cols["mode"] = j
                elif n.startswith("simulation") or n == "sim":
                    cols["sim"] = j
                elif n == "tier":
                    cols["tier"] = j
                elif n.startswith("current"):
                    cols["current"] = j
                elif n == "unit":
                    cols["unit"] = j
            return i, cols
    return None, None


def read_sim_rows(xlsx, sheet_name, label_ids=None):
    """读仿真长表 -> (records, info)，**只读不写**。
    ingest_sim 与 inspect 共用这一个解析器，保证「体检看到的」就是「入库的」。"""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws = None
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            for sn in wb.sheetnames:
                if norm(sn) == norm(sheet_name or "current_data"):
                    ws = wb[sn]
                    break
        if ws is None:  # 按名字没找到 -> 按表头内容扫
            for sn in wb.sheetnames:
                cand = wb[sn]
                if any(match_sim_header(r) for r in
                       cand.iter_rows(min_row=1, max_row=30, values_only=True)):
                    ws = cand
                    break
        if ws is None:
            raise SystemExit(f"[错误] {os.path.basename(xlsx)} 里找不到仿真 tab（按名 {sheet_name!r} "
                             f"或按表头 ID/Module/Mode/Current 都没命中），现有 tab: {wb.sheetnames}")
        hdr, cols = find_sim_header(ws)
        if hdr is None:
            raise SystemExit(f"[错误] 仿真 tab {ws.title!r} 找不到表头行（需含 ID/Mode/Current 列）")

        recs = []
        info = {"file": os.path.abspath(xlsx), "sheet": ws.title, "header_row": hdr,
                "cols": dict(cols), "n_data_rows": 0,
                "skip_no_current": 0, "skip_no_mode": 0,
                "id_from_prefix": 0, "id_from_idcol": 0, "id_missing": 0,
                "prefix_conflicts": [], "unknown_units": {}, "no_prefix_names": {},
                "modes": {}, "tiers": {}, "stages": {}, "units": {}}
        id_col_vals = []

        def bump(d, k):
            d[k] = d.get(k, 0) + 1

        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            cur = as_float(cell(row, cols.get("current")))
            mid_raw = cell(row, cols.get("id"))
            mode = cell(row, cols.get("mode"))
            if mode is None:
                info["skip_no_mode"] += 1
                continue
            if cur is None:
                info["skip_no_current"] += 1
                continue
            unit = norm(cell(row, cols.get("unit")))
            factor = UNIT_TO_UA.get(unit)
            if factor is None:  # 未知单位：按 uA 处理，但记下来让人看见（1000 倍事故的入口）
                bump(info["unknown_units"], unit)
                factor = 1.0
            stage_raw = norm(cell(row, cols.get("sim")))
            stage = "pre" if stage_raw.startswith("pre") else ("post" if stage_raw else "")
            name = str(cell(row, cols.get("module")) or "")
            tier = str(cell(row, cols.get("tier")) or "")

            try:
                id_col_i = int(mid_raw)
                id_col_vals.append(id_col_i)
            except (TypeError, ValueError):
                id_col_i = None
            mid_i = module_id_from_name(name)          # ★ 模块编号只认 Module 前缀
            if mid_i is None:
                bump(info["no_prefix_names"], name)    # 合计/标签行：靠 label_groups 按名对应
            elif id_col_i is not None and id_col_i != mid_i \
                    and len(info["prefix_conflicts"]) < 5:
                info["prefix_conflicts"].append((id_col_i, name, str(mode).strip()))

            recs.append(dict(module_id=mid_i, id_col=id_col_i, module_name=name,
                             trim=str(cell(row, cols.get("trim")) or ""),
                             mode=str(mode).strip(), stage=stage, tier=tier,
                             current_ua=cur * factor, unit_raw=unit))
            info["n_data_rows"] += 1
            bump(info["modes"], str(mode).strip())
            bump(info["tiers"], tier)
            bump(info["stages"], stage)
            bump(info["units"], unit)

        # ID 列是「全表行流水号」的判据：值全表互不重复且不止一行（真簿 4400 行 4400 个值）
        info["id_col_is_serial"] = (len(id_col_vals) > 1
                                    and len(set(id_col_vals)) == len(id_col_vals))
        # 无前缀行怎么办：ID 列是流水号时**绝不回退**——回退只会造出 521-550 这种
        # 凭空的"模块编号"（真簿实测：240 行合计/标签行会被编成 8 段假编号）。
        # 这些行的身份在名字里，由 config.label_groups 按名映射。
        for r in recs:
            if r["module_id"] is not None:
                info["id_from_prefix"] += 1
            elif not info["id_col_is_serial"] and r["id_col"] is not None:
                r["module_id"] = r["id_col"]
                info["id_from_idcol"] += 1
            else:
                info["id_missing"] += 1
            r.pop("id_col")
        info["module_ids"] = sorted({r["module_id"] for r in recs if r["module_id"] is not None})

        # 合计/标签行（无数字前缀，如 I_xxx_total）指派一个稳定编号：优先 config.sim_label_ids，
        # 其余从 901 起按名字排序自动指派（可复现、不与真实编号 1..N 冲突）。指派之后它们走
        # 与普通模块**完全相同**的对比通路，实测侧只需在 config.label_groups 里把非数字标签
        # （如 "DCO2G"）指到这个编号即可——不必为"按名查"另开一条支路。
        explicit = {}
        for k, v in (label_ids or {}).items():
            try:
                explicit[str(k)] = int(v)
            except (TypeError, ValueError):
                pass
        used, auto, nxt = set(info["module_ids"]) | set(explicit.values()), {}, 901
        for nm in sorted(info["no_prefix_names"]):
            if nm in explicit:
                continue
            while nxt in used:
                nxt += 1
            auto[nm] = nxt
            used.add(nxt)
        label_map = dict(explicit)
        label_map.update(auto)
        info["label_id_map"] = label_map
        info["label_id_auto"] = sorted(auto)
        if label_map:
            for r in recs:
                if r["module_id"] is None:
                    r["module_id"] = label_map.get(r["module_name"])
        return recs, info
    finally:
        wb.close()


def sim_warnings(info):
    """把 read_sim_rows 的统计翻译成人话警告（ingest 与 inspect 共用同一套判据）。"""
    w = []
    if info["id_col_is_serial"]:
        w.append(f"[提示] 仿真表 ID 列是全表行流水号（{info['n_data_rows']} 行值互不重复），"
                 f"不是模块编号——模块编号已改取自 Module 列数字前缀")
    if info["id_from_idcol"]:
        w.append(f"[提示] {info['id_from_idcol']} 行 Module 列没有数字前缀，退回用 ID 列当模块编号"
                 f"（ID 列不是流水号，可用）")
    if info["id_missing"]:
        n_names = len(info["no_prefix_names"])
        w.append(f"[提示] {info['id_missing']} 行没有模块编号（{n_names} 种 Module 名无数字前缀，"
                 f"多半是合计/标签行）——不回退 ID 列，靠 config.label_groups 按名映射")
    if info["unknown_units"]:
        u = " / ".join(f"{k or '(空)'}×{v}" for k, v in sorted(info["unknown_units"].items()))
        w.append(f"[警告] 未知单位按 uA 处理（可能差 1000 倍）：{u}")
    if info["skip_no_current"]:
        w.append(f"[提示] {info['skip_no_current']} 行 Current 为空被跳过"
                 f"（公式没有缓存值时会整表为空）")
    return w


def ingest_sim(conn, xlsx, sheet_name, label_ids=None):
    """返回 (入库行数, info)。"""
    recs, info = read_sim_rows(xlsx, sheet_name, label_ids)
    src = os.path.abspath(xlsx)
    conn.execute("DELETE FROM sim_current WHERE src_file=?", (src,))
    conn.executemany(
        "INSERT INTO sim_current(module_id,module_name,trim,mode,stage,tier,current_ua,unit_raw,src_file)"
        " VALUES(?,?,?,?,?,?,?,?,?)",
        [(r["module_id"], r["module_name"], r["trim"], r["mode"], r["stage"],
          r["tier"], r["current_ua"], r["unit_raw"], src) for r in recs])
    conn.commit()
    for line in sim_warnings(info):
        print(line)
    return len(recs), info


# ---------------------------------------------------------------- 实测表解析

def match_result_header(row):
    """实测表表头。带优先级：Current_mA(带单位) > Current(裸开关列)；Temperature* > temp
    （防撞 Vtemp）。返回列映射 dict（含 unit）或 None。"""
    no_col = mode_col = None
    cur_exact = cur_bare = temp_exact = temp_bare = None
    for j, c in enumerate(row):
        n = norm(c)
        if n in ("no.", "no", "no．") and no_col is None:
            no_col = j
        elif n == "mode" and mode_col is None:
            mode_col = j
        elif re.fullmatch(r"current[_\s]*[munµμ]?a", n) and cur_exact is None:
            cur_exact = j
        elif n == "current" and cur_bare is None:
            cur_bare = j
        elif n.startswith("temperature") and temp_exact is None:
            temp_exact = j
        elif n == "temp" and temp_bare is None:
            temp_bare = j
    cur_col = cur_exact if cur_exact is not None else cur_bare
    temp_col = temp_exact if temp_exact is not None else temp_bare
    if no_col is None or cur_col is None:
        return None
    if mode_col is None:
        mode_col = no_col + 1
    unit = "ma"
    m = re.fullmatch(r"current[_\s]*([munµμ]?a)", norm(row[cur_col]))
    if m and m.group(1):
        unit = m.group(1)
    return {"no": no_col, "mode": mode_col, "cur": cur_col, "temp": temp_col, "unit": unit}


def find_result_sheet(wb, sheet_name):
    """返回 (worksheet, 表头行号, 列映射)。按表头名定位，不按列字母。"""
    names = [sheet_name] if sheet_name else wb.sheetnames
    for sn in names:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            cols = match_result_header(row)
            if cols is not None:
                return ws, i, cols
    return None, None, None


def read_raw_rows(ws, hdr, cols):
    """表体 -> [(行号, no_raw, label, 电流float, 温度float)]，跳过全空行。"""
    raw = []
    for i, row in enumerate(ws.iter_rows(min_row=hdr + 1, values_only=True), hdr + 1):
        no_raw = cell(row, cols["no"])
        label = cell(row, cols["mode"])
        cur = as_float(cell(row, cols["cur"]))
        temp = as_float(cell(row, cols["temp"])) if cols["temp"] is not None else None
        if no_raw is None and label is None and cur is None:
            continue
        raw.append((i, no_raw, label, cur, temp))
    return raw


SIG_RE = re.compile(r"(.+)_sig\d+$")


def split_allmode(raw):
    """全模式单文件按行分段 -> [{mode, temp, raw:[...]}]；整表无边界时返回 []。
    开段：`*_sigN` 签名行（最可靠）/ Init 行 NO.=模式名（兜底，兼容旧单模式文件）。
    闭段：SET_TEMP 设温行、chamber 行（只控温箱，不入任何段）。
    相邻且 (canon模式,温度) 相同的段合并——签名行用 unSync、用户原行用 noSYNC 之类的
    命名分裂在这里收口，段名取带 Init 原行那段的写法（与仿真表命名一致的那个）。"""
    segs, cur_seg = [], None
    for rec in raw:
        _i, no_raw, label, _cur, _temp = rec
        ls = str(label).strip() if label is not None else ""
        ln, nn = norm(label), norm(no_raw)
        if ln.startswith("set_temp") or nn.startswith("set_temp") \
                or "chamber" in ln or "chamber" in nn:
            cur_seg = None
            continue
        m = SIG_RE.fullmatch(ls)
        if m:
            if cur_seg is None or canon_mode(cur_seg["mode"]) != canon_mode(m.group(1)):
                cur_seg = {"mode": m.group(1), "raw": []}
                segs.append(cur_seg)
        elif ln.startswith("init") and isinstance(no_raw, str) and no_raw.strip() \
                and not re.fullmatch(r"[\d ,，、;；]+", no_raw.strip()):
            name = no_raw.strip()
            if cur_seg is None or canon_mode(cur_seg["mode"]) != canon_mode(name):
                cur_seg = {"mode": name, "raw": []}
                segs.append(cur_seg)
            else:
                cur_seg["mode"] = name  # 同段：签名行段名让位给 Init 原行段名
        if cur_seg is not None:
            cur_seg["raw"].append(rec)
    out = []
    for s in segs:
        temps = [t for (_i, _n, _l, _c, t) in s["raw"] if t is not None]
        s["temp"] = temps[0] if temps else None
        if out and canon_mode(out[-1]["mode"]) == canon_mode(s["mode"]) \
                and out[-1]["temp"] == s["temp"]:
            out[-1]["raw"].extend(s["raw"])
            out[-1]["mode"] = s["mode"]
        else:
            out.append(s)
    return out


def classify_rows(ws, hdr, cols):
    """兼容入口：整表当一个序列。返回 (rows, temp)。"""
    factor_to_ma = UNIT_TO_UA.get(cols["unit"], 1000.0) / 1000.0  # 原始单位 -> mA
    return classify_raw(read_raw_rows(ws, hdr, cols), factor_to_ma)


def classify_raw(raw, factor_to_ma):
    """逐行分类并做差。返回 (rows, temp)。
    rows: dict(row_idx, no_raw, label, cur_ma, delta_ma, temp, seq, kind)
    两遍扫描：先看有没有显式 Init 行——有的话序列只从 Init 行开始，
    Init 之前的带电流行（其他测试项/签名行）留在 seq=0 不参与做差；
    全表都没有 Init 行时，才把第一个带电流行当作序列起点。"""
    has_init = any(norm(label).startswith("init") for _i, _n, label, _c, _t in raw)

    out = []
    seq = 0
    prev_cur = None
    temp_first = None
    for i, no_raw, label, cur, temp in raw:
        ln, nn = norm(label), norm(no_raw)
        if "chamber" in ln or "chamber" in nn:
            kind = "chamber"
        elif ln.startswith("init"):
            kind = "init"
            seq += 1
        elif "lock" in ln:
            kind = "lock"
        elif ln.startswith("off"):
            kind = "off"
        else:
            kind = "other"
        if seq == 0 and not has_init and cur is not None:
            seq = 1  # 整表无显式 Init 行时，第一段视为正式测量
            if kind == "other":
                kind = "init"
        delta = None
        if seq == 1 and kind in ("lock", "off", "other") and cur is not None and prev_cur is not None:
            delta = prev_cur * factor_to_ma - cur * factor_to_ma
        if seq == 1 and cur is not None and kind != "chamber":
            prev_cur = cur
        if temp is not None and temp_first is None and seq >= 1:
            temp_first = temp
        if kind == "other" and cur is None and seq == 0:
            continue  # Init 之前的测试计划行，不入库
        out.append(dict(row_idx=i, no_raw=no_raw, label=label,
                        cur_ma=(cur * factor_to_ma) if cur is not None else None,
                        delta_ma=delta, temp=temp, seq=seq, kind=kind))
    return out, temp_first


def build_groups(rows, config):
    """从 seq==1 的 OFF 行生成模块组，套用 LDO 归并。返回 (groups, absorbed_notes)。"""
    reparent = {}
    for c, p in (config.get("ldo_reparent") or {}).items():
        try:
            reparent[int(c)] = int(p)
        except (TypeError, ValueError):
            pass
    # dict 形式 {"mode":..,"ids":..}（跨模式映射）必须原样留着——早先的 list(v) 会把它
    # 拍成 ["mode","ids"]，下游 isinstance(mapped, dict) 永远为假，跨模式映射从未生效过。
    label_groups = {str(k): (v if isinstance(v, dict) else list(v))
                    for k, v in (config.get("label_groups") or {}).items()}

    steps = []
    for r in rows:
        if r["seq"] != 1 or r["kind"] != "off" or r["delta_ma"] is None:
            continue
        ids = parse_ids(r["no_raw"])
        disp = ",".join(str(i) for i in ids) if ids else str(r["no_raw"]).strip()
        step_name = re.sub(r"(?i)^off\s*", "", str(r["label"] or "")).strip()
        note = ""
        sim_ids = list(ids) if ids else None
        sim_mode = None  # 该步仿真值来自其他仿真 Mode 时（如 DCO 标签对 CK_ADPLL_*）
        if ids is None:
            mapped = label_groups.get(disp)
            if isinstance(mapped, dict):
                sim_mode = str(mapped.get("mode") or "").strip() or None
                v = mapped.get("ids")
                sim_ids = ["*"] if v in (None, "*", ["*"]) else [int(x) for x in v]
                note = (f"标签 {disp} 按 config.label_groups 映射到仿真"
                        f"{' Mode ' + sim_mode if sim_mode else ''} ID {sim_ids}")
            elif mapped:
                sim_ids = [int(x) for x in mapped]
                note = f"标签 {disp} 按 config.label_groups 映射到仿真 ID {sim_ids}"
            else:
                note = "标签未映射仿真模块（可在 current_config.json 的 label_groups 补充）"
        steps.append(dict(row_idx=r["row_idx"], ids=ids, sim_ids=sim_ids, sim_mode=sim_mode,
                          disp=disp, step_name=step_name, delta_ua=r["delta_ma"] * 1000.0,
                          note=note))

    # LDO 归并：单独成步的子模块，实测 delta 与仿真 ID 都并入父模块所在组
    absorbed = {}  # 子步 row_idx -> 父组 disp
    by_single_id = {s["ids"][0]: s for s in steps if s["ids"] and len(s["ids"]) == 1}
    for child, parent in reparent.items():
        child_step = by_single_id.get(child)
        parent_step = next((s for s in steps if s["ids"] and parent in s["ids"]), None)
        if child_step is None or parent_step is None or child_step is parent_step:
            if child_step is not None and parent_step is None:
                child_step["note"] = (child_step["note"] + "；" if child_step["note"] else "") + \
                    f"模块{child}不在被测LDO下（父模块{parent}本次未测，未归并）"
            continue
        parent_step["delta_ua"] += child_step["delta_ua"]
        add_child_sim = bool(config.get("ldo_reparent_sim_add_child", False))
        if add_child_sim:
            parent_step["sim_ids"] = (parent_step["sim_ids"] or []) + [child]
        # 编号列保持父模块本来的编号（用户定：6 就是 6），归并信息只进备注列
        parent_step["note"] = (parent_step["note"] + "；" if parent_step["note"] else "") + \
            f"含模块{child}的实测delta（{child}不在被测LDO下，仿真侧{'已并入' if add_child_sim else '不计'}{child}）"
        absorbed[child_step["row_idx"]] = parent_step["disp"]
    steps = [s for s in steps if s["row_idx"] not in absorbed]
    for order, s in enumerate(steps, 1):
        s["order"] = order
    return steps, absorbed


def _run_ts_of(xlsx):
    m = re.search(r"(\d{4}-\d{2}-\d{2})-(\d{2})-(\d{2})-(\d{2})", os.path.basename(xlsx))
    if m:
        return f"{m.group(1)} {m.group(2)}:{m.group(3)}:{m.group(4)}"
    return datetime.datetime.fromtimestamp(os.path.getmtime(xlsx)).strftime("%Y-%m-%d %H:%M:%S")


def _delete_runs_of(conn, src, chip):
    for (rid,) in conn.execute("SELECT run_id FROM runs WHERE src_file=? AND chip=?",
                               (src, chip)).fetchall():
        conn.execute("DELETE FROM meas_raw WHERE run_id=?", (rid,))
        conn.execute("DELETE FROM meas_module WHERE run_id=?", (rid,))
        conn.execute("DELETE FROM runs WHERE run_id=?", (rid,))


def _insert_run(conn, src, mode, mode_raw, chip, temp, rows, steps, absorbed, run_ts):
    cur = conn.execute(
        "INSERT INTO runs(mode,chip,temp_c,src_file,run_ts,ingested_ts,mode_raw)"
        " VALUES(?,?,?,?,?,?,?)",
        (mode, chip, temp, src, run_ts, now_iso(), mode_raw))
    run_id = cur.lastrowid
    for r in rows:
        note = ""
        if r["seq"] >= 2:
            note = "锁定复验段，忽略"
        elif r["row_idx"] in absorbed:
            note = f"并入组 {absorbed[r['row_idx']]}"
        conn.execute(
            "INSERT INTO meas_raw(run_id,row_idx,seq_idx,kind,no_raw,mode_label,current_ma,delta_ma,temp_c,note)"
            " VALUES(?,?,?,?,?,?,?,?,?,?)",
            (run_id, r["row_idx"], r["seq"], r["kind"],
             str(r["no_raw"]) if r["no_raw"] is not None else None,
             str(r["label"]) if r["label"] is not None else None,
             r["cur_ma"], r["delta_ma"], r["temp"], note))
    for s in steps:
        conn.execute(
            "INSERT INTO meas_module(run_id,step_order,group_disp,step_name,module_ids,sim_ids,"
            "current_ua,note,sim_mode) VALUES(?,?,?,?,?,?,?,?,?)",
            (run_id, s["order"], s["disp"], s["step_name"],
             json.dumps(s["ids"]) if s["ids"] else None,
             json.dumps(s["sim_ids"]) if s["sim_ids"] else None,
             s["delta_ua"], s["note"], s.get("sim_mode")))
    return run_id


def ingest_run(conn, xlsx, mode, chip, config, sheet_name=None):
    """单模式显式入库（ingest-run 子命令）：整表当一个序列，模式名由调用者给。"""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws, hdr, cols = find_result_sheet(wb, sheet_name or config.get("result_sheet"))
        if ws is None:
            raise SystemExit(f"[错误] {os.path.basename(xlsx)} 里找不到含 NO./Current 表头的 tab")
        rows, temp = classify_rows(ws, hdr, cols)
        steps, absorbed = build_groups(rows, config)
        src = os.path.abspath(xlsx)
        run_ts = _run_ts_of(xlsx)
        _delete_runs_of(conn, src, chip)
        run_id = _insert_run(conn, src, mode, mode, chip, temp, rows, steps, absorbed, run_ts)
        conn.commit()
        return run_id, len(steps), temp, run_ts
    finally:
        wb.close()


def ingest_result_file(conn, xlsx, chip, config, sim_modes, folder_mode=None, sheet_name=None):
    """一个 Result 文件 -> 若干 run（全模式单文件按 (模式,温度) 分段；旧单模式文件=1 段，
    模式名取 Init 行 NO.，退无可退才用文件夹名）。
    返回 [(run_id, mode, mode_raw, how, temp, n_steps)]。"""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws, hdr, cols = find_result_sheet(wb, sheet_name or config.get("result_sheet"))
        if ws is None:
            raise SystemExit(f"[错误] {os.path.basename(xlsx)} 里找不到含 NO./Current 表头的 tab")
        raw = read_raw_rows(ws, hdr, cols)
    finally:
        wb.close()
    factor_to_ma = UNIT_TO_UA.get(cols["unit"], 1000.0) / 1000.0
    return _ingest_raw(conn, raw, factor_to_ma, os.path.abspath(xlsx), _run_ts_of(xlsx),
                       chip, config, sim_modes, folder_mode)


def _ingest_raw(conn, raw, factor_to_ma, src, run_ts, chip, config, sim_modes, folder_mode=None):
    segs = split_allmode(raw)
    if not segs:  # 无任何段边界：整表一个序列，文件夹名当模式
        segs = [{"mode": folder_mode or "?", "raw": raw, "temp": None}]
    # 旧单模式目录（整文件一段）：文件夹名是操作者意图；段内 Init 标签见过模板复制错名
    # （DCO2G 文件夹里写着 DCO5G），此时文件夹名优先，映射表里标出来。全模式多段文件只信段标签。
    single_legacy = (len(segs) == 1 and folder_mode
                     and canon_mode(folder_mode) != canon_mode(segs[0]["mode"]))
    _delete_runs_of(conn, src, chip)
    out = []
    mode_map = config.get("mode_map") or {}
    for s in segs:
        rows, temp0 = classify_raw(s["raw"], factor_to_ma)
        temp = s["temp"] if s.get("temp") is not None else temp0
        steps, absorbed = build_groups(rows, config)
        if single_legacy:
            mode, _ = resolve_mode(folder_mode, sim_modes, mode_map)
            how = "folder"
        else:
            mode, how = resolve_mode(s["mode"], sim_modes, mode_map, folder=folder_mode)
        run_id = _insert_run(conn, src, mode, s["mode"], chip, temp, rows, steps, absorbed, run_ts)
        out.append((run_id, mode, s["mode"], how, temp, len(steps)))
    conn.commit()
    return out


def ingest_probe_json(conn, json_path, chip, config):
    """probe_allmode_result.py --json 的产物入库（黄区只带 JSON 回来时的开发机路径）。
    行流按原行号重排后走与 xlsx 完全相同的分段/分类管线。"""
    with open(json_path, "r", encoding="utf-8-sig") as f:  # BOM 容错，同 load_config
        d = json.load(f)
    recs = []
    for seg in (d.get("segments") or []):
        for r in seg.get("rows") or []:
            recs.append((r.get("row"), r.get("no"), r.get("label"),
                         as_float(r.get("current")), as_float(r.get("temp"))))
    for r in (d.get("orphans") or []):
        recs.append((r.get("row"), r.get("no"), r.get("label"),
                     as_float(r.get("current")), as_float(r.get("temp"))))
    recs.sort(key=lambda x: (x[0] if x[0] is not None else 0))
    unit = "ma"
    kc = d.get("key_cols_1based") or {}
    if isinstance(kc.get("unit"), str):
        unit = kc["unit"]
    factor_to_ma = UNIT_TO_UA.get(unit, 1000.0) / 1000.0
    src = os.path.abspath(json_path)
    m = re.search(r"(\d{4}-\d{2}-\d{2})-(\d{2})-(\d{2})-(\d{2})", str(d.get("file") or ""))
    run_ts = (f"{m.group(1)} {m.group(2)}:{m.group(3)}:{m.group(4)}" if m
              else _run_ts_of(json_path))
    sim_modes = {r[0] for r in conn.execute("SELECT DISTINCT mode FROM sim_current")}
    return _ingest_raw(conn, recs, factor_to_ma, src, run_ts, chip, config, sim_modes)


# ---------------------------------------------------------------- 导出

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def style_sheet(ws, widths=None):
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def rnd(v, n=3):
    return round(v, n) if isinstance(v, float) else v


def _inject_cached_values(xlsx_path, cache_by_sheet):
    """给 openpyxl 写出的公式单元格补上缓存值 <v>，让不自动重算的查看器也能显示结果。
    openpyxl 只写 <f> 不写 <v>，公式格在预览器/未重算的 Excel 里会空白——这里按坐标
    注入预算好的数值（保留公式，改数据后 Excel 仍会重算覆盖）。纯 stdlib（zipfile+re）。
    cache_by_sheet: {sheet标题: {坐标: 数值}}；数值为 None 的跳过（本就该显示空）。"""
    import zipfile
    if not any(cache_by_sheet.values()):
        return
    with zipfile.ZipFile(xlsx_path) as z:
        order = z.namelist()
        blob = {n: z.read(n) for n in order}
    wbxml = blob["xl/workbook.xml"].decode("utf-8")
    rels = blob["xl/_rels/workbook.xml.rels"].decode("utf-8")
    rid_target = {}   # Id 与 Target 在标签里顺序不定，分别抓；Target 可能是 /xl/.. 绝对路径
    for rel in re.findall(r"<Relationship\b[^>]*/>", rels):
        idm = re.search(r'Id="([^"]+)"', rel)
        tm = re.search(r'Target="([^"]+)"', rel)
        if idm and tm:
            t = tm.group(1)
            rid_target[idm.group(1)] = t[1:] if t.startswith("/") else "xl/" + t.lstrip("./")
    title_file = {}
    for tag in re.findall(r"<sheet\b[^>]*/>", wbxml):
        nm = re.search(r'name="([^"]+)"', tag)
        rid = re.search(r'r:id="([^"]+)"', tag)
        if nm and rid and rid.group(1) in rid_target:
            title_file[nm.group(1)] = rid_target[rid.group(1)]
    for title, cache in cache_by_sheet.items():
        fn = title_file.get(title)
        if not fn or not cache or fn not in blob:
            continue
        xml = blob[fn].decode("utf-8")
        for coord, val in cache.items():
            if val is None:
                continue
            vs = ("%.10g" % val) if isinstance(val, float) else str(val)
            pat = re.compile(r'(<c r="%s"[^>]*>)(<f[^>]*>.*?</f>)(?:<v>.*?</v>)?(</c>)'
                             % re.escape(coord), re.DOTALL)
            xml, _n = pat.subn(
                lambda m: m.group(1) + m.group(2) + "<v>" + vs + "</v>" + m.group(3),
                xml, count=1)
        blob[fn] = xml.encode("utf-8")
    tmp = xlsx_path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for n in order:
            z.writestr(n, blob[n])
    os.replace(tmp, xlsx_path)


def sim_lookup(conn, mode, ids, stage, tier="", fallback_stage=None, zero_ua=0.0):
    """返回 (合计uA, 缺失ID列表, trim集合, tier集合, 用回退阶段取值的ID列表)。
    tier 非空时只取该档位；该档位没有但有无档位('')的旧数据时退回旧数据（兼容旧仿真表）。
    ids 含 "*" = 该 Mode 全部模块合计（跨模式 label_groups 用）。

    fallback_stage：主阶段（通常 post=后仿）某模块为 0 或无行时，改用该阶段（pre=前仿）
    的值。**逐模块判断**——一个组里只补后仿为 0 的那几个成员，其余仍用后仿。
    用户侧前提：前仿已做 back annotate，多数情况下前仿≈后仿，两者可比。
    补过的 ID 会返回给调用方，报告里要标出来（口径可追溯）。

    zero_ua（config.sim_zero_ua）：**后仿漏接的模块不一定是干净的 0**——真簿里见过千分之几
    µA 的数值残渣（同模块前仿是三位数 µA）。判据写成 `== 0` 时这种残渣非零，于是它被当真值
    用进对比，一组模块的仿真合计凭空少掉一个模块，偏差% 直接假到几十个百分点。
    改成 `<= zero_ua` 才当缺项；回退阶段的值也要超过阈值才算「有值」。"""
    def rows_of(where, params):
        if tier:
            rows = conn.execute(f"SELECT current_ua,trim,tier FROM sim_current WHERE {where}"
                                " AND tier=?", params + (tier,)).fetchall()
            if rows:
                return rows
            return conn.execute(f"SELECT current_ua,trim,tier FROM sim_current WHERE {where}"
                                " AND tier=''", params).fetchall()
        return conn.execute(f"SELECT current_ua,trim,tier FROM sim_current WHERE {where}",
                            params).fetchall()

    def total_of(rows):
        return sum(r[0] for r in rows)

    def is_zero(rows):
        """空 / 合计小到等于没有（阈值 zero_ua）"""
        return not rows or abs(total_of(rows)) <= zero_ua

    total, missing, trims, tiers, fell_back = 0.0, [], set(), set(), []
    found_any = False
    if ids and any(str(x) == "*" for x in ids):
        rows = rows_of("mode=? AND stage=?", (mode, stage))
        if is_zero(rows) and fallback_stage:
            alt = rows_of("mode=? AND stage=?", (mode, fallback_stage))
            if not is_zero(alt):
                rows, _ = alt, fell_back.append("*")
        if not rows:
            return None, ["*"], set(), set(), []
        buckets = [rows]
    else:
        buckets = []
        for mid in ids:
            rows = rows_of("mode=? AND module_id=? AND stage=?", (mode, mid, stage))
            # 后仿该模块为 0/残渣（或整个没有行）时用前仿补——前仿有真值才补，
            # 否则保持 0/缺失（"两边都是 0"是真结论，不该被掩盖）
            if is_zero(rows) and fallback_stage:
                alt = rows_of("mode=? AND module_id=? AND stage=?",
                              (mode, mid, fallback_stage))
                if not is_zero(alt):
                    rows = alt
                    fell_back.append(mid)
            if not rows:
                missing.append(mid)
            else:
                buckets.append(rows)
    for rows in buckets:
        found_any = True
        for cur, trim, t in rows:
            total += cur
            if trim:
                trims.add(trim)
            if t:
                tiers.add(t)
    return (total if found_any else None), missing, trims, tiers, fell_back


def module_names(conn, ids, skip_missing=False):
    """ids -> 仿真模块名串。skip_missing=True 时略过仿真表没有的 ID（组值整组记在首
    编号上时，其余成员没有独立行，人读表里不该出现 ID15 这类噪音）；全缺才回退 IDn。"""
    names, fallback = [], []
    for mid in ids or []:
        rows = conn.execute(
            "SELECT DISTINCT module_name FROM sim_current WHERE module_id=? AND module_name!=''"
            " ORDER BY module_name", (mid,)).fetchall()
        if rows:
            n = "/".join(r[0] for r in rows)
            if n not in names:
                names.append(n)
        else:
            fallback.append(f"ID{mid}")
    if not skip_missing:
        names += fallback
    return " + ".join(names or fallback)


def latest_runs(conn, all_runs=False):
    """runs 行（含 mode_raw）；默认同 (mode,chip,温度) 取 run_ts 最新的一次。"""
    runs = conn.execute(
        "SELECT run_id,mode,chip,temp_c,src_file,run_ts,mode_raw FROM runs"
        " ORDER BY mode,chip,run_ts,run_id").fetchall()
    if not all_runs:
        latest = {}
        for r in runs:
            latest[(r[1], r[2], r[3])] = r  # 同 mode+chip+temp 取最新（已按 run_ts 升序）
        runs = sorted(latest.values(), key=lambda r: (r[1], r[2], r[3] if r[3] is not None else 0))
    return runs


def export_xlsx(conn, out_path, all_runs=False, config=None):
    tier = (config or {}).get("sim_tier") or ""
    fb_stage = bool((config or {}).get("sim_stage_fallback", True))
    zero_ua = float((config or {}).get("sim_zero_ua") or 0)
    runs = latest_runs(conn, all_runs)

    wb = openpyxl.Workbook()

    # ---- ReadMe
    ws = wb.active
    ws.title = "ReadMe"
    ws.append(["电流对比数据库 · 导出说明"])
    ws["A1"].font = Font(bold=True, size=14)
    for line in [
        "",
        f"导出时间：{now_iso()}    数据来源：current.db（current_db.py 生成）",
        "",
        "【Sheet 说明】",
        "  Compare    —— 每模式每模块组一行：仿真 pre/post、实测、偏差%（可直接用于汇报）",
        "  Long       —— 透视长表：Source 列区分 sim_pre / sim_post / meas，做透视图用",
        "  Sim_long   —— 仿真长表原样（单位已统一为 uA）",
        "  Meas_steps —— 实测逐行审计：原始电流、做差、行分类（复验段/归并行也在）",
        "  Runs       —— 本次导出包含的测试 run 列表",
        "",
        "【计算规则】",
        "  1. 基线 = 第一个 OFF 行之前最后一行的电流（通常是最后一个 Lock_step）",
        "  2. 模块电流 = 上一行电流 - 本行电流（逐级关断做差），统一为 uA",
        "  3. NO. 列多个编号（如 45,46）= 一组模块同时关断，仿真侧按组求和对比",
        "  4. LDO 归并（current_config.json 的 ldo_reparent）：子模块不在被测 LDO 下，",
        "     其实测 delta 并入父模块组（编号仍显示父模块号，归并详情见 Note 列）",
        "  5. 第二个及以后的 Init 段 = 锁定复验，忽略（Meas_steps 里有原始行）",
        "  6. 多 run 时默认每个 模式×芯片×温度 取最新一次；--all-runs 可导出全部",
        f"  7. 仿真对比只取档位 sim_tier={tier or '(未过滤)'}（current_config.json 里改）",
        "",
        "【透视建议（Long 页）】",
        "  行=Group/Modules，列=Source（或 Chip/Temp_C），值=Current_uA（用平均值，防多 run 重复计数）",
    ]:
        ws.append([line])
    ws.column_dimensions["A"].width = 100

    # ---- Compare / Long
    cmp_ws = wb.create_sheet("Compare")
    cmp_ws.append(["Mode", "Chip", "Temp_C", "Run_TS", "Step", "Group", "Step_Name", "Modules",
                   "Sim_pre_uA", "Sim_post_uA", "Meas_uA", "Meas-Post_uA", "Meas/Post", "Dev_%", "Note"])
    long_ws = wb.create_sheet("Long")
    long_ws.append(["Mode", "Chip", "Temp_C", "Run_TS", "Step", "Group", "Step_Name", "Modules",
                    "Source", "Trim", "Tier", "Current_uA", "Note"])

    sim_modes = {r[0] for r in conn.execute("SELECT DISTINCT mode FROM sim_current")}
    for run_id, mode, chip, temp, _src, run_ts, _mode_raw in runs:
        groups = conn.execute(
            "SELECT step_order,group_disp,step_name,module_ids,sim_ids,current_ua,note,sim_mode"
            " FROM meas_module WHERE run_id=? ORDER BY step_order", (run_id,)).fetchall()
        for order, disp, step_name, _mids, sim_ids_j, meas_ua, note, sim_mode in groups:
            sim_ids = json.loads(sim_ids_j) if sim_ids_j else None
            names = module_names(conn, sim_ids) if sim_ids else ""
            notes = [note] if note else []
            sim_pre = sim_post = None
            trims, tiers = set(), set()
            lk_mode = sim_mode or mode  # 跨模式 label_groups 的步查它指定的仿真 Mode
            if sim_ids and lk_mode in sim_modes:
                sim_pre, miss_pre, t1, r1, _fb1 = sim_lookup(conn, lk_mode, sim_ids, "pre", tier)
                sim_post, miss_post, t2, r2, fb2 = sim_lookup(
                    conn, lk_mode, sim_ids, "post", tier, fb_stage and "pre", zero_ua)
                if fb2:
                    notes.append(f"仿真 {','.join(str(x) for x in fb2)} 后仿为0/漏项，用前仿补")
                trims, tiers = t1 | t2, r1 | r2
                miss = sorted(set(miss_pre) & set(miss_post))
                if miss:
                    notes.append(f"仿真表缺ID: {miss}")
            elif sim_ids and lk_mode not in sim_modes:
                notes.append("仿真表未导入" if not sim_modes else f"仿真表无模式 {lk_mode}")
            note_s = "；".join(n for n in notes if n)
            diff = (meas_ua - sim_post) if (sim_post is not None) else None
            ratio = (meas_ua / sim_post) if sim_post else None
            dev = (diff / sim_post * 100.0) if sim_post else None
            cmp_ws.append([mode, chip, temp, run_ts, order, disp, step_name, names,
                           rnd(sim_pre), rnd(sim_post), rnd(meas_ua), rnd(diff),
                           rnd(ratio, 3), rnd(dev, 1), note_s])
            trim_s = ",".join(sorted(trims))
            tier_s = ",".join(sorted(tiers))
            long_ws.append([mode, chip, temp, run_ts, order, disp, step_name, names,
                            "meas", "", "", rnd(meas_ua), note_s])
            if sim_pre is not None:
                long_ws.append([mode, chip, temp, run_ts, order, disp, step_name, names,
                                "sim_pre", trim_s, tier_s, rnd(sim_pre), ""])
            if sim_post is not None:
                long_ws.append([mode, chip, temp, run_ts, order, disp, step_name, names,
                                "sim_post", trim_s, tier_s, rnd(sim_post), ""])

    style_sheet(cmp_ws, [16, 8, 8, 17, 6, 12, 22, 34, 12, 12, 12, 13, 10, 8, 40])
    style_sheet(long_ws, [16, 8, 8, 17, 6, 12, 22, 34, 10, 8, 8, 12, 40])
    if cmp_ws.max_row > 1:
        cmp_ws.conditional_formatting.add(
            f"N2:N{cmp_ws.max_row}",
            ColorScaleRule(start_type="num", start_value=-50, start_color="63BE7B",
                           mid_type="num", mid_value=0, mid_color="FFFFFF",
                           end_type="num", end_value=50, end_color="F8696B"))

    # ---- Sim_long
    ws = wb.create_sheet("Sim_long")
    ws.append(["ID", "Module", "Trim", "Mode", "Stage", "Tier", "Current_uA", "Unit_raw"])
    for r in conn.execute(
            "SELECT module_id,module_name,trim,mode,stage,tier,current_ua,unit_raw FROM sim_current"
            " ORDER BY mode,module_id,stage"):
        ws.append([r[0], r[1], r[2], r[3], r[4], r[5], rnd(r[6]), r[7]])
    style_sheet(ws, [6, 28, 8, 18, 8, 8, 12, 9])

    # ---- Meas_steps
    ws = wb.create_sheet("Meas_steps")
    ws.append(["Mode", "Chip", "Run_TS", "Row", "Seq", "Kind", "NO_raw", "Mode_label",
               "Current_mA", "Delta_mA", "Temp_C", "Note"])
    for run_id, mode, chip, _temp, _src, run_ts, _mraw in runs:
        for r in conn.execute(
                "SELECT row_idx,seq_idx,kind,no_raw,mode_label,current_ma,delta_ma,temp_c,note"
                " FROM meas_raw WHERE run_id=? ORDER BY row_idx", (run_id,)):
            ws.append([mode, chip, run_ts, r[0], r[1], r[2], r[3], r[4],
                       rnd(r[5], 4), rnd(r[6], 4), r[7], r[8]])
    style_sheet(ws, [16, 8, 17, 6, 5, 8, 12, 24, 11, 10, 8, 30])

    # ---- Runs
    ws = wb.create_sheet("Runs")
    ws.append(["Run_ID", "Mode", "Mode_raw", "Chip", "Temp_C", "Run_TS", "Src_file"])
    for r in runs:
        ws.append([r[0], r[1], r[6], r[2], r[3], r[5], r[4]])
    style_sheet(ws, [7, 16, 16, 8, 8, 17, 70])

    wb.save(out_path)


# ---- 汇总簿视觉语言（参考评审报告表：黄表头带/条件行米色/结果白/合计蓝/超差红粗/细边框） ----
C_HEADER, C_SETTING, C_RESULT, C_SEP, C_FLAG = "FFFF00", "EEECE1", "FFFFFF", "B8CCE4", "FF0000"
# 宽表的视觉分区（与 sweep_lib 同一套值）：窄灰竖栏界定"这一块是一片"，
# 浅蓝给全表唯一要跳出来的汇总组。人一眼能数清的上限是 4~5 个，超了就得数——
# 分区不是好看，是让"我现在看的是哪一颗芯片"这条信息不丢。
C_RAIL, C_SUM = "D9D9D9", "DDEBF7"
FONT_NAME = "微软雅黑"
_THIN = Side(style="thin", color="FF000000")
BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
FMT_UA, FMT_MA, FMT_PCT = "#,##0.0", "0.000", "0.0%"


def _cell(ws, r, c, val=None, bold=False, fill=None, fmt=None, align="center", size=10):
    cc = ws.cell(row=r, column=c)
    if val is not None:
        cc.value = val
    cc.font = Font(name=FONT_NAME, size=size, bold=bold)
    cc.border = BORDER_ALL
    cc.alignment = Alignment(horizontal=align, vertical="center")
    if fill:
        cc.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        cc.number_format = fmt
    return cc


def _t(v):
    """温度显示：25.0 -> '25℃'。"""
    return ("%g" % v) + "℃"


def _mode_freq(mode, config):
    """测试频率条件行文本。config.mode_freq 优先；否则按模式名首个频段 token 推断
    （2G -> 2.5GHz，5G -> 5.8GHz；BT_2G_TX_DCO5G 这类先出现 2G 算 2G 模式）。"""
    fmap = config.get("mode_freq") or {}
    if mode in fmap:
        return str(fmap[mode])
    m = re.search(r"([25])\s*G", str(mode).upper())
    if m:
        return "2.5GHz" if m.group(1) == "2" else "5.8GHz"
    return ""


def _chart_title(text, sz=1100, bold=True):
    """图表标题（显式字体/字号的富文本）。openpyxl 默认标题不带字体属性，
    Excel 对中文按默认字体度量排版会把文本框排溢出、裁掉开头几个字。"""
    rpr = CharacterProperties(sz=sz, b=bold,
                              latin=DrawFont(typeface=FONT_NAME),
                              ea=DrawFont(typeface=FONT_NAME))
    p = Paragraph(pPr=ParagraphProperties(defRPr=rpr),
                  r=[RegularTextRun(rPr=rpr, t=text)])
    return Title(tx=Text(rich=RichText(p=[p])))


def summary_data(conn, config, chips=None):
    """从库里备好两种版式共用的那份数据（run/温度轴、行列 universe、实测矩阵、仿真值）。

    单芯片版式（cmd_summary_export）和跨芯片版式（chips_sheet）只在**怎么摆**上不同，
    读数和口径必须是同一份——否则两本簿子会对同一批数据给出不同的数，这种事没法解释。"""
    d = types.SimpleNamespace()
    d.tier = tier = config.get("sim_tier") or ""
    d.stage_main = stage_main = (config.get("sim_stage") or "post").strip().lower()
    d.fb_stage = fb_stage = bool(config.get("sim_stage_fallback", True))
    d.zero_ua = zero_ua = float(config.get("sim_zero_ua") or 0)
    d.other_stage = other_stage = "pre" if stage_main == "post" else "post"
    d.sim_note = config.get("sim_temp_note") or ""
    d.thr = float(config.get("delta_flag_pct") or 20) / 100.0
    d.abs_thr = float(config.get("delta_flag_abs_ua") or 0)   # 双阈值绝对下限 µA

    runs = latest_runs(conn)
    if not runs:
        raise SystemExit("[错误] 库里没有任何实测 run")
    if chips:
        have = sorted({r[2] for r in runs})
        bad = [c for c in chips if c not in have]
        if bad:
            raise SystemExit(f"[错误] 库里没有芯片 {', '.join(bad)}；现有: {', '.join(have)}")
        runs = [r for r in runs if r[2] in chips]
    d.runs = runs
    d.sim_modes = sim_modes = {r[0] for r in conn.execute("SELECT DISTINCT mode FROM sim_current")}
    d.multi_chip = multi_chip = len({r[2] for r in runs}) > 1
    temps = sorted({r[3] for r in runs if r[3] is not None})
    if not temps:
        temps = [None]
    d.temps = temps
    d.n_t = n_t = len(temps)
    # 仿真温度：偏差把实测线性插值到该温度再对仿真，消除单点仿真 vs 多温实测的系统温差
    sim_temp_c = config.get("sim_temp_c")
    if sim_temp_c is None:
        m = re.search(r"-?\d+(?:\.\d+)?", str(config.get("sim_temp_note") or ""))
        sim_temp_c = float(m.group(0)) if m else (temps[0] if temps[0] is not None else 25)
    d.sim_temp_c = sim_temp_c

    def interp_to(pairs, target):
        """线性插值到 target 温度；范围外取最近端点（不外推）；单点直接返回。"""
        pts = sorted((t, v) for t, v in pairs if t is not None and v is not None)
        if not pts:
            return None
        if len(pts) == 1 or target <= pts[0][0]:
            return pts[0][1]
        if target >= pts[-1][0]:
            return pts[-1][1]
        for i in range(1, len(pts)):
            (t0, v0), (t1, v1) = pts[i - 1], pts[i]
            if t0 <= target <= t1:
                return v0 + (v1 - v0) * (target - t0) / (t1 - t0)
        return pts[-1][1]
    d.interp_to = interp_to

    # 列组 = (mode, chip)，按首个 run_id 排（=入库顺序=测试顺序）
    first_id = {}
    for r in runs:
        k = (r[1], r[2])
        if k not in first_id or r[0] < first_id[k]:
            first_id[k] = r[0]
    d.col_keys = col_keys = sorted(first_id, key=first_id.get)
    # 模式/芯片各自的顺序（跨芯片版式按这两根轴摆）
    d.modes = sorted({m for m, _c in col_keys}, key=lambda m: min(
        first_id[k] for k in col_keys if k[0] == m))
    d.chips = sorted({c for _m, c in col_keys}, key=lambda c: min(
        first_id[k] for k in col_keys if k[1] == c))

    def col_title(mode, chip):
        return f"{mode}({chip})" if multi_chip else mode
    d.col_title = col_title

    # 逐 run 取基线/末态/模块组
    base_ma, init_ma, end_ma, per_groups = {}, {}, {}, {}   # 键 (mode,chip,temp)
    src_files = set()
    for run_id, mode, chip, temp, src, _ts, _mraw in runs:
        k = (mode, chip, temp)
        src_files.add(os.path.basename(src))
        row = conn.execute(
            "SELECT current_ma FROM meas_raw WHERE run_id=? AND seq_idx=1 AND kind='init'"
            " AND current_ma IS NOT NULL ORDER BY row_idx LIMIT 1", (run_id,)).fetchone()
        init_ma[k] = row[0] if row else None
        row = conn.execute(
            "SELECT current_ma FROM meas_raw WHERE run_id=? AND seq_idx=1 AND kind='lock'"
            " AND current_ma IS NOT NULL ORDER BY row_idx DESC LIMIT 1", (run_id,)).fetchone()
        base_ma[k] = row[0] if row else init_ma[k]
        row = conn.execute(
            "SELECT current_ma FROM meas_raw WHERE run_id=? AND seq_idx=1 AND kind='off'"
            " AND current_ma IS NOT NULL ORDER BY row_idx DESC LIMIT 1", (run_id,)).fetchone()
        end_ma[k] = row[0] if row else None      # 全部关断后的末态电流（末个 OFF 步实测）
        per_groups[k] = conn.execute(
            "SELECT step_order,group_disp,step_name,sim_ids,sim_mode,current_ua,note"
            " FROM meas_module WHERE run_id=? ORDER BY step_order", (run_id,)).fetchall()
    d.base_ma, d.init_ma, d.end_ma = base_ma, init_ma, end_ma
    d.per_groups, d.src_files = per_groups, src_files

    # 行 universe：(disp, step_name)，按平均步序排
    matrix, order_sum, order_cnt, notes, siminfo = {}, {}, {}, {}, {}
    caliber_keys = set()   # LDO 归并父行：实测含子模块、仿真不含 -> 口径不可比，不标红
    for (mode, chip, temp), groups in per_groups.items():
        for step_order, disp, step_name, sim_ids_j, sim_mode, ua, note in groups:
            key = (disp, step_name)
            matrix.setdefault(key, {})[(mode, chip, temp)] = ua
            order_sum[key] = order_sum.get(key, 0) + step_order
            order_cnt[key] = order_cnt.get(key, 0) + 1
            if sim_ids_j and key not in siminfo:
                siminfo[key] = (json.loads(sim_ids_j), sim_mode)
            if note:
                if "不在被测LDO" in note:
                    caliber_keys.add(key)
                keep = "；".join(x for x in note.split("；")
                                 if "不在被测LDO" in x or ("映射" in x and "未映射" not in x))
                if keep:
                    notes.setdefault(key, set()).add(keep)
    def _row_sort_key(k):
        """按 buffer 编号从低到高；组合/归并组（"45,46"、"26+28"）取组内最小号；
        非数字标签（DCO2G 等）排最后，按平均步序。"""
        ids = parse_ids(str(k[0]).replace("+", ","))
        if ids:
            return (0, min(ids), order_sum[k] / order_cnt[k])
        return (1, 0, order_sum[k] / order_cnt[k])
    row_keys = sorted(matrix, key=_row_sort_key)
    d.matrix, d.notes, d.siminfo = matrix, notes, siminfo
    d.caliber_keys, d.row_keys = caliber_keys, row_keys

    # 仿真值缓存：行×模式 -> pre/post 合计 µA（tier 过滤）；行名
    sim_val, sim_pre_val, sim_names, sim_fb = {}, {}, {}, {}
    for key in row_keys:
        ids, override = siminfo.get(key, (None, None))
        sim_names[key] = module_names(conn, [i for i in (ids or []) if str(i) != "*"],
                                      skip_missing=True) if ids else ""
        for mode, chip in col_keys:
            lk_mode = override or mode
            if ids and lk_mode in sim_modes:
                # 主对比列取哪个阶段由 config.sim_stage 定（默认 post=后仿）；另一个阶段
                # 仍算出来放明细页。两者差别很大时(某阶段整片为 0)靠 inspect 的零值审计发现。
                sim_val[(key, mode)], _, _, _, fb = sim_lookup(
                    conn, lk_mode, ids, stage_main, tier, fb_stage and other_stage, zero_ua)
                sim_pre_val[(key, mode)], _, _, _, _ = sim_lookup(
                    conn, lk_mode, ids, other_stage, tier)
                if fb:
                    sim_fb[(key, mode)] = fb
    d.sim_val, d.sim_pre_val, d.sim_names, d.sim_fb = sim_val, sim_pre_val, sim_names, sim_fb
    # 标签行（DCO 等）排在末尾——Σ 分两段：LO 模块（可与仿真对）/ 总合计（含标签行，只有实测）
    d.n_label = sum(1 for k in row_keys if parse_ids(str(k[0]).replace("+", ",")) is None)
    return d


def cmd_summary_export(conn, out_path, config, mark_fb=False, chips=None):
    """人直接读的汇总簿：说明 / 总览(模块×模式×温度 + 仿真对比) / 温度趋势(图) / 对比明细。"""
    d = summary_data(conn, config, chips)
    tier, stage_main, fb_stage = d.tier, d.stage_main, d.fb_stage
    zero_ua, other_stage, sim_note = d.zero_ua, d.other_stage, d.sim_note
    thr, abs_thr = d.thr, d.abs_thr
    runs, sim_modes, multi_chip = d.runs, d.sim_modes, d.multi_chip
    temps, n_t, sim_temp_c, interp_to = d.temps, d.n_t, d.sim_temp_c, d.interp_to
    col_keys, col_title = d.col_keys, d.col_title
    base_ma, end_ma, per_groups, src_files = d.base_ma, d.end_ma, d.per_groups, d.src_files
    matrix, notes, caliber_keys, row_keys = d.matrix, d.notes, d.caliber_keys, d.row_keys
    siminfo = d.siminfo
    sim_val, sim_pre_val, sim_names, sim_fb, n_label = (
        d.sim_val, d.sim_pre_val, d.sim_names, d.sim_fb, d.n_label)

    wb = openpyxl.Workbook()
    fcache = {}   # {sheet标题: {坐标: 缓存值}}，写公式时记录，保存后注入 <v>

    def fcell(ws_, r, c, formula, value, **kw):
        """写公式并记录其缓存值（保存后注入，使不重算的查看器也能显示）。"""
        _cell(ws_, r, c, formula, **kw)
        fcache.setdefault(ws_.title, {})[f"{get_column_letter(c)}{r}"] = value
        return value

    # ================= 说明 =================
    ws = wb.active
    ws.title = "说明"
    ws["A1"] = "口径与来源"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=12)
    # ★ 只写口径和来源，不写读法。原来这一页有「行=模块…列=…」「【颜色】黄=表头…」
    #   这类整段教人怎么读表的话，还捎带 config 键名——那是写给出簿的人的，
    #   评审看到只会停下来读一句他本来不需要知道的话。
    lines = [
        "",
        f"导出 {now_iso()}  current_db.py summary  数据源 current.db",
        f"实测：{len(runs)} 个 run（模式×芯片×温度，重复测取 run_ts 最新一次）；"
        f"温度点 {'、'.join(_t(t) for t in temps if t is not None) or '未知'}",
        f"源文件：{'、'.join(sorted(src_files))}",
        f"仿真：{tier or '未过滤'} / {stage_main}-sim / {sim_note or '温度未标注'}"
        f"（{'pre' if stage_main == 'post' else 'post'}-sim 在「对比明细」页）",
        ((f"跨阶段补值：{stage_main} 为 0 或 ≤{zero_ua:g}µA 计作缺项时取 "
          f"{'pre' if stage_main == 'post' else 'post'}（前仿已做 back annotate）；两阶段都缺则留空。"
          f"共 {len(sim_fb)} 处，逐处见「对比明细」页备注列。")
         if (fb_stage and sim_fb) else "仿真列全部取自同一阶段，无跨阶段补值。")
        + ("  【本簿为自查版：补过的格子标成蓝色斜体】" if (mark_fb and sim_fb) else ""),
        "基线：每模式段第一个 OFF 行之前最后一行（末个 Lock_step）；模块电流 = 上一行 − 本行",
        "锁定后总电流为整机电流，不与仿真直接对比",
        f"偏差% = (实测线性插值到 {'%g' % sim_temp_c}℃ − 仿真) ÷ 仿真"
        f"（仿真为 {'%g' % sim_temp_c}℃ 单点，插值消除系统性温差；三温实测原值仍分列可见）",
        f"标红：|偏差%| > {thr * 100:.0f}% 且 |绝对偏差| > {abs_thr:.0f}µA（双阈值同时满足）",
        "LDO 归并 "
        + ("、".join(f"{k}→{v}" for k, v in sorted((config.get("ldo_reparent") or {}).items()))
           or "无")
        + "：子模块实测并入父组、仿真侧不计子模块 -> 父组口径不可比，偏差保留但不标红",
        "Σ LO 模块合计不含标签行（DCO 等），口径与仿真一致；Σ 总合计含标签行、仿真未覆盖",
        "偏差% / Σ 为 Excel 公式，并已写入当前结果的缓存值",
    ]
    for i, line in enumerate(lines, 2):
        ws.cell(row=i, column=1, value=line).font = Font(name=FONT_NAME, size=10)
    ws.column_dimensions["A"].width = 110

    # ================= 总览 =================
    ws = wb.create_sheet("总览")
    FIX = 3                                  # 编号/模块/单位
    grp_w = n_t + 2                          # 每模式：温度列 + 仿真 + 偏差%
    note_col = FIX + len(col_keys) * grp_w + 1

    def ref(r, c):
        return f"{get_column_letter(c)}{r}"

    # -- 表头带（3 行黄）：先整片打底，再填字做合并
    for r in range(1, 4):
        for c in range(1, note_col + 1):
            _cell(ws, r, c, fill=C_HEADER)
    for c, (title, w) in enumerate(zip(["编号", "模块 (OFF 步)", "单位"],
                                       [9, 30, 6]), 1):
        ws.merge_cells(start_row=1, start_column=c, end_row=3, end_column=c)
        _cell(ws, 1, c, title, bold=True, fill=C_HEADER)
        ws.column_dimensions[get_column_letter(c)].width = w
    for gi, (mode, chip) in enumerate(col_keys):
        c0 = FIX + 1 + gi * grp_w
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c0 + grp_w - 1)
        _cell(ws, 1, c0, col_title(mode, chip), bold=True, fill=C_HEADER)
        if n_t > 1:
            ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + n_t - 1)
        _cell(ws, 2, c0, "实测", bold=True, fill=C_HEADER)
        _cell(ws, 2, c0 + n_t, "仿真", bold=True, fill=C_HEADER)
        _cell(ws, 2, c0 + n_t + 1, "偏差", bold=True, fill=C_HEADER)
        for ti, t in enumerate(temps):
            _cell(ws, 3, c0 + ti, _t(t) if t is not None else "?", bold=True, fill=C_HEADER)
        _cell(ws, 3, c0 + n_t, (f"{sim_note} {tier}".strip() or "post"), bold=True, fill=C_HEADER)
        _cell(ws, 3, c0 + n_t + 1, f"vs{'%g' % sim_temp_c}℃*", bold=True, fill=C_HEADER)
        for ti in range(n_t):
            ws.column_dimensions[get_column_letter(c0 + ti)].width = 9.5
        ws.column_dimensions[get_column_letter(c0 + n_t)].width = 10
        ws.column_dimensions[get_column_letter(c0 + n_t + 1)].width = 8.5
    ws.merge_cells(start_row=1, start_column=note_col, end_row=3, end_column=note_col)
    _cell(ws, 1, note_col, "备注", bold=True, fill=C_HEADER)
    ws.column_dimensions[get_column_letter(note_col)].width = 42

    r_freq, r_base, r_end = 4, 5, 6
    r_mod0 = 7
    r_sum = r_mod0 + len(row_keys)          # Σ LO 模块行（不含末尾标签行）
    r_all = r_sum + 1 if n_label else None  # Σ 总合计行（含 DCO 等标签行，只有实测）
    r_lo_last = r_sum - 1 - n_label         # 最后一个 LO 模块行（SUM 范围用）
    lo_keys = row_keys[:len(row_keys) - n_label]  # LO 模块行（标签行排最后 n_label 个）

    # ★ 这里原来给「编号交叠」的行加一句备注（同一编号在不同模式被分到不同关断组，
    #   如 25 单独 vs 25,24,23）。2026-08-05 删掉——那句话的内容是"各模式列互不冲突"，
    #   等于叫人别担心，而正表上凭空出现一句"别担心"只会让人开始担心。各模式的列本来
    #   就错开填充，看数看得出来。（sim_col_sum 的 has_meas 门控才是真防线，那个还在。）

    # 偏差%/Σ 都写公式；缓存值由 fcell 记录、保存后注入，使不重算的查看器也能显示。
    def has_meas(key, mode, chip):
        return any((mode, chip, t) in matrix[key] for t in temps)

    def col_sum(keys, mode, chip, t):
        vals = [matrix[k].get((mode, chip, t)) for k in keys
                if matrix[k].get((mode, chip, t)) is not None]
        return sum(vals) if vals else None

    def sim_col_sum(keys, mode, chip):
        # 只统计该模式实际测了的行的仿真——否则 25,24,23 与单独 25 都命中 ID=25 会双算
        vals = [sim_val.get((k, mode)) for k in keys
                if has_meas(k, mode, chip) and sim_val.get((k, mode)) is not None]
        return sum(vals) if vals else None

    def dev_of(m, s):
        return (m - s) / s if (m is not None and s not in (None, 0)) else None

    def meas_at_sim(key, mode, chip):
        """该行实测插值到仿真温度（sim_temp_c）。"""
        return interp_to([(t, matrix[key].get((mode, chip, t))) for t in temps], sim_temp_c)

    def lo_meas_at_sim(mode, chip):
        return interp_to([(t, col_sum(lo_keys, mode, chip, t)) for t in temps], sim_temp_c)

    def flag_red(dev, meas, sim, key):
        """双阈值 + 口径守卫：|偏差%|>thr 且 |ΔµA|>abs_thr，且非 LDO 口径不可比行。"""
        if dev is None or meas is None or sim is None or key in caliber_keys:
            return False
        return abs(dev) > thr and abs(meas - sim) > abs_thr

    def interp_expr(rr, c0):
        """该行实测插值到 sim_temp_c 的 Excel 表达式（引用温度列单元格，随实测联动）。"""
        pts = sorted((t, i) for i, t in enumerate(temps) if t is not None)
        if not pts:
            return None
        if len(pts) == 1 or sim_temp_c <= pts[0][0]:
            return ref(rr, c0 + pts[0][1])
        if sim_temp_c >= pts[-1][0]:
            return ref(rr, c0 + pts[-1][1])
        for i in range(1, len(pts)):
            (t0, i0), (t1, i1) = pts[i - 1], pts[i]
            if t0 <= sim_temp_c <= t1:
                w = (sim_temp_c - t0) / (t1 - t0)
                a, b = ref(rr, c0 + i0), ref(rr, c0 + i1)
                return f"({a}+{w:g}*({b}-{a}))"
        return ref(rr, c0 + pts[-1][1])

    red_font = Font(name=FONT_NAME, size=10, bold=True, color=C_FLAG)

    # -- 条件/汇总行（米色）：测试频率 / 锁定后总电流 / 全关残留
    for rr, name, unit in ((r_freq, "测试频率", ""),
                           (r_base, "锁定后总电流", "mA"),
                           (r_end, "全关残留电流", "mA")):
        _cell(ws, rr, 1, fill=C_SETTING)
        _cell(ws, rr, 2, name, bold=True, fill=C_SETTING, align="left")
        _cell(ws, rr, 3, unit, fill=C_SETTING)
    for gi, (mode, chip) in enumerate(col_keys):
        c0 = FIX + 1 + gi * grp_w
        for cc in range(c0, c0 + grp_w):     # 频率行整组打底再合并
            _cell(ws, r_freq, cc, fill=C_SETTING)
        ws.merge_cells(start_row=r_freq, start_column=c0,
                       end_row=r_freq, end_column=c0 + grp_w - 1)
        _cell(ws, r_freq, c0, _mode_freq(mode, config), bold=True, fill=C_SETTING)
        for ti, t in enumerate(temps):
            v = base_ma.get((mode, chip, t))
            _cell(ws, r_base, c0 + ti, rnd(v, 3) if v is not None else "",
                  fill=C_SETTING, fmt=FMT_MA)
            v = end_ma.get((mode, chip, t))
            _cell(ws, r_end, c0 + ti, rnd(v, 3) if v is not None else "",
                  fill=C_SETTING, fmt=FMT_MA)
        # 基线/末态是含 DCO 与杂项的整机电流，仿真只覆盖 LO 模块——不做行内对比，
        # 仿真对比见底部「Σ LO 模块合计」行
        _cell(ws, r_base, c0 + n_t, "", fill=C_SETTING)
        _cell(ws, r_end, c0 + n_t, "", fill=C_SETTING)
        _cell(ws, r_base, c0 + n_t + 1, "", fill=C_SETTING)
        _cell(ws, r_end, c0 + n_t + 1, "", fill=C_SETTING)
    # 条件行不写备注：这三行原来各挂一句定义（频率怎么推的、基线是哪一行、残留是什么），
    # 都是口径不是发现，而且还捎带 config 键名。口径在说明页，正表上不重复。
    for r in (r_freq, r_base, r_end):
        _cell(ws, r, note_col, "", fill=C_SETTING, align="left")

    # -- 模块行（白，µA）
    for i, key in enumerate(row_keys):
        rr = r_mod0 + i
        disp, step_name = key
        _cell(ws, rr, 1, disp)
        _cell(ws, rr, 2, step_name, align="left")
        _cell(ws, rr, 3, "µA")
        for gi, (mode, chip) in enumerate(col_keys):
            c0 = FIX + 1 + gi * grp_w
            for ti, t in enumerate(temps):
                v = matrix[key].get((mode, chip, t))
                _cell(ws, rr, c0 + ti, rnd(v, 1) if v is not None else "", fmt=FMT_UA)
            # 仿真只在该模式实际测了这行时显示（防 25/25,24,23 交叠重复计入）
            sv = sim_val.get((key, mode)) if has_meas(key, mode, chip) else None
            _cell(ws, rr, c0 + n_t, rnd(sv, 1) if sv is not None else "", fmt=FMT_UA)
            if mark_fb and sv is not None and sim_fb.get((key, mode)):
                # 跨阶段补过的格子标成蓝色斜体。**默认不标**——总览是给别人 review 的，
                # 没解释的颜色只会让人停下来问；要自查用 summary --mark-fallback 单出一份。
                ws.cell(row=rr, column=c0 + n_t).font = Font(italic=True, color="1F4E79")
            dc = c0 + n_t + 1
            mv = meas_at_sim(key, mode, chip)   # 实测插值到仿真温度再比
            dv = dev_of(mv, sv)
            if dv is not None:
                sr = ref(rr, c0 + n_t)
                fcell(ws, rr, dc, f'=IF({sr}="","",({interp_expr(rr, c0)}-{sr})/{sr})',
                      rnd(dv, 4), fmt=FMT_PCT)
                if flag_red(dv, mv, sv, key):
                    ws.cell(row=rr, column=dc).font = red_font
            else:
                _cell(ws, rr, dc, "", fmt=FMT_PCT)
        # 备注只写「这个数不是你以为的那个数」（LDO 归并口径不可比、标签跨模式映射）。
        # 删掉的两条：编号重叠提示（"各模式列互不冲突"＝叫人别担心，反而招人担心）、
        # "仿真无对应项（标签未映射…）"（仿真格空着本身就是这个意思，还捎带 config 键名——
        # 那是写给出簿的人的）。口径都在说明页。
        _cell(ws, rr, note_col, "；".join(sorted(notes.get(key, ()))), align="left")

    # -- Σ合计（蓝）：LO 模块行（可与仿真对）+ 总合计行（含 DCO 等标签行，只有实测）
    _cell(ws, r_sum, 1, fill=C_SEP)
    _cell(ws, r_sum, 2, "Σ LO 模块合计" if n_label else "Σ 模块合计",
          bold=True, fill=C_SEP, align="left")
    _cell(ws, r_sum, 3, "µA", bold=True, fill=C_SEP)
    for gi, (mode, chip) in enumerate(col_keys):
        c0 = FIX + 1 + gi * grp_w
        for ti, t in enumerate(temps):
            col = c0 + ti
            v = col_sum(lo_keys, mode, chip, t)
            fcell(ws, r_sum, col, f"=SUM({ref(r_mod0, col)}:{ref(r_lo_last, col)})",
                  rnd(v, 1) if v is not None else None, bold=True, fill=C_SEP, fmt=FMT_UA)
        col = c0 + n_t
        sv = sim_col_sum(lo_keys, mode, chip)
        fcell(ws, r_sum, col, f"=SUM({ref(r_mod0, col)}:{ref(r_lo_last, col)})",
              rnd(sv, 1) if sv is not None else None, bold=True, fill=C_SEP, fmt=FMT_UA)
        dc = c0 + n_t + 1
        lo_mv = lo_meas_at_sim(mode, chip)
        dv = dev_of(lo_mv, sv)
        if dv is not None:
            sr = ref(r_sum, c0 + n_t)
            fcell(ws, r_sum, dc, f'=IF({sr}=0,"",({interp_expr(r_sum, c0)}-{sr})/{sr})',
                  rnd(dv, 4), bold=True, fill=C_SEP, fmt=FMT_PCT)
            if flag_red(dv, lo_mv, sv, None):
                ws.cell(row=r_sum, column=dc).font = red_font
        else:
            _cell(ws, r_sum, dc, "", bold=True, fill=C_SEP, fmt=FMT_PCT)
    # 行名已经写着「Σ LO 模块合计」/「Σ 总合计（含 DCO 等标签行）」，
    # 再补一句"不含下方标签行"「仿真未覆盖」是把同一件事说第二遍
    _cell(ws, r_sum, note_col, "", fill=C_SEP, align="left")
    if r_all:
        _cell(ws, r_all, 1, fill=C_SEP)
        _cell(ws, r_all, 2, "Σ 总合计（含 DCO 等标签行）", bold=True, fill=C_SEP, align="left")
        _cell(ws, r_all, 3, "µA", bold=True, fill=C_SEP)
        for gi, (mode, chip) in enumerate(col_keys):
            c0 = FIX + 1 + gi * grp_w
            for ti, t in enumerate(temps):
                col = c0 + ti
                v = col_sum(row_keys, mode, chip, t)
                fcell(ws, r_all, col, f"=SUM({ref(r_mod0, col)}:{ref(r_sum - 1, col)})",
                      rnd(v, 1) if v is not None else None, bold=True, fill=C_SEP, fmt=FMT_UA)
            _cell(ws, r_all, c0 + n_t, "", fill=C_SEP)
            _cell(ws, r_all, c0 + n_t + 1, "", fill=C_SEP)
        _cell(ws, r_all, note_col, "", fill=C_SEP, align="left")

    # 偏差红标改为写单元格时的双阈值静态染色（见 flag_red）——绝对偏差不在表内，
    # 条件格式无法做「|Δ%|>阈 且 |ΔµA|>阈」的联合判断，故用 Python 直接染红粗。
    ws.freeze_panes = ws.cell(row=r_freq, column=FIX + 1)

    # ================= 温度趋势 =================
    n_charts = 0
    if len([t for t in temps if t is not None]) >= 2:
        ws = wb.create_sheet("温度趋势")
        ws.sheet_view.showGridLines = False
        cur_row = 1

        def chart_block(title, row_names, values_of, unit):
            """左侧数据块 + 右侧折线图。values_of(name, temp) -> 数值或 None。"""
            nonlocal cur_row, n_charts
            _cell(ws, cur_row, 1, title, bold=True, size=12).border = Border()
            hdr = cur_row + 1
            _cell(ws, hdr, 1, "系列", bold=True, fill=C_HEADER)
            for ti, t in enumerate(temps):
                _cell(ws, hdr, 2 + ti, _t(t), bold=True, fill=C_HEADER)
            for ri, name in enumerate(row_names):
                _cell(ws, hdr + 1 + ri, 1, name, align="left")
                for ti, t in enumerate(temps):
                    v = values_of(name, t)
                    _cell(ws, hdr + 1 + ri, 2 + ti, rnd(v, 2) if v is not None else "")
            last = hdr + len(row_names)
            ch = LineChart()
            ch.title = _chart_title(title)
            ch.style = 12
            # 只有一行指标时这张图就是单系列，Excel 会"按点着色"并给每个点发一条
            # 图例（varyColors 省略时默认为真）。同一处兜底在 sweep_lib.blank_policy
            ch.varyColors = False
            ch.y_axis.title = _chart_title(unit, sz=1000, bold=False)
            ch.x_axis.title = _chart_title("温度", sz=1000, bold=False)
            ch.height, ch.width = max(7.5, 1.2 + 0.42 * len(row_names)), 16
            data = Reference(ws, min_col=1, min_row=hdr + 1, max_col=1 + n_t, max_row=last)
            ch.add_data(data, from_rows=True, titles_from_data=True)
            ch.set_categories(Reference(ws, min_col=2, min_row=hdr, max_col=1 + n_t))
            for s in ch.series:
                s.marker.symbol = "circle"
                s.marker.size = 5
                s.smooth = False
            ws.add_chart(ch, ref(cur_row + 1, n_t + 3))
            n_charts += 1
            cur_row = last + max(3, int(ch.height * 2) - len(row_names) - 1)

        ws.column_dimensions["A"].width = 38
        for ti in range(n_t):
            ws.column_dimensions[get_column_letter(2 + ti)].width = 10
        chart_block("各模式锁定后总电流 vs 温度 (mA)",
                    [col_title(m, c) for m, c in col_keys],
                    lambda name, t: next((base_ma.get((m, c, t)) for m, c in col_keys
                                          if col_title(m, c) == name), None), "mA")
        for mode, chip in col_keys:
            names = [k[1] for k in row_keys
                     if any((mode, chip, t) in matrix[k] for t in temps)]
            keyof = {k[1]: k for k in row_keys}
            chart_block(f"{col_title(mode, chip)} 各模块电流 vs 温度 (µA)", names,
                        lambda name, t, _m=mode, _c=chip, _ko=keyof:
                            matrix[_ko[name]].get((_m, _c, t)), "µA")

    # ================= 对比明细 =================
    ws = wb.create_sheet("对比明细")
    hdrs = ["模式", "芯片", "温度", "编号", "模块 (OFF 步)",
            "实测_µA", "仿真pre_µA", "仿真post_µA", "Δ_µA", "Δ%", "备注"]
    widths = [20, 7, 8, 9, 30, 11, 11, 11, 10, 8, 40]
    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        _cell(ws, 1, c, h, bold=True, fill=C_HEADER)
        ws.column_dimensions[get_column_letter(c)].width = w
    rr = 2
    for mode, chip in col_keys:
        for t in temps:
            k3 = (mode, chip, t)
            if k3 not in per_groups:
                continue
            for key in row_keys:
                if k3 not in matrix.get(key, {}):
                    continue
                disp, step_name = key
                _cell(ws, rr, 1, col_title(mode, chip), align="left")
                _cell(ws, rr, 2, chip)
                _cell(ws, rr, 3, t if t is not None else "")
                _cell(ws, rr, 4, disp)
                _cell(ws, rr, 5, step_name, align="left")
                mv = matrix[key][k3]
                _cell(ws, rr, 6, rnd(mv, 1), fmt=FMT_UA)
                pv, sv = sim_pre_val.get((key, mode)), sim_val.get((key, mode))
                _cell(ws, rr, 7, rnd(pv, 1) if pv is not None else "", fmt=FMT_UA)
                _cell(ws, rr, 8, rnd(sv, 1) if sv is not None else "", fmt=FMT_UA)
                diff = (mv - sv) if sv is not None else None
                fcell(ws, rr, 9, f'=IF(H{rr}="","",F{rr}-H{rr})',
                      rnd(diff, 1) if diff is not None else None, fmt=FMT_UA)
                dv = dev_of(mv, sv)
                # 仿真两阶段都是真 0 时 H 列写的就是 0（不是空）——不加 =0 这道闸，
                # 这一格在 Excel 里是 #DIV/0!，出现在给人 review 的簿子上
                fcell(ws, rr, 10, f'=IF(OR(H{rr}="",H{rr}=0),"",(F{rr}-H{rr})/H{rr})',
                      rnd(dv, 4) if dv is not None else None, fmt=FMT_PCT)
                if flag_red(dv, mv, sv, key):   # 双阈值 + 口径守卫，同总览
                    ws.cell(row=rr, column=10).font = red_font
                # 说明页承诺"跨阶段补值逐项见本页备注列"——这里就是那一项，缺了说明页就是假话
                row_notes = sorted(notes.get(key, ()))
                fb_ids = sim_fb.get((key, mode))
                if fb_ids:
                    row_notes.append(f"仿真 {','.join(str(x) for x in fb_ids)} "
                                     f"{stage_main} 为 0/缺项，取{other_stage}")
                _cell(ws, rr, 11, "；".join(row_notes), align="left")
                rr += 1
    if rr > 2:
        ws.auto_filter.ref = f"A1:K{rr - 1}"
    ws.freeze_panes = "A2"

    wb.save(out_path)
    _inject_cached_values(out_path, fcache)   # 给公式补缓存值，不重算的查看器也能显示
    return len(runs), len(row_keys), n_charts


def cmd_chips_export(conn, out_path, config, chips=None, audit=True):
    """跨芯片评审版汇总簿：一页总览（+ 隐藏的 _审计 页存口径）。

    与单芯片版（cmd_summary_export）的区别只在**摆法**，读数走同一个 summary_data：
      - 仿真列只出现一次（仿真与芯片无关，单芯片版每个模式×芯片块都重复一遍）
      - 一颗芯片一竖条（n_t 列），加片进来是加竖条不是加块，宽度线性且慢
      - 右侧直接给「片间极差」——多片测试要看的是一致性，横着比是人眼干的活
      - 模式做成行分区（band 行），不再占一根列轴；一行 = 一个模式的一个模块组，
        所以跨阶段补值的备注能落到行上，不需要再开一页明细
    """
    d = summary_data(conn, config, chips)
    temps, n_t, chip_ids, modes = d.temps, d.n_t, d.chips, d.modes
    matrix, row_keys, sim_val, sim_names = d.matrix, d.row_keys, d.sim_val, d.sim_names
    n_chip = len(chip_ids)
    lo_keys = set(row_keys[:len(row_keys) - d.n_label])   # 标签行(DCO)排在最后 n_label 个

    def has(key, mode, chip):
        return any((mode, chip, t) in matrix[key] for t in temps)

    rows_of_mode = {m: [k for k in row_keys if any(has(k, m, c) for c in chip_ids)]
                    for m in modes}

    # ---- 列几何。★组与组之间插一条窄灰竖栏（宽 2 的空列），不是为了好看：
    #      40 列宽的表里最先丢的信息是「我现在看的是哪一颗芯片」，一条实心竖栏比给整块
    #      上底色更省视觉预算（Gestalt common region），灰度打印下照样成立。
    #      编号/模块/单位 ▏仿真 ▏片1 ▏片2 … ▏[极差 极差%] 均值 偏差% ▏备注
    #      ★仿真值、各片均值、偏差% 是**一件事**（仿测对比），必须挨在一起：把仿真放最左、
    #      偏差甩到最右，等于让人横跨整张表去对一个比值。片间极差反而该贴着实测区右边——
    #      它是从那几竖条算出来的。
    FIX = 3
    C_R0 = FIX + 1
    C_SIM, C_MEAN, C_DEV = C_R0 + 1, C_R0 + 2, C_R0 + 3        # 仿测对比组
    C_R1 = C_DEV + 1
    C_CHIP = C_R1 + 1
    CHIP_W = n_t + 1                         # 每片：n_t 个温度列 + 1 条竖栏
    after = C_CHIP + n_chip * CHIP_W         # 最后一片的竖栏之后
    spread = n_chip > 1                      # 只有一颗芯片时「片间极差」是废列，不出
    C_SPREAD, C_SPCT = (after, after + 1) if spread else (None, None)
    C_NOTE = (after + 3) if spread else after
    C_R2 = (after + 2) if spread else None   # 片间组与备注之间的竖栏
    SUM_C0, SUM_C1 = C_SIM, C_DEV            # 加粗框住的是仿测对比组（判断对着它做）

    def cc(chip_i, ti):
        return C_CHIP + chip_i * CHIP_W + ti

    rails = [c for c in (C_R0, C_R1, C_R2) if c] + \
            [C_CHIP + j * CHIP_W + n_t for j in range(n_chip)]

    def paint_rails(ws_, r):
        """把竖栏画到这一行上。band 行例外——它要横贯整表，不留缝。"""
        for c in rails:
            _cell(ws_, r, c, fill=C_RAIL)

    wb = openpyxl.Workbook()
    fcache = {}

    def fcell(ws_, r, c, formula, value, **kw):
        _cell(ws_, r, c, formula, **kw)
        fcache.setdefault(ws_.title, {})[f"{get_column_letter(c)}{r}"] = value
        return value

    def ref(r, c):
        return f"{get_column_letter(c)}{r}"

    # ================= 总览（这本簿唯一可见的页） =================
    # ★ 不做"说明/读法"页：簿子是给评审看的，评审要的是数。教人怎么读表、
    #   阈值还没定这类话是写给操作者的，一律出簿——口径进隐藏的 _审计 页，
    #   给操作者的话回控制台。（同一条规矩在 summarize_chips 已经踩过一次）
    ws = wb.active
    ws.title = "总览"
    ws.sheet_view.showGridLines = False
    for r in (1, 2):
        for c in range(1, C_NOTE + 1):
            _cell(ws, r, c, fill=C_HEADER)

    def head_single(c, title, width):
        ws.merge_cells(start_row=1, start_column=c, end_row=2, end_column=c)
        _cell(ws, 1, c, title, bold=True, fill=C_HEADER)
        ws.column_dimensions[get_column_letter(c)].width = width

    def head_group(c0, c1, title, sub, widths, fill=C_HEADER):
        """两行表头：上行组名（跨列合并），下行分轴名。表头本身也按组切开，
        眼睛先看到几个组、再看组里几列——比一排 15 个平铺的列名好数。"""
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        _cell(ws, 1, c0, title, bold=True, fill=C_HEADER)
        for k, (s, w) in enumerate(zip(sub, widths)):
            _cell(ws, 2, c0 + k, s, bold=True, fill=fill, size=9)
            ws.column_dimensions[get_column_letter(c0 + k)].width = w

    for c, (title, w) in enumerate(zip(["编号", "模块 (OFF 步)", "单位"], [9, 30, 6]), 1):
        head_single(c, title, w)
    # 括号里是定义不是说明——列名自带口径，评审就不用去翻别的页
    head_group(C_SIM, C_DEV, "仿测对比",
               [f"仿真\n{d.sim_note} {d.tier}", "各片均值\n@%g℃" % d.sim_temp_c, "偏差%"],
               [11, 11, 10])
    for j, chip in enumerate(chip_ids):
        head_group(cc(j, 0), cc(j, n_t - 1), chip,
                   [_t(t) if t is not None else "?" for t in temps], [10] * n_t)
    if spread:
        head_group(C_SPREAD, C_SPCT, "片间一致性",
                   ["极差 µA\n(全温最坏)", "极差 %\n(÷常温)"], [11, 10])
    head_single(C_NOTE, "备注", 34)
    for c in rails:
        ws.column_dimensions[get_column_letter(c)].width = 2
    for r in (1, 2):
        paint_rails(ws, r)
        for c in range(1, C_NOTE + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = ref(3, C_R0)

    red_font = Font(name=FONT_NAME, size=10, bold=True, color=C_FLAG)
    i25 = temps.index(_closest(temps, 25)) if temps[0] is not None else 0

    def interp_expr(rr, j):
        """该片这一行插值到 sim_temp_c 的表达式（引用温度列，随实测联动）。"""
        pts = sorted((t, ti) for ti, t in enumerate(temps) if t is not None)
        if not pts:
            return ref(rr, cc(j, 0))
        if len(pts) == 1 or d.sim_temp_c <= pts[0][0]:
            return ref(rr, cc(j, pts[0][1]))
        if d.sim_temp_c >= pts[-1][0]:
            return ref(rr, cc(j, pts[-1][1]))
        for i in range(1, len(pts)):
            (t0, i0), (t1, i1) = pts[i - 1], pts[i]
            if t0 <= d.sim_temp_c <= t1:
                w = (d.sim_temp_c - t0) / (t1 - t0)
                a, b = ref(rr, cc(j, i0)), ref(rr, cc(j, i1))
                return f"({a}+{w:g}*({b}-{a}))"
        return ref(rr, cc(j, pts[-1][1]))

    rr, guides = 3, []
    for mode in modes:
        # -- band 行：模式起点。**横贯整表、盖掉竖栏**——竖栏是分"片"的，
        #    band 是分"模式"的，两者交叉会把表切成棋盘格，反而没有层次
        for c in range(1, C_NOTE + 1):
            _cell(ws, rr, c, fill=C_SEP)
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=FIX)
        _cell(ws, rr, 1, mode, bold=True, fill=C_SEP, align="left", size=11)
        ws.merge_cells(start_row=rr, start_column=C_R0, end_row=rr, end_column=C_NOTE)
        freq = _mode_freq(mode, config)
        _cell(ws, rr, C_R0, f"测试频率 {freq}" if freq else "", fill=C_SEP, align="left")
        ws.row_dimensions[rr].height = 18
        rr += 1

        # -- 条件行：锁定后总电流 / 全关残留（mA，不与仿真对比）
        for name, src in (("锁定后总电流", d.base_ma), ("全关残留电流", d.end_ma)):
            for c in range(1, C_NOTE + 1):
                _cell(ws, rr, c, fill=C_SETTING)
            paint_rails(ws, rr)
            _cell(ws, rr, 2, name, fill=C_SETTING, align="left")
            _cell(ws, rr, 3, "mA", fill=C_SETTING)
            present = []
            for j, chip in enumerate(chip_ids):
                got = False
                for ti, t in enumerate(temps):
                    v = src.get((mode, chip, t))
                    _cell(ws, rr, cc(j, ti), rnd(v, 3) if v is not None else "",
                          fill=C_SETTING, fmt=FMT_MA)
                    got = got or v is not None
                if got:
                    present.append(j)
            _spread_cells(ws, fcell, rr, present, cc, n_t,
                          getval=lambda j, ti, _m=mode, _s=src: _s.get((_m, chip_ids[j], temps[ti])),
                          C_SPREAD=C_SPREAD, C_SPCT=C_SPCT, i25=i25, spread=spread,
                          fill=C_SETTING, fmt=FMT_MA, ref=ref)
            # 备注留空：「基线=末个 Lock_step」这类是口径不是发现，写在正表上
            # 只会让评审停下来读一句他本来不需要知道的话（口径在隐藏的 _审计 页）
            _cell(ws, rr, C_NOTE, "", fill=C_SETTING, align="left")
            rr += 1

        # -- 结果行：每个模块组一行
        mode_rows = rows_of_mode[mode]
        first_row = rr
        lo_rows, all_rows = [], []
        for n_in_block, key in enumerate(mode_rows):
            disp, step_name = key
            paint_rails(ws, rr)
            # 两个派生组（仿测对比、片间一致性）浅蓝，实测竖条留白：
            # 底色区分的是"算出来的"和"测出来的"，不是区分组
            for c in list(range(SUM_C0, SUM_C1 + 1)) + ([C_SPREAD, C_SPCT] if spread else []):
                _cell(ws, rr, c, fill=C_SUM)
            _cell(ws, rr, 1, disp)
            _cell(ws, rr, 2, sim_names.get(key) or step_name, align="left")
            _cell(ws, rr, 3, "µA")
            sv = sim_val.get((key, mode))
            _cell(ws, rr, C_SIM, rnd(sv, 1) if sv is not None else "", fmt=FMT_UA)
            present, vals = [], {}
            for j, chip in enumerate(chip_ids):
                got = False
                for ti, t in enumerate(temps):
                    v = matrix[key].get((mode, chip, t))
                    vals[(j, ti)] = v
                    _cell(ws, rr, cc(j, ti), rnd(v, 1) if v is not None else "", fmt=FMT_UA)
                    got = got or v is not None
                if got:
                    present.append(j)
            _spread_cells(ws, fcell, rr, present, cc, n_t,
                          getval=lambda j, ti, _v=vals: _v.get((j, ti)),
                          C_SPREAD=C_SPREAD, C_SPCT=C_SPCT, i25=i25, spread=spread,
                          fill=None, fmt=FMT_UA, ref=ref)
            # 均值@sim_temp / 偏差%
            mv = _mean_at(d, [[vals.get((j, ti)) for ti in range(n_t)] for j in present])
            if present:
                expr = "+".join(interp_expr(rr, j) for j in present)
                guard = ",".join(ref(rr, cc(j, ti)) for j in present for ti in range(n_t))
                fcell(ws, rr, C_MEAN,
                      f"=IF(COUNT({guard})<{len(present) * n_t},\"\",({expr})/{len(present)})",
                      rnd(mv, 1), fmt=FMT_UA)
            else:
                _cell(ws, rr, C_MEAN, "", fmt=FMT_UA)
            dv = (mv - sv) / sv if (mv is not None and sv not in (None, 0)) else None
            m_ref, s_ref = ref(rr, C_MEAN), ref(rr, C_SIM)
            fcell(ws, rr, C_DEV,
                  f'=IF(OR({s_ref}="",{s_ref}=0,{m_ref}=""),"",({m_ref}-{s_ref})/{s_ref})',
                  rnd(dv, 4) if dv is not None else None, fmt=FMT_PCT)
            if (dv is not None and key not in d.caliber_keys
                    and abs(dv) > d.thr and abs(mv - sv) > d.abs_thr):
                ws.cell(row=rr, column=C_DEV).font = red_font
            # 备注只写「这个数不是你以为的那个数」——LDO 归并导致口径不可比、
            # 仿真值取自另一个仿真阶段。空仿真格本身就说明没有仿真值，不必再写一句
            # 「标签行，仿真未映射」；解释性的话一律不进正表。
            note = sorted(d.notes.get(key, ()))
            fb = d.sim_fb.get((key, mode))
            if fb:
                note.append(f"仿真 {','.join(str(x) for x in fb)} 取自{d.other_stage}仿")
            _cell(ws, rr, C_NOTE, "；".join(note), align="left")
            all_rows.append(rr)
            if key in lo_keys:
                lo_rows.append(rr)
            # 每 4 个模块行一条横导引线（末行不画——紧接着就是 Σ，它自带边界）
            if (n_in_block + 1) % 4 == 0 and n_in_block + 1 < len(mode_rows):
                guides.append(rr)
            rr += 1

        # -- 段末 Σ
        for title, rws, with_sim in (("Σ LO 模块合计", lo_rows, True),
                                     ("Σ 总合计（含标签行）", all_rows, False)):
            if not rws:
                continue
            for c in range(1, C_NOTE + 1):
                _cell(ws, rr, c, fill=C_SUM)
            paint_rails(ws, rr)
            _cell(ws, rr, 2, title, bold=True, fill=C_SUM, align="left")
            _cell(ws, rr, 3, "µA", fill=C_SUM)
            sum_of = lambda c: (f"=SUM({ref(rws[0], c)}:{ref(rws[-1], c)})"
                                if rws == list(range(rws[0], rws[-1] + 1))
                                else "=" + "+".join(ref(x, c) for x in rws))
            sv_sum = (sum(v for v in (sim_val.get((k, mode)) for k in mode_rows
                                      if k in lo_keys) if v is not None)
                      if with_sim else None)
            if with_sim:
                fcell(ws, rr, C_SIM, sum_of(C_SIM), rnd(sv_sum, 1), fill=C_SUM, fmt=FMT_UA)
            else:
                _cell(ws, rr, C_SIM, "", fill=C_SUM)
            present, sums = [], {}
            for j, chip in enumerate(chip_ids):
                got = False
                for ti, t in enumerate(temps):
                    vals = [matrix[k].get((mode, chip, t)) for k in mode_rows
                            if (k in lo_keys or not with_sim)]
                    vals = [v for v in vals if v is not None]
                    sums[(j, ti)] = sum(vals) if vals else None
                    fcell(ws, rr, cc(j, ti), sum_of(cc(j, ti)),
                          rnd(sum(vals), 1) if vals else None, fill=C_SUM, fmt=FMT_UA)
                    got = got or bool(vals)
                if got:
                    present.append(j)
            # Σ 行的片间极差 = 这颗片整段加起来比别人高/低多少，是一致性最直接的一个数
            _spread_cells(ws, fcell, rr, present, cc, n_t,
                          getval=lambda j, ti, _s=sums: _s.get((j, ti)),
                          C_SPREAD=C_SPREAD, C_SPCT=C_SPCT, i25=i25, spread=spread,
                          fill=C_SUM, fmt=FMT_UA, ref=ref)
            if with_sim and present:
                expr = "+".join(interp_expr(rr, j) for j in present)
                mv = _mean_at(d, [[
                    sum(v for v in (matrix[k].get((mode, chip_ids[j], temps[ti]))
                                    for k in mode_rows if k in lo_keys) if v is not None)
                    for ti in range(n_t)] for j in present])
                fcell(ws, rr, C_MEAN, f"=({expr})/{len(present)}", rnd(mv, 1),
                      fill=C_SUM, fmt=FMT_UA)
                dv = (mv - sv_sum) / sv_sum if (mv is not None and sv_sum) else None
                m_ref, s_ref = ref(rr, C_MEAN), ref(rr, C_SIM)
                fcell(ws, rr, C_DEV,
                      f'=IF(OR({s_ref}="",{s_ref}=0,{m_ref}=""),"",({m_ref}-{s_ref})/{s_ref})',
                      rnd(dv, 4) if dv is not None else None, fill=C_SUM, fmt=FMT_PCT)
                if dv is not None and abs(dv) > d.thr and abs(mv - sv_sum) > d.abs_thr:
                    ws.cell(row=rr, column=C_DEV).font = red_font
                _cell(ws, rr, C_NOTE, "", fill=C_SUM, align="left")
            else:
                # 行名已经写着「Σ 总合计（含标签行）」，仿真格空着——再补一句
                # 「仿真未覆盖，无对比」是把同一件事说第二遍
                _cell(ws, rr, C_MEAN, "", fill=C_SUM)
                _cell(ws, rr, C_DEV, "", fill=C_SUM)
                _cell(ws, rr, C_NOTE, "", fill=C_SUM, align="left")
            rr += 1
        rr += 1     # 段间空行

    # 边框最后统一加——_cell 每次写都会重置 border，先画会被后面的写覆盖掉
    _vedges(ws, 1, rr, SUM_C0, SUM_C1)
    for g in guides:
        _hguide(ws, g, 1, C_NOTE)

    # ================= _审计（隐藏页：口径与来源，不是读法） =================
    # 要追口径的人取消隐藏就能看到，评审打开簿子看到的仍然只有数。
    if audit:
        aw = wb.create_sheet("_审计")
        aw.sheet_state = "hidden"
        aw["A1"] = "口径与来源"
        aw["A1"].font = Font(name=FONT_NAME, bold=True, size=12)
        one_temp = _t(_closest(temps, 25)) if temps[0] is not None else "?"
        alines = [
            f"导出 {now_iso()}  current_db.py summary-chips  数据源 current.db",
            f"芯片 {n_chip} 颗：{'、'.join(chip_ids)}",
            f"源文件：{'、'.join(sorted(d.src_files))}",
            f"模式 {len(modes)} 个；温度点 {'、'.join(_t(t) for t in temps if t is not None) or '未知'}；"
            f"实测 {len(d.runs)} 个 run；同一 (模式,芯片,温度) 重复测取 run_ts 最新一次",
            f"仿真：{d.tier or '未过滤'} / {d.stage_main}-sim / {d.sim_note or '温度未标注'}；"
            "与芯片无关，各片共用同一列",
            (f"跨阶段补值：{d.stage_main} 为 0 或 ≤{d.zero_ua:g}µA 计作缺项时取 {d.other_stage}"
             f"（前仿已做 back annotate）；两阶段都缺则留空。共 {len(d.sim_fb)} 处，"
             "逐处写在该行备注列" if (d.fb_stage and d.sim_fb) else "仿真列全部取自同一阶段，无补值"),
            "基线：每模式段第一个 OFF 行之前最后一行（末个 Lock_step）；模块电流 = 上一行 − 本行",
            "LDO 归并 "
            + ("、".join(f"{k}→{v}" for k, v in sorted((config.get("ldo_reparent") or {}).items()))
               or "无")
            + "：子模块实测并入父组、仿真侧不计子模块 -> 父组口径不可比，偏差保留但不标红",
            f"偏差% = (各片均值线性插值到 {'%g' % d.sim_temp_c}℃ − 仿真) ÷ 仿真",
            f"标红：|偏差%| > {d.thr * 100:.0f}% 且 |绝对偏差| > {d.abs_thr:.0f}µA（双阈值同时满足）",
        ]
        if spread:
            alines += [
                f"片间极差：同一温度下各片 max−min，取全温最大者；极差% = 极差 ÷ 常温({one_temp})各片均值",
                "某片该组未测时极差留空（覆盖度不同的片之间不做减法）；片间一致性无 spec，不设标色判据",
            ]
        alines += [
            "Σ LO 模块合计不含标签行（DCO 等），口径与仿真一致；Σ 总合计含标签行、仿真未覆盖",
            "极差 / 均值 / 偏差 / Σ 为 Excel 公式，并已写入当前结果的缓存值",
        ]
        for i, line in enumerate(alines, 3):
            aw.cell(row=i, column=1, value=line).font = Font(name=FONT_NAME, size=10)
        aw.column_dimensions["A"].width = 110

    wb.save(out_path)
    _inject_cached_values(out_path, fcache)
    return {"runs": len(d.runs), "modes": len(modes), "chips": chip_ids,
            "n_fb": len(d.sim_fb), "spread": spread, "cols": C_NOTE}


_MEDIUM = Side(style="medium", color="FF000000")


def _vedges(ws, r0, r1, c_first, c_last):
    """给一个列组的左右两侧加中等粗细竖边框。
    汇总组是全表唯一要「跳出来」的（判断就是对着它做的）。只靠底色不够——
    灰度打印和色弱下几种浅色会撞，边框不会。"""
    for r in range(r0, r1 + 1):
        for c, which in ((c_first, "l"), (c_last, "r")):
            b = ws.cell(row=r, column=c).border
            ws.cell(row=r, column=c).border = Border(
                left=_MEDIUM if which == "l" else b.left,
                right=_MEDIUM if which == "r" else b.right, top=b.top, bottom=b.bottom)


def _hguide(ws, r, c0, c1):
    """一条横向导引线（中粗下边框）。
    一眼能数清的上限是 4~5 行，再多横着扫就会串行。每 4 个模块行给一条线，
    不上斑马底色——竖向已经分区了，再加行底色就成了网格噪声。"""
    for c in range(c0, c1 + 1):
        b = ws.cell(row=r, column=c).border
        ws.cell(row=r, column=c).border = Border(left=b.left, right=b.right,
                                                 top=b.top, bottom=_MEDIUM)


def _closest(temps, target):
    """最接近 target 的温度点（常温列用它定位，别写死 25）。"""
    pts = [t for t in temps if t is not None]
    return min(pts, key=lambda t: abs(t - target)) if pts else None


def _mean_at(d, per_chip_series):
    """各片先各自插值到 sim_temp_c，再取平均；有片缺值就返回 None。"""
    vals = []
    for series in per_chip_series:
        v = d.interp_to(list(zip(d.temps, series)), d.sim_temp_c)
        if v is None:
            return None
        vals.append(v)
    return sum(vals) / len(vals) if vals else None


def _spread_cells(ws, fcell, rr, present, cc, n_t, getval,
                  C_SPREAD, C_SPCT, i25, spread, fill, fmt, ref):
    """片间极差两列：同温度 max−min 取三温最大；极差% = 极差 ÷ 常温各片均值。
    少于两片有数就留空——覆盖度不同的片之间做减法没有意义。"""
    if not spread:
        return
    if len(present) < 2:
        _cell(ws, rr, C_SPREAD, "", fill=fill, fmt=fmt)
        _cell(ws, rr, C_SPCT, "", fill=fill)
        return
    per_t_f, per_t_v = [], []
    for ti in range(n_t):
        cells = [ref(rr, cc(j, ti)) for j in present]
        per_t_f.append(f"MAX({','.join(cells)})-MIN({','.join(cells)})")
        vs = [getval(j, ti) for j in present]
        per_t_v.append(max(vs) - min(vs) if all(v is not None for v in vs) else None)
    guard = ",".join(ref(rr, cc(j, ti)) for j in present for ti in range(n_t))
    got = [v for v in per_t_v if v is not None]
    fcell(ws, rr, C_SPREAD,
          f'=IF(COUNT({guard})<{len(present) * n_t},"",MAX({",".join(per_t_f)}))',
          rnd(max(got), 3 if fmt == FMT_MA else 1) if len(got) == n_t else None,
          fill=fill, fmt=fmt)
    base_cells = ",".join(ref(rr, cc(j, i25)) for j in present)
    base_vals = [getval(j, i25) for j in present]
    base = (sum(base_vals) / len(base_vals)
            if all(v is not None for v in base_vals) else None)
    pct = (max(got) / base) if (len(got) == n_t and base) else None
    fcell(ws, rr, C_SPCT,
          f'=IF(OR({ref(rr, C_SPREAD)}="",AVERAGE({base_cells})=0),"",'
          f'{ref(rr, C_SPREAD)}/AVERAGE({base_cells}))',
          rnd(pct, 4) if pct is not None else None, fill=fill, fmt=FMT_PCT)


def db_chip_roster(conn):
    """库里现有哪些芯片、各几个 run、数据从哪个文件来。"""
    return conn.execute(
        "SELECT chip, COUNT(*), COUNT(DISTINCT src_file), MIN(run_ts), MAX(run_ts)"
        " FROM runs GROUP BY chip ORDER BY chip").fetchall()


def db_chip_conflicts(conn):
    """同一个 (芯片,模式,温度) 被多个源文件占着 -> 出簿时只留 run_ts 最新的那次，
    另一次静默消失。**两颗芯片被贴成同一个名字**就长这样（真实事故），
    而老版式一模式一文件、同芯片多源是正常的，不能只看源文件数。"""
    return dict(conn.execute(
        "SELECT chip, COUNT(*) FROM (SELECT chip, mode, temp_c FROM runs"
        "   GROUP BY chip, mode, temp_c HAVING COUNT(DISTINCT src_file) > 1)"
        " GROUP BY chip").fetchall())


def print_chip_roster(conn, title="库里现有"):
    rows = db_chip_roster(conn)
    if not rows:
        print(f"[{title}] 一个 run 都没有")
        return rows
    conflicts = db_chip_conflicts(conn)
    print(f"[{title}] 芯片 {len(rows)} 颗：")
    for chip, n_run, n_src, ts0, ts1 in rows:
        span = ts0 if ts0 == ts1 else f"{ts0} ~ {ts1}"
        n_bad = conflicts.get(chip, 0)
        warn = (f"   ⚠ {n_bad} 个(模式,温度)被多个文件占着，出簿只留最新的一次"
                "——两颗芯片贴成同一个名字就长这样" if n_bad else "")
        print(f"    {chip:<12} {n_run:>3} 个 run   {n_src} 个源文件   {span}{warn}")
    return rows


def cmd_add_chip(args):
    """往已有库里加一颗芯片的实测，不动仿真、不动别的芯片。

    build 是全量重建（库=目录的纯函数，没有"库里有、目录里没有"的幽灵数据），
    代价是每次都要重读那本几千行的仿真簿。新到一颗芯片只想把它并进来时用这个。
    同一个文件重复 add 是幂等的——按 (源文件, 芯片) 先删后插。"""
    if not os.path.exists(args.db):
        raise SystemExit(f"[错误] 库不存在: {args.db}（第一次请用 build 建库，仿真表要先进去）")
    root = os.path.dirname(os.path.abspath(args.db))
    config, cfg_path, cfg_created = load_config(root, args.config)
    conn = open_db(args.db)
    sim_modes = {r[0] for r in conn.execute("SELECT DISTINCT mode FROM sim_current")}
    if not sim_modes:
        print("[警告] 库里没有仿真数据（仿真列会全空）——先跑一次 build，或用 ingest-sim 补")

    xlsx = os.path.abspath(args.xlsx)
    if not os.path.isfile(xlsx):
        raise SystemExit(f"[错误] 文件不存在: {xlsx}")
    # 芯片号：--chip 优先，否则取文件所在目录名（与 build 同一条规则：
    # 目录名对得上某个仿真 Mode 就说明这是老的"按模式分目录"版式，此时必须显式给 --chip）
    chip, folder_mode = chip_of_path(xlsx, os.path.dirname(os.path.dirname(xlsx)),
                                     sim_modes, config.get("mode_map"), None)
    if args.chip:
        chip, folder_mode = args.chip, folder_mode
    if not chip:
        raise SystemExit("[错误] 认不出芯片号：文件所在目录名对上了一个仿真 Mode"
                         f"（{folder_mode}），这是按模式分目录的老版式 -> 请显式给 --chip")

    before = {c for c, *_ in db_chip_roster(conn)}
    results = ingest_result_file(conn, xlsx, chip, config, sim_modes, folder_mode=folder_mode)
    print(f"[实测] {os.path.basename(xlsx)}  芯片 {chip} -> {len(results)} 个 (模式,温度) 段")
    bad_map = set()
    for run_id, mode, mode_raw, how, temp, n_steps in results:
        print(f"        {mode:<22} run#{run_id}  {n_steps} 个模块组  "
              f"{temp if temp is not None else '?'}°C  {how}")
        if how in ("none", "ambig"):
            bad_map.add((mode_raw, how))
    if bad_map:
        # 用的是 db 同目录的 config（除非 --config）。db 和 config 不在一起时 mode_map
        # 读不到，同一个文件在 build 和 add-chip 下会落到不同的模式名上——不喊出来看不见
        print(f"[警告] {len(bad_map)} 个段没匹配上仿真 Mode: "
              + "、".join(f"{m}({h})" for m, h in sorted(bad_map)))
        print(f"       配置用的是 {cfg_path}"
              + ("（本次现建的默认配置，mode_map 是空的——db 和 config 不在同一个目录？）"
                 if cfg_created else " -> 确认 mode_map 是否覆盖了这些段"))
    print_chip_roster(conn)
    if chip in before:
        print(f"[提示] {chip} 原本就在库里：同源文件的旧 run 已被本次替换（幂等）")
    conn.close()
    print("[下一步] 出簿："
          f"python current_db.py summary-chips --db {args.db} --out <输出.xlsx>")


def cmd_chips(args):
    root = os.path.dirname(os.path.abspath(args.db))
    config, _, _ = load_config(root, args.config)
    conn = open_db(args.db)
    chips = None
    if args.chip:
        chips = [c.strip() for arg in args.chip for c in str(arg).split(",") if c.strip()]
    r = cmd_chips_export(conn, args.out, config, chips=chips, audit=not args.no_audit)
    conn.close()
    print(f"[完成] 跨芯片汇总簿: {args.out}")
    print(f"       {len(r['chips'])} 颗芯片（{'、'.join(r['chips'])}）× {r['modes']} 个模式，"
          f"{r['runs']} 个 run，总览 {r['cols']} 列")
    # 下面这些是给出簿的人看的，**所以只在控制台**：簿子是给评审的，
    # 里面出现"阈值还没定"这类话，评审会停下来问这行字是干嘛的
    if r["n_fb"]:
        print(f"       仿真跨阶段补值 {r['n_fb']} 处（逐处写在总览备注列；口径在隐藏的 _审计 页）")
    if r["spread"]:
        print("       片间极差未设标色判据——片间一致性没有 spec，定了阈值再加")
    print("       正簿只有「总览」一页；口径与来源在隐藏页 _审计（--no-audit 可不出）")


def cmd_summary(args):
    root = os.path.dirname(os.path.abspath(args.db))
    config, _, _ = load_config(root, args.config)
    conn = open_db(args.db)
    chips = None
    if args.chip:
        chips = [c.strip() for arg in args.chip for c in str(arg).split(",") if c.strip()]
    n_runs, n_rows, n_charts = cmd_summary_export(conn, args.out, config,
                                                  mark_fb=args.mark_fallback, chips=chips)
    conn.close()
    print(f"[完成] 功耗汇总簿: {args.out}（{n_runs} 个 run，矩阵 {n_rows} 行，{n_charts} 张趋势图）")


# ---------------------------------------------------------------- 命令

def walk_xlsx(root, skip_dirs, exclude_globs=()):
    """递归列出 root 下所有 xlsx（跳过 skip_dirs 子树、~$ 锁文件和排除名单）。"""
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in skip_dirs]
        for f in sorted(filenames):
            if not f.lower().endswith((".xlsx", ".xlsm")) or f.startswith("~$"):
                continue
            if any(fnmatch.fnmatch(f, pat) for pat in exclude_globs):
                continue
            yield os.path.join(dirpath, f)


def result_globs(config):
    """result_glob 允许写成字符串或数组。

    ★ 实测文件的命名已经换过两次词序（`*_all_mode_Current_*` -> `*_Current_all_mode_*`），
      一条通配符钉不住。给数组，加一条就完事，别改代码。"""
    g = (config or {}).get("result_glob", "Result*.xlsx")
    return [g] if isinstance(g, str) else [str(x) for x in g]


def is_result_file(name, globs):
    return any(fnmatch.fnmatch(name, p) for p in globs)


def report_skipped(files, globs, root, sim_wb=None):
    """扫到但没匹配 result_glob 的 xlsx —— 必须报出来。

    ★ 真实事故：新芯片的文件名换了词序，通配符匹配不上，build 静默跳过，
      簿子上少一颗芯片而控制台干干净净。"跳过"是要出声的事。"""
    sim = os.path.abspath(sim_wb) if sim_wb else None
    odd = [f for f in files
           if not is_result_file(os.path.basename(f), globs) and os.path.abspath(f) != sim]
    if not odd:
        return odd
    print(f"[跳过] {len(odd)} 个 xlsx 不匹配 result_glob {globs}（不是实测文件就不用管；"
          "是的话把它的命名加进 config.result_glob 数组）:")
    for f in odd[:8]:
        print(f"    {os.path.relpath(f, root)}")
    if len(odd) > 8:
        print(f"    …还有 {len(odd) - 8} 个")
    return odd


def chip_of_path(path, root, sim_modes, mode_map, fallback):
    """实测文件 -> (芯片号, 当模式名用的文件夹名)。

    两种目录版式并存，靠「文件夹名对不对得上一个仿真 Mode」区分，不另加开关：
      <root>/<模式>/Result*.xlsx    老版式：文件夹名=模式名（单模式文件靠它兜底/纠错）
      <root>/<芯片号>/*.xlsx        新版式：文件夹名=芯片号（全模式单文件自带模式标签）
    判错的代价不对称：把芯片号误当模式名，只是让段标签兜底失效（多段文件根本不看它）；
    把模式名误当芯片号，会让老数据整批改名。所以只在「对不上任何仿真 Mode」时才认作芯片号。
    root 直放的文件没有文件夹可依，用 fallback（--chip，默认 C1）。"""
    rel = os.path.relpath(os.path.dirname(path), root)
    top = "" if rel in ("", ".", os.curdir) else rel.replace("\\", "/").split("/")[0]
    if not top:
        return fallback, None
    _mode, how = resolve_mode(top, sim_modes, mode_map)
    if how in ("config", "auto"):
        return fallback, top
    return top, None


def find_sim_candidates(root, config, verbose=False):
    """扫 root 下所有工作簿，返回**全部**疑似仿真长表 [(mtime, path, sheet, how)]，
    how ∈ name/header。刻意不 break——同一个簿里可能有多个页命中，inspect 要把落选的
    也列出来给人看（选错仿真快照是静默事故，已经发生过一次）。
    排序：修改时间新→旧；同一簿内按名命中优先。"""
    sim_sheet = norm(config.get("sim_sheet", "Current_data"))
    skip = set(config.get("skip_dirs") or [])
    excl = list(config.get("exclude_globs") or [])
    globs = result_globs(config)
    out = []
    for path in walk_xlsx(root, skip, excl):
        # 实测文件不会是仿真长表；跳过省一次开簿（openpyxl 开一个大 xlsm 很贵，
        # 且这些文件稍后还要再开一次读数据）
        if is_result_file(os.path.basename(path), globs):
            continue
        if verbose:
            print(f"  [扫描] {os.path.relpath(path, root)}", flush=True)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            mt = os.path.getmtime(path)
            # 先只比 sheet 名（不碰单元格）——名字命中就定案，不再逐页解析表头。
            # 真簿有 30 个页，逐页扫前 30 行是这条命令最慢的一步。
            hit = [sn for sn in wb.sheetnames if norm(sn) == sim_sheet]
            if hit:
                out.append((mt, path, hit[0], "按名"))
            else:
                for sn in wb.sheetnames:
                    try:
                        by_hdr = any(match_sim_header(r) for r in
                                     wb[sn].iter_rows(min_row=1, max_row=30, values_only=True))
                    except Exception:
                        by_hdr = False
                    if by_hdr:
                        out.append((mt, path, sn, "按表头"))
        except Exception:
            pass
        finally:
            wb.close()
    out.sort(key=lambda c: (-c[0], c[1], c[3] != "按名"))
    return out


def find_sim_workbook(root, config):
    """返回 (工作簿路径, tab名) 或 (None, None)。不认文件名，按表头内容全递归扫。"""
    if config.get("sim_workbook"):
        p = config["sim_workbook"]
        return (p if os.path.isabs(p) else os.path.join(root, p)), config.get("sim_sheet")
    cands = find_sim_candidates(root, config)
    if not cands:
        return None, None
    best = cands[0]
    others = [c for c in cands if (c[1], c[2]) != (best[1], best[2])]
    if others:
        print("[提示] 发现多个疑似仿真长表，取最新修改的；其余可在 config.sim_workbook 里指定：")
        for _mt, p, sn, how in others:
            print(f"       {os.path.relpath(p, root)} / {sn} ({how})")
    return best[1], best[2]


def cmd_build(args):
    root = os.path.abspath(args.root)
    if not os.path.isdir(root):
        raise SystemExit(f"[错误] 根目录不存在: {root}")
    config, cfg_path, created = load_config(root, args.config)
    if created:
        print(f"[提示] 首次运行，已生成配置 {cfg_path}（LDO 归并/模式映射/标签映射都在里面改）")

    db_path = args.db or os.path.join(root, "current.db")
    if os.path.exists(db_path):
        os.remove(db_path)  # build = 全量重建；增量请用 ingest-* 子命令
    conn = open_db(db_path)

    if args.sim:
        sim_wb, sim_sheet = args.sim, config.get("sim_sheet")
    else:
        sim_wb, sim_sheet = find_sim_workbook(root, config)
    if sim_wb and os.path.exists(sim_wb):
        n, sim_info = ingest_sim(conn, sim_wb, sim_sheet, config.get("sim_label_ids"))
        print(f"[仿真] {os.path.relpath(sim_wb, root)} / {sim_info['sheet']} -> {n} 行，"
              f"{len(sim_info['module_ids'])} 个模块编号")
    else:
        print("[警告] 未找到仿真长表（任意工作簿中表头含 ID/Module/Mode/Current+simulation|Unit 的 tab），"
              "只导入实测；可用 --sim 或 config.sim_workbook 指定")

    skip = set(config.get("skip_dirs") or [])
    excl = list(config.get("exclude_globs") or [])
    out = args.out or os.path.join(root, "Current_compare_pivot.xlsx")
    excl.append(os.path.basename(out))
    globs = result_globs(config)
    sim_modes = {r[0] for r in conn.execute("SELECT DISTINCT mode FROM sim_current")}
    n_runs = 0
    mapping = {}  # mode_raw -> (mode, how)
    fallback_chip = args.chip or "C1"
    all_xlsx = list(walk_xlsx(root, skip, excl))
    report_skipped(all_xlsx, globs, root, sim_wb)
    plan = [(f,) + chip_of_path(f, root, sim_modes, config.get("mode_map"), fallback_chip)
            for f in all_xlsx if is_result_file(os.path.basename(f), globs)]
    dir_chips = sorted({c for _f, c, fm in plan if fm is None})
    # 目录已经按芯片分好时 --chip 是有害的：它会把几颗芯片的 run 全贴成同一个名字，
    # 而汇总簿按 (模式,芯片,温度) 只取最新一次 -> 后测的那颗静默顶掉先测的，簿子上看不出来
    if args.chip and (len(dir_chips) > 1 or (dir_chips and dir_chips != [args.chip])):
        raise SystemExit(
            f"[错误] 子目录名已经是芯片号（{'/'.join(dir_chips)}），此时不要传 --chip"
            f"（会把它们全贴成 {args.chip}，同名 run 互相顶掉且无痕）。去掉 --chip 重跑。")
    for f, chip, folder_mode in plan:
        results = ingest_result_file(conn, f, chip, config, sim_modes, folder_mode=folder_mode)
        n_runs += len(results)
        rel = os.path.relpath(f, root)
        if len(results) > 1:
            print(f"[实测] {rel}  芯片 {chip}  全模式单文件 -> {len(results)} 个 (模式,温度) 段:")
        for run_id, mode, mode_raw, how, temp, n_steps in results:
            mapping[(mode_raw, mode)] = how
            pre = "        " if len(results) > 1 else f"[实测] {rel}  芯片 {chip}  "
            print(f"{pre}{mode:<22} run#{run_id}  {n_steps} 个模块组  "
                  f"{temp if temp is not None else '?'}°C")
    if n_runs == 0:
        print("[警告] 没有扫到任何 Result 文件")
    if dir_chips:
        print(f"[芯片] 按子目录名认出 {len(dir_chips)} 颗: {', '.join(dir_chips)}")
    if n_runs:
        print_chip_roster(conn, "入库后")   # 库里到底有什么，以这行为准

    if mapping:
        print("[模式映射] 实测段标签 -> 仿真 Mode（config.mode_map 可强制指定）:")
        how_disp = {"config": "config指定", "auto": "自动匹配", "ambig": "⚠多个候选未映射",
                    "none": "⚠仿真表无此模式", "folder": "⚠按文件夹名(段内标签与文件夹不符)"}
        for raw_name, mode in sorted(mapping):
            how = mapping[(raw_name, mode)]
            arrow = "=" if raw_name == mode else "->"
            print(f"    {raw_name:<24} {arrow} {mode:<24} {how_disp.get(how, how)}")

    export_xlsx(conn, out, all_runs=args.all_runs, config=config)
    conn.close()
    print(f"[完成] 数据库: {db_path}")
    print(f"[完成] 导出:   {out}")


def _fmt_counts(d, limit=12):
    items = sorted(d.items(), key=lambda kv: (-kv[1], str(kv[0])))
    s = " / ".join(f"{k if k != '' else '(空)'}×{v}" for k, v in items[:limit])
    if len(items) > limit:
        s += f"  …还有 {len(items) - limit} 种"
    return s


def _ranges(nums):
    """[1,2,3,5,6,9] -> '1-3,5-6,9'（模块编号有没有缺口，一眼看出来）"""
    out, start, prev = [], None, None
    for n in nums:
        if start is None:
            start = prev = n
        elif n == prev + 1:
            prev = n
        else:
            out.append(str(start) if start == prev else f"{start}-{prev}")
            start = prev = n
    if start is not None:
        out.append(str(start) if start == prev else f"{start}-{prev}")
    return ",".join(out)


def cmd_inspect(args):
    """只读体检：不建库、不写任何文件（配置也不生成）。回答三个问题——
      ① 我把哪个簿的哪个页当成了仿真表？还有哪些候选落选了？
      ② 我从里面认出了什么？（模块编号取自哪、Mode/Tier/simulation/Unit 各有哪些取值）
      ③ 实测的模式名和 NO. 编号，跟仿真表对不对得上？对不上的逐条列出。
    出簿之前先跑它——静默错值都在这三个问题里。"""
    root = os.path.abspath(args.root)
    if not os.path.isdir(root):
        raise SystemExit(f"[错误] 根目录不存在: {root}")
    config, cfg_path, _ = load_config(root, args.config, create=False)
    tier = config.get("sim_tier") or ""
    stage_main = (config.get("sim_stage") or "post").strip().lower()
    zero_ua = float(config.get("sim_zero_ua") or 0)
    fb_stage = bool(config.get("sim_stage_fallback", True))
    stage_alt = "pre" if stage_main == "post" else "post"
    print(f"根目录: {root}")
    print(f"配置:   {cfg_path}"
          f"{'' if os.path.exists(cfg_path) else '   ← 不存在，本次用内置默认值（build 时才会生成）'}")
    print(f"        sim_tier={tier or '(不过滤——多档共存会重复求和!)'}  sim_stage={stage_main}  "
          f"sim_sheet={config.get('sim_sheet')!r}  result_glob={result_globs(config)!r}")
    print(f"        sim_zero_ua={zero_ua:g}µA（≤此值计作缺项）  "
          f"sim_stage_fallback={'开 -> 缺项取 ' + stage_alt if fb_stage else '关 -> 缺项保持空'}")

    problems = []

    # ---------------------------------------------------------- ① 仿真表选谁
    print("\n" + "=" * 72)
    print("① 仿真长表")
    sim_wb = sim_sheet = None
    if args.sim:
        sim_wb, sim_sheet = args.sim, config.get("sim_sheet")
        print(f"  选用（--sim 指定）: {sim_wb}")
    elif config.get("sim_workbook"):
        p = config["sim_workbook"]
        sim_wb = p if os.path.isabs(p) else os.path.join(root, p)
        sim_sheet = config.get("sim_sheet")
        print(f"  选用（config.sim_workbook 指定）: {sim_wb}")
    else:
        print("  正在扫描工作簿（大簿较慢）…", flush=True)
        cands = find_sim_candidates(root, config, verbose=True)
        if not cands:
            print("  ⚠ 一个候选都没扫到（需要某页表头含 ID+Module+Mode+Current*，"
                  "且有 simulation 或 Unit 列）")
            problems.append("没找到仿真长表 -> 报告的仿真列会全空")
        else:
            best = cands[0]
            sim_wb, sim_sheet = best[1], best[2]
            for mt, p, sn, how in cands:
                mark = "★选用" if (p, sn) == (best[1], best[2]) else "  落选"
                ts = datetime.datetime.fromtimestamp(mt).strftime("%Y-%m-%d %H:%M")
                print(f"  {mark}  {os.path.relpath(p, root)} / {sn}  ({how}, 改于 {ts})")
            if len(cands) > 1:
                print("        ↑ 多个候选时按修改时间取最新。选错仿真快照是静默事故，"
                      "确认无误后建议在 config.sim_workbook 里钉死。")
                problems.append(f"有 {len(cands)} 个仿真表候选，靠修改时间自动选 -> 建议钉死 sim_workbook")

    # ---------------------------------------------------------- ② 认出了什么
    recs, info = [], None
    if sim_wb and os.path.exists(sim_wb):
        recs, info = read_sim_rows(sim_wb, sim_sheet, config.get("sim_label_ids"))
        colmap = "  ".join(f"{k}={get_column_letter(v + 1)}"
                           for k, v in sorted(info["cols"].items(), key=lambda kv: kv[1]))
        print(f"\n  页={info['sheet']}  表头行={info['header_row']}")
        print(f"  列映射: {colmap}")
        print(f"  数据行 {info['n_data_rows']}   跳过: 无电流值 {info['skip_no_current']} / "
              f"无模式 {info['skip_no_mode']}")
        src = []
        if info["id_from_prefix"]:
            src.append(f"Module 列数字前缀 {info['id_from_prefix']} 行")
        if info["id_from_idcol"]:
            src.append(f"ID 列回退 {info['id_from_idcol']} 行")
        if info["id_missing"]:
            src.append(f"无法确定 {info['id_missing']} 行")
        print(f"  模块编号来源: {' / '.join(src) or '(无)'}")
        mids = info["module_ids"]
        print(f"  模块编号: {len(mids)} 个  {_ranges(mids) if mids else '(无)'}")
        if info["prefix_conflicts"]:
            print("  ID 列 vs Module 前缀不一致的样例（前 5 条，证明 ID 列不是模块编号）:")
            for idv, name, md in info["prefix_conflicts"]:
                print(f"      ID={idv:<6} Module={name:<24} Mode={md}")
        if info["no_prefix_names"]:
            lm = info["label_id_map"]
            auto = set(info["label_id_auto"])
            print(f"  无数字前缀的 Module 名 {len(info['no_prefix_names'])} 种（合计/标签行），"
                  f"已指派编号:")
            for nm, cnt in sorted(info["no_prefix_names"].items(),
                                  key=lambda kv: (-kv[1], kv[0])):
                src = "自动" if nm in auto else "config.sim_label_ids"
                print(f"      {nm!r:<28} -> 编号 {lm.get(nm)}   {cnt} 行  ({src})")
            tot = [nm for nm in info["no_prefix_names"]
                   if re.search(r"(?i)total|sum|合计", nm)]
            if tot:
                print("      ⚠ 名字带 total/sum 的是**合计行**：通常是对上面已逐项列出的模块求和。"
                      "把实测标签指到它 = 拿单项测量值去对一整片的合计，会造出巨大的假偏差"
                      "（本项目真实事故：DCO 标签行被对到 buffer 合计，假偏差 -25~-49% 追查了两轮）。"
                      "除非确认它对应实测里某个独立测量步，否则保持不映射。")
            print("      不映射的行只是躺在库里，不进任何对比，也不会影响 Σ 合计。")
        print(f"  Mode       ({len(info['modes'])}): {_fmt_counts(info['modes'])}")
        print(f"  Tier       ({len(info['tiers'])}): {_fmt_counts(info['tiers'])}")
        print(f"  simulation ({len(info['stages'])}): {_fmt_counts(info['stages'])}")
        print(f"  Unit       ({len(info['units'])}): {_fmt_counts(info['units'])}")
        for line in sim_warnings(info):
            print("  " + line)
        if tier and tier not in info["tiers"]:
            print(f"  ⚠ config.sim_tier={tier!r} 在 Tier 列里一个都匹配不上"
                  f"（Tier 是 SQL 精确比对，差一个字符即全空）")
            problems.append(f"sim_tier={tier!r} 对不上仿真表 Tier 取值 -> 仿真列会全空")
        if info["unknown_units"]:
            problems.append("有未知单位按 uA 处理 -> 可能差 1000 倍")
        if not info["n_data_rows"]:
            problems.append("仿真表一行都没读到")
    elif sim_wb:
        print(f"  ⚠ 文件不存在: {sim_wb}")
        problems.append(f"仿真簿路径不存在: {sim_wb}")

    # 该 Mode 下（按 sim_tier 过滤后）实际有哪些模块编号 —— 用来对实测 NO.
    sim_ids_by_mode = {}
    for r in recs:
        if r["module_id"] is None or (tier and r["tier"] != tier):
            continue
        sim_ids_by_mode.setdefault(r["mode"], set()).add(r["module_id"])

    # (模式,编号,档位,阶段) -> 合计值。summary 的仿真列固定取 post×sim_tier 这一格；
    # 零值审计要看另外 7 格有没有值，才能分清"簿里本来就是 0(该模式不工作)"
    # 和"取错格子(值在别的 Tier/stage 上)"。同前缀的子块(7_1/7_2)求和，与 sim_lookup 一致。
    sim_cell = {}
    for r in recs:
        if r["module_id"] is None:
            continue
        k = (r["mode"], r["module_id"], r["tier"], r["stage"])
        sim_cell[k] = sim_cell.get(k, 0.0) + r["current_ua"]
    all_tiers = sorted(t for t in info["tiers"]) if info else []
    all_stages = sorted(s for s in info["stages"]) if info else []

    def cell_is_zero(v):
        """≤ sim_zero_ua 一律算「等于没有」——与 sim_lookup.is_zero 同一条判据。
        两边判据必须一致，否则 inspect 说没事、build 却补了值（或反过来）。"""
        return v is None or abs(v) <= zero_ua

    def zero_audit(mode, mid):
        """返回 (报告用格子的值, [其他有值格子描述])。"""
        target = sim_cell.get((mode, mid, tier, stage_main))
        others = []
        for t in all_tiers:
            for st in all_stages:
                if (t, st) == (tier, stage_main):
                    continue
                v = sim_cell.get((mode, mid, t, st))
                if not cell_is_zero(v):
                    others.append(f"{t}/{st}={v:.1f}")
        return target, others

    # ---------------------------------------------------------- ③ 实测对不对得上
    print("\n" + "=" * 72)
    print("② 实测 Result 文件 / 分段 / 与仿真的对应")
    sim_modes = set(info["modes"]) if info else set()
    mode_map = config.get("mode_map") or {}
    skip = set(config.get("skip_dirs") or [])
    excl = list(config.get("exclude_globs") or [])
    globs = result_globs(config)
    how_disp = {"config": "config 指定", "auto": "自动匹配", "folder": "⚠ 按文件夹名",
                "ambig": "⚠ 多个候选未映射", "none": "⚠ 仿真表无此模式"}
    n_files = 0
    chips_seen = {}
    all_xlsx = list(walk_xlsx(root, skip, excl))
    if report_skipped(all_xlsx, globs, root, sim_wb):
        problems.append("有 xlsx 不匹配 result_glob 被跳过 -> 确认里面没有实测文件")
    for f in all_xlsx:
        if not is_result_file(os.path.basename(f), globs):
            continue
        n_files += 1
        # 与 build 同一条规则：文件夹名对得上仿真 Mode 就当模式名，否则当芯片号
        chip, folder = chip_of_path(f, root, sim_modes, mode_map, "C1")
        chips_seen.setdefault(chip, []).append(os.path.relpath(f, root))
        print(f"\n  {os.path.relpath(f, root)}   芯片={chip}"
              + (f"  文件夹名当模式用={folder}" if folder else ""))
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        try:
            ws, hdr, cols = find_result_sheet(wb, config.get("result_sheet"))
            if ws is None:
                print("    ⚠ 找不到含 NO./Current 表头的 tab")
                problems.append(f"{os.path.basename(f)} 认不出实测表头")
                continue
            raw = read_raw_rows(ws, hdr, cols)
        finally:
            wb.close()
        print(f"    页={ws.title} 表头行={hdr} 单位列={cols['unit'] or '(空,按mA)'}  原始行 {len(raw)}")
        factor_to_ma = UNIT_TO_UA.get(cols["unit"], 1000.0) / 1000.0
        segs = split_allmode(raw) or [{"mode": folder or "?", "raw": raw, "temp": None}]
        # 与 _ingest_raw 保持同一条规则：旧单模式文件（整文件一段）段内标签与文件夹不符时
        # 文件夹名优先（见过模板复制错名）。体检必须复刻它，否则 inspect 与 build 结论会打架。
        single_legacy = (len(segs) == 1 and folder
                         and canon_mode(folder) != canon_mode(segs[0]["mode"]))
        for s in segs:
            rows, temp0 = classify_raw(s["raw"], factor_to_ma)
            temp = s["temp"] if s.get("temp") is not None else temp0
            steps, _absorbed = build_groups(rows, config)
            if single_legacy:
                mode, _how0 = resolve_mode(folder, sim_modes, mode_map)
                how = "folder"
            else:
                mode, how = resolve_mode(s["mode"], sim_modes, mode_map, folder=folder)
            ids_used = sorted({i for st in steps for i in (st["ids"] or [])})
            labels = sorted({st["disp"] for st in steps if not st["ids"]})
            arrow = "=" if s["mode"] == mode else "->"
            print(f"    段 {s['mode']:<22} {arrow} 仿真 {mode:<22} "
                  f"{how_disp.get(how, how):<12} {temp if temp is not None else '?'}°C  "
                  f"{len(steps)} 组")
            if how in ("ambig", "none"):
                problems.append(f"模式 {s['mode']!r} 匹配不到仿真 Mode（{how}）"
                                f" -> 在 config.mode_map 里指定")
            elif how == "folder":
                problems.append(f"{folder} 目录里段内标签写的是"
                                f" {s['mode']!r}（模板复制错名？）-> 已按文件夹名当 {mode}")
            have = sim_ids_by_mode.get(mode)
            if have is None:
                if sim_modes:
                    print(f"        ⚠ 仿真表在该 Mode + Tier={tier or '不限'} 下没有任何模块")
            else:
                miss = [i for i in ids_used if i not in have]
                print(f"        NO. 编号 {len(ids_used)} 个 {_ranges(ids_used)}；"
                      f"仿真表缺 {len(miss)} 个" + (f": {_ranges(miss)}" if miss else ""))
                if miss:
                    problems.append(f"{mode}: {len(miss)}/{len(ids_used)} 个 NO. 编号仿真表里没有")
                # 零值审计：报告只用 post×sim_tier 一格，为 0/残渣 时把其他 7 格摊开
                zeros = [(i, zero_audit(mode, i)) for i in ids_used
                         if cell_is_zero(sim_cell.get((mode, i, tier, stage_main)))]
                if zeros:
                    n_elsewhere = sum(1 for _i, (_v, o) in zeros if o)
                    # 同档位另一阶段有值 = 跨阶段补值正好覆盖的情形，与"值在别的 Tier 上"分开报，
                    # 否则已被 config 自动处理的事也进问题清单，把真问题淹了
                    covered = [i for i, (_v, o) in zeros
                               if fb_stage and not cell_is_zero(
                                   sim_cell.get((mode, i, tier, stage_alt)))]
                    print(f"        ⚠ 仿真为 0/缺项的编号 {len(zeros)}/{len(ids_used)} 个"
                          f"（报告取 {tier or '不限'}×{stage_main} 这一格）:")
                    for i, (v, others) in zeros:
                        residue = "" if not v else f"={v:g}(≤{zero_ua:g}µA 计作缺项)"
                        tag = "  ✔ 同档位另一阶段有值，将自动补" if i in covered else ""
                        if others:
                            print(f"            编号 {i:<4} 本格{residue or '=0'}，"
                                  f"但别的档位有值: {' '.join(others)}{tag}")
                        else:
                            print(f"            编号 {i:<4} 全 8 个档位×阶段皆为 0/无值"
                                  f" -> 簿里认为该模块在此模式不工作")
                    n_open = n_elsewhere - len(covered)
                    if n_open > 0:
                        problems.append(f"{mode}: {n_open} 个编号在 {tier}×{stage_main} 为 0/缺项，"
                                        f"{stage_alt} 也没有、但别的 Tier 有值 -> 可能取错档位")
                    if n_elsewhere < len(zeros):
                        problems.append(f"{mode}: {len(zeros) - n_elsewhere} 个编号仿真全档位皆 0，"
                                        f"而实测有值 -> Σ 仿真与偏差% 不可信")
                    if covered:
                        print(f"            （其中 {len(covered)} 个由 sim_stage_fallback 用 "
                              f"{stage_alt} 补上，不算问题）")
            if labels:
                print(f"        非数字标签 {len(labels)} 个: {', '.join(labels[:8])}"
                      + (" …" if len(labels) > 8 else ""))
    if n_files == 0:
        print(f"\n  ⚠ 没扫到任何匹配 {result_glob} 的文件")
        problems.append(f"没有扫到 {globs}")
    elif len(chips_seen) > 1:
        print(f"\n  芯片 {len(chips_seen)} 颗（按子目录名）: "
              + " / ".join(f"{c}×{len(v)} 个文件" for c, v in sorted(chips_seen.items())))
        print("    build 不要传 --chip（传了会把它们贴成同一个名字，同名 run 互相顶掉）；"
              "summary 可用 --chip 只出其中一颗。")

    # ---------------------------------------------------------- 结论
    print("\n" + "=" * 72)
    problems = list(dict.fromkeys(problems))  # 同一问题每个温度段会重复报一次，去重保序
    if problems:
        print(f"结论：发现 {len(problems)} 个需要处理的问题")
        for i, p in enumerate(problems, 1):
            print(f"  {i}. {p}")
    else:
        print("结论：没发现问题，可以跑 build 了。")


def cmd_ingest_sim(args):
    conn = open_db(args.db)
    n, info = ingest_sim(conn, args.xlsx, args.sheet)
    conn.close()
    print(f"[仿真] {args.xlsx} / {info['sheet']} -> {n} 行，"
          f"{len(info['module_ids'])} 个模块编号")


def cmd_ingest_run(args):
    root = os.path.dirname(os.path.abspath(args.db))
    config, _, _ = load_config(root, args.config)
    conn = open_db(args.db)
    run_id, n_steps, temp, run_ts = ingest_run(conn, args.xlsx, args.mode, args.chip, config,
                                               sheet_name=args.sheet)
    conn.close()
    print(f"[实测] run#{run_id} mode={args.mode} {n_steps} 个模块组 {temp}°C {run_ts}")


def cmd_ingest_probe(args):
    root = os.path.dirname(os.path.abspath(args.db))
    config, _, _ = load_config(root, args.config)
    conn = open_db(args.db)
    results = ingest_probe_json(conn, args.json, args.chip, config)
    conn.close()
    for run_id, mode, mode_raw, how, temp, n_steps in results:
        print(f"[实测] {mode:<22} (段标签 {mode_raw}, {how})  run#{run_id}  "
              f"{n_steps} 个模块组  {temp if temp is not None else '?'}°C")


def cmd_export(args):
    root = os.path.dirname(os.path.abspath(args.db))
    config, _, _ = load_config(root, args.config)
    conn = open_db(args.db)
    export_xlsx(conn, args.out, all_runs=args.all_runs, config=config)
    conn.close()
    print(f"[完成] 导出: {args.out}")


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    ap = argparse.ArgumentParser(description="电流数据库：仿真+实测 -> SQLite -> pivot 长表")
    sub = ap.add_subparsers(dest="cmd", required=True)

    b = sub.add_parser("build", help="一键：扫描数据根目录，全量重建库并导出")
    b.add_argument("--root", required=True, help="数据根目录，如 D:\\Excel")
    b.add_argument("--chip", help="芯片编号；**子目录名已经是芯片号时不要传**"
                                  "（默认按子目录名逐文件认；root 直放的文件归 C1）")
    b.add_argument("--sim", help="仿真工作簿路径（默认自动找 Current_all_mode*.xlsx）")
    b.add_argument("--db", help="SQLite 输出路径（默认 root/current.db）")
    b.add_argument("--out", help="Excel 输出路径（默认 root/Current_compare_pivot.xlsx）")
    b.add_argument("--config", help="配置文件路径（默认 root/current_config.json）")
    b.add_argument("--all-runs", action="store_true", help="导出全部 run（默认每模式×芯片取最新）")
    b.set_defaults(func=cmd_build)

    i = sub.add_parser("inspect", help="只读体检：不建库不写文件，看清仿真表选了谁/认出了什么/"
                                       "实测编号对不对得上（build 之前先跑它）")
    i.add_argument("--root", required=True, help="数据根目录")
    i.add_argument("--sim", help="指定仿真工作簿（默认自动扫）")
    i.add_argument("--config", help="配置文件路径（默认 root/current_config.json；不存在也不生成）")
    i.set_defaults(func=cmd_inspect)

    s = sub.add_parser("ingest-sim", help="导入仿真长表")
    s.add_argument("--db", required=True)
    s.add_argument("--xlsx", required=True)
    s.add_argument("--sheet", default="Current_data")
    s.set_defaults(func=cmd_ingest_sim)

    r = sub.add_parser("ingest-run", help="导入单个实测 Result 文件")
    r.add_argument("--db", required=True)
    r.add_argument("--xlsx", required=True)
    r.add_argument("--mode", required=True, help="模式名（需与仿真表 Mode 一致）")
    r.add_argument("--chip", default="C1")
    r.add_argument("--sheet", help="tab 名（默认自动扫描）")
    r.add_argument("--config", help="配置文件路径（默认取 db 同目录 current_config.json）")
    r.set_defaults(func=cmd_ingest_run)

    a = sub.add_parser("add-chip", help="往已有库里加一颗芯片的全模式实测文件（增量，不动仿真"
                                        "和别的芯片；同一文件重复加是幂等的）")
    a.add_argument("--db", required=True)
    a.add_argument("--xlsx", required=True, help="该芯片的全模式实测文件")
    a.add_argument("--chip", help="芯片编号；默认取文件所在目录名")
    a.add_argument("--config", help="配置文件路径（默认取 db 同目录 current_config.json）")
    a.set_defaults(func=cmd_add_chip)

    p = sub.add_parser("ingest-probe", help="导入 probe_allmode_result.py --json 的产物")
    p.add_argument("--db", required=True)
    p.add_argument("--json", required=True)
    p.add_argument("--chip", default="C1")
    p.add_argument("--config", help="配置文件路径（默认取 db 同目录 current_config.json）")
    p.set_defaults(func=cmd_ingest_probe)

    e = sub.add_parser("export", help="从库导出 Excel")
    e.add_argument("--db", required=True)
    e.add_argument("--out", required=True)
    e.add_argument("--all-runs", action="store_true")
    e.add_argument("--config", help="配置文件路径（默认取 db 同目录 current_config.json）")
    e.set_defaults(func=cmd_export)

    m = sub.add_parser("summary", help="导出人直接读的功耗汇总簿（总览矩阵+温度趋势图+对比明细）")
    m.add_argument("--db", required=True)
    m.add_argument("--out", required=True)
    m.add_argument("--config", help="配置文件路径（默认取 db 同目录 current_config.json）")
    m.add_argument("--chip", action="append",
                   help="只出这些芯片（可重复或逗号分隔）；默认库里有几颗出几颗")

    mc = sub.add_parser("summary-chips", help="跨芯片评审版汇总簿：仿真列只一列、一片一竖条、"
                                              "带片间极差（多颗芯片时用这个）")
    mc.add_argument("--db", required=True)
    mc.add_argument("--out", required=True)
    mc.add_argument("--config", help="配置文件路径（默认取 db 同目录 current_config.json）")
    mc.add_argument("--chip", action="append",
                    help="只出这些芯片（可重复或逗号分隔）；默认库里有几颗出几颗")
    mc.add_argument("--no-audit", action="store_true",
                    help="连隐藏的 _审计 页也不出（正簿本来就只有总览一页）")
    mc.set_defaults(func=cmd_chips)
    m.add_argument("--mark-fallback", action="store_true",
                   help="把跨阶段补值的仿真格标成蓝色斜体（自查版用；默认不标，"
                        "保持给人 review 的总览干净）")
    m.set_defaults(func=cmd_summary)

    args = ap.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
