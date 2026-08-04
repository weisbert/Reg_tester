#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""电流测试簿探查（只读体检）。

    python probe_current.py <目录或单个 xlsx> [--sheet 名] [--json 出.json]
                            [--emit-groups 清单.json] [--max-list 80]

**为什么先探再写出簿脚本**：电流文件的模板不是温扫那一套，列位 / 表头 / 单位 /
一行是什么，全都可能不一样。先把结构看清楚，出簿脚本才不会写死列字母——
这条线上"照着样例写死列位、换份文件就散架"已经发生过。

**只读**：不写、不改、不动任何被探的文件。

零 IP：本文件里不出现任何真实模块名 / 芯片名 / 地址。要报的名字全部来自被探的
文件本身。`--emit-groups` 落盘的清单**含真名，只能放 private/ 或黄区本地**。
"""

import argparse
import json
import os
import re
import sys
import time
from collections import Counter, OrderedDict

import openpyxl

try:
    from summarize_chips import discover, KIND_CUR, KIND_LABEL, natkey
except Exception:                                     # noqa: B902
    discover = None

NA = {"-", "--", "---", "n/a", "na", "nan", "none", "null", ""}
# 表头里出现这些词就认为该列大概是什么——只做提示，不做判据
HINT = OrderedDict([
    ("电流", ("current", "curr", "idd", "icc", "ma", "ua", "µa", "电流")),
    ("温度", ("temp", "℃", "degc", "温度")),
    ("电压", ("volt", "vdd", "vbat", "vsup", "电压")),
    ("模式", ("mode", "state", "cond", "模式", "状态")),
    ("序号", ("no.", "no", "idx", "index", "序")),
])


def fmt(v):
    if v is None:
        return "?"
    return str(int(v)) if float(v).is_integer() else str(v)


def _cl(i):
    from openpyxl.utils import get_column_letter
    return get_column_letter(i)


def _s(v):
    return "" if v is None else str(v).strip()


def _num(v):
    """能当数用就返回 float。文本型数字（'1.23'）也算——原簿里很常见。"""
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = _s(v).replace(",", "")
    if not t or t.lower() in NA:
        return None
    try:
        return float(t)
    except ValueError:
        return None


def read_grid(ws, max_rows):
    """整页一次性流式读成二维表，之后所有分析都在内存里做。

    ★★ 不能用 `ws.cell(r, c)` 逐格扫：openpyxl 每次调用都要构造一个 Cell 对象，
      而这类导出簿的"已使用区域"动不动就是几百列 × 几千行（列宽/底色刷过整行整列
      也会把 max_column 顶到很大）。逐格扫 = 上百万次对象构造，看起来就是**卡死**。
      `iter_rows(values_only=True)` 只吐元组，快两个数量级。
      （2026-08-04：第一版就是逐格扫，用户在真文件上跑没有任何输出。）
    """
    rows = []
    capped = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i + 1 >= max_rows:
            capped = True
            break
    while rows and not any(_s(v) for v in rows[-1]):     # 掐掉末尾空行
        rows.pop()
    ncol = 0
    for row in rows:
        for j in range(len(row) - 1, ncol - 1, -1):
            if _s(row[j]):
                ncol = j + 1
                break
    return rows, ncol, capped


def at(grid, r, c):
    """1 基取值，越界给 None（各行长度不一定齐）。"""
    if 1 <= r <= len(grid):
        row = grid[r - 1]
        if 1 <= c <= len(row):
            return row[c - 1]
    return None


def find_header(grid, ncol, limit=40):
    """猜表头行：短文本最多的那一行。返回 (行号, 得分)。

    ★ 不按"第 1 行就是表头"处理：这类导出簿常在上面压几行标题 / 空行 / 合并单元格。
    """
    best, best_score = 1, -1
    for r in range(1, min(limit, len(grid)) + 1):
        n_txt = n_num = 0
        for c in range(1, ncol + 1):
            v = at(grid, r, c)
            if v is None:
                continue
            if _num(v) is not None and not isinstance(v, str):
                n_num += 1
            elif _s(v) and len(_s(v)) <= 40:
                n_txt += 1
        score = n_txt - n_num          # 表头是文字多、数字少的那一行
        if score > best_score:
            best, best_score = r, score
    return best, best_score


def last_data_row(grid, ncol, hdr):
    """真正有内容的最后一行。

    ★ 探查器报的 n_data_rows ≠ 工作表行数：导出脚本常在末尾追加页脚行
      （只有前一两列有字、没有任何结果值）。这条在温扫那边踩过，这里直接报出来。
    """
    last = hdr
    for r in range(len(grid), hdr, -1):
        if any(_s(at(grid, r, c)) for c in range(1, ncol + 1)):
            last = r
            break
    return last


def profile(grid, hdr, r0, r1, c):
    """一列的体检：名字、类型构成、样例、去重个数。"""
    name = _s(at(grid, hdr, c))
    vals = [at(grid, r, c) for r in range(r0, r1 + 1)]
    txt = [_s(v) for v in vals if _s(v) and _s(v).lower() not in NA]
    nums = [v for v in vals if _num(v) is not None]
    # 文本型数字：Python 解得动，Excel 的 COUNT/MIN/MATCH 一律不认
    as_text = [v for v in nums if isinstance(v, str)]
    na_hits = [_s(v) for v in vals if _s(v).lower() in NA and _s(v)]
    uniq = list(OrderedDict.fromkeys(txt))
    kind = ""
    low = name.lower()
    for tag, keys in HINT.items():
        if any(k in low for k in keys):
            kind = tag
            break
    return {
        "col": c, "letter": _cl(c), "name": name, "kind_hint": kind,
        "n_filled": len(txt), "n_num": len(nums), "n_text_num": len(as_text),
        "n_na": len(na_hits), "n_uniq": len(uniq),
        "sample": uniq[:6] if len(nums) < len(txt) else
                  [_s(v) for v in vals if _num(v) is not None][:6],
        "all_text": uniq,
    }


FLAGS = {"yes", "no", "y", "n", "true", "false", "on", "off", "-", "--"}


def _is_flag(p):
    """YES/NO 这种开关列，不管有多少行都不是标签列。"""
    return bool(p["all_text"]) and \
        {t.lower() for t in p["all_text"]} <= FLAGS


def guess_roles(cols, n_rows, force_label=None, force_value=None):
    """猜每列的角色。只给建议，最终以人看到的为准。"""
    value, cond, cand = [], [], []
    for p in cols:
        if not p["n_filled"]:
            continue
        numeric = p["n_num"] >= max(3, p["n_filled"] * 0.8)
        if numeric:
            value.append(p) if p["kind_hint"] in ("电流", "") else cond.append(p)
        else:
            cond.append(p)
            # ★★ 标签列判据不能要求"不同值 ≈ 行数"：同一批标签会**在每个温度下
            #   重复一遍**（真文件：56 个步骤 × 3 个温度 = 167 行），按行数一半去卡
            #   就一个都认不出来（2026-08-04 真文件上就是这么漏的）。
            #   改成：文字列里**不同值最多**的那一列，开关列（YES/NO）排除。
            if p["n_uniq"] >= 4 and not _is_flag(p):
                cand.append(p)
    label = None
    if force_label:
        label = next((p for p in cols if force_label.upper() in
                      (p["letter"], p["name"].upper())), None)
    if label is None and cand:
        label = max(cand, key=lambda p: (p["n_uniq"], -p["col"]))
    if force_value:
        v = next((p for p in cols if force_value.upper() in
                  (p["letter"], p["name"].upper())), None)
        if v is not None:
            value = [v] + [x for x in value if x is not v]
    return label, value, cond


def _pick(cols, want_hint, numeric=None, max_uniq=None):
    best = None
    for p in cols:
        if p["kind_hint"] != want_hint or not p["n_filled"]:
            continue
        if numeric is True and p["n_num"] < p["n_filled"] * 0.8:
            continue
        if numeric is False and p["n_num"] > p["n_filled"] * 0.2:
            continue
        if max_uniq is not None and p["n_uniq"] > max_uniq:
            continue
        if best is None or p["n_uniq"] > best["n_uniq"]:
            best = p
    return best


def build_ladder(grid, r0, r1, label, cur, temp, mode, cap):
    """按文件顺序列出「标签 → 电流」的阶梯，并给出相邻做差。

    ★ 这类逐级关断簿里，值列装的是**每关掉一个模块之后的总电流**，
      不是那个模块自己的电流。模块电流＝相邻两个测量点相减。把两列并排打出来，
      口径对不对一眼就能看出来（差出负数、差出 0，都是当场就该发现的事）。
    """
    if not (label and cur):
        return [], None
    t0 = None
    if temp:
        for r in range(r0, r1 + 1):
            v = _num(at(grid, r, temp["col"]))
            if v is not None:
                t0 = v
                break
    out, prev = [], None
    for r in range(r0, r1 + 1):
        if temp and t0 is not None and _num(at(grid, r, temp["col"])) != t0:
            continue
        lab = _s(at(grid, r, label["col"]))
        md = _s(at(grid, r, mode["col"])) if mode else ""
        iv = _num(at(grid, r, cur["col"]))
        d = None if (iv is None or prev is None) else round(prev - iv, 4)
        out.append((r, lab, md, iv, d))
        if iv is not None:
            prev = iv
        if len(out) >= cap:
            break
    return out, t0


def probe_sheet(title, grid, ncol, capped, force_label=None, force_value=None,
                ladder=40):
    hdr, _ = find_header(grid, ncol)
    r0 = hdr + 1
    r1 = last_data_row(grid, ncol, hdr)
    cols = [profile(grid, hdr, r0, r1, c) for c in range(1, ncol + 1)]
    used = [p for p in cols if p["n_filled"] or p["name"]]
    dup = [n for n, k in Counter(p["name"] for p in used if p["name"]).items() if k > 1]
    label, value, cond = guess_roles(used, r1 - r0 + 1, force_label, force_value)
    # 尾部页脚行：只有零星几个格子有字、没有任何数
    foot = []
    for r in range(r1, max(r0, r1 - 8), -1):
        filled = [c for c in range(1, ncol + 1) if _s(at(grid, r, c))]
        nums = [c for c in filled if _num(at(grid, r, c)) is not None]
        if filled and not nums and len(filled) <= 3:
            foot.append((r, [_s(at(grid, r, c)) for c in filled][:3]))
        else:
            break
    # 温度列：像温度、是数、档数不多；电流值列：像电流、不同值最多的那一个
    temp = _pick(used, "温度", numeric=True, max_uniq=12)
    cur = _pick(used, "电流", numeric=True)
    if cur is None and value:
        cur = max(value, key=lambda p: p["n_uniq"])
    # 第二文字列当"这一步在干嘛"的说明
    mode = None
    for p in used:
        if p is not label and p["n_filled"] and p["n_num"] < p["n_filled"] * 0.5 \
                and p["n_uniq"] >= 4 and not _is_flag(p):
            if mode is None or p["n_uniq"] > mode["n_uniq"]:
                mode = p
    lad, t0 = build_ladder(grid, r0, r1, label, cur, temp, mode, ladder)
    # 标签重复几遍＝多半是几个温度各跑了一趟
    rep = None
    if label and label["n_uniq"]:
        rep = round((r1 - r0 + 1) / label["n_uniq"], 2)
    return {
        "sheet": title, "max_row": len(grid), "max_col": ncol, "capped": capped,
        "header_row": hdr, "data_rows": [r0, r1], "n_data_rows": r1 - r0 + 1,
        "cols": used, "dup_names": dup, "footer_rows": foot,
        "label_col": label, "value_cols": value, "cond_cols": cond,
        "temp_col": temp, "cur_col": cur, "mode_col": mode,
        "ladder": lad, "ladder_temp": t0, "repeat": rep,
    }


def probe_book(path, only_sheet=None, max_rows=20000,
               force_label=None, force_value=None, ladder=40):
    size = os.path.getsize(path) / 1e6
    print(f"  读 {os.path.basename(path)}（{size:.1f} MB）…", end="", flush=True)
    t0 = time.time()
    # read_only=True：只取值、不建 Cell 对象树。大簿子上是能不能跑完的区别
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {"path": path, "name": os.path.basename(path), "sheets": []}
    try:
        for ws in wb.worksheets:
            if only_sheet and ws.title != only_sheet:
                continue
            grid, ncol, capped = read_grid(ws, max_rows)
            print(f" [{ws.title}: {len(grid)}×{ncol}]", end="", flush=True)
            if len(grid) < 2 or ncol < 2:
                out["sheets"].append({"sheet": ws.title, "empty": True,
                                      "max_row": len(grid), "max_col": ncol})
                continue
            out["sheets"].append(probe_sheet(ws.title, grid, ncol, capped,
                                             force_label, force_value,
                                             ladder))
    finally:
        wb.close()
    print(f"  {time.time() - t0:.1f}s")
    return out


def signature(sh):
    """格式指纹：表头文字序列。两份文件指纹一样 = 出簿脚本可以一视同仁。"""
    return "|".join(p["name"] for p in sh.get("cols", []) if p["name"])


def show(rep, max_list):
    print()
    print("=" * 78)
    print(f"文件: {rep['name']}")
    for sh in rep["sheets"]:
        if sh.get("empty"):
            print(f"  [页 {sh['sheet']}] 空的（{sh['max_row']}×{sh['max_col']}）")
            continue
        print(f"  [页 {sh['sheet']}] {sh['max_row']} 行 × {sh['max_col']} 列，"
              f"表头在第 {sh['header_row']} 行，数据 {sh['data_rows'][0]}~"
              f"{sh['data_rows'][1]}（{sh['n_data_rows']} 行）"
              + ("  ★只读了前 %d 行（--max-rows 调）" % sh["max_row"]
                 if sh.get("capped") else ""))
        if sh["max_row"] != sh["data_rows"][1]:
            print(f"     · 工作表行数 {sh['max_row']} ≠ 有内容的最后一行 "
                  f"{sh['data_rows'][1]}（末尾有空行）")
        for r, txt in sh["footer_rows"]:
            print(f"     · 第 {r} 行像页脚（只有 {txt} 有字、没有数）——不是数据")
        if sh["dup_names"]:
            for n in sh["dup_names"]:
                where = [p["letter"] for p in sh["cols"] if p["name"] == n]
                print(f"     ⚠ 重名列「{n}」出现在 {', '.join(where)}——"
                      f"按名字取只会拿到第一个")

        lab = sh["label_col"]
        print("     ---- 列 ----")
        for p in sh["cols"]:
            if not p["n_filled"] and not p["name"]:
                continue
            role = ("标签列" if p is lab else
                    "值列" if p in sh["value_cols"] else "条件/其它")
            bits = [f"{p['n_filled']} 个非空"]
            if p["n_num"]:
                bits.append(f"{p['n_num']} 个数")
            if p["n_text_num"]:
                bits.append(f"★{p['n_text_num']} 个是**文本型数字**")
            if p["n_na"]:
                bits.append(f"{p['n_na']} 个 -/N/A")
            if p["n_uniq"]:
                bits.append(f"{p['n_uniq']} 个不同值")
            print(f"       {p['letter']:>3} {p['name'][:26]:<26} [{role}] "
                  f"{'，'.join(bits)}"
                  + (f"   {p['kind_hint']}?" if p["kind_hint"] else ""))
            if p["sample"]:
                print(f"           样例: {', '.join(str(x)[:18] for x in p['sample'])}")

        if lab:
            key = [f"标签列 {lab['letter']}「{lab['name']}」"]
            for tag, p in (("温度列", sh.get("temp_col")),
                           ("电流值列", sh.get("cur_col")),
                           ("说明列", sh.get("mode_col"))):
                if p is not None:
                    key.append(f"{tag} {p['letter']}「{p['name']}」")
            print("     ---- 认出来的关键列 ----")
            print("       " + "； ".join(key))
            if sh.get("repeat"):
                print(f"       标签 {lab['n_uniq']} 个 / 数据 {sh['n_data_rows']} 行"
                      f" ≈ 重复 {sh['repeat']} 遍"
                      + (f"（温度列有 {sh['temp_col']['n_uniq']} 档——"
                         f"多半是每个温度各跑一趟）" if sh.get("temp_col") else ""))
            lad = sh.get("ladder") or []
            if lad:
                print(f"     ---- 逐级阶梯（{fmt(sh.get('ladder_temp'))}℃ 那一趟，"
                      f"前 {len(lad)} 行）----")
                print(f"       {'行':>4}  {'标签':<22} {'电流':>10} {'与上一点之差':>12}"
                      f"  说明")
                for r, l_, md, iv, d in lad:
                    print(f"       {r:>4}  {l_[:22]:<22} "
                          f"{('' if iv is None else f'{iv:.4f}'):>10} "
                          f"{('' if d is None else f'{d:+.4f}'):>12}  {md[:34]}")
                print("       （值列装的是每一步之后的**总电流**；模块电流＝相邻做差，"
                      "上面右边那一列就是）")
            print(f"     ---- 标签列的全部取值（{len(lab['all_text'])} 个）----")
            for i, t in enumerate(lab["all_text"][:max_list]):
                print(f"       {i + 1:>3}. {t}")
            if len(lab["all_text"]) > max_list:
                print(f"       …还有 {len(lab['all_text']) - max_list} 个"
                      f"（--max-list 调）")
        else:
            print("     ⚠ 没认出标签列（哪一列是「一行一个模块」的名字）——"
                  "用 --label-col A 手指一列，或者把这段输出发我")


def main():
    ap = argparse.ArgumentParser(description="电流测试簿探查（只读）")
    ap.add_argument("path", help="芯片目录的根，或单个 xlsx")
    ap.add_argument("--sheet", default=None, help="只看这一页")
    ap.add_argument("--max-list", type=int, default=80, help="标签最多列几个")
    ap.add_argument("--max-rows", type=int, default=20000,
                    help="每页最多读多少行（防超大簿子）")
    ap.add_argument("--label-col", default=None,
                    help="手指标签列（列字母或表头文字），认错时用")
    ap.add_argument("--value-col", default=None, help="手指电流值列")
    ap.add_argument("--ladder", type=int, default=40,
                    help="逐级阶梯打前几行（0=不打）")
    ap.add_argument("--json", default=None, help="结构写一份 JSON（不含数据值）")
    ap.add_argument("--emit-groups", default=None,
                    help="把标签列写成分组清单模板（★含真实名字，只能放 private/）")
    args = ap.parse_args()

    files = []
    if os.path.isfile(args.path):
        files = [(None, None, args.path)]
    elif discover is None:
        sys.exit("summarize_chips 没导入成功，只能探单个文件")
    else:
        picked, _dropped, unknown, loose = discover(args.path)
        for (chip, mod, kind), b in sorted(picked.items(),
                                           key=lambda x: natkey(x[0][0])):
            if kind == KIND_CUR:
                files.append((chip, mod, b.path))
        print(f"扫到 {len(files)} 份电流文件"
              f"（{KIND_LABEL[KIND_CUR]}）:")
        for chip, mod, p in files:
            print(f"  {chip} / {mod}   {os.path.basename(p)}")
        if unknown:
            print(f"  另有 {len(unknown)} 份认不出类型/模块的文件")
        if loose:
            print(f"  ⚠ 根目录下还散着 {len(loose)} 份 xlsx 没扫（芯片目录以外的不读）")
    if not files:
        sys.exit("没找到电流文件（文件名里要有 current）")

    reps = []
    for chip, mod, p in files:
        try:
            rep = probe_book(p, args.sheet, args.max_rows,
                             args.label_col, args.value_col, args.ladder)
        except Exception as e:                        # noqa: B902
            print(f"\n✗ 读失败 {os.path.basename(p)}: {type(e).__name__}: {e}")
            continue
        rep["chip"], rep["module"] = chip, mod
        reps.append(rep)
        show(rep, args.max_list)

    if len(reps) > 1:
        print()
        print("=" * 78)
        print("格式兼容性（表头指纹相同 = 出簿脚本可以一视同仁）")
        sigs = {}
        for rep in reps:
            for sh in rep["sheets"]:
                if sh.get("empty"):
                    continue
                sigs.setdefault(signature(sh), []).append(
                    f"{rep.get('chip') or '?'}/{rep.get('module') or '?'}"
                    f"[{sh['sheet']}]")
        if len(sigs) == 1:
            print(f"  ✓ {len(list(sigs.values())[0])} 份全一致")
        else:
            print(f"  ⚠ 有 {len(sigs)} 种不同的表头，出簿前得先对齐：")
            base = None
            for i, (sig, who) in enumerate(sigs.items(), 1):
                names = sig.split("|")
                print(f"    第{i}种（{len(who)} 份）: {', '.join(who)}")
                print(f"       {len(names)} 列: {', '.join(names[:12])}"
                      + ("…" if len(names) > 12 else ""))
                if base is None:
                    base = set(names)
                else:
                    only_this = [n for n in names if n not in base]
                    only_base = [n for n in base if n not in set(names)]
                    if only_this:
                        print(f"       只有它有: {', '.join(only_this[:8])}")
                    if only_base:
                        print(f"       它缺: {', '.join(only_base[:8])}")

    if args.json:
        slim = []
        for rep in reps:
            slim.append({
                "chip": rep.get("chip"), "module": rep.get("module"),
                "name": rep["name"],
                "sheets": [{k: v for k, v in sh.items()
                            if k not in ("label_col", "value_cols", "cond_cols")}
                           for sh in rep["sheets"]],
            })
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(slim, f, ensure_ascii=False, indent=1)
        print(f"\n结构已写: {os.path.abspath(args.json)}")

    if args.emit_groups:
        groups = OrderedDict()
        for rep in reps:
            for sh in rep["sheets"]:
                lab = sh.get("label_col") if not sh.get("empty") else None
                if lab:
                    key = rep.get("module") or rep["name"]
                    groups.setdefault(key, [])
                    for t in lab["all_text"]:
                        if t not in groups[key]:
                            groups[key].append(t)
        with open(args.emit_groups, "w", encoding="utf-8") as f:
            json.dump({"__说明__": "每个模块下面是要进电流页的行，按这里的顺序排；"
                                   "删掉不要的行、调整顺序即可。",
                       "groups": groups}, f, ensure_ascii=False, indent=1)
        print(f"\n分组清单模板已写: {os.path.abspath(args.emit_groups)}")
        print("  ★ 这份文件含真实模块名 —— 放 private/ 或黄区本地，别提交。")


if __name__ == "__main__":
    main()
