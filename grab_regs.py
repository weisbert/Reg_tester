#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""grab_regs.py — 从「寄存器写序 Excel」按地址抓行（保持写入顺序）。

表单结构假设（自动探测表头行，列名匹配不到时按前四列兜底）：
    Register Name | Address | Value | Description

用法：
    python grab_regs.py <xlsx/xlsm> --prefix 5980D [--prefix 5980E] \
        --addr AAAA0000,BBBB0004 [--sheet Sheet1] [--json out.json] [--max-desc 60]

  --prefix   地址前缀（可多次）：命中所有以此开头的地址（大小写/0x 不敏感）
  --addr     明确地址清单（可多次，逗号分隔）
  --sheet    只处理指定 sheet（默认全部 sheet）
  --json     结果另存 JSON（含完整 Description，便于落盘比对）
  --max-desc 控制台 Description 截断长度（默认 60，0=不截断）

输出两段：
  1) 命中行按原始顺序的完整写序（写入顺序就是执行顺序，别重排）
  2) 被写过多次的地址 → 按时间序的值历史（分级使能/先清后置类操作一眼可见）

依赖 openpyxl；只读不写。
"""
import argparse
import io
import json
import re
import sys


def norm_addr(x):
    """地址归一：去空白/0x/下划线，大写。非字符串也转。"""
    if x is None:
        return ""
    s = str(x).strip().replace("_", "")
    s = re.sub(r"(?i)^0x", "", s)
    return s.upper()


def find_header(rows, scan=20):
    """在前 scan 行里找含 'address' 的表头行，返回 (行号idx, {列名->列idx})。找不到 -> (None, 兜底映射)。"""
    for i, row in enumerate(rows[:scan]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(c == "address" or c.replace(" ", "") == "registeraddress" for c in cells):
            m = {}
            for j, c in enumerate(cells):
                key = c.replace(" ", "")
                if key in ("registername", "regname", "name"):
                    m["name"] = j
                elif key in ("address", "registeraddress", "addr"):
                    m["addr"] = j
                elif key in ("value", "registervalue", "writevalue", "val"):
                    m["value"] = j
                elif key in ("description", "desc", "note", "comment"):
                    m["desc"] = j
            if "addr" in m:
                m.setdefault("name", 0)
                m.setdefault("value", m["addr"] + 1)
                m.setdefault("desc", m["value"] + 1)
                return i, m
    return None, {"name": 0, "addr": 1, "value": 2, "desc": 3}


def main(argv=None):
    ap = argparse.ArgumentParser(description="从寄存器写序 Excel 按地址抓行（保持顺序）")
    ap.add_argument("xlsx")
    ap.add_argument("--prefix", action="append", default=[], help="地址前缀，可多次")
    ap.add_argument("--addr", action="append", default=[], help="明确地址，逗号分隔，可多次")
    ap.add_argument("--sheet", help="只处理该 sheet（默认全部）")
    ap.add_argument("--json", help="结果另存 JSON")
    ap.add_argument("--max-desc", type=int, default=60)
    args = ap.parse_args(argv)

    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

    prefixes = [norm_addr(p) for p in args.prefix if p]
    explicit = set()
    for a in args.addr:
        for x in a.split(","):
            if x.strip():
                explicit.add(norm_addr(x))
    if not prefixes and not explicit:
        sys.exit("必须给 --prefix 或 --addr（否则整表都抓，没意义）")

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    sheets = [args.sheet] if args.sheet else wb.sheetnames

    hits = []          # 顺序命中: {seq, sheet, row, name, addr, value, desc}
    total_rows = 0
    for sn in sheets:
        if sn not in wb.sheetnames:
            sys.exit("找不到 sheet: %s（有: %s）" % (sn, ", ".join(wb.sheetnames)))
        ws = wb[sn]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        total_rows += len(rows)
        hidx, cm = find_header(rows)
        start = (hidx + 1) if hidx is not None else 0
        for i in range(start, len(rows)):
            row = rows[i]
            if not row or all(c is None for c in row):
                continue
            addr = norm_addr(row[cm["addr"]] if cm["addr"] < len(row) else None)
            if not addr:
                continue
            hit = addr in explicit or any(addr.startswith(p) for p in prefixes)
            if not hit:
                continue
            def cell(k):
                j = cm[k]
                return ("" if j >= len(row) or row[j] is None else str(row[j]).strip())
            hits.append({"seq": len(hits) + 1, "sheet": sn, "row": i + 1,
                         "name": cell("name"), "addr": addr,
                         "value": cell("value"), "desc": cell("desc")})

    print("扫描 %d sheet / %d 行，命中 %d 行（前缀=%s；明确地址 %d 个）"
          % (len(sheets), total_rows, len(hits), ",".join(prefixes) or "-", len(explicit)))
    print()
    print("== 命中写序（按表单原始顺序） ==")
    for h in hits:
        d = h["desc"]
        if args.max_desc and len(d) > args.max_desc:
            d = d[:args.max_desc] + "…"
        print("%4d | r%-5d %-28s %-10s %-10s %s" % (h["seq"], h["row"], h["name"], h["addr"], h["value"], d))

    # 多次写历史
    from collections import OrderedDict
    hist = OrderedDict()
    for h in hits:
        hist.setdefault(h["addr"], []).append(h)
    multi = {a: hs for a, hs in hist.items() if len(hs) > 1}
    if multi:
        print()
        print("== 同一地址多次写（按时间序） ==")
        for a, hs in multi.items():
            print("  %-10s %-28s: %s" % (a, hs[0]["name"], " -> ".join(x["value"] for x in hs)))

    # 明确地址里没命中的（表单根本没写它）
    missed = sorted(x for x in explicit if x not in hist)
    if missed:
        print()
        print("== 明确地址中表单没写的（重要线索：工作流程根本不碰它） ==")
        for a in missed:
            print("  " + a)

    if args.json:
        with io.open(args.json, "w", encoding="utf-8") as f:
            json.dump({"file": args.xlsx, "hits": hits, "missed_explicit": missed},
                      f, ensure_ascii=False, indent=1)
        print()
        print("JSON 已写: %s" % args.json)


if __name__ == "__main__":
    main()
