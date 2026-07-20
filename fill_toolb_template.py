#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
fill_toolb_template.py —— 把一批 test item 行填进【工具B 原厂 PLL测试模板】原件。

通用零 IP 引擎(代码⊥数据): 芯片专属的地址/值/模式全在 --data 指定的 JSON 里,
本脚本只有"开模板→按表头文字定位列→写行"的通用逻辑。可进公开仓库。

用法:
    python fill_toolb_template.py 模板.xlsx --data rows.json
    python fill_toolb_template.py 模板.xlsx --data rows.json -o 输出.xlsx
    python fill_toolb_template.py 模板.xlsx --data rows.json --dry-run     # 只探测列映射
    python fill_toolb_template.py 模板.xlsx --data rows.json --sheet "..." --header-row 4

data JSON 结构:
{
  "defaults": {"system":"PLL","fxo":26,"temperature":25,"addr_bare":true},
  "rows": [
    {"no":"...","mode":"...","testitem":"PLL_Test","test":"YES","current":"NO",
     "regs":[["1234ABCD","0x0FF0"], ...],       # [地址16进制, 值16进制]
     "switches":{"ReadBack":"YES"}          # 可选, 覆盖某些开关
    }, ...
  ]
}
  · regs: [[地址, 值], ...] 十六进制字符串, 每行最多写满模板 REG 对数(工具B=11)。
  · test/current: YES/NO; 未给时 test 默认 YES、current 默认 NO。
  · 其余测量开关(IPN/SpotPN/ReadBack/... /Chamber)默认 NO, 可用 switches 覆盖。

设计: 只写单元格 .value, 不动模板样式/下拉/其它 sheet(填写说明等), 保留原厂格式。
"""
import argparse, json, os, re, sys
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
try: sys.stdout.reconfigure(encoding="utf-8")
except Exception: pass

# 工具B 主表可识别字段 (归一化后精确匹配; test 与 testitem 靠精确匹配区分)
FIELDS = {
    "no":{"no","序号"}, "mode":{"mode"}, "system":{"system","system(pll/vco)"},
    "flo":{"flomhz","flo"}, "fvco":{"fvcomhz","fvco"}, "fxo":{"fxomhz","fxo"},
    "testitem":{"testitem"}, "test":{"test"}, "ipn":{"ipn"}, "spotpn":{"spotpn"},
    "readback":{"readback"}, "otherspur":{"otherspur"}, "vtune":{"vtune"},
    "vtemp":{"vtemp"}, "current":{"current"}, "pntrace":{"pntrace"},
    "spurlist":{"spurlist"}, "chamber":{"chamber"}, "temperature":{"temperature"},
}
# 测量开关列 (data 行的 switches 键名 -> 字段名)
SWITCH_FIELDS = ["test","ipn","spotpn","readback","otherspur","vtune","vtemp",
                 "current","pntrace","spurlist","chamber"]
SWITCH_KEY = {"test":"Test","ipn":"IPN","spotpn":"SpotPN","readback":"ReadBack",
              "otherspur":"OtherSpur","vtune":"Vtune","vtemp":"Vtemp",
              "current":"Current","pntrace":"PNTrace","spurlist":"SpurList","chamber":"Chamber"}

def norm(x):
    if x is None: return ""
    s = str(x).strip().lower().replace("℃","").replace("°c","")
    return re.sub(r"[\s_\.\-]+","", s)

def find_sheet_and_header(wb, force_sheet=None, force_hrow=None):
    best=None
    sheets=[wb[force_sheet]] if force_sheet else wb.worksheets
    for ws in sheets:
        maxc=min(ws.max_column,260)
        rows_try=[force_hrow] if force_hrow else range(1,min(ws.max_row,15)+1)
        for hr in rows_try:
            if hr is None or hr>ws.max_row: continue
            cm={}
            for c in range(1,maxc+1):
                nh=norm(ws.cell(hr,c).value)
                if not nh: continue
                for f,al in FIELDS.items():
                    if f in cm: continue
                    if nh in al or (f=="system" and nh.startswith("system")) \
                       or (f=="temperature" and nh.startswith("temperature")):
                        cm[f]=c
            score=len(cm)+(3 if {"no","mode","testitem","current"}<=set(cm) else 0)
            if best is None or score>best[0]: best=(score,ws,hr,cm)
    return best[1],best[2],best[3]

def find_reg_addr_cols(ws,hrow,maxc):
    tag={}
    for c in range(1,maxc+1):
        nh=norm(ws.cell(hrow,c).value)
        if not nh or "sweep" in nh: continue
        if re.search(r"addr",nh): tag[c]="a"
        elif re.search(r"val",nh): tag[c]="v"
    best=[]; cur=[]; c=1
    while c<=maxc:
        if tag.get(c)=="a" and tag.get(c+1)=="v": cur.append(c); c+=2
        else:
            if len(cur)>len(best): best=cur
            cur=[]; c+=1
    if len(cur)>len(best): best=cur
    return best[:11]

def fmt_addr(a,bare): return f"{int(a,16):08X}" if bare else f"0x{int(a,16):08X}"
def fmt_val(v):
    n=int(v,16); w=max(4,len(f"{n:X}")); return f"0x{n:0{w}X}"

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("template")
    ap.add_argument("--data",required=True,help="行数据 JSON")
    ap.add_argument("-o","--out",default=None)
    ap.add_argument("--sheet",default=None)
    ap.add_argument("--header-row",type=int,default=None)
    ap.add_argument("--style-row",type=int,default=None,
                    help="拿哪一行当'数据行格式'模板往下套(默认=表头下第一行)")
    ap.add_argument("--no-style",action="store_true",help="不复制模板数据行格式")
    ap.add_argument("--dry-run",action="store_true")
    args=ap.parse_args()

    for p in (args.template,args.data):
        if not os.path.exists(p): sys.exit("找不到: "+p)
    data=json.load(open(args.data,encoding="utf-8"))
    dfl=data.get("defaults",{})
    bare=dfl.get("addr_bare",True)
    rows=data["rows"]

    wb=openpyxl.load_workbook(args.template,keep_vba=args.template.lower().endswith(".xlsm"))
    ws,hrow,cm=find_sheet_and_header(wb,args.sheet,args.header_row)
    maxc=ws.max_column
    reg_cols=find_reg_addr_cols(ws,hrow,maxc)
    if len(reg_cols)<2:
        reg_cols=[42+2*i for i in range(11)]; reg_note="[!] 未识别 REG 区, 回退 AP(42)起 11 对"
    else: reg_note=f"识别 REG 区 {len(reg_cols)} 对"

    print("="*60)
    print("主 sheet :",ws.title)
    print("表头行   :",hrow,"→ 数据从第",hrow+1,"行写")
    for f in FIELDS:
        if f in cm: print(f"    {f:12s}-> {get_column_letter(cm[f])}({cm[f]})  '{ws.cell(hrow,cm[f]).value}'")
    miss=[f for f in ("no","mode","testitem","test","current") if f not in cm]
    if miss: print("    [!] 关键列缺:",miss," (用 --sheet/--header-row 指定)")
    print("REG 区   :",reg_note,"→",",".join(get_column_letter(c) for c in reg_cols))
    print("将写入   :",len(rows),"行")
    print("="*60)
    if args.dry_run:
        print("dry-run: 未写文件。列映射无误后去掉 --dry-run。"); return

    start=hrow+1
    managed=set(cm.values())
    for c in reg_cols: managed.add(c); managed.add(c+1)
    no_col=cm.get("no",1)
    existing_last=start-1
    for r in range(start,start+400):
        if ws.cell(r,no_col).value not in (None,""): existing_last=r

    # 抓"数据行格式"模板(默认表头下第一行): 每列的填充/字体/边框/对齐/数字格式,
    # 写完值后套到我写的每一行——否则超出模板已有格式区的行是裸单元格(格式丢失)。
    style_by_col=None
    if not args.no_style:
        srow=args.style_row or start
        style_by_col={}
        for c in range(1,maxc+1):
            s=ws.cell(srow,c)
            if s.has_style:
                style_by_col[c]=(copy(s.font),copy(s.fill),copy(s.border),
                                 copy(s.alignment),s.number_format,copy(s.protection))
    def apply_style(rr):
        if not style_by_col: return
        for c,(fo,fi,bo,al,nf,pr) in style_by_col.items():
            cell=ws.cell(rr,c)
            cell.font=copy(fo); cell.fill=copy(fi); cell.border=copy(bo)
            cell.alignment=copy(al); cell.number_format=nf; cell.protection=copy(pr)

    r=start
    for row in rows:
        def put(field,val):
            if field in cm and val is not None: ws.cell(r,cm[field],val)
        put("no",row.get("no")); put("mode",row.get("mode"))
        put("system",dfl.get("system","PLL")); put("testitem",row.get("testitem"))
        if "fxo" in dfl: put("fxo",dfl["fxo"])
        if "temperature" in dfl: put("temperature",dfl["temperature"])
        sw=row.get("switches",{})
        for f in SWITCH_FIELDS:
            if f not in cm: continue
            if f=="test":     v=row.get("test","YES")
            elif f=="current":v=row.get("current","NO")
            else:             v=sw.get(SWITCH_KEY[f],"NO")
            ws.cell(r,cm[f],v)
        regs=row.get("regs",[])
        for k in range(len(reg_cols)):
            ac=reg_cols[k]
            if k<len(regs):
                a,v=regs[k]; ws.cell(r,ac,fmt_addr(a,bare)); ws.cell(r,ac+1,fmt_val(v))
            else: ws.cell(r,ac,None); ws.cell(r,ac+1,None)
        apply_style(r)
        r+=1
    my_last=r-1
    cleared=0
    for rr in range(my_last+1,existing_last+1):
        for c in managed: ws.cell(rr,c,None)
        cleared+=1

    out=args.out or os.path.splitext(args.template)[0]+"_filled"+os.path.splitext(args.template)[1]
    wb.save(out)
    print(f"[OK] 写入 {len(rows)} 行 (第{start}~{my_last}行); 清理多余样例 {cleared} 行")
    print("输出:",out)

if __name__=="__main__":
    main()
