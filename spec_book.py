#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
spec_book.py — 汇总簿上「Spec / 仿真 / Limit」那七列：读出来存成数据，下次生成时填回去

汇总页上 Limit、Spec(Min/Typ/Max)、仿真(Min/Typ/Max) 这七列是留给人手填的
（浅黄底），填完判定列的公式自动出 PASS/FAIL。问题在于**每加一颗芯片就要重出
一份簿子**，手填的东西全没了——于是要么不敢重出，要么每次照着旧簿子再填一遍。

这个模块把那七列变成数据：
    读 read_specs()  填好的簿子 → 一份 JSON（哪张表哪一行填了什么）
    写 SpecBook      生成时按 (页, 表, 分组带, 测试项) 找回来，填进同样的格子

一行的身份 ＝ (页名, 表键, 分组带, 测试项)。
表键 ＝ 大标题**去掉折算/DSB 那截注解**（table_key）：折算倍数换了、SSB/DSB 换了，
标题会变，但那还是同一张表，spec 不该因此丢掉。
★ 反过来说，折算换了 spec 的**数值口径**也就变了（×4 折算 ＝ 相噪 +12 dB），
  所以读的时候把当时的完整标题一起记下来，填回去时对不上就当场喊一句——
  这类"名字对得上、口径对不上"的错最贵：表还是那张表，PASS/FAIL 却是错的。

零 IP：这里一个真实模块名/指标名都不写死，全部从簿子里读、按名字对。
"""

import re

# scale_title() 加在大标题末尾的那截注解，只有这两种开头
SCALE_TAG_RE = re.compile(r"（(?:折算|积分相噪)[^）]*）\s*$")

AXES_KEY = ("min", "typ", "max")
LIMITS = ("≤", "≥", "range")


def table_key(title):
    """大标题 → 稳定的表键（去掉折算/DSB 注解）。

    生成侧不用它——生成侧手里本来就有没加注解的那截标题，直接拿来当键；
    `SpecBook.table()` 会拿这个函数复核两边算出来的键一不一样，
    不一样就说明有张表的名字自己带着（…）尾巴，被这里误剥了。
    """
    return SCALE_TAG_RE.sub("", str(title or "")).strip()


def _nearest(name, cands):
    """名字对不上时，最像的那个是谁。"""
    import difflib
    t = re.sub(r"[^0-9a-z]", "", str(name).lower())
    if not t:
        return None
    best, score = None, 0.0
    for c in cands:
        n = re.sub(r"[^0-9a-z]", "", str(c).lower())
        if not n:
            continue
        r = 1.0 if (t in n or n in t) else difflib.SequenceMatcher(None, t, n).ratio()
        if r > score:
            best, score = c, r
    return best if score >= 0.6 else None


# ---------------------------------------------------------------- 读

def _cv(ws, r, c):
    v = ws.cell(r, c).value
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def read_specs(path):
    """填好的汇总簿 → (data, titles, stats, warns)

    data   {页: {表键: {分组带: {测试项: {"limit":…, "spec":{min/typ/max}, "sim":{…}}}}}}
    titles {页: {表键: 读的时候那张表的完整大标题}}
    stats  [(页, 表键, 结论行数, 填了spec的行数, 填了仿真的行数)]
    warns  [人话告警]

    表结构不写死列号，全靠表头自己说：col1＝"测试项"的那一行就是表头行，
    行内找 Spec / 仿真 / Limit / 判定 / 备注 这几个组名，组下面那一行找
    Min/Typ/Max。这样换了版本、挪了列，照样读得动。
    """
    import openpyxl
    # data_only=False：判定列的**公式**就是"这一行是结论行"的判据
    wb = openpyxl.load_workbook(path, data_only=False)
    data, titles, stats, warns = {}, {}, [], []

    for ws in wb.worksheets:
        maxr, maxc = ws.max_row, ws.max_column
        r = 1
        while r <= maxr:
            if _cv(ws, r, 1) != "测试项":
                r += 1
                continue
            hr, ar = r, r + 1
            title = str(_cv(ws, hr - 1, 1) or "") if hr > 1 else ""
            tkey = table_key(title) or f"{ws.title}!{hr}"

            groups = {}
            for c in range(1, maxc + 1):
                v = _cv(ws, hr, c)
                if isinstance(v, str) and v not in groups:
                    groups[v] = c
            miss = [k for k in ("Spec", "仿真", "判定", "备注") if k not in groups]
            if miss:
                warns.append(f"{ws.title} 第 {hr} 行的表头缺 {'/'.join(miss)}，"
                             f"这张表没读")
                r = ar + 1
                continue
            gcols = sorted(groups.values())

            def axes_of(c0):
                nxt = min([c for c in gcols if c > c0], default=maxc + 1)
                out = {}
                for c in range(c0, nxt):
                    lb = _cv(ws, ar, c)
                    if isinstance(lb, str) and lb.lower() in AXES_KEY:
                        out[lb.lower()] = c
                return out

            spec_ax, sim_ax = axes_of(groups["Spec"]), axes_of(groups["仿真"])
            c_limit, c_judge, width = groups.get("Limit"), groups["判定"], groups["备注"]

            tbl, cat, n_res, n_sp, n_si = {}, "", 0, 0, 0
            rr = ar + 1
            while rr <= maxr:
                vals = [_cv(ws, rr, c) for c in range(1, width + 1)]
                if not any(v is not None for v in vals):
                    break                       # 表与表之间空两行
                jd = _cv(ws, rr, c_judge)
                is_res = isinstance(jd, str) and jd.startswith("=")
                if not is_res:
                    # 分组带 ＝ 只有第一格有字、别的整行皆空；条件行/非结论行
                    # 是有值的，不能拿它当分组带（否则后面的行会挂错组）
                    if vals[0] is not None and not any(v is not None for v in vals[1:]):
                        cat = str(vals[0]).strip()
                    else:
                        stray = [c for _ax, c in
                                 list(spec_ax.items()) + list(sim_ax.items())
                                 if _cv(ws, rr, c) is not None]
                        if stray:
                            warns.append(
                                f"{ws.title} / {tkey} / 第 {rr} 行「{vals[0]}」"
                                f"不是结论行（没有判定公式），填在这行的 spec "
                                f"不会参与任何判定，也没被读走")
                    rr += 1
                    continue

                n_res += 1
                item = str(vals[0]).strip() if vals[0] is not None else ""
                got = {}
                for name, ax in (("spec", spec_ax), ("sim", sim_ax)):
                    d = {}
                    for k, c in ax.items():
                        v = _cv(ws, rr, c)
                        if v is not None:
                            d[k] = v
                    if d:
                        got[name] = d
                if got:
                    if c_limit is not None:
                        lim = _cv(ws, rr, c_limit)
                        if lim is not None:
                            got["limit"] = lim
                            if lim not in LIMITS:
                                warns.append(f"{ws.title} / {tkey} / {item}: "
                                             f"Limit 填的是「{lim}」，只认 "
                                             f"{' / '.join(LIMITS)}")
                    where = f"{ws.title} / {tkey} / {item}"
                    sp = got.get("spec", {})
                    lo, hi = sp.get("min"), sp.get("max")
                    if isinstance(lo, (int, float)) and isinstance(hi, (int, float)) \
                            and lo > hi:
                        warns.append(f"{where}: Spec Min {lo} > Max {hi}，填反了？"
                                     f"（这样填的话两头都会判 FAIL）")
                    for k, v in list(sp.items()) + list(got.get("sim", {}).items()):
                        if not isinstance(v, (int, float)):
                            warns.append(f"{where}: {k} 填的是文字「{v}」——"
                                         f"判定公式用 COUNT() 认数，文字这一头不判")
                    tbl.setdefault(cat, {})[item] = got
                    n_sp += 1 if "spec" in got else 0
                    n_si += 1 if "sim" in got else 0
                rr += 1

            if tbl:
                data.setdefault(ws.title, {})[tkey] = tbl
                titles.setdefault(ws.title, {})[tkey] = title
            if n_res == 0 and rr > ar + 1:
                # 整张表一行结论行都认不出来。最常见的原因是判定列被"选择性粘贴
                # 成值"了——那一列没了公式，这份簿子里就再没有任何东西说得清
                # 哪些行是结论行。填在上面的 spec 会一条都读不出来。
                warns.append(f"{ws.title} / {tkey}: 一行结论行都没认出来——"
                             f"判定列还有公式吗（整列被粘成值了？）。"
                             f"这张表填的 spec 一条都没读走")
            stats.append((ws.title, tkey, n_res, n_sp, n_si))
            r = rr
    return data, titles, stats, warns


def merge_into(cfg, data, titles, replace=False):
    """把读出来的 spec 并进配置字典（原地改），返回 (新增/覆盖的表, 原样留着的表)。

    ★ 语义：**这本簿子覆盖到的表，以簿子为准**——表里被清空的格子就此消失，
      不然"填错了想删掉"就没法做。簿子里没有的表原样留着：
      拿 --no-vco 出的簿子回填，不该把 VCO 那页的 spec 一起抹掉。
      要整块换掉用 replace=True。
    """
    old = cfg.get("spec") or {}
    old_t = (cfg.get("spec_meta") or {}).get("titles") or {}
    if replace:
        new, new_t, kept = {}, {}, []
    else:
        new = {sh: dict(tb) for sh, tb in old.items()}
        new_t = {sh: dict(tb) for sh, tb in old_t.items()}
        kept = [(sh, tk) for sh, tb in old.items() for tk in tb
                if tk not in (data.get(sh) or {})]
    hit = []
    for sh, tb in data.items():
        for tk, rows in tb.items():
            new.setdefault(sh, {})[tk] = rows
            new_t.setdefault(sh, {})[tk] = (titles.get(sh) or {}).get(tk, "")
            hit.append((sh, tk))
    cfg["spec"] = new
    meta = cfg.setdefault("spec_meta", {})
    meta["titles"] = new_t
    return hit, kept


# ---------------------------------------------------------------- 写

class SpecTable:
    """一张表的 spec。生成侧逐行问它要，问过的记下来，剩下的就是没对上的。"""

    __slots__ = ("rows", "sheet", "tkey", "asked", "renamed", "n_hit", "flat")

    def __init__(self, rows, sheet, tkey):
        self.rows = rows or {}
        self.sheet, self.tkey = sheet, tkey
        self.asked, self.renamed, self.n_hit = set(), [], 0
        self.flat = {}
        for cat, items in self.rows.items():
            for item in items:
                self.flat.setdefault(item, []).append(cat)

    def row(self, cat, item):
        self.asked.add((cat, item))
        d = (self.rows.get(cat) or {}).get(item)
        if d is None:
            # 分组带改过名（"Frequency Range" 那种整组挪走的事这条线已经出过），
            # 测试项在这张表里唯一就照它来，并且说一句
            cats = self.flat.get(item)
            if not cats or len(cats) != 1 or cats[0] == cat:
                return None
            d = self.rows[cats[0]][item]
            self.asked.add((cats[0], item))
            self.renamed.append((cats[0], cat, item))
        self.n_hit += 1
        return d


class SpecBook:
    """全簿的 spec。用法：每张表开头 table()，每个结论行 .row()，最后 report()。"""

    def __init__(self, data, titles=None, src=""):
        self.data = data or {}
        self.titles = titles or {}
        self.src = src
        self.tables = {}
        self.drift, self.badkey = [], []

    def __bool__(self):
        return bool(self.data)

    def table(self, sheet, base_title, full_title=None):
        tkey = str(base_title).strip()
        t = self.tables.get((sheet, tkey))
        if t is not None:
            return t
        t = SpecTable((self.data.get(sheet) or {}).get(tkey), sheet, tkey)
        self.tables[(sheet, tkey)] = t
        # 复核：读侧靠 table_key() 从完整标题剥出表键，写侧直接用没加注解那截。
        # 两边算不到一块去，说明有张表的名字自己带着（…）尾巴被误剥了——
        # 那样这张表的 spec 会**一次都对不上**，而且不报错。
        if full_title and table_key(full_title) != tkey:
            self.badkey.append((sheet, tkey, table_key(full_title)))
        old = (self.titles.get(sheet) or {}).get(tkey)
        if t.rows and old and full_title and old != full_title:
            self.drift.append((sheet, tkey, old, full_title))
        return t

    def _rows_total(self):
        return sum(1 for sh in self.data.values() for tb in sh.values()
                   for items in tb.values() for _ in items)

    def absent(self):
        """JSON 里有、这次**整张表都没出**的表。[(页, 表, 行数)]

        跟"某一行没对上"分开报：整页没出（--no-vco / 那类文件这次没有）是
        一句话的事，摊成几十行"这行没对上"只会把真正对不上的那几行冲掉。
        """
        out = []
        for sheet, tb in self.data.items():
            for tkey, cats in tb.items():
                if (sheet, tkey) not in self.tables:
                    out.append((sheet, tkey, sum(len(v) for v in cats.values())))
        return out

    def leftovers(self):
        """表出了、但这一行没对上的。[(页, 表, 组, 项, 猜)]"""
        out = []
        for sheet, tb in self.data.items():
            for tkey, cats in tb.items():
                t = self.tables.get((sheet, tkey))
                if t is None:
                    continue
                for cat, items in cats.items():
                    for item in items:
                        if (cat, item) in t.asked:
                            continue
                        out.append((sheet, tkey, cat, item,
                                    _nearest(item, sorted({i for _c, i in t.asked}))))
        return out

    def report(self):
        """控制台要打的话。没有 spec 就返回空。"""
        if not self.data:
            return []
        hit = sum(t.n_hit for t in self.tables.values())
        tot = self._rows_total()
        out = [f"Spec 套用: {hit}/{tot} 行填回了 Limit / Spec / 仿真 列"
               + (f"   ←{self.src}" if self.src else "")]
        for sheet, tkey, got in self.badkey:
            out.append(f"  ⚠⚠ {sheet} / 表名「{tkey}」自己带着（…）尾巴，"
                       f"读回来会被剥成「{got}」——这张表的 spec 存了也套不上。"
                       f"把表名末尾的括号去掉")
        for sheet, tkey, old, new in self.drift:
            out.append(f"  ⚠⚠ {sheet} / {tkey}: 填 spec 那会儿这张表的标题是"
                       f"「{old}」，现在是「{new}」——**折算/SSB-DSB 口径变了，"
                       f"spec 的数还是老口径的**。spec 照填了，判定先别信，"
                       f"确认一下这些数该不该跟着换")
        for t in self.tables.values():
            for old_cat, new_cat, item in t.renamed:
                out.append(f"  · {t.sheet} / {t.tkey} / {item}: 分组带从"
                           f"「{old_cat}」挪到了「{new_cat}」，按测试项对上了")
        for sheet, tkey, n in self.absent():
            out.append(f"  · {sheet} / {tkey}: 这次没出这张表，它的 {n} 行 spec "
                       f"原样留在配置里没动")
        miss = self.leftovers()
        if miss:
            out.append(f"  ⚠ {len(miss)} 行 spec 没对上（表上没有这一行，"
                       f"它们的数就此不参与判定）:")
            for sheet, tkey, cat, item, g in miss[:12]:
                out.append(f"      {sheet} / {tkey} / {cat or '—'} / {item}"
                           + (f"（是不是 {g}？）" if g else ""))
            if len(miss) > 12:
                out.append(f"      …共 {len(miss)} 行")
        return out

    def audit_rows(self):
        """给隐藏的 _审计 页：(页, 表, 分组带, 测试项, 状态)。"""
        out = []
        for sheet, tkey, n in self.absent():
            out.append((sheet, tkey, "—", "整张表",
                        f"这次没出这张表，{n} 行 spec 没用上"))
        for sheet, tkey, cat, item, g in self.leftovers():
            out.append((sheet, tkey, cat or "—", item,
                        "没对上" + (f"（是不是 {g}？）" if g else "")))
        for sheet, tkey, old, new in self.drift:
            out.append((sheet, tkey, "—", "整张表",
                        f"填 spec 时标题是「{old}」，现在是「{new}」，口径可能变了"))
        for t in self.tables.values():
            for old_cat, new_cat, item in t.renamed:
                out.append((t.sheet, t.tkey, new_cat, item,
                            f"分组带原来叫「{old_cat}」，按测试项对上的"))
        return out
