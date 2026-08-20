#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
sweep_lib.py — 扫描簿报表的公共层（读数 / 分段 / 统计 / 画格子 / 图表原语）

从 summarize_pll_sweep.py 抽出来的。抽的理由：同一份原厂宽表现在有两个消费者
（单簿脚本 summarize_pll_sweep.py，跨芯片汇总 summarize_chips.py），
读取和统计的口径必须只有一份——两份必然漂移，而漂移出来的是"两份报表同一个
指标报不同的数"，那种账最难对。

这里只放**通用引擎**：没有任何真实模块名/寄存器名/频点常量。
真实的东西全靠表头名和命令行参数进来。

分四段：
    ① 取值      is_blank / num / txt / fmt_num / fmt_hz
    ② 结构      Columns（按表头定位、重名列报出来）/ Item / build_items
    ③ 行与段    Row / Leg / segment / room_temp / leg_series / stats
    ④ 装载      load_sweep()  ——  一行把一份簿子读成 (rows, legs, items)
    ⑤ 出稿      styles / put / 图表原语（配色、纵轴范围、图例）

依赖：只用 openpyxl（且只在真的要读写 xlsx 时才 import）。
"""

import re
from collections import OrderedDict

# report-forge compliance 表的配色（黄表头 / 米色条件行 / 白结果行 / 红超规）
FILL_HEADER = "FFFF00"
FILL_GROUP = "EEECE1"
FILL_RESULT = "FFFFFF"
FILL_SEP = "B8CCE4"
COLOR_FLAG = "FF0000"
COLOR_PASS = "006100"
FILL_FAIL = "FFC7CE"
FILL_PASS = "C6EFCE"
FILL_INPUT = "FFF2CC"       # 要人填的格子（Spec / 仿真值 / 判据），浅黄一眼看得出
# 宽表分区用的灰阶与低饱和色。★色相总数压在 3 个以内（黄=表头 / 浅黄=要人填 /
# 浅蓝=判定依据），其余分区一律用灰阶深浅 + 分隔，否则 40 列的表变成花布，
# 而且灰度打印和色弱下几种色相会互相撞。
FILL_ZONE = "F2F2F2"        # 组内的第二个功能区（极值列）——极浅灰
FILL_RAIL = "D9D9D9"        # 组与组之间的竖栏（中灰），界定"这一块是一片"
FILL_SUM = "DDEBF7"         # 汇总组：全表唯一需要跳出来的东西
COLOR_MUTED = "595959"      # 注释性文字（@℃ 这类），要退到背景里去

# 表里表示"没测/不适用"的占位符，一律当空值
BLANK_TOKENS = {"", "-", "--", "—", "n/a", "na", "null", "none", "#n/a"}


class SweepError(Exception):
    """簿子读不下去（缺列/没数据行）。调用方决定是退出还是跳过这一份。"""


# ---------------------------------------------------------------- ① 取值

def is_blank(v):
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip().lower() in BLANK_TOKENS
    return False


def num(v):
    """数值化；取不到数就返回 None。'-' 这类占位符算没测。"""
    if is_blank(v) or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    return "" if v is None else str(v).strip()


def fmt_num(x, nd=3):
    if x is None:
        return None
    if abs(x - round(x)) < 1e-12:
        return int(round(x))
    return round(x, nd)


def fmt_hz(mhz):
    """offset 频率 MHz -> 好读的标签：0.001->1kHz, 1->1MHz。"""
    if mhz is None:
        return "?"
    if mhz < 1:
        khz = mhz * 1000.0
        return f"{fmt_num(khz)}kHz"
    return f"{fmt_num(mhz)}MHz"


def median(vals):
    v = sorted(x for x in vals if x is not None)
    if not v:
        return None
    n = len(v)
    return v[n // 2] if n % 2 else (v[n // 2 - 1] + v[n // 2]) / 2.0


# ---------------------------------------------------------------- ② 结构

class Columns:
    """按表头名定位列；重名列会报出来。

    这类原厂模板扩列时复制粘贴不改序号很常见（同名列出现两遍）。
    按名字取只会一直拿到第一份，第二份永远读不到且不报错——
    所以这里把重名的列位置全记下来，取第一份并留警告。
    """

    def __init__(self, header):
        self.header = list(header)
        self.pos = OrderedDict()
        for i, h in enumerate(header):
            k = txt(h)
            if k:
                self.pos.setdefault(k, []).append(i)
        self.duplicates = {k: v for k, v in self.pos.items() if len(v) > 1}

    def idx(self, name):
        v = self.pos.get(name)
        return v[0] if v else None

    def find(self, *patterns, exclude=()):
        """按正则找第一个匹配的表头，返回 (名字, 列下标)。"""
        for k in self.pos:
            kl = k.lower()
            if any(re.search(p, kl) for p in patterns) and \
               not any(re.search(p, kl) for p in exclude):
                return k, self.pos[k][0]
        return None, None

    def dup_report(self):
        from openpyxl.utils import get_column_letter as gl
        return [f"重复列名 {k}：出现在 {', '.join(gl(i + 1) for i in v)}，"
                f"按名字只取到第一个（{gl(v[0] + 1)}）"
                for k, v in self.duplicates.items()]


class Item:
    __slots__ = ("cat", "label", "unit", "col", "src", "text_src",
                 "pick", "pick_note", "pick_stats")

    def __init__(self, cat, label, unit, col, src):
        self.cat, self.label, self.unit, self.col, self.src = cat, label, unit, col, src
        self.text_src = False       # 原始列是文本型数字？是的话引用要加双负号
        # pick(raw) -> 这一行该读哪一列。给"值不在固定列上"的指标用（杂散清单）。
        # 挑出来的仍然是**原表上的一个真格子**，所以引用照样指得回去。
        self.pick = None
        self.pick_note = ""
        self.pick_stats = None

    @property
    def key(self):
        """跨簿子对齐用的键。列位置在不同芯片的簿子里可能不同，名字才是身份。"""
        return self.label

    @property
    def has_cell(self):
        """这一行在原表里有没有自己的列。

        ★★ False ＝ 值只能逐行从表尾杂散清单里挑（`spur_targets` 加出来的频点，
          模板里根本没有这一列）。此时 `col` 是个**超出表宽的合成号**，只当
          `vals` / `src` 的键用，不是列号：
            · 拿它去 `raw[it.col]` 会 IndexError（守卫都写着 `it.col < len(raw)`）；
            · 要拼引用只能走 `row.col_of(it)`（那里记着这一行真正读的那一格）；
            · 它也没有"模板那一格"可以退回去，所以整份没搜到就不出这一行。
        ★ 判据别各处自己写 `not it.src`：它有 7 处要用，散着写就是 7 处各自
          记住一条没名字的约定——而这份代码里所有"静默错值"都是这么来的。
        """
        return bool(self.src)



SIMPLE_ITEMS = [
    # (表头, 分类, 单位)
    ("Freq_MHz", "Frequency", "MHz"),
    ("Power_dBm", "Output", "dBm"),
    ("IPN_SSB", "Phase Noise", "dBc"),
    ("IPN_Omit_SSB", "Phase Noise", "dBc"),
    ("Vtune_V", "Tuning", "V"),
    ("Vtemp_V", "Temp Sensor", "V"),
    ("Current_mA", "Current", "mA"),
]
# 成对列：<前缀>Freq<i> 给频点、<前缀>Result<i> 给结果
PAIRED_ITEMS = [
    ("SpotPN", "Phase Noise", "dBc/Hz", "SpotPN@{f}"),
    ("OtherSpur", "Spur", "dBc", "Spur@{f}"),
]


def find_spur_list(header, rows, min_pairs=2):
    """表尾那段**没有表头**的成对数据 ＝ 仪器搜出来的杂散清单 (偏移MHz, 电平dBc)。

    ★ 为什么要它：模板里的 `OtherSpurFreq<i>/Result<i>` 是**在标称频点上量一个数**。
      真实杂散只要偏出去几百 kHz，那一格量到的就是**噪底**，
      不是杂散——比真值低十几个 dB，报出去等于说"这颗片子杂散很干净"。
      真值在这段清单里。

    ★ 判据全是结构性的，不认列字母也不认表头名（不同批次导出列位会挪）：
      ① 落在最后一个有表头的列**右边** ② 自己没有表头 ③ 成对出现、
      两列的非空行数相等 ④ **奇数列为负**（dBc 电平），偶数列是数（偏移频率）。
      四条都过才认。认错了是**安静地把两列错位读成杂散**，比读不到糟得多。
    ★ 偏移**允许为负**：有的簿子把载波下边那一侧记成 −26 MHz。原来要求
      偶数列非负，于是带负偏移的那份直接认不出来（真数据上 18/62 反例）。
      判据仍然挡得住"整体错开一列"——那种情况下奇数列会拿到正的频率值。

    返回 (pairs, why)：pairs=[(偏移列, 电平列), …]，认不出来时 pairs 为空、
    why 说明卡在哪一条。
    """
    width = max((len(r) for r in rows), default=0)
    last_head = max((i for i, h in enumerate(header) if not is_blank(h)), default=-1)
    if last_head + 1 >= width:
        return [], "表尾没有多余的列"

    def has_data(c):
        return any(c < len(r) and not is_blank(r[c]) for r in rows)

    start = next((c for c in range(last_head + 1, width) if has_data(c)), None)
    if start is None:
        return [], "表头右边那些列全是空的"
    end = start
    for c in range(start + 1, width):
        if has_data(c):
            end = c
        else:
            break                       # 中间断开就到此为止
    n_pairs = (end - start + 1) // 2
    if n_pairs < min_pairs:
        return [], f"只有 {end - start + 1} 列有数据，凑不满 {min_pairs} 对"

    pairs, ok, bad = [], 0, 0
    for k in range(n_pairs):
        fc, lc = start + 2 * k, start + 2 * k + 1
        nf = sum(1 for r in rows if fc < len(r) and not is_blank(r[fc]))
        nl = sum(1 for r in rows if lc < len(r) and not is_blank(r[lc]))
        if nf != nl:
            return [], (f"第 {k + 1} 对两列的非空行数对不上（{nf} vs {nl}），"
                        f"不像 (偏移, 电平) 成对")
        for r in rows:
            f = num(r[fc]) if fc < len(r) else None
            v = num(r[lc]) if lc < len(r) else None
            if f is None or v is None:
                continue
            ok, bad = (ok + 1, bad) if v < 0 else (ok, bad + 1)
        pairs.append((fc, lc))
    if not ok or bad > ok * 0.02:
        return [], (f"奇数列该是负的 dBc，这条不成立（{bad}/{ok + bad} 反例），"
                    f"这段多半不是杂散清单")
    return pairs, ""


def spur_picker(pairs, target, tol):
    """在标称频点 ±tol 的窗口里，挑**幅度最大**（dBc 最靠近 0）的那一条。

    取最大不取最近：这一格是拿去跟 spec 比的，窗口里有两条时该报最坏的那条。
    挑出来的是原表上的真格子，所以返回列号——引用还能指回去。
    """
    def pick(raw):
        best = None
        for fc, lc in pairs:
            f = num(raw[fc]) if fc < len(raw) else None
            v = num(raw[lc]) if lc < len(raw) else None
            # 偏移可能记成负数（载波下边那一侧），按绝对值对标称频点
            if f is None or v is None or abs(abs(f) - target) > tol:
                continue
            if best is None or v > best[1]:
                best = (lc, v)
        return best[0] if best else None
    return pick


# 窗口不许超过标称频点自身的这个比例（见 spur_windows 第 ③ 条）
SPUR_REL_TOL = 0.25


def spur_windows(targets, tol):
    """每个标称偏移各自的窗口半宽 {标称: 半宽}。三条取最小：

    ① `tol`（默认 ±2 MHz，是从 26/52 那两个频点来的老默认值）。
    ② **两个标称频点的窗口不许重叠**：2/4/6 这种彼此只差 2 MHz 的分量，用 ±2
      会让同一条杂散**同时落进相邻两行**——两行报同一个数，表上一点看不出来。
      相邻间距 d 时半宽最多 0.45d（留一点缝，正好落在中间的那条谁也不进）。
    ③ ★★**不许伸到标称自身的 25% 之外**。②只挡得住"另一个标称频点"，挡不住
      "清单里真有、但你没配"的那一条：只配 `[4, 20]` 时，`Spur@4MHz` 用 ±2 会把
      **2.006 MHz 那条**捞进来（|2.006−4|<2）——这一行报的根本不是 4M，
      **取了别人的数，还挂着 4M 的名字**。按标称成比例收之后 4M 窗口是 ±1，
      够容下 4.011 那种偏差，又够把 2.006 挡在外面。
    ★ 三条都只收窄不放宽：26/52（0.25×26=6.5、0.25×52=13 都大于 tol）照旧 ±tol
      —— 老簿子的数一格不动。
    ★ 宁可漏报不要错报：窗口收紧顶多让某一行"0/N 行命中"，那句控制台会打；
      而报错一条只在备注的"实测偏移"上留一行痕，很容易滑过去。
    """
    ts = sorted({float(t) for t in targets if t is not None})
    out = {}
    for i, t in enumerate(ts):
        gaps = ([t - ts[i - 1]] if i else []) + \
               ([ts[i + 1] - t] if i + 1 < len(ts) else [])
        out[t] = min([tol, abs(t) * SPUR_REL_TOL] + [g * 0.45 for g in gaps])
    return out


def spur_off_warn(it):
    """取到的那条离标称有多远——远到半个窗口以外就喊一句。

    ★ 窗口收紧之后这种事应该很少，但"我叫它 4M、它其实是别的那条"这个错
      **只在备注的实测偏移上留一行痕**，不主动喊就会被当成 4M 的数发出去。
    """
    st = it.pick_stats or {}
    t, off, w = st.get("target"), st.get("off"), st.get("tol")
    if t is None or off is None or not w:
        return ""
    if abs(abs(off) - t) <= w * 0.5:
        return ""
    return ("%s: 取到的那条在 %s MHz，离标称 %s 有 %s MHz——确认这是不是你要的"
            "那个分量（窗口现在是 ±%s，要更严就调 --spur-tol）"
            % (it.label, fmt_num(off), fmt_num(t), fmt_num(abs(abs(off) - t), 2),
               fmt_num(w, 2)))


def attach_spur_list(cols, rows, items, tol=2.0, targets=()):
    """把 Spur@<标称> 那几行的取值改成"从尾部杂散清单里挑"，并按需要**加几行**。

    标称频点有两个来源：
      ① 模板的 `OtherSpurFreq<i>`（这份测试原本声明要看哪几个，真数据是 26/52）；
      ② `targets`＝额外要报的分量（2/4/6/20…）。**模板里根本没有这几列**，
         但它们在仪器搜出来的清单里，所以照样报得出来：给每个标称建一行，
         逐行在窗口里挑最大的那一条。
    ★★ 某一行清单里没有这个分量，那一格就**留空**——不是 0、也不是噪底。
      "有些行只有 2M、有些只有 4M、有些两个都有"就是这么落到表上的。
    找不到清单就什么都不做，退回原来的行为。

    返回 (why, notes)：why＝认不出清单时卡在哪一条；notes＝要打给人看的几句
    （窗口被收窄了 / 清单里还有哪些偏移没人认领）。
    """
    from openpyxl.utils import get_column_letter as gl
    pairs, why = find_spur_list(cols.header, rows)
    # ★★ 认不出清单也照样把 targets 那几行建出来（只是没有值）：两个模块的表
    #   要能横着比，行集合就得一样。少一行看着像 bug，空一行至少能问"为什么空"
    #   ——而"为什么空"控制台正好写着。
    where = (f"{gl(pairs[0][0] + 1)}..{gl(pairs[-1][1] + 1)}（{len(pairs)} 对）"
             if pairs else "")

    # ① 模板声明的标称频点（每个对应一个已经建好的 Spur@ 行）
    declared = OrderedDict()
    for it in items:
        if not it.label.startswith("Spur@"):
            continue
        fc = cols.idx(it.src.replace("Result", "Freq")) if it.has_cell else None
        if fc is None:
            continue
        tgts = {num(r[fc]) for r in rows if fc < len(r) and num(r[fc]) is not None}
        if len(tgts) != 1:
            continue                    # 标称频点在这份簿里不唯一，不动它
        declared[it] = tgts.pop()

    extra = sorted({float(t) for t in (targets or ()) if t is not None})
    wins = spur_windows(list(declared.values()) + extra, tol)

    def attach(it, t):
        it.pick = spur_picker(pairs, t, wins[t])
        it.pick_note = (f"杂散清单 {where} 里 {fmt_num(t)}±{fmt_num(wins[t], 2)} MHz "
                        f"内最大的一条")
        # 实测偏移落在哪儿 —— 报表备注要写它（"你这 26M 是在哪测的"）。
        # 在这里算是因为两个读取层都调这个函数，算一次两边都有。
        offs = [num(r[c - 1]) for r in rows for c in (it.pick(r),)
                if c is not None and c - 1 < len(r)]
        it.pick_stats = {"target": t, "tol": wins[t], "off": median(offs),
                         "n": len(offs)}
        # 文本型数字要重判：现在读的是清单那几列，不是原来那一格
        it.text_src = any(
            isinstance(r[c], str) and num(r[c]) is not None
            for r in rows for c in (it.pick(r),) if c is not None and c < len(r))

    if pairs:
        for it, t in declared.items():
            attach(it, t)

    # ② 模板里没有的分量：新建一行。col 给一个**超出表宽**的合成号——
    #    它只当 vals/src 的键用（不是真列），取值走 pick，引用走 row.src 里
    #    记下的那个真格子。★ 合成号必须超出表宽：负数会被 raw[-1] 绕回最后一列，
    #    而 `it.col < len(raw)` 这种判据在别处到处都是。
    nxt = max([len(cols.header)] + [len(r) for r in rows]
              + [it.col + 1 for it in items])
    for t in extra:
        if any(abs(t - d) <= wins[t] for d in declared.values()):
            continue                    # 模板已经声明过这个频点，别报两遍
        nxt += 1
        # 用现有 item 的类来建：两个读取层各有一份 Item（同形不同类），
        # 这里写死哪一个都会在另一边埋一颗雷。
        it = type(items[0])("Spur", f"Spur@{fmt_hz(t)}", "dBc", nxt, "")
        if pairs:
            attach(it, t)
        else:
            # 没有清单就没有值，但这一行照样存在（pick_stats 留着，
            # 备注和控制台才说得出"标称多少、为什么空"）
            it.pick_stats = {"target": t, "tol": wins[t], "off": None, "n": 0}
        items.append(it)

    # ★ Spur@ 那一族按频点排好再交出去：模板声明的（26/52）和加出来的
    #   （2/4/6/20）本来一前一后，不排就是 26 → 52 → 2 → 20。频点跳着走的表，
    #   评审第一句就是"这表怎么回事"。整族原地重排，别的行一个都不动。
    sp = [i for i, it in enumerate(items) if it.label.startswith("Spur@")]
    if sp:
        block = sorted((items[i] for i in sp),
                       key=lambda it: (it.pick_stats or {}).get("target", 1e9))
        rest = [it for i, it in enumerate(items) if i not in set(sp)]
        items[:] = rest[:sp[0]] + block + rest[sp[0]:]

    notes = []
    if not pairs:
        if extra:
            notes.append("认不出表尾的杂散清单，配的 %s MHz 这几行只能**整列留空**"
                         "（模板声明的那几个还能退回模板那一格）"
                         % "/".join(str(fmt_num(t)) for t in extra))
        return why, notes
    narrowed = [t for t in wins if wins[t] < tol - 1e-9]
    if narrowed:
        notes.append("杂散窗口：%s MHz 这几个彼此挨得近，窗口各自收到 ±%s"
                     "（不收的话同一条杂散会同时进相邻两行）"
                     % ("/".join(str(fmt_num(t)) for t in sorted(narrowed)),
                        fmt_num(min(wins[t] for t in narrowed), 2)))
    # ★ 反方向也要说：清单里有、却没有任何标称频点收走的偏移。
    #   不说的话，"我要的那个分量到底在不在这份数据里"只能靠猜。
    claimed = [(t, w) for t, w in wins.items()]
    loose = {}
    for r in rows:
        for fc, lc in pairs:
            f = num(r[fc]) if fc < len(r) else None
            v = num(r[lc]) if lc < len(r) else None
            if f is None or v is None:
                continue
            a = abs(f)
            if any(abs(a - t) <= w for t, w in claimed):
                continue
            k = round(a, 1)
            loose[k] = loose.get(k, 0) + 1
    if loose:
        top = sorted(loose.items(), key=lambda kv: -kv[1])[:8]
        notes.append("清单里还有这些偏移没被任何标称频点收走（要报的话加进 "
                     "spur_targets）: "
                     + "，".join("%s MHz×%d 行" % (fmt_num(k), n) for k, n in top)
                     + ("，…共 %d 种" % len(loose) if len(loose) > 8 else ""))
    return "", notes


def build_items(cols, rows):
    """识别有数据的结果列。结果列整片是空的（占位列）自动丢掉，并记在 dropped 里。"""
    items, dropped = [], []

    def nonempty(ci):
        return sum(1 for r in rows if ci < len(r) and not is_blank(r[ci]))

    for name, cat, unit in SIMPLE_ITEMS:
        ci = cols.idx(name)
        if ci is None:
            continue
        if nonempty(ci) == 0:
            dropped.append((name, "整列没有数据"))
            continue
        items.append(Item(cat, name, unit, ci, name))

    for prefix, cat, unit, tpl in PAIRED_ITEMS:
        i = 1
        while True:
            fc, rc = cols.idx(f"{prefix}Freq{i}"), cols.idx(f"{prefix}Result{i}")
            if fc is None and rc is None:
                break
            i += 1
            if rc is None:
                continue
            n = nonempty(rc)
            if n == 0:
                dropped.append((f"{prefix}Result{i-1}", "整列没有数据"))
                continue
            freqs = {num(r[fc]) for r in rows
                     if fc is not None and fc < len(r) and num(r[fc]) is not None}
            f = fmt_hz(sorted(freqs)[0]) if len(freqs) == 1 else (
                "/".join(fmt_hz(x) for x in sorted(freqs)[:3]) or f"#{i-1}")
            items.append(Item(cat, tpl.format(f=f), unit, rc, f"{prefix}Result{i-1}"))

    # 标出「文本型数字」列：Python 解得动、Excel 不认。不标出来的话，
    # 引用过去 COUNT/MIN 全落空，极值温度那格直接 #N/A。
    for it in items:
        n_txt = sum(1 for r in rows
                    if it.col < len(r) and isinstance(r[it.col], str)
                    and num(r[it.col]) is not None)
        it.text_src = n_txt > 0
    return items, dropped


# ---------------------------------------------------------------- ③ 行与段

class Row:
    __slots__ = ("xl", "temp", "leg", "kind", "vals", "raw", "src")

    def __init__(self, xl, temp, kind, raw):
        self.xl, self.temp, self.kind, self.raw = xl, temp, kind, raw
        self.leg, self.vals = None, {}
        self.src = {}          # {指标键: 这一行实际读的那一列}，只有 pick 类指标有

    def col_of(self, item):
        """这一行里，这个指标的值出自哪一列——引用要指到这里，不是 item.col。

        对固定列的指标就是 item.col；对杂散清单这种"每行挑一格"的指标，
        挑中的列逐行不同（不同行的最大杂散落在清单里不同的位置）。
        """
        return self.src.get(item.col, item.col)


class Leg:
    def __init__(self, n, lock_temp):
        self.n, self.lock_temp = n, lock_temp
        self.rows = []

    @property
    def temps(self):
        return [r.temp for r in self.rows if r.temp is not None]

    @property
    def direction(self):
        t = self.temps
        if len(t) < 2:
            return ""
        return "↑" if t[-1] > t[0] else ("↓" if t[-1] < t[0] else "")

    @property
    def title(self):
        lt = fmt_num(self.lock_temp)
        return f"段{self.n} 锁@{lt}℃" if lt is not None else f"段{self.n}"

    @property
    def stage(self):
        t = self.temps
        if not t:
            return ""
        return f"{fmt_num(t[0])}→{fmt_num(t[-1])}℃ {self.direction}".strip()


def segment(rows, leg_col_i, lock_re):
    """按重锁事件切段。重锁行本身也是测量点（它有结果），算作本段第一个点。"""
    legs, cur = [], None
    orphan = []
    for r in rows:
        mode = txt(r.raw[leg_col_i]) if leg_col_i is not None and leg_col_i < len(r.raw) else ""
        if lock_re.search(mode):
            cur = Leg(len(legs) + 1, r.temp)
            legs.append(cur)
            r.kind = "lock"
        if cur is None:
            orphan.append(r)          # 第一次重锁之前的行
            continue
        r.leg = cur
        cur.rows.append(r)
    return legs, orphan


def room_temp(legs, target=25.0):
    """"常温"取实际测过的、最接近 target 的那个温度。

    别用"排序后的中位数"——那是统计意义的中间，不是工程上的常温（曾挑到 35℃）。
    """
    ts = sorted({r.temp for lg in legs for r in lg.rows if r.temp is not None})
    return min(ts, key=lambda t: abs(t - target)) if ts else None


# ---------------------------------------------------------------- 统计

def leg_series(leg, item, with_lock=False):
    """本段的「温度 -> 值」去重视图：同段同温有多个点时取最后一个。

    ★ 默认剔掉重锁行。重锁是这组数据的**取得条件**，不是被考核的性能：
    把锁定瞬间的读数混进"全温最坏值"里，等于拿锁的过程去判性能的规格。
    （实际上每个重锁点后面都紧跟一个同温的稳定测量点，所以剔掉不丢温度。）

    ★ 汇总统计和温度明细页**共用这一份**。否则汇总按全部行取极值、明细一格
    只放得下一个值，就会出现「汇总报的极值在明细里查无此值」——报表一旦对不上账
    就没人敢信了。
    """
    out = OrderedDict()
    for r in leg.rows:
        if r.kind == "lock" and not with_lock:
            continue
        v = r.vals.get(item.col)
        if r.temp is not None and v is not None:
            out[r.temp] = v
    return out


def _extremes(pairs):
    if not pairs:
        return None
    lo = min(pairs, key=lambda x: x[0])
    hi = max(pairs, key=lambda x: x[0])
    return {"min": lo[0], "min_t": lo[1], "max": hi[0], "max_t": hi[1],
            "delta": hi[0] - lo[0], "n": len(pairs)}


def stats(leg, item):
    return _extremes([(v, t) for t, v in leg_series(leg, item).items()])


def stats_all(legs, item, room_t=None):
    """全温统计。TYP 取常温点的中位数——常温在温巡里被经过好几次，
    取中位数比随便挑一次稳。"""
    s = _extremes([(v, t) for lg in legs
                   for t, v in leg_series(lg, item).items()])
    if s and room_t is not None:
        rv = [v for lg in legs for t, v in leg_series(lg, item).items() if t == room_t]
        if rv:
            s["typ"] = median(rv)
            s["typ_n"] = len(rv)
    return s


# ---------------------------------------------------------------- ④ 装载

class Sweep:
    """一份读完的扫描簿。字段名跟原来 main() 里的局部变量一一对应。"""

    def __init__(self, **kw):
        self.__dict__.update(kw)

    def item(self, *labels):
        """按标签取指标（给多个名字＝依次尝试）。"""
        for n in labels:
            for it in self.items:
                if it.label == n:
                    return it
        return None

    @property
    def temps(self):
        return sorted({r.temp for lg in self.legs for r in lg.rows
                       if r.temp is not None})


# `<v ...>` / `<is ...>` 也认：漏认一种写法＝静默少读一段数据。
# sheetData 里以 v / is 开头的元素只有这两个，不会误命中。
_MARK_RE = re.compile(rb'<row r="(\d+)"|<v[ >]|<is[ >]')


def last_value_row(path, part, chunk=1 << 20):
    """扫出「最后一个带值的行号」。只在字节流上找，**不 parse**。拿不准就返回 None。

    ★★ 为什么值得单独扫一遍：这条线的簿子真数据几百行，表里却写着一万行——
      多出来的全是**只有格式、没有值**的空格子（`<c r="HG9999" s="0"/>`，一个 17 字节，
      215 列 × 一万行 ≈ 40 MB）。openpyxl 就算 read_only 也得把这两百万个格子
      一个个 parse 出来：真数据 139 行的一份要 3.1s，先扫出真数据到哪儿、
      再只 parse 到那儿，0.08s——**37×**。真数据一跑 15 份簿子，读占 98% 的时间。
    ★ 判据是**精确的**，不是猜一个"连续多少空行就停"的阈值：一行里没有 `<v>`／`<is>`
      就是一个值都没有，读进来也是整行 None，会被"掐掉尾部空行"那一步掐掉——
      结果一个字节都不差。（只有公式没有缓存值的行同理：data_only=True 下它本来
      就读成 None。）
    ★ 任何一步不成立（拿不到 part、行号没写 r、正则对不上）就返回 None，
      调用方退回全量读。**宁可慢，不可少读。**
    ★ 分块扫、不把整个 sheet1.xml 读进内存（40 MB 一份，这条线正为内存发愁）。
      块与块之间留 40 字节接缝，免得标签正好被切断；重复扫到的标记不影响结果
      （状态机对同一个标记重复处理是幂等的）。
    """
    import zipfile
    cur = best = None
    try:
        with zipfile.ZipFile(path) as z:
            with z.open(part) as f:
                tail = b""
                while True:
                    buf = f.read(chunk)
                    if not buf:
                        break
                    buf = tail + buf
                    for m in _MARK_RE.finditer(buf):
                        if m.group(1):
                            cur = int(m.group(1))
                        elif cur is not None:
                            best = cur
                    tail = buf[-40:]
    except Exception:                     # noqa: B902
        return None                       # 扫不动就当没扫过
    return best


def read_values(path, sheet=None):
    """读一张表的**值**（不要样式、不要公式原文）→ (表名, 行列表, 一句形状备注)。

    取值这件事两个 loader（load_sweep / load_vco）必须是同一份实现——
    "两份实现必然漂移，漂出来的是两份报表同一个指标报不同的数"。

    ★★ 用 read_only：非 read_only 会把工作簿里**每一张表**都完整解析成 Cell 对象，
      而我们只用第一张。合成基准（400 行 × 230 列）：一张表 2.1×，六张表 12.7×。
      顺带把内存也压下来了——同一个成因还制造过 MemoryError（见 xlsx_shape）。
    ★★ `reset_dimensions()` 是**方法**，不是属性。写成 `ws.reset_dimensions = True`
      只会挂一个没人看的属性：不但没生效，还会让 iter_rows **按表自己声明的宽度
      截列**——整列套过格式的簿子声明 `A1:HG1048576`，实测就把 230 列截成 215 列。
      静默丢列比慢严重得多，这一句写错是不会报错的。
    ★★ read_only 的 iter_rows 是"有多少写多少"，跟非 read_only 有两处不同，
      都得抹平，否则行号会错位（行号是要逐条打给人看的）：
        · 行可能长短不齐 → 一律补齐到最宽那行（非 read_only 是补齐的）。
        · 尾部全空行也照给 → 掐掉；非 read_only 的 max_row 只数到最后一个有值的格。
      中间的全空行**必须留着**（它们占着行号）——已在合成用例上逐格对过。
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        declared = ws.max_row             # 表**自己声明**的行数，可能是虚胖的
        part = (getattr(ws, "_worksheet_path", "") or "").lstrip("/")
        ws.reset_dimensions()             # 别信表自己声明的范围
        title = ws.title
        stop = last_value_row(path, part) if part else None
        rows = [list(r) for r in ws.iter_rows(max_row=stop, values_only=True)]
    finally:
        wb.close()                        # read_only 抓着 zip 不放，读完就还
    while rows and all(v is None for v in rows[-1]):
        rows.pop()
    w = max((len(r) for r in rows), default=0)
    for r in rows:
        if len(r) < w:
            r.extend([None] * (w - len(r)))
    note = ""
    if declared and len(rows) and declared >= 2 * len(rows):
        note = (f"表声明 {declared} 行，实际有值到第 {len(rows)} 行"
                f"（多出来的是只有格式、没有值的空格子，没读）")
    return title, rows, note


def load_sweep(path, sheet=None, header_row=1, leg_col="Mode",
               lock_pattern=r"_lock$", temp_col=None,
               keep_test_item=None, keep_mode=None, keep_original=True,
               spur_tol=2.0, spur_targets=()):
    """把一份扫描簿读成 Sweep（行已过滤、段已切好、指标已识别）。

    keep_original=True 时额外用 data_only=False 再读一遍工作簿并挂在 .wb 上，
    给"输出簿第一页保留原表"用。跨芯片汇总不需要那份，给 False 省一半内存。
    """
    import openpyxl

    # 读两份：一份取缓存值用来算，一份原封不动用来存。
    # 只用 data_only=True 那份去存的话，原表里若有公式会被替换成计算结果——
    # 「第 1 页保留原始 excel」就不成立了。
    src_title, all_rows, shape_note = read_values(path, sheet)
    n_rows = len(all_rows)
    n_cols = max((len(r) for r in all_rows), default=0)
    wb, ws = None, None
    if keep_original:
        wb = openpyxl.load_workbook(path, data_only=False)
        ws = wb[src_title]

    if len(all_rows) < header_row + 1:
        raise SweepError("表里没有数据行")
    header = all_rows[header_row - 1]
    data = all_rows[header_row:]

    cols = Columns(header)
    warnings = list(cols.dup_report())

    if temp_col:
        tname, tcol = temp_col, cols.idx(temp_col)
    else:
        tname, tcol = cols.find(r"temperature", r"^temp")
    if tcol is None:
        raise SweepError("找不到温度列，用 --temp-col 指定")

    leg_i = cols.idx(leg_col)
    if leg_i is None:
        warnings.append(f"没有 {leg_col} 列，无法按重锁切段——全部行算作一段")
    ti_name, ti_col = (None, None)
    if cols.idx("Test Item") is not None:
        ti_name, ti_col = "Test Item", cols.idx("Test Item")

    # 行过滤：Test Item 少数派（收尾行/模板遗留行）踢掉，但逐行记原因
    excluded = []
    keep_ti = keep_test_item
    if ti_col is not None and keep_ti is None:
        keep_ti = _majority(data, ti_col)

    lock_re = re.compile(lock_pattern, re.I)

    # 主模式：默认取出现最多的那个值（重锁行不参与统计，它带 _lock 后缀）
    if leg_i is not None and keep_mode is None:
        keep_mode = _majority(data, leg_i, skip_re=lock_re)

    rows, cut = [], []          # cut: (行号, 原因, 原始行) —— 稍后补"带几个结果值"
    for n, raw in enumerate(data):
        xl = header_row + 1 + n
        if all(is_blank(v) for v in raw):
            continue
        if ti_col is not None and keep_ti is not None:
            v = txt(raw[ti_col]) if ti_col < len(raw) else ""
            if v != keep_ti:
                cut.append((xl, f"{ti_name} = {v!r}，不是主测试项 {keep_ti!r}", raw))
                continue
        if leg_i is not None and keep_mode is not None:
            v = txt(raw[leg_i]) if leg_i < len(raw) else ""
            if v != keep_mode and not lock_re.search(v):
                cut.append((xl, f"{leg_col} = {v!r}，不是主模式 {keep_mode!r}", raw))
                continue
        rows.append(Row(xl, num(raw[tcol]) if tcol < len(raw) else None, "meas", raw))

    items, dropped = build_items(cols, [r.raw for r in rows])
    if not items:
        raise SweepError("没识别出任何有数据的结果列")
    # ★ 杂散：模板固定频点那一格量到的常常是噪底，真值在表尾的杂散清单里。
    #   认不出清单就退回原来的行为，并把卡在哪一条说出来。
    spur_why, spur_extra = attach_spur_list(cols, [r.raw for r in rows], items,
                                            tol=spur_tol, targets=spur_targets)
    for r in rows:
        for it in items:
            if it.pick is not None:
                c = it.pick(r.raw)
                if c is not None:
                    r.src[it.col] = c
                r.vals[it.col] = num(r.raw[c]) if (c is not None and c < len(r.raw)) \
                    else None
            else:
                r.vals[it.col] = num(r.raw[it.col]) if it.col < len(r.raw) else None
    # 换了取值口径就得报出来差多少：这一改动的全部意义就在那个差值上，
    # 不打出来没人知道它真的生效了、也没法判断窗口开得合不合适。
    spur_notes = []
    for it in items:
        if it.pick is None:
            continue
        new = [r.vals[it.col] for r in rows if r.vals.get(it.col) is not None]
        old = [num(r.raw[it.col]) for r in rows
               if it.col < len(r.raw) and num(r.raw[it.col]) is not None]
        it.pick_stats.update(new=median(new), old=median(old) if old else None)
        offs = it.pick_stats.get("off")
        d = (median(new) - median(old)) if (new and old) else None
        if not it.has_cell:
            # 模板里压根没有这个频点：没有"原来那一格"可比，只报取自哪儿、命中几行
            # ★ 一条都没搜到也**保留这一行**（用户原话「没有的就留空」）：
            #   两个模块的表行集合一样才横着比得了。
            if not new:
                spur_notes.append(
                    f"{it.label}: 清单里 {fmt_num(it.pick_stats['target'])}±"
                    f"{fmt_num(it.pick_stats['tol'], 2)} MHz 内**一条都没有**"
                    f"（0/{len(rows)} 行）——表上这一行留空")
                continue
            spur_notes.append(
                f"{it.label}: 模板没有这个频点，只从清单取（"
                f"{fmt_num(it.pick_stats['target'])}±"
                f"{fmt_num(it.pick_stats['tol'], 2)} MHz 内最大的一条，实测偏移中位 "
                f"{fmt_num(offs)} MHz），中位 {fmt_num(median(new), 2)} dBc；"
                f"{len(new)}/{len(rows)} 行命中，其余行留空")
            continue
        spur_notes.append(
            f"{it.label}: 取清单里 {fmt_num(offs)} MHz 那条，中位 "
            f"{fmt_num(median(new), 2)} dBc（{it.src} 那一格中位 "
            f"{fmt_num(median(old), 2) if old else '空'}"
            + (f"，差 {d:+.2f} dB）" if d is not None else "）")
            + f"；{len(new)}/{len(rows)} 行命中")
    for it in items:
        _w = spur_off_warn(it)
        if _w:
            spur_notes.append("⚠ " + _w)
    spur_notes += spur_extra
    if spur_why:
        spur_notes.append(f"没认出表尾的杂散清单（{spur_why}），"
                          f"Spur 仍按模板固定频点那一格取")

    # ★ 被过滤掉的行到底有没有带测量结果，必须说出来。
    #   "排除了 6 行"这句话本身不足以判断有没有丢数据——原厂模板常在数据后面
    #   追加页脚/图例行（条件列全空），那种排掉是对的；但旁路自检行是**带部分
    #   结果值**的，把它算进相邻那一段会把段的走向和极值带偏。两者只看行号
    #   分不出来，得看它带不带结果。
    # 分母只数**表上真有那一列**的指标：杂散清单加出来的行没有自己的列
    #   （值是逐行从清单里挑的），把它们算进分母会让"4/16"莫名其妙变成"4/20"。
    n_sheet = sum(1 for it in items if it.has_cell)
    for xl, why, raw in cut:
        k = sum(1 for it in items
                if it.col < len(raw) and num(raw[it.col]) is not None)
        # 带 4/16 跟带 16/16 是两件事：前者是配置行顺手回读了几个量，
        # 后者是**另一个配置下的一整套测量**（比如关掉 test mux 再测一遍）。
        # 只打绝对个数分不出来，分母必须给。
        excluded.append((xl, why + (f"（这行带 {k}/{n_sheet} 个结果值）" if k else
                                    "（这行没有任何结果值）")))

    # 没有任何结果值的行（纯配置行）也踢掉
    keep = []
    for r in rows:
        if all(v is None for v in r.vals.values()):
            excluded.append((r.xl, "所有结果列都是空的（配置/开关行，不是测量点）"))
        else:
            keep.append(r)
    rows = keep

    legs, orphan = segment(rows, leg_i, lock_re)
    for r in orphan:
        excluded.append((r.xl, f"排在第一次重锁之前（"
                               f"{leg_col}={txt(r.raw[leg_i]) if leg_i is not None else '?'}）"))
    if not legs:
        legs = [Leg(1, None)]
        legs[0].rows = rows
        warnings.append(f"没有匹配 {lock_pattern!r} 的行，整表按一段处理")
    excluded.sort(key=lambda x: x[0])

    return Sweep(path=path, wb=wb, ws=ws, wb_val=None, ws_val=None,
                 src_title=src_title, header=header, data=data, cols=cols,
                 temp_name=tname, temp_col=tcol, leg_col=leg_col, leg_i=leg_i,
                 lock_pattern=lock_pattern, lock_re=lock_re,
                 ti_name=ti_name, ti_col=ti_col, keep_ti=keep_ti, keep_mode=keep_mode,
                 rows=rows, items=items, dropped=dropped, legs=legs, orphan=orphan,
                 excluded=excluded, warnings=warnings, spur_notes=spur_notes,
                 room_t=room_temp(legs), shape_note=shape_note,
                 n_rows=n_rows, n_cols=n_cols)


def _majority(data, ci, skip_re=None):
    cnt = {}
    for r in data:
        v = txt(r[ci]) if ci < len(r) else ""
        if v and not (skip_re and skip_re.search(v)):
            cnt[v] = cnt.get(v, 0) + 1
    return max(cnt, key=lambda k: cnt[k]) if cnt else None


# ---------------------------------------------------------------- ⑤ 出稿

def styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style="thin", color="FF000000")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        # ★ 数字列右对齐：居中的数字列小数点不对齐，比大小得逐个读；右对齐后
        #   位数差直接变成视觉长度差，扫一眼就知道谁大。这条比任何配色都管用。
        "right": Alignment(horizontal="right", vertical="center", wrap_text=False),
        "f_head": PatternFill("solid", fgColor=FILL_HEADER, bgColor=FILL_HEADER),
        "f_group": PatternFill("solid", fgColor=FILL_GROUP, bgColor=FILL_GROUP),
        "f_res": PatternFill("solid", fgColor=FILL_RESULT, bgColor=FILL_RESULT),
        "f_sep": PatternFill("solid", fgColor=FILL_SEP, bgColor=FILL_SEP),
        "f_in": PatternFill("solid", fgColor=FILL_INPUT, bgColor=FILL_INPUT),
        "f_zone": PatternFill("solid", fgColor=FILL_ZONE, bgColor=FILL_ZONE),
        "f_rail": PatternFill("solid", fgColor=FILL_RAIL, bgColor=FILL_RAIL),
        "f_sum": PatternFill("solid", fgColor=FILL_SUM, bgColor=FILL_SUM),
        "Font": Font,
    }


def put(ws, r, c, v, st, fill=None, bold=False, color=None, align="center", size=10):
    cell = ws.cell(row=r, column=c)
    cell.value = v
    cell.border = st["border"]
    cell.alignment = st.get(align) or st["center"]
    cell.font = st["Font"](bold=bold, color=color, size=size)
    if fill is not None:
        cell.fill = fill
    return cell


def as_text(v):
    """要当纯文本写进去的串。

    ★ 以 = 开头的字符串 openpyxl 会当公式写进 <f>：备注里写
    "= Fmax − Fmin" 这种算式样子，Excel 就去解析它，整列变成 #REF!。
    垫一个空格最省事，看不出来也不会被解析。
    只垫 `=`：实测 openpyxl 只按 startswith("=") 判公式，`+ - @` 照常存成字符串，
    多垫的话 "-40 ~ 105" 这种条件值会平白多一个前导空格。
    """
    if isinstance(v, str) and v[:1] == "=":
        return " " + v
    return v


# ---- 版面几何 ------------------------------------------------------------
# Excel 的列宽单位＝默认字体里 '0' 的宽度，换算成像素是 round(w*7)+5（Calibri 11）；
# 行高单位是磅，96 dpi 下 1 磅 = 4/3 px。图的 width/height 单位是厘米。
# ★ 这三套单位不换算到一起，就只能靠眼睛猜"这张图放不放得下"——
#   2026-08-04 就是这么撞的：图宽写死 14.5cm、竖条列宽写死出来 14.15cm，
#   每张图往右边邻居里压 13px，三颗芯片的图连成一片。

PX_PER_CM = 96 / 2.54
DEFAULT_ROW_PT = 15.0


def col_px(w):
    return round(w * 7) + 5


def cols_cm(widths):
    return sum(col_px(w) for w in widths) / PX_PER_CM


def rows_cm(n, row_pt=DEFAULT_ROW_PT):
    return n * row_pt * (4.0 / 3.0) / PX_PER_CM


def fit_strip(base, chart_cm, margin_cm=0.15):
    """按目标图宽把数据列**等比撑开**，保证竖条一定放得下图。

    撑数据列而不是撑间隔列：间隔拉宽只是留白，把数据列一起撑开，
    图和它下面那张表就是同一个宽度，看着是一整条。
    返回 (撑开后的列宽表, 实际能用的图宽cm)。
    """
    need = (chart_cm + margin_cm) * PX_PER_CM
    have = sum(col_px(w) for w in base)
    k = max(1.0, need / have)
    out = [round(w * k, 2) for w in base]
    return out, cols_cm(out) - margin_cm


def chart_rows(h_cm, margin_px=16, row_pt=DEFAULT_ROW_PT):
    """一张图该占几行——**按图高算**，不写死。写死的话图一变大就上下压在一起。"""
    import math
    return max(2, int(math.ceil((h_cm * PX_PER_CM + margin_px) /
                                (row_pt * 4.0 / 3.0))))


# ---- 图表原语 ------------------------------------------------------------

# 每段/每条线一个颜色 + 一个记号形状：几条线叠在一张图上，光靠颜色分不开
# （打印和色弱都糊），形状也得不一样。
LEG_STYLE = [("1F77B4", "circle"), ("D62728", "square"),
             ("2CA02C", "triangle"), ("FF7F0E", "diamond"),
             ("9467BD", "x"), ("8C564B", "plus")]


def nice_step(span):
    """给定跨度挑一个好看的刻度步长（1/2/2.5/5 × 10^n）。"""
    import math
    if span <= 0:
        return 1.0
    raw = span / 5.0
    e = 10.0 ** math.floor(math.log10(raw))
    for f in (1, 2, 2.5, 5, 10):
        if raw <= f * e:
            return f * e
    return 10 * e


def axis_bounds(vals, pad=0.10):
    """按数据自己算 (min, max, 步长)，两头留一点余量再对齐到整刻度。

    ★ 不要指望 Excel 自动缩放。它常把值轴从 0 起画，量本身只在 0.69~0.73
    晃的话就被压成一条平线，什么都看不出来。范围必须由数据算出来钉死。
    """
    import math
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    lo, hi = min(vals), max(vals)
    # ★ 判"平"要用相对量。写 hi == lo 只挡得住完全相等的情况：一条常数曲线经过
    #   浮点运算后跨度常是 1e-13 这种量级，于是下面 nice_step 算出个 1e-13 的步长、
    #   floor/ceil 又把上下界压回同一个数 —— Excel 拿到 min == max 的坐标轴，
    #   整张图画不出来。（ΔF-vs-CT 这种"步长恒定"的曲线、以及某个温度压控没生效
    #   压成一条平线的情况，都会踩上。）
    if hi - lo <= abs(hi) * 1e-9:
        d = abs(hi) * 0.05 or 1.0
        lo, hi = lo - d, hi + d
    m = (hi - lo) * pad
    lo, hi = lo - m, hi + m
    step = nice_step(hi - lo)
    return math.floor(lo / step) * step, math.ceil(hi / step) * step, step


def apply_y(chart, bounds):
    if not bounds:
        return
    chart.y_axis.scaling.min, chart.y_axis.scaling.max = bounds[0], bounds[1]
    chart.y_axis.majorUnit = bounds[2]


def legend_bottom(chart):
    """图例放下面，并且**所有标题都别盖在绘图区上**。

    ★★ OOXML 里 `<c:title>` 的 `<c:overlay>` 缺省＝盖在绘图区上，openpyxl 压根
      不写这个元素。后果：图标题被画进绘图框**里面**，两根轴标题直接压在刻度
      文字上（"Kvco (MHz/V)" 盖住 "275"、"Vtune (V)" 盖住 "0.4"）。
      2026-08-04 用户报"看样子你没有做重叠检测"——真正的原因是这个默认值，
      不是标注算法。显式置 False，Excel 才会给标题留出自己的位置。
      放这里是因为每个画图的地方都调它，漏不掉。
    """
    if chart.legend is not None:
        chart.legend.position = "b"
        chart.legend.overlay = False
    for t in (chart.title, getattr(chart.x_axis, "title", None),
              getattr(chart.y_axis, "title", None)):
        if t is not None and hasattr(t, "overlay"):
            t.overlay = False


def blank_policy(chart, data_sheet_hidden=False):
    """断点留空 + 数据页隐藏了也照画 + 关掉"按点着色"。

    ★ 属性名必须是 openpyxl 的 display_blanks / visible_cells_only。
      写成 OOXML 里的 dispBlanksAs / plotVisOnly **不报错**，只是给对象挂了个
      没人读的属性：dispBlanksAs 缺省是 "zero"，于是没测的格子被当 0 画进图，
      曲线在测点之间扎到零成一串尖刺。
    ★★ varyColors 省略时 OOXML 默认是**真**，而 Excel 的"按点着色"**只对
      单系列图生效**——于是同一份代码画出来的图，三条温度曲线那些正常，
      只画一条曲线的（ΔF vs CT 码只画常温）会**给每个数据点发一条图例**：
      几十条「—■— 0.5 / —■— 1.5 …」把绘图区挤成顶上一条，纵轴刻度叠成一坨。
      2026-08-04 用户报的就是这张图。openpyxl 不写这个元素 ≠ 关掉它，
      必须显式 False。
    """
    chart.display_blanks = "gap"
    chart.varyColors = False
    if data_sheet_hidden:
        chart.visible_cells_only = False


def style_series(series, color, symbol, line=True, dash=None, size=6):
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    gp = GraphicalProperties()
    gp.line = LineProperties(noFill=True) if not line else \
        LineProperties(solidFill=color, w=20000, prstDash=dash)
    series.graphicalProperties = gp
    if symbol:
        m = Marker(symbol=symbol, size=size)
        m.graphicalProperties = GraphicalProperties(solidFill=color)
        series.marker = m
    else:
        series.marker = Marker(symbol="none")
    series.smooth = False
