#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""probe_allmode_result.py — 探查「全模式单文件」的实测 Result Excel。

单模式时代 模式=文件夹名（current_db 的假设）；全模式合并跑完后 7+ 个模式在
同一个 sheet 里，必须按行分段。本脚本做三件事：
  1) sheet 清单（谁把文件撑大的一目了然）
  2) 定位主表（NO./Mode/Current_mA 表头），按行重建模式分段：
     `*_sigN` 签名行 / Init 行的 NO. 列（模式名）都当段边界
  3) 逐模式打印电流序列（OFF 步 → mA）+ lock 行的 ReadBack 判锁码 + 异常清单
     （Current 开关=YES 但没测到值、归不到模式段的测量行）

用法：
    python probe_allmode_result.py <Result*.xlsx> [--json out.json]
  --json 落盘完整逐行数据（含 gz），带回开发机喂后续入库工具。

只读；依赖 openpyxl；需与 probe_current_data.py 同目录（复用其表头定位）。
"""
import argparse
import gzip
import io
import json
import re
import sys

from probe_current_data import match_result_header, norm, jval, cell

MAX_DUMP_ROWS = 5000


def find_named_cols(header_row, names):
    """表头行 → {名字: 0-based列号}（strip 精确匹配，找不到的名字缺席）。"""
    out = {}
    for j, c in enumerate(header_row):
        s = str(c).strip() if c is not None else ""
        if s in names and s not in out:
            out[s] = j
    return out


def main(argv=None):
    ap = argparse.ArgumentParser(description="全模式 Result Excel 探查（分段+电流序列+判锁）")
    ap.add_argument("xlsx")
    ap.add_argument("--json", help="完整逐行数据落盘（同时生成 .gz）")
    args = ap.parse_args(argv)

    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    except Exception:
        pass

    import openpyxl
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    # ---- 1) sheet 清单 ----
    print("== sheet 清单 ==")
    inventory = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        try:
            dim = ws.calculate_dimension()
        except Exception:
            dim = "?"
        inventory.append({"sheet": sn, "dimension": dim})
        print("  %-30s %s" % (sn, dim))

    # ---- 2) 定位主表 ----
    main_ws = hdr_idx = cols = header_row = None
    for sn in wb.sheetnames:
        ws = wb[sn]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            c = match_result_header(row)
            if c is not None:
                main_ws, hdr_idx, cols, header_row = sn, i, c, row
                break
        if main_ws:
            break
    if not main_ws:
        sys.exit("没找到含 NO./Current 表头的主表——把上面的 sheet 清单贴回来人工看。")

    extra_names = ["Test Item", "Test", "Current", "ReadBack", "Chamber"] + \
                  ["ReadBack_VALUE%d" % k for k in range(1, 8)]
    named = find_named_cols(header_row, set(extra_names))
    rb_cols = [named.get("ReadBack_VALUE%d" % k) for k in range(1, 8)]
    print()
    print("== 主表 ==  sheet=%s 表头行=%d  关键列(1基)=%s  ReadBack列=%s"
          % (main_ws, hdr_idx, {k: v + 1 for k, v in cols.items() if isinstance(v, int)},
             [c + 1 for c in rb_cols if c is not None]))

    # ---- 3) 逐行走 + 模式分段 ----
    ws = wb[main_ws]
    segments = []          # {mode, first_row, last_row, rows:[...]}
    orphans, flag_miss = [], []
    cur_seg = None

    def new_seg(mode, r):
        segments.append({"mode": mode, "first_row": r, "last_row": r, "rows": []})
        return segments[-1]

    n_data = 0
    for i, row in enumerate(ws.iter_rows(min_row=hdr_idx + 1, values_only=True), hdr_idx + 1):
        no_raw = cell(row, cols["no"])
        label = cell(row, cols["mode"])
        cur = cell(row, cols["cur"])
        temp = cell(row, cols["temp"]) if cols["temp"] is not None else None
        if no_raw is None and label is None and cur is None:
            continue
        n_data += 1
        ln = norm(label)
        m = re.fullmatch(r"(.+)_sig\d+", str(label).strip()) if label else None
        if m:                                              # 签名行 = 段边界（最可靠）
            if cur_seg is None or cur_seg["mode"] != m.group(1):
                cur_seg = new_seg(m.group(1), i)
        elif ln.startswith("init") and isinstance(no_raw, str) and no_raw.strip() \
                and not re.fullmatch(r"[\d ,]+", no_raw.strip()):
            if cur_seg is None or (cur_seg["mode"] != no_raw.strip()
                                   and not cur_seg["mode"].startswith(no_raw.strip())):
                cur_seg = new_seg(no_raw.strip(), i)       # 兜底：Init 行 NO.=模式名
        rec = {"row": i, "no": jval(no_raw), "label": jval(label),
               "test_item": jval(cell(row, named.get("Test Item"))),
               "test_flag": jval(cell(row, named.get("Test"))),
               "cur_flag": jval(cell(row, named.get("Current"))),
               "current": jval(cur), "temp": jval(temp),
               "readback": [jval(cell(row, c)) if c is not None else None for c in rb_cols],
               "mode": cur_seg["mode"] if cur_seg else None}
        if cur_seg:
            cur_seg["rows"].append(rec)
            cur_seg["last_row"] = i
        elif "chamber" in ln or "chamber" in norm(no_raw):
            pass                                           # 段外 chamber 行正常
        elif cur is not None:
            orphans.append(rec)
        if (norm(rec["cur_flag"]) == "yes" and cur is None
                and norm(rec["test_flag"]) != "no"
                and "chamber" not in ln and "chamber" not in norm(no_raw)):
            flag_miss.append(rec)
        if n_data >= MAX_DUMP_ROWS:
            print("  ⚠ 数据行超过 %d，截断" % MAX_DUMP_ROWS)
            break
    wb.close()

    # ---- 4) 报告 ----
    print()
    print("== 模式分段（%d 段）==" % len(segments))
    for s in segments:
        n_cur = sum(1 for r in s["rows"] if r["current"] is not None)
        temps = sorted(set(str(r["temp"]) for r in s["rows"] if r["temp"] is not None))
        print("  %-24s 行 %d-%d  共 %d 行 / 带电流 %d 行 / 温度 %s"
              % (s["mode"], s["first_row"], s["last_row"], len(s["rows"]), n_cur, ",".join(temps) or "-"))

    print()
    print("== 各模式电流序列（label: mA）==")
    for s in segments:
        print("◆ %s" % s["mode"])
        for r in s["rows"]:
            if r["current"] is not None:
                print("   r%-4d %-26s %s" % (r["row"], str(r["label"])[:26], r["current"]))
        locks = [r for r in s["rows"] if norm(r["label"]).startswith("lock")]
        if locks:
            last = locks[-1]
            rb = [str(x) for x in last["readback"] if x is not None]
            print("   ReadBack(末个lock行 r%d): %s" % (last["row"], " | ".join(rb) or "（空）"))

    print()
    if flag_miss:
        print("⚠ Current 开关=YES 但没测到电流值的行（%d）：%s"
              % (len(flag_miss), [(r["row"], r["label"]) for r in flag_miss[:20]]))
    if orphans:
        print("⚠ 归不到任何模式段的测量行（%d）：%s"
              % (len(orphans), [(r["row"], r["label"]) for r in orphans[:20]]))
    if not flag_miss and not orphans:
        print("✔ 无异常：所有带电流行都归到了模式段，Current=YES 的行都有值。")

    if args.json:
        blob = json.dumps({"file": args.xlsx, "sheets": inventory,
                           "main_sheet": main_ws, "header_row": hdr_idx,
                           "key_cols_1based": {k: (v + 1 if isinstance(v, int) else v)
                                               for k, v in cols.items()},
                           "readback_cols_1based": [c + 1 for c in rb_cols if c is not None],
                           "segments": segments, "orphans": orphans,
                           "flag_yes_no_value": flag_miss},
                          ensure_ascii=False, separators=(",", ":"), default=str)
        with io.open(args.json, "w", encoding="utf-8") as f:
            f.write(blob)
        with gzip.open(args.json + ".gz", "wt", encoding="utf-8") as f:
            f.write(blob)
        print()
        print("完整数据已写: %s（+.gz，带回开发机）" % args.json)


if __name__ == "__main__":
    main()
