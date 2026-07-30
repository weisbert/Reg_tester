#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
summarize_vco_sweep.py — 开环压控扫描簿 → 带汇总页的 Excel

输入：一份仪器/脚本导出的 VCO 开环特性宽表（一行 = 一个测量点）。典型跑法：
    · 某温度下把环路打开，把调谐电压 Vtune 从低扫到高，记 频率/功率/相噪/电流；
    · 再把 Vtune 钉在中间某个值，改电容阵列码（CT，通过一个寄存器位段写入），
      看频段怎么搬；
    · 换温度重来（换温度前一般先闭环锁一次再打开）。

输出：一份新的 .xlsx——
    第 1 页   原始数据（原封不动，就是输入那一页）
    ★ 结论    这页才是报告：把两种扫描合起来算，只放能下判断的行——
              有没有锁不上的盲区（Band overlap 比）、目标频率覆盖余量、
              温漂折合几个 CT 码、工作点 Kvco 与带宽漂移倍数、单调性、各指标最差点。
              Spec 两列留空给人填，「关注」列提示该填哪边，填完 PASS/FAIL 自动出、超规标红。
    汇总      逐点极值大表（Vtune 扫一张、CT 扫一张）：查数用，不是结论
    Vtune明细 指标 × Vtune 矩阵（按温度分列），既能翻数也是图表的数据源
    Kvco明细  逐点斜率 ΔF/ΔV（区间中点 × 温度）
    CT明细    指标 × CT 码
    图表      频率 vs Vtune / vs CT（每温一条线，另叠一条目标 fVCO 水平参考线）、
              Kvco vs Vtune、ΔF/ΔCT vs CT、相噪 vs offset
    闭环锁定点 闭环/锁定那几行单列出来，跟开环曲线对照着看

★ 交出去的簿子里不写任何操作说明、告警、排除记录——那是给做表的人对账用的，
  评审的人看见「⚠ 告警」第一反应是数据有问题。这些信息每次运行都打在控制台上，
  要在簿子里留档用 --notes 单出一页。

★ 所有数字都是公式，不是算好的死值。三级引用：
      结论 / 汇总  ──►  明细、斜率、温漂  ──►  原始表那一格
  评审时点开任何一个数，Excel 的「追踪引用单元格」能一路追到源头；原始数据
  改一格，整本跟着重算。要纯数值版用 --static。
  ★ 公式格必须自己补「缓存值」：openpyxl 写出来的是 `<f>公式</f><v></v>`，
  那个**空的** `<v>` 等于告诉 Excel "结果就是空"——自己机器上因为
  fullCalcOnLoad 重算了看着正常，**发给别人就是一片空白**，得在格子里敲一次
  回车才出数。所以 emit() 把公式和算好的值成对交给 put()，存盘后由
  `xlsx_formula_cache` 把值补进 `<v>`（结果未知的判定类公式则删掉空 `<v>`，
  逼 Excel 自己算）。补完之后别的工具直接读单元格也能读到数。

为什么单独有一页「结论」
    把每个指标的 Min/Max/Δ 都铺一遍，是数据搬运不是结论：十几个指标乘几个温度
    乘三列，没人读得完，而真正要回答的问题一个都没答。这类测试的核心问题是
    「单个电容码用调谐电压能盖多宽 ÷ 相邻码差多少」——这个比值 <1 就有锁不上的
    频率，而它必须把两张扫描合起来才算得出来。所以结论页在前，大表降级到后面查数。

为什么按「扫描组」而不是按温度汇总
    这类簿里自变量是 Vtune 和 CT，温度只是分组维度。同一个温度下会有两种扫描
    （先扫 Vtune 再扫 CT），横轴根本不是一回事，混在一起统计出来的极值没有意义。
    所以本脚本先按「(温度, 扫的是谁) 变一次就切一组」分组，组内再按横轴排。

依赖：只用 openpyxl。   pip install openpyxl

用法：
    python summarize_vco_sweep.py <扫描簿.xlsx>                 # 出 <原名>_summary.xlsx
    python summarize_vco_sweep.py <扫描簿.xlsx> --dry-run       # 只打印识别结果，不写文件
    python summarize_vco_sweep.py <扫描簿.xlsx> -o 汇总.xlsx
    python summarize_vco_sweep.py <扫描簿.xlsx> --vtune-col Vtune_V --ct-col "REG Value1"
    python summarize_vco_sweep.py <扫描簿.xlsx> --ref-temp 25   # 温漂参考温度

先跑 --dry-run 看四件事对不对：列映射 / 横轴选了哪列 / 分了几组 / 排除了哪些行。
排除的行不会被悄悄丢掉，汇总页底部会逐行列出原因。

⚠ CT 是靠某个寄存器位段扫的，那个地址是 IP：默认只打印列名不打印地址值，
   要看加 --show-addr（输出的 xlsx 里本来就有原始数据，那份别往公开处传）。
"""

import argparse
import math
import os
import re
import sys
from collections import OrderedDict

from xlsx_formula_cache import Formula, FormulaCache

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

# report-forge compliance 表的配色（黄表头 / 米色分组行 / 白结果行 / 红超规）
FILL_HEADER = "FFFF00"
FILL_GROUP = "EEECE1"
FILL_RESULT = "FFFFFF"
FILL_SEP = "B8CCE4"
COLOR_FLAG = "FF0000"
COLOR_PASS = "006100"
FILL_FAIL = "FFC7CE"
FILL_PASS = "C6EFCE"

# 表里表示"没测/不适用"的占位符，一律当空值
BLANK_TOKENS = {"", "-", "--", "—", "n/a", "na", "null", "none", "#n/a"}


# ---------------------------------------------------------------- 取值

def is_blank(v):
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip().lower() in BLANK_TOKENS
    return False


def num(v):
    """数值化；取不到数就返回 None。'-' 这类占位符算没测。"""
    if is_blank(v) or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def txt(v):
    return "" if v is None else str(v).strip()


# 格子里写引用公式（默认）还是写算好的死值（--static）。
# 默认写公式：评审时点开任何一个数都能一路追回原始表那一格，原始数据改了整本跟着变。
LIVE_REFS = True
# 公式格的缓存值：不补真值的话别人打开是一片空白（见 xlsx_formula_cache）
VCACHE = FormulaCache()


def emit(formula, value):
    """公式态下返回 (公式, 值) 一对，交给 put() 落盘并记缓存值。

    ★ 为什么要把值一起带上：openpyxl 给公式格写的是 `<f>公式</f><v></v>`——
      一个**空的**缓存值，等于告诉 Excel "这公式的结果就是空"。自己机器上
      看着正常（fullCalcOnLoad 触发了重算），**发给别人就是一片空白**，
      得在格子里敲一次回车才出数。这里本来就两样都算好了，顺手把值
      记进缓存，存盘后补进 <v>，任何环境打开都直接看见数。
      细节见 xlsx_formula_cache 的模块说明。
    """
    if LIVE_REFS and formula is not None:
        return Formula(formula, value)
    return value


def as_text(v):
    """说明文字写进格子前过一道。

    ★ 以 = 开头的字符串会被 openpyxl 按**公式**写进去，Excel 一解析就变成
      #NAME? / #REF!——备注列整列报错或空白。踩过一次：备注本来写成
      "= Fmax − Fmin" 这种算式样子，结果全成了错误值。
      开头是 = + - @ 的一律前面垫个空格挡住。
    """
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@"):
        return " " + v
    return v


def q(title):
    """工作表名放进公式里：一律加单引号（表名带空格/中文/减号都不会炸）。"""
    return "'%s'" % str(title).replace("'", "''")


def cref(title, col0, row):
    """单元格引用 'Sheet1'!$DX$15。col0 是 0 基列号。"""
    from openpyxl.utils import get_column_letter
    return "%s!$%s$%d" % (q(title), get_column_letter(col0 + 1), row)


def rref(title, col0, r0, r1):
    """列方向区间引用 'Sheet1'!$B$3:$B$16。"""
    from openpyxl.utils import get_column_letter
    L = get_column_letter(col0 + 1)
    return "%s!$%s$%d:$%s$%d" % (q(title), L, r0, L, r1)


def vref(ref):
    """引用原始表的一格，并且强制变成数值。

    ★ 原始表里"数字存成文本"是常事（仪器脚本导出、或者列被设成文本格式）。
      直接写 ='Sheet1'!$X$9 的话，Python 这边 num() 照样解析得出来，
      但 Excel 的 MIN/MAX/COUNT 会**跳过文本**——公式版静默算错、静态版却是对的，
      两边对不上还查不出原因。VALUE() 把文本数字转成数，
      真数字原样通过，"-"/空/非数字则落到 "" 当没测。
    """
    return '=IFERROR(VALUE(%s),"")' % ref


def f_minmax(which, refs):
    """MIN/MAX，区间全空时给空串而不是 0。refs 可以是多段（跨表也行）。"""
    a = ",".join(refs)
    return '=IF(COUNT(%s)=0,"",%s(%s))' % (a, which, a)


def f_span(refs):
    a = ",".join(refs)
    return '=IF(COUNT(%s)=0,"",MAX(%s)-MIN(%s))' % (a, a, a)


def f_absmax(ref):
    """区间里绝对值最大的那个数（不用数组公式）。"""
    return '=IF(COUNT(%s)=0,"",MAX(MAX(%s),-MIN(%s)))' % (ref, ref, ref)


def f_absmin(ref):
    """区间里绝对值最小的那个数。全同号时精确；跨零时给 0（确实存在斜率为 0 的点）。"""
    return ('=IF(COUNT({r})=0,"",IF(MIN({r})>=0,MIN({r}),IF(MAX({r})<=0,-MAX({r}),0)))'
            .format(r=ref))


def f_signed_absmax(ref):
    """绝对值最大、但保留正负号（温漂要看是往上漂还是往下漂）。"""
    return ('=IF(COUNT({r})=0,"",IF(MAX({r})>=-MIN({r}),MAX({r}),MIN({r})))'
            .format(r=ref))


def f_guard(expr, *cells):
    """依赖的格子都是数才算，否则给空串——上游给了 "" 的话别算出 #VALUE!。"""
    cond = ",".join("ISNUMBER(%s)" % c for c in cells)
    if len(cells) > 1:
        cond = "AND(%s)" % cond
    return '=IF(%s,%s,"")' % (cond, expr)


def f_div(a, b):
    return '=IF(N(%s)=0,"",%s/%s)' % (b, a, b)


def qx(v, nd):
    """横轴取值量化。

    仪器脚本常用累加生成扫描点，于是同一个设定值在表里长这样：
    0.15000000000000002 / 0.39999999999999997 / 0.7000000000000001。
    不量化的话，同一个设定值在不同温度段会变成两个不同的字典键——
    明细页多出半空的行、温漂对不上点，而且一点报错都没有。
    """
    return None if v is None else round(v, nd)


def fmt_num(x, nd=3):
    if x is None:
        return None
    if abs(x - round(x)) < 1e-12:
        return int(round(x))
    return round(x, nd)


def fmt_hz(mhz):
    """offset 频率 MHz -> 好读的标签：0.001->1kHz, 1->1MHz。"""
    if mhz is None:
        return "?"
    if mhz < 1:
        return f"{fmt_num(mhz * 1000.0)}kHz"
    return f"{fmt_num(mhz)}MHz"


# ---------------------------------------------------------------- 列定位

class Columns:
    """按表头名定位列；重名列会报出来。

    这类原厂模板扩列时复制粘贴不改序号很常见（同名列出现两遍）。
    按名字取只会一直拿到第一份，第二份永远读不到且不报错——
    所以这里把重名的列位置全记下来，取第一份并留警告。
    """

    def __init__(self, header):
        self.pos = OrderedDict()
        for i, h in enumerate(header):
            k = txt(h)
            if k:
                self.pos.setdefault(k, []).append(i)
        self.duplicates = {k: v for k, v in self.pos.items() if len(v) > 1}

    def idx(self, name):
        v = self.pos.get(name)
        return v[0] if v else None

    def find(self, *patterns, **kw):
        """按正则找第一个匹配的表头，返回 (名字, 列下标)。"""
        exclude = kw.get("exclude", ())
        for k in self.pos:
            kl = k.lower()
            if any(re.search(p, kl) for p in patterns) and \
               not any(re.search(p, kl) for p in exclude):
                return k, self.pos[k][0]
        return None, None

    def match_all(self, pattern):
        """所有匹配的表头，按列序返回 [(名字, 列下标)]。"""
        out = [(k, v[0]) for k, v in self.pos.items() if re.search(pattern, k.lower())]
        return sorted(out, key=lambda x: x[1])


# ---------------------------------------------------------------- 指标识别

class Item:
    __slots__ = ("cat", "label", "unit", "col", "src")

    def __init__(self, cat, label, unit, col, src):
        self.cat, self.label, self.unit, self.col, self.src = cat, label, unit, col, src


SIMPLE_ITEMS = [
    # (表头, 分类, 单位)
    ("Freq_MHz", "Frequency", "MHz"),
    ("Power_dBm", "Output", "dBm"),
    ("IPN_SSB", "Phase Noise", "dBc"),
    ("IPN_Omit_SSB", "Phase Noise", "dBc"),
    ("Vtune_V", "Tuning", "V"),
    ("Vtemp_V", "Temp Sensor", "V"),
    ("Current_mA", "Current", "mA"),
]
# 成对列：<前缀>Freq<i> 给频点、<前缀>Result<i> 给结果
PAIRED_ITEMS = [
    ("SpotPN", "Phase Noise", "dBc/Hz", "SpotPN@{f}"),
    ("OtherSpur", "Spur", "dBc", "Spur@{f}"),
]


def build_items(cols, rows, skip_cols=()):
    """识别有数据的结果列。整片是空的（占位列）自动丢掉，并记在 dropped 里。

    skip_cols 里的列是横轴（Vtune / CT 码），它是自变量不是指标，不进 items。
    """
    items, dropped = [], []

    def nonempty(ci):
        return sum(1 for r in rows if ci < len(r) and not is_blank(r[ci]))

    for name, cat, unit in SIMPLE_ITEMS:
        ci = cols.idx(name)
        if ci is None or ci in skip_cols:
            continue
        if nonempty(ci) == 0:
            dropped.append((name, "整列没有数据"))
            continue
        items.append(Item(cat, name, unit, ci, name))

    for prefix, cat, unit, tpl in PAIRED_ITEMS:
        i = 1
        while True:
            fc, rc = cols.idx("%sFreq%d" % (prefix, i)), cols.idx("%sResult%d" % (prefix, i))
            if fc is None and rc is None:
                break
            i += 1
            if rc is None or rc in skip_cols:
                continue
            if nonempty(rc) == 0:
                dropped.append(("%sResult%d" % (prefix, i - 1), "整列没有数据"))
                continue
            freqs = set()
            if fc is not None:
                freqs = {num(r[fc]) for r in rows
                         if fc < len(r) and num(r[fc]) is not None}
            f = fmt_hz(sorted(freqs)[0]) if len(freqs) == 1 else (
                "/".join(fmt_hz(x) for x in sorted(freqs)[:3]) or "#%d" % (i - 1))
            items.append(Item(cat, tpl.format(f=f), unit, rc, "%sResult%d" % (prefix, i - 1)))
    return items, dropped


# ---------------------------------------------------------------- 行 / 分组

class Row:
    __slots__ = ("xl", "temp", "mode", "vt", "ct", "group", "vals", "raw")

    def __init__(self, xl, temp, mode, vt, ct, raw):
        self.xl, self.temp, self.mode, self.raw = xl, temp, mode, raw
        self.vt, self.ct = vt, ct
        self.group, self.vals = None, {}


KIND_LABEL = {"vtune": "Vtune扫", "ct": "CT扫", "point": "单点"}


class Group:
    def __init__(self, n, kind, temp):
        self.n, self.kind, self.temp = n, kind, temp
        self.rows = []
        self.tag = ""                     # 同温同类出现多次时的 #2 #3

    @property
    def x_of(self):
        return (lambda r: r.ct) if self.kind == "ct" else (lambda r: r.vt)

    @property
    def x_label(self):
        return "CT 码" if self.kind == "ct" else "Vtune (V)"

    @property
    def x_unit(self):
        return "code" if self.kind == "ct" else "V"

    @property
    def title(self):
        t = fmt_num(self.temp)
        head = "%s℃" % t if t is not None else "?℃"
        return "%s %s%s" % (head, KIND_LABEL.get(self.kind, self.kind), self.tag)

    @property
    def stage(self):
        xs = [x for x in (self.x_of(r) for r in self.rows) if x is not None]
        if not xs:
            return ""
        return "%s→%s  %d 点" % (fmt_num(min(xs)), fmt_num(max(xs)), len(self.rows))


def _changed(a, b):
    if a is None and b is None:
        return False
    if a is None or b is None:
        return True
    return abs(a - b) > 1e-12


def segment(rows):
    """切扫描组。

    规则：(温度, 模式) 一变就切；组内再看「这一步动的是谁」——
    只有 Vtune 动 = Vtune 扫，只有 CT 动 = CT 扫，两个一起动 = 换扫法了，硬切。
    组的类型取组内多数票（单行组没有票，记 point）。
    """
    runs, cur = [], []
    for r in rows:
        if cur and (_changed(cur[-1].temp, r.temp) or cur[-1].mode != r.mode):
            runs.append(cur)
            cur = []
        cur.append(r)
    if cur:
        runs.append(cur)

    groups = []
    for run in runs:
        seg, votes = [run[0]], []
        for prev, row in zip(run, run[1:]):
            dv, dc = _changed(prev.vt, row.vt), _changed(prev.ct, row.ct)
            if dv and dc:                      # 换扫法：硬切
                groups.append((seg, votes))
                seg, votes = [row], []
                continue
            seg.append(row)
            if dc:
                votes.append("ct")
            elif dv:
                votes.append("vtune")
        groups.append((seg, votes))

    out, seen = [], {}
    for seg, votes in groups:
        if not seg:
            continue
        kind = "point"
        if votes:
            kind = "ct" if votes.count("ct") > votes.count("vtune") else "vtune"
        g = Group(len(out) + 1, kind, seg[0].temp)
        g.rows = seg
        key = (g.temp, g.kind)
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            g.tag = " #%d" % seen[key]
        for r in seg:
            r.group = g
        out.append(g)
    return out


# ---------------------------------------------------------------- 统计

def group_series(g, item):
    """本组的「横轴 -> 值」去重视图：同一个横轴值有多个点时取最后一个。

    什么时候会同一横轴多点：CT 扫经常先粗扫一遍（0/64/128/192/255）再回头细扫
    （0/1/2…），0 就出现了两次。取后者。

    ★ 汇总统计和明细页**共用这一份**。否则汇总按全部行取极值、明细一格只放得下
    一个值，就会出现「汇总报的极值在明细里查无此值」——报表一旦对不上账就没人敢信。
    """
    return OrderedDict((x, sv[0]) for x, sv in group_series_src(g, item).items())


def group_series_src(g, item):
    """同上，但连来源行一起给：x -> (值, Row)。

    明细页的格子要写成引用原始表的公式，就得知道这个值是从原表哪一行来的。
    """
    xf = g.x_of
    out = OrderedDict()
    for r in g.rows:
        x, v = xf(r), r.vals.get(item.col)
        if x is not None and v is not None:
            out[x] = (v, r)
    return OrderedDict(sorted(out.items()))


def _extremes(pairs):
    if not pairs:
        return None
    lo = min(pairs, key=lambda x: x[0])
    hi = max(pairs, key=lambda x: x[0])
    return {"min": lo[0], "min_x": lo[1], "max": hi[0], "max_x": hi[1],
            "delta": hi[0] - lo[0], "n": len(pairs)}


def items_with_data(groups, items):
    """这一批组里真的量到了的指标。

    两种扫法记的东西不一样：Vtune 扫每点都测点相噪/杂散/电流，CT 扫只记
    频率/功率/积分相噪。不按组筛一遍，CT 那几页会多出十几列整片空白。
    """
    return [it for it in items if any(group_series(g, it) for g in groups)]


def stats(g, item):
    return _extremes([(v, x) for x, v in group_series(g, item).items()])


def stats_all(groups, item):
    return _extremes([(v, "%s@%s" % (fmt_num(x), g.title.split()[0]))
                      for g in groups for x, v in group_series(g, item).items()])


def slopes(g, freq_item):
    """逐区间斜率：[(区间中点, ΔF/Δx, 左端, 右端)]。Vtune 扫就是 Kvco。"""
    ser = list(group_series(g, freq_item).items())
    out = []
    for (x0, f0), (x1, f1) in zip(ser, ser[1:]):
        if abs(x1 - x0) < 1e-12:
            continue
        out.append(((x0 + x1) / 2.0, (f1 - f0) / (x1 - x0), x0, x1))
    return out


def monotonic(g, freq_item):
    ser = list(group_series(g, freq_item).values())
    if len(ser) < 2:
        return None
    ups = sum(1 for a, b in zip(ser, ser[1:]) if b > a)
    dns = sum(1 for a, b in zip(ser, ser[1:]) if b < a)
    if ups and dns:
        return "否"
    return "是↑" if ups else ("是↓" if dns else "平")


def drift_vs_ref(g, ref, freq_item):
    """同一横轴值上，本组频率相对参考组的差：(最小, 最大, 重合点数)。"""
    a, b = group_series(g, freq_item), group_series(ref, freq_item)
    shared = [a[x] - b[x] for x in a if x in b]
    if not shared:
        return None
    return min(shared), max(shared), len(shared)


# ---------------------------------------------------------------- 画格子

def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style="thin", color="FF000000")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "f_head": PatternFill("solid", fgColor=FILL_HEADER, bgColor=FILL_HEADER),
        "f_group": PatternFill("solid", fgColor=FILL_GROUP, bgColor=FILL_GROUP),
        "f_res": PatternFill("solid", fgColor=FILL_RESULT, bgColor=FILL_RESULT),
        "f_sep": PatternFill("solid", fgColor=FILL_SEP, bgColor=FILL_SEP),
        "Font": Font,
    }


def put(ws, r, c, v, st, fill=None, bold=False, color=None, align="center", size=10):
    cell = ws.cell(row=r, column=c)
    if isinstance(v, Formula):          # emit() 交来的 (公式, 值) 对
        VCACHE.remember(ws, r, c, v.value)
        v = v.formula
    cell.value = v
    cell.border = st["border"]
    cell.alignment = st["center"] if align == "center" else st["left"]
    cell.font = st["Font"](bold=bold, color=color, size=size)
    if fill is not None:
        cell.fill = fill
    # 公式格子由 Excel 算，值是全精度的（102.11419753086425 这种）。给个显示格式，
    # 否则报告上一列数字长短不一没法看。0.### = 最多 3 位小数、整数不拖 .000
    if isinstance(v, float) or (isinstance(v, str) and v.startswith("=")):
        cell.number_format = "0.###"
    return cell


def _pass_fail_cf(ws, col_letter, r0, r1, st):
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill
    if r1 < r0:
        return
    rng = "%s%d:%s%d" % (col_letter, r0, col_letter, r1)
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['%s%d="FAIL"' % (col_letter, r0)],
        fill=PatternFill("solid", fgColor=FILL_FAIL, bgColor=FILL_FAIL),
        font=st["Font"](bold=True, color=COLOR_FLAG), stopIfTrue=False))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['%s%d="PASS"' % (col_letter, r0)],
        fill=PatternFill("solid", fgColor=FILL_PASS, bgColor=FILL_PASS),
        font=st["Font"](bold=True, color=COLOR_PASS), stopIfTrue=False))


# ---------------------------------------------------------------- 汇总页

def _range_table(ws, r0, groups, items, st, title, lay):
    """逐点极值表：Category|Item|Unit|Spec Min|Spec Max| |组i Min/Max/Δ| … | |合计| |判定

    格子里全是公式（对明细页的列取 MIN/MAX），不是算好的死值：
    明细页的每一格又指回原始表，所以这一页的任何数字都能一路追回源头。
    """
    from openpyxl.utils import get_column_letter as L

    def rng(g, it):
        if it.col not in lay["block"]:
            return None
        f, l = lay["block"][it.col]
        return rref(lay["sheet"], lay["col_of"][id(g)], f, l)

    plan = [("cat", "Category", 16), ("item", "Item", 22), ("unit", "Unit", 9),
            ("spec_min", "Min", 10), ("spec_typ", "Typ", 10), ("spec_max", "Max", 10),
            ("sep", "", 2)]
    blocks = [("Spec", "", [3, 4, 5])]
    for g in groups:
        base = len(plan)
        plan += [("min", "Min", 11), ("max", "Max", 11), ("delta", "Δ", 10), ("sep", "", 2)]
        blocks.append((g.title, g.stage, [base, base + 1, base + 2]))
    base = len(plan)
    plan += [("min", "Min", 11), ("max", "Max", 11), ("delta", "Δ", 10), ("sep", "", 2)]
    blocks.append(("合计", "%d 组" % len(groups), [base, base + 1, base + 2]))
    col_judge = len(plan)
    plan += [("judge", "判定", 12)]

    for i, (_k, _l, w) in enumerate(plan):
        cur = ws.column_dimensions[L(i + 1)].width
        if not cur or cur < w:
            ws.column_dimensions[L(i + 1)].width = w

    put(ws, r0, 1, title, st, st["f_group"], bold=True, align="left")
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=min(14, len(plan)))
    h0 = r0 + 1
    for r in range(h0, h0 + 3):
        for c in range(1, len(plan) + 1):
            put(ws, r, c, None, st, st["f_head"])
    for c0, label in ((0, "Category"), (1, "Item"), (2, "Unit")):
        ws.merge_cells(start_row=h0, start_column=c0 + 1, end_row=h0 + 2, end_column=c0 + 1)
        put(ws, h0, c0 + 1, label, st, st["f_head"], bold=True)
    ws.merge_cells(start_row=h0, start_column=col_judge + 1, end_row=h0 + 2,
                   end_column=col_judge + 1)
    put(ws, h0, col_judge + 1, "判定", st, st["f_head"], bold=True)
    for name, stage, cc in blocks:
        ws.merge_cells(start_row=h0, start_column=cc[0] + 1, end_row=h0, end_column=cc[-1] + 1)
        put(ws, h0, cc[0] + 1, name, st, st["f_head"], bold=True)
        ws.merge_cells(start_row=h0 + 1, start_column=cc[0] + 1, end_row=h0 + 1,
                       end_column=cc[-1] + 1)
        put(ws, h0 + 1, cc[0] + 1, stage, st, st["f_head"], bold=True, size=9)
        for ci in cc:
            put(ws, h0 + 2, ci + 1, plan[ci][1], st, st["f_head"], bold=True, size=9)
    for i, (k, _l, _w) in enumerate(plan):
        if k == "sep":
            for r in range(h0, h0 + 3):
                put(ws, r, i + 1, None, st, st["f_sep"])

    d0 = h0 + 3
    for n, it in enumerate(items):
        r = d0 + n
        for ci in range(len(plan)):
            put(ws, r, ci + 1, None, st,
                st["f_sep"] if plan[ci][0] == "sep" else st["f_res"])
        put(ws, r, 2, it.label, st, st["f_res"], align="left")
        put(ws, r, 3, it.unit, st, st["f_res"])
        alls = []
        for gi, g in enumerate(groups):
            a = rng(g, it)
            if not a:
                continue
            alls.append(a)
            cc, s = blocks[gi + 1][2], stats(g, it)
            put(ws, r, cc[0] + 1, emit(f_minmax("MIN", [a]),
                                       fmt_num(s["min"]) if s else None), st, st["f_res"])
            put(ws, r, cc[1] + 1, emit(f_minmax("MAX", [a]),
                                       fmt_num(s["max"]) if s else None), st, st["f_res"])
            put(ws, r, cc[2] + 1, emit(f_span([a]),
                                       fmt_num(s["delta"]) if s else None), st, st["f_res"])
        cc, s = blocks[-1][2], stats_all(groups, it)
        if alls:
            put(ws, r, cc[0] + 1, emit(f_minmax("MIN", alls),
                                       fmt_num(s["min"]) if s else None), st, st["f_res"])
            put(ws, r, cc[1] + 1, emit(f_minmax("MAX", alls),
                                       fmt_num(s["max"]) if s else None), st, st["f_res"])
            put(ws, r, cc[2] + 1, emit(f_span(alls),
                                       fmt_num(s["delta"]) if s else None), st, st["f_res"])
        smin, smax = "$D%d" % r, "$F%d" % r
        amin, amax = "%s%d" % (L(cc[0] + 1), r), "%s%d" % (L(cc[1] + 1), r)
        put(ws, r, col_judge + 1,
            '=IF(AND(%s="",%s=""),"",IF(AND(OR(%s="",%s>=%s),OR(%s="",%s<=%s)),"PASS","FAIL"))'
            % (smin, smax, smin, amin, smin, smax, amax, smax),
            st, st["f_res"], bold=True)

    i = 0
    while i < len(items):
        j = i
        while j + 1 < len(items) and items[j + 1].cat == items[i].cat:
            j += 1
        if j > i:
            ws.merge_cells(start_row=d0 + i, start_column=1, end_row=d0 + j, end_column=1)
        put(ws, d0 + i, 1, items[i].cat, st, st["f_res"], bold=True)
        i = j + 1

    _pass_fail_cf(ws, L(col_judge + 1), d0, d0 + len(items) - 1, st)
    return d0 + len(items)


def worst_dir(item):
    """这个指标往哪个方向算"差"。None = 不谈好坏（频率、调谐电压之类）。

    借 report-forge 的 limit le/ge 那个思路：一个结果只关心单边，
    相噪只有上限、功率只有下限，两边都要人填反而没人填。
    """
    src = item.src.lower()
    lab = item.label.lower()
    if "power" in src:
        return "min"                       # 输出功率越低越差
    if "current" in src:
        return "max"                       # 电流越大越差
    if "ipn" in src or lab.startswith("spotpn@") or lab.startswith("spur@"):
        return "max"                       # 相噪/杂散越接近 0 越差
    return None


def _at(ser, x0):
    """在「横轴->值」视图里取离 x0 最近的那个点，返回 (值, 实际横轴值)。"""
    if not ser or x0 is None:
        return None, None
    x = min(ser, key=lambda k: abs(k - x0))
    return ser[x], x


def build_conclusion(by_kind, items, freq_item, ref_temp, fvco, fvco_ref, LAY,
                     op_vtune=None):
    """把两种扫描合起来，推出真能下判断的几行。

    ★ 为什么必须合起来：这份测试的核心问题是「有没有锁不上的盲区」——
    单个电容码用调谐电压全程能盖多宽（只有 Vtune 扫知道）除以相邻码差多少
    （只有 CT 扫知道），比值 >1 才不留缝。两张扫描各算各的，这个数永远出不来。

    每行同时带两样：fn 算出来的值（脚本自己核对用）和 fml 生成的公式（真正写进
    格子的）。写进去的是公式，一路引用回明细页再回原始表——评审时点开任何一个
    数都能追到源头，原始数据改了整页跟着变。
    """
    vg, cg = {}, {}
    for kind, groups in by_kind:
        for g in groups:
            if g.temp is None:
                continue
            d = vg if kind == "vtune" else (cg if kind == "ct" else None)
            if d is not None:
                d.setdefault(g.temp, g)
    temps = sorted(set(vg) | set(cg))
    if not temps:
        return [], []

    rows = []

    def add(cat, item, unit, direction, fn, fml=None, kind="result", note="",
            key=False, rid=None):
        vals = {}
        for t in temps:
            try:
                vals[t] = fn(t)
            except Exception:
                vals[t] = None
        if any(v is not None for v in vals.values()):
            rows.append({"cat": cat, "item": item, "unit": unit, "dir": direction,
                         "kind": kind, "vals": vals, "note": note, "key": key,
                         "id": rid, "fml": fml})

    # ---- 引用小工具：把「哪个温度的哪种扫描的哪个指标」翻成单元格/区间地址 ----
    def _lay(k):
        return (LAY.get("vtune" if k == "v" else "ct") or {}).get("detail")

    def _g(k, t):
        return (vg if k == "v" else cg).get(t)

    def rng(k, t, it):
        lay, g = _lay(k), _g(k, t)
        if not lay or g is None or it is None or it.col not in lay["block"]:
            return None
        f, l = lay["block"][it.col]
        return rref(lay["sheet"], lay["col_of"][id(g)], f, l)

    def cell(k, t, it, x):
        lay, g = _lay(k), _g(k, t)
        if not lay or g is None or it is None or x is None:
            return None
        rr = (lay["xrow"].get(it.col) or {}).get(x)
        return cref(lay["sheet"], lay["col_of"][id(g)], rr) if rr else None

    def xcell(k, t, it, x):
        lay = _lay(k)
        if not lay or it is None or x is None:
            return None
        rr = (lay["xrow"].get(it.col) or {}).get(x)
        return cref(lay["sheet"], 0, rr) if rr else None

    def slp(k, t):
        s = (LAY.get("vtune" if k == "v" else "ct") or {}).get("slope")
        g = _g(k, t)
        if not s or g is None or id(g) not in s["col_of"]:
            return None
        return s

    def srng(k, t):
        s, g = slp(k, t), _g(k, t)
        return rref(s["sheet"], s["col_of"][id(g)], s["first"], s["last"]) if s else None

    def drng(k, t):
        d = (LAY.get("vtune" if k == "v" else "ct") or {}).get("drift")
        g = _g(k, t)
        if not d or g is None or id(g) not in d["col_of"]:
            return None
        return rref(d["sheet"], d["col_of"][id(g)], d["first"], d["last"])

    def ser(k, t):
        g = _g(k, t)
        return group_series(g, freq_item) if (g and freq_item) else {}

    def ends(k, t):
        s = sorted(ser(k, t))
        return (s[0], s[-1]) if len(s) >= 2 else (None, None)

    def v_ct(t):
        g = cg.get(t)
        xs = [r.vt for r in g.rows if r.vt is not None] if g else []
        return xs[0] if xs else None

    def near_x(k, t, x0):
        s = ser(k, t)
        return min(s, key=lambda z: abs(z - x0)) if (s and x0 is not None) else None

    def dfv(t):
        s = ser("v", t)
        return (max(s.values()) - min(s.values())) if len(s) >= 2 else None

    def steps(k, t):
        g = _g(k, t)
        return [abs(s[1]) for s in slopes(g, freq_item)] if (g and freq_item) else []

    # ---- 条件行 ----
    def f_temp(t, R):
        g = _g("v", t) or _g("c", t)
        if g is None or LAY.get("tcol") is None:
            return None
        return vref(cref(LAY["src"], LAY["tcol"], g.rows[0].xl))

    def f_vrange(t, R):
        a, b = ends("v", t)
        ca, cb = xcell("v", t, freq_item, a), xcell("v", t, freq_item, b)
        return ('=%s&"~"&%s' % (ca, cb)) if (ca and cb) else None

    def f_crange(t, R):
        a, b = ends("c", t)
        ca, cb = xcell("c", t, freq_item, a), xcell("c", t, freq_item, b)
        rr = rng("c", t, freq_item)
        return ('=%s&"~"&%s&" ("&COUNT(%s)&")"' % (ca, cb, rr)) if (ca and cb and rr) else None

    def f_vct(t, R):
        if t not in cg or LAY.get("vtcol") is None:
            return None
        return vref(cref(LAY["src"], LAY["vtcol"], cg[t].rows[0].xl))

    add("Condition", "Temperature", "℃", "", lambda t: t, fml=f_temp, kind="cond",
        note="原表 Temperature 列")
    add("Condition", "Vtune Range", "V", "",
        lambda t: "%s~%s" % (fmt_num(ends("v", t)[0]), fmt_num(ends("v", t)[1]))
        if ends("v", t)[0] is not None else None, fml=f_vrange, kind="cond",
        note="Vtune 扫的起止设定值")
    add("Condition", "CT Code Range (points)", "code", "",
        lambda t: "%s~%s (%d)" % (fmt_num(ends("c", t)[0]), fmt_num(ends("c", t)[1]),
                                  len(ser("c", t))) if ends("c", t)[0] is not None else None,
        fml=f_crange, kind="cond", note="CT 扫的起止码值（括号内为测点数）")
    add("Condition", "Vtune @ CT sweep", "V", "",
        lambda t: v_ct(t), fml=f_vct, kind="cond", note="CT 扫全程把 Vtune 钉在这个值")
    if fvco is not None:
        add("Condition", "Target fVCO", "MHz", "", lambda t: fvco,
            fml=lambda t, R: vref(fvco_ref) if fvco_ref else None,
            kind="cond", rid="fvco", note="原表 fVCO_MHz 列")

    # ---- 频率范围 / 调谐范围 / 目标余量 ----
    # 备注里写清每个数的算法：报告上出现「估计」这种字眼，评审第一句就是
    # 「这怎么算出来的」。把算式写在旁边，就没有这一问。
    _t0 = next((t for t in temps if t in vg), None)
    _vmin, _vmax = ends("v", _t0) if _t0 is not None else (None, None)
    _tc = next((t for t in temps if t in cg), None)
    _vc = v_ct(_tc) if _tc is not None else None
    _V = lambda x: ("%sV" % fmt_num(x)) if x is not None else "?"

    def fr(t, side):
        """Fmin / Fmax：CT 扫的端点，再用 Vtune 相对 CT 扫那个电压还能拉的量补上。

        CT 扫是在固定 Vtune 上做的，Vtune 扫是在固定 CT 码上做的，
        所以两个角点没有直接测；这里就是把两段量拼起来，算式写在备注里。
        """
        sv, sc = ser("v", t), ser("c", t)
        if len(sv) < 2 or len(sc) < 2:
            return None
        xn = near_x("v", t, v_ct(t))
        if xn is None:
            return None
        f0 = sv[xn]
        return (min(sc.values()) - (f0 - min(sv.values()))) if side == "low" \
            else (max(sc.values()) + (max(sv.values()) - f0))

    def fr_f(t, side):
        rv, rc = rng("v", t, freq_item), rng("c", t, freq_item)
        c0 = cell("v", t, freq_item, near_x("v", t, v_ct(t)))
        if not (rv and rc and c0):
            return None
        return ("=MIN(%s)-(%s-MIN(%s))" % (rc, c0, rv)) if side == "low" \
            else ("=MAX(%s)+(MAX(%s)-%s)" % (rc, rv, c0))

    add("Frequency Range", "Fmin", "MHz", "≤", lambda t: fr(t, "low"),
        fml=lambda t, R: fr_f(t, "low"), rid="fmin",
        note="CT 扫最低频 − [F(%s) − F(%s)]" % (_V(_vc), _V(_vmin)))
    add("Frequency Range", "Fmax", "MHz", "≥", lambda t: fr(t, "high"),
        fml=lambda t, R: fr_f(t, "high"), rid="fmax",
        note="CT 扫最高频 + [F(%s) − F(%s)]" % (_V(_vmax), _V(_vc)))
    add("Frequency Range", "Tuning Range", "MHz", "≥",
        lambda t: (fr(t, "high") - fr(t, "low"))
        if (fr(t, "low") is not None and fr(t, "high") is not None) else None,
        fml=lambda t, R: f_guard("%s-%s" % (R("fmax", t), R("fmin", t)),
                                 R("fmin", t), R("fmax", t)),
        note="Fmax − Fmin")
    if fvco is not None:
        add("Frequency Range", "Margin to fVCO (low)", "MHz", "≥",
            lambda t: (fvco - fr(t, "low")) if fr(t, "low") is not None else None,
            fml=lambda t, R: f_guard("%s-%s" % (R("fvco", t), R("fmin", t)),
                                     R("fmin", t), R("fvco", t)),
            note="目标 fVCO − Fmin")
        add("Frequency Range", "Margin to fVCO (high)", "MHz", "≥",
            lambda t: (fr(t, "high") - fvco) if fr(t, "high") is not None else None,
            fml=lambda t, R: f_guard("%s-%s" % (R("fmax", t), R("fvco", t)),
                                     R("fmax", t), R("fvco", t)),
            note="Fmax − 目标 fVCO")

    # ---- 频段搭接：有没有锁不上的盲区 ----
    add("CT Band", "Sub-band Tuning Range", "MHz", "≥", dfv,
        fml=lambda t, R: f_span([rng("v", t, freq_item)]) if rng("v", t, freq_item) else None,
        rid="dfv",
        note="F(%s) − F(%s)，CT 码固定不动" % (_V(_vmax), _V(_vmin)))
    add("CT Band", "CT Band Coverage", "MHz", "≥",
        lambda t: (max(ser("c", t).values()) - min(ser("c", t).values()))
        if len(ser("c", t)) >= 2 else None,
        fml=lambda t, R: f_span([rng("c", t, freq_item)]) if rng("c", t, freq_item) else None,
        note="CT 扫最高频 − 最低频，Vtune 钉在 %s" % _V(_vc))
    add("CT Band", "CT Band Step (avg)", "MHz/code", "",
        lambda t: (sum(steps("c", t)) / len(steps("c", t))) if steps("c", t) else None,
        fml=lambda t, R: ('=IF(COUNT(%s)=0,"",ABS(AVERAGE(%s)))'
                          % (srng("c", t), srng("c", t))) if srng("c", t) else None,
        rid="step_avg", note="相邻两个码的频率差 |ΔF/ΔCT| 的平均")
    add("CT Band", "CT Band Step (max)", "MHz/code", "≤",
        lambda t: max(steps("c", t)) if steps("c", t) else None,
        fml=lambda t, R: f_absmax(srng("c", t)) if srng("c", t) else None, rid="step_max",
        note="相邻两个码的频率差 |ΔF/ΔCT| 的最大值")

    # ---- 温漂 ----
    ref_t = min(temps, key=lambda t: abs(t - ref_temp)) if temps else None
    if ref_t is not None and len(temps) > 1:
        def drift(t):
            # 参考温对自己的漂移恒等于 0，那一格不是测量结果，留空
            if t == ref_t:
                return None
            a, b = ser("v", t), ser("v", ref_t)
            sh = [a[x] - b[x] for x in a if x in b]
            return max(sh, key=abs) if sh else None

        def f_drift(t, R):
            d = drng("v", t)
            return f_signed_absmax(d) if (d and t != ref_t) else None

        rt = fmt_num(ref_t)
        add("Temp Drift", "Freq Drift vs %s℃" % rt, "MHz", "≤", drift,
            fml=f_drift, rid="drift",
            note="同一 Vtune 下 F(T) − F(%s℃)，取 |ΔF| 最大的那个点，带符号" % rt)
        add("Temp Drift", "Drift in CT Codes", "code", "≤",
            lambda t: (abs(drift(t)) / (sum(steps("c", ref_t)) / len(steps("c", ref_t))))
            if (t != ref_t and drift(t) is not None and steps("c", ref_t)) else None,
            fml=lambda t, R: '=IF(N(%s)=0,"",ABS(%s)/%s)' % (R("step_avg", ref_t),
                                                             R("drift", t),
                                                             R("step_avg", ref_t)),
            note="|Freq Drift| ÷ CT Band Step 平均（%s℃）" % rt)

    # ---- 压控增益 ----
    def kv_work(t):
        g = vg.get(t)
        ss = slopes(g, freq_item) if (g and freq_item) else []
        if not ss:
            return None
        return abs(sorted(ss, key=lambda z: abs(z[0] - (v_ct(t) or 0.4)))[0][1])

    def kv_work_f(t, R):
        s, g = slp("v", t), vg.get(t)
        ss = slopes(g, freq_item) if (g and freq_item) else []
        if not s or not ss:
            return None
        best = sorted(ss, key=lambda z: abs(z[0] - (v_ct(t) or 0.4)))[0]
        try:
            i = s["mids"].index(round(best[0], 9))
        except (KeyError, ValueError):
            return None
        return "=ABS(%s)" % cref(s["sheet"], s["col_of"][id(g)], s["first"] + i)

    add("Kvco", "Kvco @ %s" % _V(_vc), "MHz/V", "", kv_work, fml=kv_work_f,
        note="ΔF/ΔVtune，取 %s 所在的那个区间" % _V(_vc))
    add("Kvco", "Kvco min", "MHz/V", "",
        lambda t: min((abs(s[1]) for s in slopes(vg[t], freq_item)), default=None)
        if (t in vg and freq_item) else None,
        fml=lambda t, R: f_absmin(srng("v", t)) if srng("v", t) else None, rid="kv_min",
        note="逐区间 ΔF/ΔVtune 里 |值| 最小的")
    add("Kvco", "Kvco max", "MHz/V", "",
        lambda t: max((abs(s[1]) for s in slopes(vg[t], freq_item)), default=None)
        if (t in vg and freq_item) else None,
        fml=lambda t, R: f_absmax(srng("v", t)) if srng("v", t) else None, rid="kv_max",
        note="逐区间 ΔF/ΔVtune 里 |值| 最大的")
    add("Kvco", "Kvco max/min", "-", "≤",
        lambda t: (max(abs(s[1]) for s in slopes(vg[t], freq_item))
                   / min(abs(s[1]) for s in slopes(vg[t], freq_item)))
        if (t in vg and freq_item and slopes(vg[t], freq_item)
            and min(abs(s[1]) for s in slopes(vg[t], freq_item)) > 1e-12) else None,
        fml=lambda t, R: f_div(R("kv_max", t), R("kv_min", t)),
        note="Kvco max ÷ Kvco min")

    # ---- 工作点性能：Vtune 钉在一个电压上，看三个温度 ----
    # ★ 相噪/功率这些不能"跨整个 Vtune 扫取最差"：不同 Vtune = 不同振荡频率，
    #   把它们并成一个集合取极值等于把不同工作条件混在一起，那个数没有意义。
    #   钉在一个工作点（默认 = CT 扫用的那个 Vtune）看温度差异才是要问的。
    #   同一个 Vtune 在不同温度对应的频率并不一样，所以频率也列出来当参照。
    _op = op_vtune if op_vtune is not None else _vc
    if _op is None and _t0 is not None:
        xs0 = sorted(ser("v", _t0))
        _op = xs0[len(xs0) // 2] if xs0 else None

    def op_x(t):
        return near_x("v", t, _op)

    if _op is not None:
        cat = "@ Vtune %s" % _V(_op)
        add(cat, "Freq", "MHz", "",
            lambda t: ser("v", t).get(op_x(t)),
            fml=lambda t, R: ("=" + cell("v", t, freq_item, op_x(t)))
            if cell("v", t, freq_item, op_x(t)) else None,
            note="Vtune 扫里 %s 那一点的输出频率（各温度不同）" % _V(_op))
        for it in items:
            d = worst_dir(it)
            if d is None:
                continue

            def at_op(t, it=it):
                g = _g("v", t)
                return group_series(g, it).get(op_x(t)) if g else None

            def at_op_f(t, R, it=it):
                c = cell("v", t, it, op_x(t))
                return ("=" + c) if c else None

            add(cat, it.label, it.unit, "≤" if d == "max" else "≥",
                at_op, fml=at_op_f,
                note="Vtune 扫里 %s 那一点的实测值" % _V(_op))
    return rows, temps


def _conclusion_table(ws, r0, temps, rows, st, title):
    """结论表：列 = 各温度；每行一个能下判断的结论。"""
    from openpyxl.utils import get_column_letter as L

    # Category|Item|Unit|Limit|Spec Min|Spec Typ|Spec Max
    # Typ 只是给人对照的标称值，不参与判定——判定还是看 Min/Max 两个边界。
    n_fix = 7
    c_sim = n_fix + len(temps) + 1     # 仿真值：留空给人填，不参与判定
    c_note = c_sim + 1
    c_judge = c_note + 1
    widths = [16, 34, 10, 6, 10, 10, 10] + [13] * len(temps) + [12, 52, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[L(i + 1)].width = w

    put(ws, r0, 1, title, st, st["f_group"], bold=True, align="left")
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=c_judge)
    h = r0 + 1
    for c in range(1, c_judge + 1):
        put(ws, h, c, None, st, st["f_head"])
        put(ws, h + 1, c, None, st, st["f_head"])
    for c, lab in ((1, "Category"), (2, "Item"), (3, "Unit"), (4, "Limit"),
                   (c_sim, "仿真"), (c_note, "备注"), (c_judge, "判定")):
        ws.merge_cells(start_row=h, start_column=c, end_row=h + 1, end_column=c)
        put(ws, h, c, lab, st, st["f_head"], bold=True, size=9)
    ws.merge_cells(start_row=h, start_column=5, end_row=h, end_column=7)
    put(ws, h, 5, "Spec", st, st["f_head"], bold=True, size=9)
    for c, lab in ((5, "Min"), (6, "Typ"), (7, "Max")):
        put(ws, h + 1, c, lab, st, st["f_head"], bold=True, size=9)
    ws.merge_cells(start_row=h, start_column=n_fix + 1, end_row=h,
                   end_column=n_fix + len(temps))
    put(ws, h, n_fix + 1, "测试", st, st["f_head"], bold=True, size=9)
    for j, t in enumerate(temps):
        put(ws, h + 1, n_fix + 1 + j, "%s℃" % fmt_num(t), st, st["f_head"], bold=True, size=9)

    d0 = h + 2
    # 行内互相引用（例如「目标余量」= 目标 fVCO − 覆盖下限）靠这个把 id 翻成地址
    idx_of = {row["id"]: i for i, row in enumerate(rows) if row.get("id")}

    def R(rid, t):
        if rid not in idx_of or t not in temps:
            return '""'
        return "$%s$%d" % (L(n_fix + 1 + temps.index(t)), d0 + idx_of[rid])

    for n, row in enumerate(rows):
        r = d0 + n
        fill = st["f_group"] if row["kind"] == "cond" else st["f_res"]
        for c in range(1, c_judge + 1):
            put(ws, r, c, None, st, fill)
        put(ws, r, 2, row["item"], st, fill, align="left", bold=row.get("key", False))
        put(ws, r, c_sim, None, st, st["f_res"])          # 仿真值：手工填
        put(ws, r, 3, row["unit"], st, fill)
        put(ws, r, 4, row["dir"], st, fill, bold=True)
        numeric = True
        for j, t in enumerate(temps):
            v = row["vals"].get(t)
            if isinstance(v, str):
                numeric = False
            f = None
            if v is not None and row.get("fml"):
                try:
                    f = row["fml"](t, R)           # ★ 优先写公式，不写死值
                except Exception:
                    f = None
            put(ws, r, n_fix + 1 + j,
                emit(f, v if isinstance(v, str) else fmt_num(v)), st, fill)
        put(ws, r, c_note, as_text(row["note"]), st, fill, align="left", size=8)
        if numeric and row["kind"] == "result":
            rng = "%s%d:%s%d" % (L(n_fix + 1), r, L(n_fix + len(temps)), r)
            put(ws, r, c_judge,
                '=IF(COUNT(%s)=0,"",IF(AND($E%d="",$G%d=""),"",'
                'IF(AND(OR($E%d="",MIN(%s)>=$E%d),OR($G%d="",MAX(%s)<=$G%d)),"PASS","FAIL")))'
                % (rng, r, r, r, rng, r, r, rng, r), st, fill, bold=True)

    i = 0                                   # Category 纵向合并
    while i < len(rows):
        j = i
        while j + 1 < len(rows) and rows[j + 1]["cat"] == rows[i]["cat"]:
            j += 1
        if j > i:
            ws.merge_cells(start_row=d0 + i, start_column=1, end_row=d0 + j, end_column=1)
        put(ws, d0 + i, 1, rows[i]["cat"], st,
            st["f_group"] if rows[i]["kind"] == "cond" else st["f_res"], bold=True)
        i = j + 1
    _pass_fail_cf(ws, L(c_judge), d0, d0 + len(rows) - 1, st)
    ws.freeze_panes = "%s%d" % (L(n_fix + 1), d0)
    return d0 + len(rows)


def write_conclusion(wb, rows, temps, st, title):
    ws = wb.create_sheet("结论")
    _conclusion_table(ws, 1, temps, rows, st, title)
    return ws


def write_summary(wb, by_kind, items, st, LAY):
    """每种扫法一张逐点极值页（查数用；结论在「结论」页）。

    Vtune 扫和 CT 扫的组数不一样、列宽也不一样，硬塞进一页会互相把列撑歪，
    所以分页：第一种扫法占「汇总」，其余的另起「汇总-CT扫」这样的页。
    派生指标（Kvco / 覆盖 / 温漂那些）不在这里重复一遍——结论页已经给全了。
    """
    first_ws = None
    for kind, groups in by_kind:
        lay = (LAY.get(kind) or {}).get("detail")
        if not groups or not lay:
            continue
        name = "汇总" if first_ws is None else "汇总-%s" % KIND_LABEL.get(kind, kind)
        ws = wb.create_sheet(name)
        if first_ws is None:
            first_ws = ws
        _range_table(ws, 1, groups, items_with_data(groups, items), st,
                     "逐点极值 · %s" % KIND_LABEL.get(kind, kind), lay)
        ws.freeze_panes = "D5"
    return first_ws


def write_diag(wb, meta, st):
    """数据处理记录：排除了哪些行、探查到什么异常。

    ★ 默认不写进簿子（--notes 才出）。这页是给做表的人对账用的，不是报告内容；
    交出去的簿子里出现「⚠ 告警」「没算进汇总的行」这类字样，评审的人第一反应
    是数据有问题。信息本身不能丢——脚本每次运行都会把同样的内容打在控制台上。
    """
    ws = wb.create_sheet("数据处理记录")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 90
    r = 1
    put(ws, r, 1, "分组规则", st, st["f_group"], bold=True, align="left")
    put(ws, r, 2, meta["why_groups"], st, None, align="left")
    r += 2
    put(ws, r, 1, "未纳入统计的行", st, st["f_group"], bold=True, align="left")
    r += 1
    put(ws, r, 1, "原表行号", st, st["f_head"], bold=True)
    put(ws, r, 2, "原因", st, st["f_head"], bold=True)
    for xl, why in meta["excluded"]:
        r += 1
        put(ws, r, 1, xl, st)
        put(ws, r, 2, as_text(why), st, align="left")
    if meta["warnings"]:
        r += 2
        put(ws, r, 1, "探查告警", st, st["f_group"], bold=True, align="left")
        for w in meta["warnings"]:
            r += 1
            put(ws, r, 2, w, st, None, align="left")
    return ws


# ---------------------------------------------------------------- 明细页

def write_detail(wb, name, groups, items, st, src_title, target=None,
                 target_item=None, target_ref=None):
    """每个指标一块：行=横轴取值（各组并集，升序），列=组。图表就吃这个。

    target 给了的话，在频率那一块多加一列常数（目标 fVCO），图上就成了一条
    水平参考线——目标频率有没有落在覆盖范围里，扫一眼就知道，不用去查数。
    """
    from openpyxl.utils import get_column_letter as L
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 12
    for i in range(len(groups)):
        ws.column_dimensions[L(2 + i)].width = 14

    xs = sorted({g.x_of(r) for g in groups for r in g.rows if g.x_of(r) is not None})
    blocks, r = [], 1
    layout = {"sheet": ws.title, "col_of": {}, "block": {}, "xrow": {}, "xs": xs}
    for i, g in enumerate(groups):
        layout["col_of"][id(g)] = 1 + i          # 0 基列号，0 号是横轴列
    for it in items:
        # ★ 目标频率参考线只在它落进本图数据范围时才画。落在外面的话，
        #   Y 轴被拉到目标值那么高，真正的曲线被压成底部一条细带，形状全看不见——
        #   一条"有用的参考线"能把整张图毁掉。差多少在结论页有精确数。
        want_target = target is not None and it is target_item
        if want_target:
            allv = [v for g in groups for v in group_series(g, it).values()]
            lo, hi = (min(allv), max(allv)) if allv else (None, None)
            pad = (hi - lo) * 0.15 if (lo is not None and hi > lo) else 0
            want_target = lo is not None and (lo - pad) <= target <= (hi + pad)
        n_col = len(groups) + (1 if want_target else 0)
        put(ws, r, 1, "%s  [%s]" % (it.label, it.unit), st, st["f_group"], bold=True, align="left")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1 + n_col)
        r += 1
        put(ws, r, 1, groups[0].x_label, st, st["f_head"], bold=True)
        for i, g in enumerate(groups):
            put(ws, r, 2 + i, g.title, st, st["f_head"], bold=True, size=9)
        if want_target:
            put(ws, r, 2 + len(groups), "目标 fVCO", st, st["f_head"], bold=True, size=9)
        head, r = r, r + 1
        first = r
        series = [group_series_src(g, it) for g in groups]
        xrow = {}
        for x in xs:
            put(ws, r, 1, fmt_num(x), st, st["f_res"])   # 横轴是量化后的刻度，不引用
            xrow[x] = r
            for i in range(len(groups)):
                sv = series[i].get(x)
                if sv is None:
                    put(ws, r, 2 + i, None, st, st["f_res"])
                else:
                    # ★ 不写死：指回原始表那一格。评审时点开就能追到源头。
                    put(ws, r, 2 + i,
                        emit(vref(cref(src_title, it.col, sv[1].xl)), fmt_num(sv[0])),
                        st, st["f_res"])
            if want_target:
                put(ws, r, 2 + len(groups), emit(target_ref, target), st, st["f_res"])
            r += 1
        blocks.append((it, head, first, r - 1, n_col, target if want_target else None))
        layout["block"][it.col] = (first, r - 1)
        layout["xrow"][it.col] = xrow
        r += 1
    ws.freeze_panes = "B1"
    return ws, blocks, layout


def write_slope(wb, name, groups, freq_item, st, unit, xlabel, layout):
    """逐点斜率页：行=区间中点，列=组。Vtune 扫的就是 Kvco vs Vtune。"""
    from openpyxl.utils import get_column_letter as L
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 14
    for i in range(len(groups)):
        ws.column_dimensions[L(2 + i)].width = 14

    per = [dict(((round(m, 9), (a, b, v)) for m, v, a, b in slopes(g, freq_item)))
           for g in groups]
    mids = sorted({m for d in per for m in d})
    put(ws, 1, 1, "Δ%s / Δ%s，横坐标取区间中点  [%s]"
        % (freq_item.label, xlabel.split()[0], unit), st, st["f_group"], bold=True, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(groups))
    put(ws, 2, 1, "%s 中点" % xlabel, st, st["f_head"], bold=True)
    for i, g in enumerate(groups):
        put(ws, 2, 2 + i, g.title, st, st["f_head"], bold=True, size=9)
    dt, xrow = layout["sheet"], layout["xrow"][freq_item.col]
    r = 3
    for m in mids:
        put(ws, r, 1, fmt_num(m), st, st["f_res"])
        for i, g in enumerate(groups):
            ab = per[i].get(m)
            if ab is None:
                put(ws, r, 2 + i, None, st, st["f_res"])
                continue
            # 斜率也不写死：(右端-左端)/(右横轴-左横轴)，四个格子全指向明细页
            c = layout["col_of"][id(g)]
            r0, r1 = xrow[ab[0]], xrow[ab[1]]
            put(ws, r, 2 + i, emit("=(%s-%s)/(%s-%s)" % (
                cref(dt, c, r1), cref(dt, c, r0), cref(dt, 0, r1), cref(dt, 0, r0)),
                fmt_num(ab[2])), st, st["f_res"])
        r += 1
    ws.freeze_panes = "B2"
    yv = [v for d in per for (_a, _b, v) in d.values()]
    sers = [dict((m, ab[2]) for m, ab in d.items()) for d in per]
    xrow = dict((m, 3 + k) for k, m in enumerate(mids))
    return ws, (2, 3, r - 1), {"sheet": ws.title,
                               "col_of": {id(g): 1 + i for i, g in enumerate(groups)},
                               "first": 3, "last": r - 1, "mids": mids,
                               "xvals": mids, "yvals": yv, "sers": sers, "xrow": xrow}


def write_drift(wb, name, groups, ref_g, freq_item, st, layout, xlabel):
    """温漂明细：逐横轴点，各组相对参考温的频差。

    单独成页有两个用处：结论页的「最大频漂」可以直接对这一列取极值（不用数组
    公式），而且这条 ΔF-vs-横轴 曲线本身就该看——温漂在调谐范围两端往往不一样。
    """
    from openpyxl.utils import get_column_letter as L
    others = [g for g in groups if g is not ref_g]
    if not others or freq_item is None:
        return None, None
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 14
    for i in range(len(others)):
        ws.column_dimensions[L(2 + i)].width = 16
    put(ws, 1, 1, "ΔF 相对 %s  [MHz]" % ref_g.title, st, st["f_group"], bold=True, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(others))
    put(ws, 2, 1, xlabel, st, st["f_head"], bold=True)
    for i, g in enumerate(others):
        put(ws, 2, 2 + i, g.title, st, st["f_head"], bold=True, size=9)

    dt, xrow = layout["sheet"], layout["xrow"][freq_item.col]
    cref_ = layout["col_of"]
    sref = group_series(ref_g, freq_item)
    sers = [group_series(g, freq_item) for g in others]
    r = 3
    for x in layout["xs"]:
        if x not in xrow:
            continue
        put(ws, r, 1, fmt_num(x), st, st["f_res"])
        rr = xrow[x]
        b = cref(dt, cref_[id(ref_g)], rr)
        for i, g in enumerate(others):
            va, vb = sers[i].get(x), sref.get(x)
            if va is None or vb is None:
                # ★ 必须留真空格，不能写一个返回 "" 的公式。
                #   Excel 的"空值按间断处理"只认真正的空单元格；公式返回的空串
                #   是有内容的格子，散点图会把它当 0 画进去——常温扫了全部 256 个
                #   码、高低温只扫 5 个，中间 251 格全成 0，曲线在 5 个测点之间
                #   来回扎到零，看着像一串尖刺。
                put(ws, r, 2 + i, None, st, st["f_res"])
                continue
            a = cref(dt, cref_[id(g)], rr)
            put(ws, r, 2 + i,
                emit("=%s-%s" % (a, b), fmt_num(va - vb)), st, st["f_res"])
        r += 1
    ws.freeze_panes = "B2"
    return ws, {"sheet": ws.title, "col_of": {id(g): 1 + i for i, g in enumerate(others)},
                "first": 3, "last": r - 1, "ref": ref_g}


# ---------------------------------------------------------------- 图表

# 曲线配色：不给显式颜色的话，Excel 会用主题强调色的**深浅变体**去区分系列——
# 屏幕上勉强能分，打印或投影基本分不开。这里点名指定对比度足够的颜色。
SERIES_COLORS = ["1F77B4", "D62728", "2CA02C", "FF7F0E", "9467BD", "8C564B"]
TARGET_COLOR = "808080"


def _nice(lo, hi, pad=0.06):
    """坐标轴范围：贴着数据留一点边距，再对齐到整刻度。

    不显式给 min/max 的话由 Excel 自己定，同一组图的范围会各挑各的，
    几张图并排看时曲线高低没法直接比；某些情况还会把基线拉到 0，
    把一条 2000~2180 MHz 的调谐曲线压成顶上一条平线。
    """
    if lo is None or hi is None or hi < lo:
        return None
    # 整条线是常数：给个对称的窄带，别让轴退化。
    # 判"相等"要用相对容差——斜率这类算出来的值常差在 1e-13 上，不是严格相等，
    # 但那点差异画出来就是坐标轴退化成 1e-13 量级、刻度全是科学计数法。
    if hi - lo <= max(abs(lo), abs(hi), 1e-12) * 1e-9:
        d = max(abs(hi) * 0.05, 1e-6)
        return lo - d, hi + d
    span = hi - lo
    lo, hi = lo - span * pad, hi + span * pad
    step = 10 ** math.floor(math.log10((hi - lo) / 8.0))
    return math.floor(lo / step) * step, math.ceil(hi / step) * step


def _scatter(title, xtitle, ytitle, logx=False, xlim=None, ylim=None):
    from openpyxl.chart import ScatterChart
    from openpyxl.chart.axis import ChartLines
    ch = ScatterChart()
    ch.title = title
    ch.x_axis.title = xtitle
    ch.y_axis.title = ytitle
    ch.x_axis.delete = False          # openpyxl 老毛病：不显式置 False 坐标轴不显示
    ch.y_axis.delete = False
    ch.x_axis.majorGridlines = ChartLines()
    ch.y_axis.majorGridlines = ChartLines()
    ch.x_axis.numFmt = "General"
    ch.y_axis.numFmt = "General"
    ch.height, ch.width = 8.5, 14
    # ★ 属性名必须是 openpyxl 的 display_blanks / visible_cells_only。
    #   写成 OOXML 里的 dispBlanksAs / plotVisOnly 不会报错，只是挂了个没人读的
    #   属性，序列化时压根不看——图照样生成，设置却一个都没生效。
    ch.display_blanks = "gap"          # 断点留空，不连成直线
    ch.visible_cells_only = False      # 数据页是隐藏的，不关掉这个图会整片空白
    if ch.legend is not None:
        ch.legend.position = "b"       # 默认在右边，会吃掉四分之一绘图区
        ch.legend.overlay = False
    if logx:
        ch.x_axis.scaling.logBase = 10
    else:
        if xlim:
            ch.x_axis.scaling.min, ch.x_axis.scaling.max = xlim
        if ylim:
            ch.y_axis.scaling.min, ch.y_axis.scaling.max = ylim
    return ch


def _style_series(s, color, dashed=False, marker="circle", size=5):
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    ln = LineProperties(solidFill=color, w=22000)
    if dashed:
        ln.prstDash = "dash"
    s.graphicalProperties = GraphicalProperties(ln=ln)
    mk = Marker(symbol=marker, size=size)
    if marker != "none":
        mk.graphicalProperties = GraphicalProperties(
            solidFill=color, ln=LineProperties(solidFill=color))
    s.marker = mk


def _label_points(s, idxs, numfmt="0.###"):
    """只给指定的几个点打数值标签。

    ★ 标的是**全图的最低点和最高点**，不是每条线各标首末点：三条温度曲线在
      同一个横轴端点上值挨得很近，各标各的会叠成一坨看不清。全图两个极值点
      各一个标签，位置天然分得开，而且那两个点正是要看的。
    """
    from openpyxl.chart.label import DataLabel, DataLabelList
    if not idxs:
        return
    s.dLbls = DataLabelList()
    s.dLbls.showVal = False
    s.dLbls.showSerName = False
    s.dLbls.showCatName = False
    s.dLbls.showLegendKey = False
    for i in sorted(set(idxs)):
        dl = DataLabel(idx=i)
        dl.showVal = True
        dl.showSerName = False
        dl.showCatName = False
        dl.showLegendKey = False
        dl.numFmt = numfmt
        dl.dLblPos = "t"
        s.dLbls.dLbl.append(dl)


def _add_series(ch, ws, n_series, head, first, last, counts=None, labels=None,
                target_idx=None):
    """挂数据系列。

    几处必须显式设，否则图会骗人：
    · smooth：OOXML 散点图默认走平滑曲线，压控曲线会被画出假拐点（读者
      会以为 Kvco 在中间有突变）；点多的时候更是扭成波浪。一律关掉。
    · marker 按**每条线自己的点数**定，不能按整块的行数：同一张图上常温扫了
      256 个码、高低温只扫了 5 个，用整块行数判断的话那 5 个点的线就没有
      marker，被画成一条跨满全程的直线——看着像"全程都测过"。
    · 颜色点名给，别让 Excel 用主题色深浅去分。
    """
    from openpyxl.chart import Reference, Series
    xref = Reference(ws, min_col=1, min_row=first, max_row=last)
    for i in range(n_series):
        n_pts = counts[i] if (counts and i < len(counts)) else (last - first + 1)
        yref = Reference(ws, min_col=2 + i, min_row=head, max_row=last)
        s = Series(yref, xref, title_from_data=True)
        s.smooth = False
        if target_idx is not None and i == target_idx:
            _style_series(s, TARGET_COLOR, dashed=True, marker="none")
        else:
            _style_series(s, SERIES_COLORS[i % len(SERIES_COLORS)],
                          marker="circle" if n_pts <= 60 else "none",
                          size=5 if n_pts <= 30 else 3)
            if labels:
                _label_points(s, labels.get(i) or [])
        ch.series.append(s)


class ChartGrid:
    """图一张挨一张往下摆：两列（A / I），每行带 16 行高。

    别用 ws.max_row 当锚点——图不算单元格内容，max_row 一直是 1，
    后加的图会全叠在第一张上面。
    """

    def __init__(self, ws):
        self.ws, self.n = ws, 0

    def add(self, ch):
        self.ws.add_chart(ch, "%s%d" % ("A" if self.n % 2 == 0 else "I",
                                        3 + (self.n // 2) * 16))
        self.n += 1


# 默认出图的指标。宁缺毋滥：逐 offset 的点相噪/杂散有十几个，趋势看
# 「相噪 vs offset」那一张就够；CT 全码扫只关心频段怎么搬和相噪跟不跟得上。
CHART_ITEMS = {"vtune": ("Freq_MHz", "Power_dBm", "IPN_SSB", "Current_mA"),
               "ct": ("Freq_MHz", "IPN_SSB")}


def _extreme_labels(sers, xrow, first):
    """挑出全图的最低点和最高点，返回 {系列下标: [该系列里要标的点下标]}。"""
    lo = hi = None
    for i, d in enumerate(sers):
        for x, v in d.items():
            if x not in xrow:
                continue
            idx = xrow[x] - first
            if lo is None or v < lo[0]:
                lo = (v, i, idx)
            if hi is None or v > hi[0]:
                hi = (v, i, idx)
    out = {}
    for b in (lo, hi):
        if b:
            out.setdefault(b[1], []).append(b[2])
    return out


def write_charts(wb, panels, st, all_charts=False):
    """panels: [(sheet, blocks, groups, xlabel)]，每块一张图。"""
    ws = wb.create_sheet("图表")
    grid = ChartGrid(ws)
    for ws_src, blocks, groups, xlabel, kind, lay in panels:
        for it, head, first, last, n_col, tgt in blocks:
            if not all_charts and it.src not in CHART_ITEMS.get(kind, ()):
                continue
            sers = [group_series(g, it) for g in groups]
            ys = [v for d in sers for v in d.values()] + ([tgt] if tgt is not None else [])
            xsv = [x for d in sers for x in d]
            ch = _scatter("%s vs %s" % (it.label, xlabel), xlabel, it.unit,
                          xlim=_nice(min(xsv), max(xsv)) if xsv else None,
                          ylim=_nice(min(ys), max(ys)) if ys else None)
            # 关键点 = 每条线自己的首末点，标上数值，不用对着坐标轴猜
            xrow = lay["xrow"].get(it.col, {})
            labels = _extreme_labels(sers, xrow, first)
            _add_series(ch, ws_src, n_col, head, first, last,
                        [len(d) for d in sers], labels,
                        target_idx=len(groups) if tgt is not None else None)
            grid.add(ch)
    return ws, grid


def write_slope_chart(grid, ws_slope, groups, anchor, title, xlabel, unit,
                      counts=None, rangevals=None, sers=None, xrow=None):
    head, first, last = anchor[0], anchor[1], anchor[2]
    if last < first:
        return
    # 坐标范围得用 Python 侧算好的值：公式版里格子里是字符串，读不出数
    xs, ys = (rangevals or ([], []))
    ch = _scatter(title, xlabel, unit,
                  xlim=_nice(min(xs), max(xs)) if xs else None,
                  ylim=_nice(min(ys), max(ys)) if ys else None)
    _add_series(ch, ws_slope, len(groups), head, first, last, counts,
                _extreme_labels(sers or [], xrow or {}, first))
    grid.add(ch)


def write_pn_chart(wb, grid, groups, pn_items, st, src_title, op_x=None):
    """相噪 vs offset：数据另开一小块（行=offset，列=各组的代表点），再挂散点图。

    取的是工作点那一点，跟结论页「@ Vtune 0.4V」用的是同一个点——
    相噪随 Vtune 变，表和图要是各取各的点，两处数字对不上就没人信了。
    """
    from openpyxl.chart import Reference, Series
    from openpyxl.chart.marker import Marker
    if not pn_items or not groups:
        return
    ws = wb.create_sheet("相噪曲线")
    picks = []
    for g in groups:
        xs = sorted({g.x_of(r) for r in g.rows if g.x_of(r) is not None})
        if not xs:
            continue
        xm = min(xs, key=lambda z: abs(z - op_x)) if op_x is not None             else xs[len(xs) // 2]
        row = None
        for r in g.rows:
            if g.x_of(r) is not None and abs(g.x_of(r) - xm) < 1e-12:
                row = r
        if row is not None:
            picks.append((g, xm, row))
    if not picks:
        return

    put(ws, 1, 1, "相噪 vs offset（每组取横轴中点处的测量点）",
        st, st["f_group"], bold=True, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(picks))
    put(ws, 2, 1, "offset_MHz", st, st["f_head"], bold=True)
    for j, (g, xm, _r) in enumerate(picks):
        put(ws, 2, 2 + j, "%s @%s%s" % (g.title.split()[0], fmt_num(xm), g.x_unit),
            st, st["f_head"], bold=True, size=9)
    for i, it in enumerate(pn_items):
        m = re.search(r"@([\d.]+)(k|M)Hz", it.label)
        off = float(m.group(1)) / (1000.0 if m and m.group(2) == "k" else 1.0) if m else None
        put(ws, 3 + i, 1, off, st, st["f_res"])
        for j, (_g, _xm, row) in enumerate(picks):
            v = row.vals.get(it.col)
            put(ws, 3 + i, 2 + j,
                emit(vref(cref(src_title, it.col, row.xl)), fmt_num(v))
                if v is not None else None, st, st["f_res"])
    ws.column_dimensions["A"].width = 12

    # 标注同样只给全图最低/最高点，各线各标会在左端叠成一坨
    pn_sers, pn_row = [], dict((i, 3 + i) for i in range(len(pn_items)))
    for _g, _xm, row in picks:
        pn_sers.append(dict((i, row.vals[it.col]) for i, it in enumerate(pn_items)
                            if row.vals.get(it.col) is not None))
    pn_labels = _extreme_labels(pn_sers, pn_row, 3)

    ch = _scatter("相噪 vs offset", "offset (MHz)", "dBc/Hz", logx=True)
    xref = Reference(ws, min_col=1, min_row=3, max_row=2 + len(pn_items))
    for j in range(len(picks)):
        yref = Reference(ws, min_col=2 + j, min_row=2, max_row=2 + len(pn_items))
        s = Series(yref, xref, title_from_data=True)
        s.smooth = False
        _style_series(s, SERIES_COLORS[j % len(SERIES_COLORS)])
        _label_points(s, pn_labels.get(j) or [])
        ch.series.append(s)
    grid.add(ch)


# ---------------------------------------------------------------- 锁定点页

def write_locked(wb, locked, items, st, temp_name):
    """扫描序列之外、但确实量到东西的行，单独列出来。

    这类簿里每换一次温度都有一小段 闭环 → 锁定 → 开环 的切换动作，其中「锁定」
    那行是带完整测量的：它就是闭环最后落在压控曲线的哪个位置，跟开环曲线对照着看
    才有意义。还有自检 / 换模式那几行也可能带部分结果——一并列在这里，
    免得「排除了 N 行」看着像凭空丢了数据。
    """
    ws = wb.create_sheet("闭环锁定点")
    put(ws, 1, 1, "闭环 / 锁定测量点", st, st["f_group"], bold=True, align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5 + len(items))
    heads = ["原表行号", "类型", "Mode", temp_name or "温度", "Vtune"] + \
            ["%s\n[%s]" % (it.label, it.unit) for it in items]
    for j, h in enumerate(heads):
        put(ws, 2, 1 + j, h, st, st["f_head"], bold=True, size=9)
    for i, (r, kind) in enumerate(locked):
        put(ws, 3 + i, 1, r.xl, st, st["f_res"])
        put(ws, 3 + i, 2, kind, st, st["f_res"])
        put(ws, 3 + i, 3, r.mode, st, st["f_res"], align="left")
        put(ws, 3 + i, 4, fmt_num(r.temp), st, st["f_res"])
        put(ws, 3 + i, 5, fmt_num(r.vt), st, st["f_res"])
        for j, it in enumerate(items):
            put(ws, 3 + i, 6 + j, fmt_num(r.vals.get(it.col)), st, st["f_res"])
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "C3"
    return ws


# ---------------------------------------------------------------- 横轴列识别

VTUNE_CANDIDATES = [r"^dc_start_v$", r"vtune.*set", r"^vtune_v$", r"^vtune$"]


def pick_vtune_col(cols, data):
    """挑 Vtune 横轴列：优先设定值列（DC_Start_V 之类），没有才用读回的 Vtune_V。"""
    for pat in VTUNE_CANDIDATES:
        for name, ci in cols.match_all(pat):
            vals = {num(r[ci]) for r in data if ci < len(r) and num(r[ci]) is not None}
            if len(vals) >= 3:
                return name, ci, "有 %d 个不同取值" % len(vals)
    return None, None, ""


def pick_ct_col(cols, data):
    """挑 CT 码列：REG Value<n> 里那个"只在一部分行有值、且值在变"的。

    ★ 优先"只覆盖一部分行"的列：每行都写的多半是常驻配置（整份簿都要写的那几组
    寄存器），扫描列一般只在扫描那几行才填。选错了会把分组切碎——所以 --dry-run
    一定要看一眼这里选中的是哪列，不对就 --ct-col 指定。
    """
    n_rows = sum(1 for r in data if any(not is_blank(v) for v in r))
    best = None
    for name, ci in cols.match_all(r"^reg[ _]?value\d+$"):
        good = [num(r[ci]) for r in data if ci < len(r) and num(r[ci]) is not None]
        distinct = set(good)
        if len(distinct) < 3:
            continue
        partial = 1 if len(good) < n_rows else 0
        cand = (partial, len(distinct), -len(good), name, ci)
        if best is None or cand > best:
            best = cand
    if best is None:
        return None, None, ""
    return best[3], best[4], "有 %d 个不同取值、%d 行非空%s" % (
        best[1], -best[2], "" if best[0] else "（整表都有值，确认一下是不是常驻配置列）")


# ---------------------------------------------------------------- main

class VcoError(Exception):
    """这份簿读不下去。调用方决定是退出还是跳过。"""


class VcoSweep(object):
    """一份读完的开环扫描簿。字段名跟原来 main() 里的局部变量一一对应。"""

    def __init__(self, **kw):
        self.__dict__.update(kw)


def load_vco(path, sheet=None, header_row=1, mode_col="Mode",
             lock_pattern=r"lock$|close.?loop", sweep_mode_opt=None,
             temp_col=None, vtune_col=None, ct_col_opt=None,
             keep_test_item=None, ref_temp=25.0, fvco_opt=None, x_round=6,
             show_addr=False, keep_original=True):
    """把一份开环压控扫描簿读成 VcoSweep（行已过滤、组已切好、指标已识别）。

    从 main() 原样抽出来的——跨芯片汇总要用同一套识别与分组口径，
    两份实现必然漂移，漂出来的是"两份报表同一个指标报不同的数"。
    keep_original=False 时不再读第二份工作簿（跨芯片汇总不需要"原表原样保留"）。
    """
    import openpyxl
    # 读两份：一份取缓存值用来算，一份原封不动用来存。
    # 只用 data_only=True 那份去存的话，原表里若有公式会被替换成计算结果——
    # 「第 1 页保留原始 excel」就不成立了。
    wb_val = openpyxl.load_workbook(path, data_only=True)
    ws_val = wb_val[sheet] if sheet else wb_val[wb_val.sheetnames[0]]
    wb = openpyxl.load_workbook(path, data_only=False) if keep_original else None
    ws = wb[ws_val.title] if wb is not None else ws_val
    all_rows = [list(r) for r in ws_val.iter_rows(values_only=True)]
    if len(all_rows) < header_row + 1:
        raise VcoError("表里没有数据行")
    header = all_rows[header_row - 1]
    data = all_rows[header_row:]

    cols = Columns(header)
    warnings = []
    if cols.duplicates:
        from openpyxl.utils import get_column_letter as gl
        for k, v in cols.duplicates.items():
            warnings.append("重复列名 %s：出现在 %s，按名字只取到第一个（%s）"
                            % (k, ", ".join(gl(i + 1) for i in v), gl(v[0] + 1)))

    if temp_col:
        tname, tcol = temp_col, cols.idx(temp_col)
    else:
        tname, tcol = cols.find(r"temperature", r"^temp")
    if tcol is None:
        warnings.append("找不到温度列，全部行按同一个温度处理（--temp-col 可指定）")
    mode_i = cols.idx(mode_col)
    if mode_i is None:
        warnings.append("没有 %s 列，无法分辨闭环/开环行" % mode_col)
    ti_col = cols.idx("Test Item")

    # 横轴列
    if vtune_col:
        vt_name, vt_col, vt_why = vtune_col, cols.idx(vtune_col), "命令行指定"
        if vt_col is None:
            raise VcoError("找不到 --vtune-col 指定的列: %s" % vtune_col)
    else:
        vt_name, vt_col, vt_why = pick_vtune_col(cols, data)
    if vt_col is None:
        raise VcoError("没找到 Vtune 横轴列，用 --vtune-col 指定")

    ct_name = ct_col = None
    ct_why = ""
    if txt(ct_col_opt).lower() != "none":
        if ct_col_opt:
            ct_name, ct_col, ct_why = ct_col_opt, cols.idx(ct_col_opt), "命令行指定"
            if ct_col is None:
                raise VcoError("找不到 --ct-col 指定的列: %s" % ct_col_opt)
        else:
            ct_name, ct_col, ct_why = pick_ct_col(cols, data)
    ct_addr = None
    if ct_name:
        m = re.search(r"(\d+)$", ct_name)
        if m:
            ai = cols.idx("REG ADDR%s" % m.group(1))
            if ai is not None:
                for r in data:
                    if ai < len(r) and not is_blank(r[ai]):
                        ct_addr = txt(r[ai])
                        break

    lock_re = re.compile(lock_pattern, re.I)

    keep_ti = keep_test_item
    if ti_col is not None and keep_ti is None:
        cnt = {}
        for r in data:
            v = txt(r[ti_col]) if ti_col < len(r) else ""
            if v:
                cnt[v] = cnt.get(v, 0) + 1
        if cnt:
            keep_ti = max(cnt, key=lambda k: cnt[k])

    sweep_mode = sweep_mode_opt
    if mode_i is not None and sweep_mode is None:
        cnt = {}
        for r in data:
            v = txt(r[mode_i]) if mode_i < len(r) else ""
            if v and not lock_re.search(v):
                cnt[v] = cnt.get(v, 0) + 1
        if cnt:
            sweep_mode = max(cnt, key=lambda k: cnt[k])

    excluded, rows, locked, others = [], [], [], []
    cut = []                    # (行号, 原因, 原始行) —— 稍后补"带几个结果值"
    for n, raw in enumerate(data):
        xl = header_row + 1 + n
        if all(is_blank(v) for v in raw):
            continue
        mode = txt(raw[mode_i]) if mode_i is not None and mode_i < len(raw) else ""
        temp = num(raw[tcol]) if tcol is not None and tcol < len(raw) else None
        vt = qx(num(raw[vt_col]) if vt_col < len(raw) else None, x_round)
        ct = qx(num(raw[ct_col]) if ct_col is not None and ct_col < len(raw) else None,
                x_round)
        r = Row(xl, temp, mode, vt, ct, raw)
        if mode and lock_re.search(mode):
            locked.append(r)
            cut.append((xl, "%s = %r，闭环/锁定行" % (mode_col, mode), raw))
            continue
        if ti_col is not None and keep_ti is not None:
            v = txt(raw[ti_col]) if ti_col < len(raw) else ""
            if v != keep_ti:
                cut.append((xl, "Test Item = %r，不是主测试项 %r" % (v, keep_ti), raw))
                continue
        if sweep_mode is not None and mode != sweep_mode:
            others.append(r)
            cut.append((xl, "%s = %r，不是扫描模式 %r" % (mode_col, mode, sweep_mode), raw))
            continue
        rows.append(r)

    skip = {vt_col} | ({ct_col} if ct_col is not None else set())
    items, dropped = build_items(cols, [r.raw for r in rows + locked + others],
                                 skip_cols=skip)
    if not items:
        raise VcoError("没识别出任何有数据的结果列")
    for r in rows + locked + others:
        for it in items:
            r.vals[it.col] = num(r.raw[it.col]) if it.col < len(r.raw) else None

    # ★ 被过滤掉的行到底有没有带测量结果，必须说出来（同 sweep_lib.load_sweep）。
    #   "排除了 N 行"这句话本身不足以判断有没有丢数据。
    for xl0, why, raw0 in cut:
        k = sum(1 for it in items
                if it.col < len(raw0) and num(raw0[it.col]) is not None)
        excluded.append((xl0, why + ("（这行带 %d 个结果值）" % k if k else
                                     "（这行没有任何结果值）")))

    # 扫描序列之外但确实量到东西的行：单列一页，别让「排除了 N 行」看着像丢了数据
    extra = [(r, "锁定") for r in locked if any(v is not None for v in r.vals.values())]
    extra += [(r, "其他模式") for r in others if any(v is not None for v in r.vals.values())]
    extra.sort(key=lambda x: x[0].xl)

    keep = []
    for r in rows:
        if all(v is None for v in r.vals.values()):
            excluded.append((r.xl, "所有结果列都是空的（配置/开关行，不是测量点）"))
        else:
            keep.append(r)
    rows = keep
    if not rows:
        raise VcoError("过滤完没有测量点了，检查 --sweep-mode / --keep-test-item")

    groups = segment(rows)
    # 横轴值大面积重复 = 组里还藏着另一个在动的变量（最常见：CT 列没识别出来，
    # 于是 CT 扫那几行被并进 Vtune 扫，去重后只剩最后一个，把曲线上那个点悄悄换掉）
    for g in groups:
        xs = [g.x_of(r) for r in g.rows if g.x_of(r) is not None]
        dup = len(xs) - len(set(xs))
        # 少量重复是正常的（先粗扫一遍再回头细扫，几个点会重复量）；
        # 大面积重复才是「组里还藏着另一个在动的变量」。
        if dup >= 2 and dup > 0.2 * len(xs):
            warnings.append("组「%s」有 %d 个点但横轴只有 %d 个不同取值——"
                            "是不是还有一个没识别出来的扫描变量？(--ct-col 指定)"
                            % (g.title, len(xs), len(set(xs))))

    freq_item = next((it for it in items if it.src == "Freq_MHz"), None)
    if freq_item is None:
        freq_item = next((it for it in items if it.cat == "Frequency"), None)
        if freq_item is None:
            warnings.append("没有频率列，Kvco / 覆盖范围这些派生指标出不来")

    by_kind = []
    for kind in ("vtune", "ct", "point"):
        gs = [g for g in groups if g.kind == kind]
        if gs:
            by_kind.append((kind, gs))
    ref_by_kind = {}
    for kind, gs in by_kind:
        withtemp = [g for g in gs if g.temp is not None]
        if withtemp:
            ref_by_kind[kind] = min(withtemp, key=lambda g: abs(g.temp - ref_temp))

    # 目标 VCO 频率：算覆盖余量要用。表里就有这一列，不写死在脚本里；
    # 连结论页那一格也是指回原表的引用，不是抄过来的数。
    fvco, fvco_ref = fvco_opt, None
    _n, fc = cols.find(r"^fvco")
    if fc is not None:
        for r in rows:
            v = num(r.raw[fc]) if fc < len(r.raw) else None
            if v:
                if fvco_opt is None:
                    fvco = v
                    fvco_ref = cref(ws.title, fc, r.xl)
                break

    excluded.sort(key=lambda x: x[0])
    meta = {
        "excluded": excluded, "warnings": warnings, "freq_item": freq_item,
        "ref_by_kind": ref_by_kind,
        "why_groups": ("(温度, 模式) 一变就切；组内再看这一步动的是 Vtune 还是 CT，"
                       "两个一起动就是换扫法了，也切。横轴不是一回事的点混在一起统计没有意义。"),
    }

    return VcoSweep(**{k: v for k, v in locals().items()
                      if k not in ("openpyxl",)})


# ---------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(description="开环压控扫描簿 → 带汇总页的 Excel")
    ap.add_argument("path", help="扫描簿 .xlsx")
    ap.add_argument("-o", "--out", default=None, help="输出路径（默认 <原名>_summary.xlsx）")
    ap.add_argument("--sheet", default=None, help="数据所在 sheet（默认第一个）")
    ap.add_argument("--header-row", type=int, default=1, help="表头行号（1 基，默认 1）")
    ap.add_argument("--mode-col", default="Mode", help="模式列（默认 Mode）")
    ap.add_argument("--lock-pattern", default=r"lock$|close.?loop",
                    help="模式列匹配这个正则的行 = 闭环/锁定行，不进开环统计"
                         "（默认 lock$|close.?loop）")
    ap.add_argument("--sweep-mode", default=None,
                    help="只统计该模式的行（默认取非锁定行里出现最多的那个值）。"
                         "挡的是夹在扫描序列里的旁路/自检行——它们也带部分结果值，"
                         "不挡就会被算进相邻那一组，把曲线和极值都带偏")
    ap.add_argument("--temp-col", default=None, help="温度列（默认自动找含 Temperature 的列）")
    ap.add_argument("--vtune-col", default=None, help="Vtune 横轴列（默认自动挑）")
    ap.add_argument("--ct-col", default=None,
                    help="CT 码横轴列（默认自动挑；写 none 关掉 CT 扫识别）")
    ap.add_argument("--keep-test-item", default=None,
                    help="只保留 Test Item 等于该值的行（默认取出现最多的那个值）")
    ap.add_argument("--ref-temp", type=float, default=25.0,
                    help="温漂参考温度（默认 25）")
    ap.add_argument("--fvco", type=float, default=None,
                    help="目标 VCO 频率 MHz（默认从 fVCO_MHz 列读）。用来算覆盖余量")
    ap.add_argument("--op-vtune", type=float, default=None,
                    help="工作点调谐电压 V（默认取 CT 扫用的那个值）。相噪/功率这些"
                         "按这个点报，不跨整个扫描取极值——不同 Vtune 是不同振荡频率")
    ap.add_argument("--x-round", type=int, default=6,
                    help="横轴取值四舍五入到几位小数（默认 6）。扫描点常是累加出来的，"
                         "表里会是 0.39999999999999997 这种，不量化就跟别的段对不上点")
    ap.add_argument("--show-addr", action="store_true",
                    help="打印 CT 扫用的寄存器地址（默认不打印，地址是 IP）")
    ap.add_argument("--all-charts", action="store_true",
                    help="每个指标都出图（默认跳过逐 offset 的点相噪/杂散那十几张）")
    ap.add_argument("--tables", action="store_true",
                    help="多出「汇总」页（各指标逐点 Min/Max/Δ 的大表）。默认不出——"
                         "同样的信息图上看得见，表只是多一页要翻的东西")
    ap.add_argument("--show-data-sheets", action="store_true",
                    help="不隐藏明细/斜率/温漂那几页（默认隐藏，它们是图的数据源）")
    ap.add_argument("--static", action="store_true",
                    help="格子里写算好的死值，不写引用公式。默认是公式——"
                         "每个数都能一路追回原始表那一格，原始数据改了整本跟着变")
    ap.add_argument("--notes", action="store_true",
                    help="多写一页「数据处理记录」（排除了哪些行/探查告警）。"
                         "默认不写——那是对账用的，不是报告内容；同样的内容每次运行都打在控制台")
    ap.add_argument("--title", default=None,
                    help="结论页的标题（默认「VCO 开环特性 · 结论」）")
    ap.add_argument("--dry-run", action="store_true", help="只打印识别结果，不写文件")
    args = ap.parse_args()

    global LIVE_REFS
    LIVE_REFS = not args.static

    if not os.path.isfile(args.path):
        sys.exit("找不到文件: %s" % args.path)
    try:
        import openpyxl
    except ImportError:
        sys.exit("缺少 openpyxl，请先: pip install openpyxl")




    # 读取/过滤/分组/识别指标全在 load_vco 里（跨芯片汇总用同一份）
    try:
        sw = load_vco(args.path, sheet=args.sheet, header_row=args.header_row,
                      mode_col=args.mode_col, lock_pattern=args.lock_pattern,
                      sweep_mode_opt=args.sweep_mode, temp_col=args.temp_col,
                      vtune_col=args.vtune_col, ct_col_opt=args.ct_col,
                      keep_test_item=args.keep_test_item, ref_temp=args.ref_temp,
                      fvco_opt=args.fvco, x_round=args.x_round,
                      show_addr=args.show_addr)
    except VcoError as e:
        sys.exit(str(e))
    by_kind, ct_addr, ct_name, ct_why = sw.by_kind, sw.ct_addr, sw.ct_name, sw.ct_why
    data, dropped, excluded, extra = sw.data, sw.dropped, sw.excluded, sw.extra
    freq_item, fvco, fvco_ref = sw.freq_item, sw.fvco, sw.fvco_ref
    groups, header, items, keep_ti = sw.groups, sw.header, sw.items, sw.keep_ti
    meta, others, ref_by_kind = sw.meta, sw.others, sw.ref_by_kind
    sweep_mode, tcol, tname = sw.sweep_mode, sw.tcol, sw.tname
    vt_col, vt_name, vt_why = sw.vt_col, sw.vt_name, sw.vt_why
    warnings, wb, ws = sw.warnings, sw.wb, sw.ws

    # ---- 打印识别结果 ----
    # ws.max_row/max_column 是「声明尺寸」——模板预设过格式的空区域也算进去，
    # 常见到 10000 行 × 205 列这种数字。真正有内容的行才是要报的。
    n_filled = sum(1 for r in data if any(not is_blank(v) for v in r))
    print("源文件 : %s   sheet=%s" % (os.path.basename(args.path), ws.title))
    print("规模   : 有内容的数据行 %d（表头第 %d 行，表头列 %d 个；"
          "工作表声明尺寸 %d 行 × %d 列，其余是模板预留的空区域）"
          % (n_filled, args.header_row, sum(1 for h in header if txt(h)),
             ws.max_row, ws.max_column))
    print("温度列 : %s" % (tname or "(没有)"))
    print("目标fVCO: %s MHz%s" % (fmt_num(fvco) if fvco else "(没读到，覆盖余量出不来)",
                                 "" if args.fvco is None else "  (命令行指定)"))
    print("Vtune  : %s   %s" % (vt_name, vt_why))
    if ct_name:
        addr = ("  地址=%s" % ct_addr) if (ct_addr and args.show_addr) else \
               ("  (地址见原表，--show-addr 才打印)" if ct_addr else "")
        print("CT 码  : %s   %s%s" % (ct_name, ct_why, addr))
    else:
        print("CT 码  : 没识别到（--ct-col 可指定）")
    if keep_ti is not None:
        print("主测试项: %r（其余行排除）" % keep_ti)
    if sweep_mode is not None:
        print("扫描模式: %r（闭环/锁定行另列；其余行排除）" % sweep_mode)
    print("识别指标 %d 个:" % len(items))
    for it in items:
        print("    [%-12s] %-18s %-7s <- 列 %s" % (it.cat, it.label, it.unit, it.src))
    if dropped:
        print("  跳过的空列 %d 个: %s" % (len(dropped), ", ".join(k for k, _ in dropped)))
    print("分组 %d 组:" % len(groups))
    for g in groups:
        print("    %-16s %-22s 行 %d~%d" % (g.title, g.stage, g.rows[0].xl, g.rows[-1].xl))
    if extra:
        print("扫描序列之外带测量值的 %d 行（列在「闭环锁定点」页）: %s"
              % (len(extra), ", ".join("%d/%s" % (r.xl, k) for r, k in extra[:12])))
    if excluded:
        print("排除 %d 行:" % len(excluded))
        for xl, why in excluded:
            print("    行%d: %s" % (xl, why))
    for w in warnings:
        print("  ⚠ %s" % w)
    if args.dry_run:
        print("\n--dry-run：没有写文件。")
        return

    # ---- 写出 ----
    # 顺序有讲究：明细页先写，上面几层才知道该引用哪一格。页序最后再排。
    st = _styles()
    LAY = {"src": ws.title, "tcol": tcol, "vtcol": vt_col}
    panels, slope_jobs = [], []
    for kind, gs in by_kind:
        if kind == "point":
            continue
        nm = "Vtune明细" if kind == "vtune" else "CT明细"
        ws_d, blocks, lay = write_detail(wb, nm, gs, items_with_data(gs, items), st,
                                         ws.title, target=fvco, target_item=freq_item,
                                         target_ref=vref(fvco_ref) if fvco_ref else None)
        LAY.setdefault(kind, {})["detail"] = lay
        panels.append((ws_d, blocks, gs, gs[0].x_label, kind, lay))
        if freq_item is not None and any(len(group_series(g, freq_item)) >= 2 for g in gs):
            sn = "Kvco明细" if kind == "vtune" else "CT斜率明细"
            unit = "MHz/V" if kind == "vtune" else "MHz/code"
            ws_s, anchor, slay = write_slope(wb, sn, gs, freq_item, st, unit,
                                             gs[0].x_label, lay)
            LAY[kind]["slope"] = slay
            slope_jobs.append((ws_s, gs, anchor,
                               "Kvco vs Vtune" if kind == "vtune" else "ΔF/ΔCT vs CT",
                               gs[0].x_label, unit,
                               [max(0, len(group_series(g, freq_item)) - 1) for g in gs],
                               (slay.get("xvals") or [], slay.get("yvals") or []),
                               slay.get("sers"), slay.get("xrow")))
        refg = ref_by_kind.get(kind)
        if refg is not None and len(gs) > 1 and freq_item is not None:
            dn = "温漂明细" if kind == "vtune" else "温漂明细-CT"
            _wsd, dlay = write_drift(wb, dn, gs, refg, freq_item, st, lay, gs[0].x_label)
            if dlay:
                LAY[kind]["drift"] = dlay

    concl_rows, concl_temps = build_conclusion(by_kind, items, freq_item,
                                               args.ref_temp, fvco, fvco_ref, LAY,
                                               args.op_vtune)
    if concl_rows:
        write_conclusion(wb, concl_rows, concl_temps, st,
                         args.title or "VCO 开环特性 · 结论")
    if args.tables:
        write_summary(wb, by_kind, items, st, LAY)

    _ws_chart, grid = write_charts(wb, panels, st, all_charts=args.all_charts)
    for ws_s, gs, anchor, title, xlabel, unit, cnts, rv, sr, xr in slope_jobs:
        write_slope_chart(grid, ws_s, gs, anchor, title, xlabel, unit, cnts, rv, sr, xr)
    for kind, gs in by_kind:
        d = (LAY.get(kind) or {}).get("drift")
        if not d or d["last"] < d["first"]:
            continue
        others = [g for g in gs if g is not d["ref"]]
        sref = group_series(d["ref"], freq_item)
        dxs, dys = [], []
        for g in others:
            for x, v in group_series(g, freq_item).items():
                if x in sref:
                    dxs.append(x)
                    dys.append(v - sref[x])
        ch = _scatter("ΔF vs %s（相对 %s）" % (gs[0].x_label, d["ref"].title),
                      gs[0].x_label, "MHz",
                      xlim=_nice(min(dxs), max(dxs)) if dxs else None,
                      ylim=_nice(min(dys), max(dys)) if dys else None)
        dsers, drow, dr = [], {}, d["first"]
        xs_all = sorted({x for g in others for x in group_series(g, freq_item)
                         if x in sref})
        for k, x in enumerate(xs_all):
            drow[x] = dr + k
        for g in others:
            sg = group_series(g, freq_item)
            dsers.append({x: sg[x] - sref[x] for x in sg if x in sref})
        _add_series(ch, wb[d["sheet"]], len(d["col_of"]), 2, d["first"], d["last"],
                    [len(group_series(g, freq_item)) for g in others],
                    _extreme_labels(dsers, drow, dr))
        grid.add(ch)
    pn_items = [it for it in items if it.label.startswith("SpotPN@")]
    vt_groups = next((gs for k, gs in by_kind if k == "vtune"), [])
    op_x = args.op_vtune
    if op_x is None:
        cg0 = next((g for k, gs in by_kind if k == "ct" for g in gs), None)
        if cg0 is not None:
            xv = [r.vt for r in cg0.rows if r.vt is not None]
            op_x = xv[0] if xv else None
    write_pn_chart(wb, grid, vt_groups or [g for _k, gs in by_kind for g in gs],
                   pn_items, st, ws.title, op_x)
    if extra:
        write_locked(wb, extra, items, st, tname)
    if args.notes:
        write_diag(wb, meta, st)

    # 页序：能看的在前，数据源在后
    order = [ws.title, "结论", "图表", "闭环锁定点", "汇总", "汇总-CT扫",
             "Vtune明细", "Kvco明细", "温漂明细", "CT明细", "CT斜率明细",
             "温漂明细-CT", "相噪曲线", "数据处理记录"]
    have = [n for n in order if n in wb.sheetnames]
    wb._sheets = [wb[n] for n in have] + [s for s in wb._sheets if s.title not in have]
    # ★ 明细/斜率/温漂/相噪那几页是图的数据源，不是给人翻的：默认藏起来，
    #   打开簿子只看见 原表 / 结论 / 图表 / 闭环锁定点。要翻数在 Excel 里
    #   右键取消隐藏就行，一个数都没少。
    if not args.show_data_sheets:
        for n in ("Vtune明细", "Kvco明细", "温漂明细", "CT明细", "CT斜率明细",
                  "温漂明细-CT", "相噪曲线"):
            if n in wb.sheetnames:
                wb[n].sheet_state = "hidden"
    # 让 Excel 一打开就重算（图才有数据）。⚠ 这一条**不足以**解决"发给别人是
    # 空格子"——那是 openpyxl 写了空的 <v></v> 缓存值，得靠下面 VCACHE.inject()
    wb.calculation.fullCalcOnLoad = True

    out = args.out or os.path.splitext(args.path)[0] + "_summary.xlsx"
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    n_fill, n_strip = VCACHE.inject(out)
    print("\n已写出: %s" % os.path.abspath(out))
    print("  页：%s" % " / ".join(wb.sheetnames))
    print("  公式格补了 %d 个缓存值、清掉 %d 个空缓存 —— 发给别人打开就能看见数，"
          "不用敲回车重算。" % (n_fill, n_strip))
    # 下面这些是给做表的人看的，所以只打在控制台，不写进簿子
    print("\n怎么用（不写进簿子，只在这里说）：")
    print("  · 结论页 Spec 的 Min / Max 两列自己填；Limit 列提示该填哪边"
          "（≥ 填 Min、≤ 填 Max，另一边留空）。填完判定列自动出 PASS/FAIL 并标红。")
    print("  · Fmin/Fmax 是把 CT 扫和 Vtune 扫两段拼出来的（算式写在结论页备注列里）；"
          "两个角点没有直接测过，前提是 Kvco 与 CT 码基本无关。要精确值得做二维扫描。")
    print("  · 同一组同一横轴值测了两次的取后者（例如 CT 先粗扫再回头细扫）。")
    if not args.notes:
        print("  · 排除的行与探查告警见上面的控制台输出；要在簿子里留档加 --notes。")


if __name__ == "__main__":
    main()
