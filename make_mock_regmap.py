#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
make_mock_regmap.py — 用抓回的 目标寄存器 sheet 行 + 控制信号 list，
                      (1) 解析寄存器模型、把每个控制信号解析到 寄存器/地址/bit/默认值，
                      (2) 本地生成一个结构一模一样的寄存器 .xlsx（nManager 布局）供开发。

输入(都在 private/，含 IP，不入库)：
    --rows      pll_rows.json         explore_excel --rowdump 抓的寄存器 sheet 行区间
    --signals   control_signals.json  控制信号 list(取 reg_net)
    --aliases   aliases.json          reg_net -> 实际字段名 的变体映射(可选)
    --schema    <REG>.schema.json     取 sheet 头几行元信息(Base Address 等)重建结构(可选)

输出：
    --out-xlsx  <REG>_mock.xlsx       本地复刻的寄存器表(结构一致)
    --out-map   signal_reg_map.json   控制信号 -> 寄存器/地址/bit/默认/off值 解析结果

块基址 / sheet 名等项目专属值：优先读 gitignore 的 private/tool_config/make_mock_regmap.json，
或用 --base 显式传。本脚本本身不含任何真实信号名/地址，只读 private/ 输入，故可入库。
"""
import argparse
import json
import os
import re
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


def _local_cfg():
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                     "private", "tool_config", "make_mock_regmap.json")
    if os.path.exists(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def load_project(proj_dir):
    """读工程包 project.json（schema/2）。None -> 走传统 --rows/--signals/--aliases + tool_config。"""
    if not proj_dir:
        return None
    p = os.path.join(proj_dir, "project.json")
    if not os.path.exists(p):
        sys.exit("工程包缺 project.json: %s" % p)
    with open(p, encoding="utf-8") as f:
        return json.load(f)


# nManager 列布局（0 基）——默认；项目专属列映射从 project.json regbook.column_map 读。
DEFAULT_COLMAP = {"reg_name": 0, "regtype": 1, "offset": 3, "width": 4, "reset": 7,
                  "field_name": 9, "bit": 10, "attr": 11, "default": 12, "comment": 14}


def parse_registers(rows, base, cm=None):
    cm = cm or DEFAULT_COLMAP
    pad = max(cm.values()) + 1
    regs, cur = [], None
    for r in rows:
        r = list(r) + [None] * max(0, pad - len(r))
        name, off = r[cm["reg_name"]], r[cm["offset"]]
        if name and off:                      # 寄存器头行
            try:
                addr = base + int(str(off), 16)
            except ValueError:
                addr = None
            cur = {"reg_name": name, "regtype": r[cm["regtype"]], "offset": str(off),
                   "addr": (hex(addr) if addr is not None else None),
                   "reset": r[cm["reset"]], "width": r[cm["width"]], "fields": []}
            regs.append(cur)
        if r[cm["field_name"]]:               # 字段行(头行也带首字段)
            f = {"name": r[cm["field_name"]], "bit": r[cm["bit"]], "attr": r[cm["attr"]],
                 "default": r[cm["default"]], "comment": r[cm["comment"]]}
            if cur is None:
                cur = {"reg_name": "(above_range)", "offset": None, "addr": None,
                       "reset": None, "width": None, "fields": []}
                regs.append(cur)
            cur["fields"].append(f)
    return regs


def collect_reg_nets(sig_obj):
    nets = []

    def walk(o):
        if isinstance(o, dict):
            for k, v in o.items():
                if k == "reg_net" and isinstance(v, str):
                    nets.append((v, o))
                else:
                    walk(v)
        elif isinstance(o, list):
            for x in o:
                walk(x)
    walk(sig_obj)
    return nets


def infer_off(field):
    """据注释里的 高有效/低有效 猜 1-bit enable 的关断值。返回 (active_high, off_value) 或 (None,None)。"""
    c = str(field.get("comment") or "")
    bit = str(field.get("bit") or "")
    if ":" in bit:
        return (None, None)                   # 多 bit, 不猜
    if "高有效" in c or "high active" in c.lower() or "1: en" in c.lower():
        return (True, 0)
    if "低有效" in c or "low active" in c.lower():
        return (False, 1)
    return (None, None)


def main():
    ap = argparse.ArgumentParser(description="解析控制信号->寄存器 + 生成复刻 Excel")
    ap.add_argument("--project", help="工程包目录（project.json schema/2）：base/sheet/column_map 读 regbook 段，"
                                       "alias/logic_derived 读 matching 段，IO 按 artifacts 从包内解析。")
    ap.add_argument("--rows", default=None)
    ap.add_argument("--signals", default=None)
    ap.add_argument("--aliases", default=None)
    ap.add_argument("--schema", default=None)
    ap.add_argument("--base", default=None, help="块基址(默认读工程包/本地 config，无则 0x0)")
    ap.add_argument("--sheet-name", default=None, help="生成表的 sheet 名(默认读工程包/本地 config，无则 REGS)")
    ap.add_argument("--out-xlsx", default=None)
    ap.add_argument("--out-map", default=None)
    args = ap.parse_args()

    proj = load_project(args.project)
    rb = proj.get("regbook", {}) if proj else {}
    art = proj.get("artifacts", {}) if proj else {}
    cfg = _local_cfg()

    def ppath(name):
        return os.path.join(args.project, name)

    base_str = args.base or (rb.get("base_address") if proj else None) or cfg.get("base", "0x0")
    base = int(base_str, 16)
    sheet_name = args.sheet_name or (rb.get("sheet_name") if proj else None) or cfg.get("sheet_name", "REGS")
    colmap = rb.get("column_map") if proj else None
    rows_path = args.rows or (ppath(art.get("pll_rows", "pll_rows.json")) if proj else None)
    signals_path = args.signals or (ppath(art.get("control_signals", "control_signals.json")) if proj else None)
    out_map = args.out_map or (ppath(art.get("signal_reg_map", "signal_reg_map.json")) if proj else None)
    if not rows_path or not signals_path:
        ap.error("需要 --rows 和 --signals（或 --project 从工程包解析）")

    rows_doc = json.load(open(rows_path, encoding="utf-8"))
    rows = rows_doc["rows"]
    regs = parse_registers(rows, base, colmap)

    # 字段索引（精确 + 小写）
    fidx, ci = {}, {}
    for reg in regs:
        for f in reg["fields"]:
            fidx.setdefault(f["name"], (reg, f))
            ci.setdefault(f["name"].lower(), (reg, f))

    alias, logic_set = {}, set()
    if proj is not None:                       # 工程包权威：alias/logic_derived 来自 matching 段
        mt = proj.get("matching", {})
        alias = mt.get("alias", {})
        logic_set = set(mt.get("logic_derived", []))
    elif args.aliases and os.path.isfile(args.aliases):
        ad = json.load(open(args.aliases, encoding="utf-8"))
        alias = ad.get("alias", {})
        logic_set = set(ad.get("logic_derived", []))

    def resolve(net):
        if net in fidx:
            return fidx[net], "exact"
        if net in alias and alias[net] in fidx:
            return fidx[alias[net]], "alias"
        if net.lower() in ci:
            return ci[net.lower()], "case"
        if net in logic_set:
            return None, "logic-derived"
        return None, "unresolved"

    sig_obj = json.load(open(signals_path, encoding="utf-8"))
    nets = collect_reg_nets(sig_obj)

    entries, counts = [], {}
    for net, meta in nets:
        res, how = resolve(net)
        counts[how] = counts.get(how, 0) + 1
        e = {"reg_net": net, "match": how,
             "category": meta.get("category"), "drives": meta.get("drives")}
        if res:
            reg, f = res
            ah, offv = infer_off(f)
            e.update({"field_name": f["name"], "reg_name": reg["reg_name"],
                      "offset": reg["offset"], "addr": reg["addr"],
                      "bit": f["bit"], "attr": f["attr"], "default": f["default"],
                      "comment": f["comment"], "active_high": ah, "off_value": offv})
        entries.append(e)

    print(f"寄存器 {len(regs)} 个; 控制信号 {len(nets)} 个")
    print("解析: " + ", ".join(f"{k}={v}" for k, v in sorted(counts.items())))
    un = [e["reg_net"] for e in entries if e["match"] == "unresolved"]
    if un:
        print("未解析: " + ", ".join(un))
    print()
    print(f"{'控制信号(reg_net)':40} {'match':7} {'addr':11} {'bit':6} {'dflt':6} {'off'}")
    for e in entries:
        if e["match"] in ("unresolved", "logic-derived"):
            print(f"{e['reg_net']:40} {e['match']}")
        else:
            print(f"{e['reg_net']:40} {e['match']:7} {str(e.get('addr')):11} "
                  f"{str(e.get('bit')):6} {str(e.get('default')):6} {e.get('off_value')}")

    if out_map:
        json.dump({"base": base_str, "n_regs": len(regs), "counts": counts,
                   "registers": regs, "signals": entries},
                  open(out_map, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
        print(f"\n解析映射已写: {out_map}")

    if args.out_xlsx:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        # 元信息头几行：优先用 schema 的 sample_rows(前6行)，否则最简重建
        meta_rows = None
        if args.schema and os.path.isfile(args.schema):
            sd = json.load(open(args.schema, encoding="utf-8"))
            meta_rows = sd.get("sheet", {}).get("sample_rows")
        if not meta_rows:
            meta_rows = [
                ["Module Name", sheet_name], ["CPU Data Width", 16],
                ["Base Address", base_str], ["Memory Name", "Memory"],
                ["Table/Register Information(Item has * mean must)"],
                ["*Reg Name", "Description", "C_Reserved", "*Offset Addr", "*Width",
                 "Table Length", "Attribute", "Default Value", "I_Reserved",
                 "*Field Name", "*Field Range", "*Field Attribute",
                 "Field Default Value", "N_Reserved", "Field Comments"],
            ]
        for row in meta_rows:
            ws.append(list(row))
        for row in rows:
            ws.append(list(row))
        os.makedirs(os.path.dirname(args.out_xlsx) or ".", exist_ok=True)
        wb.save(args.out_xlsx)
        print(f"复刻 Excel 已写: {args.out_xlsx}  ({ws.max_row} 行 x {ws.max_column} 列)")


if __name__ == "__main__":
    main()
